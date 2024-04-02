
# Standard library imports.

import collections
import ctypes
import datetime
import gzip
import hashlib
import io
import itertools
import logging
import math
import os.path
import pickle
import pprint
import random
import re
import shutil
import string
import textwrap
import time
import traceback
import uuid

# Library imports.

import graphene
import graphene.types

# So that matplotlib does not require display stuff, in particular, tkinter. See e.g. https://
# stackoverflow.com/questions/4931376/generating-matplotlib-graphs-without-a-running-x-server.
import matplotlib
matplotlib.use('Agg', warn = False)

from matplotlib import pyplot
from matplotlib.collections import LineCollection

from mpl_toolkits.mplot3d.art3d import Line3DCollection

import numpy
import openpyxl
import pandas as pd
import pathvalidate

from pretty_html_table import build_table

from pyramid.httpexceptions import HTTPOk

import scipy.optimize
import scipy.sparse.csgraph

import sklearn.decomposition
import sklearn.manifold
import sklearn.metrics
import sklearn.mixture

import sqlalchemy

from sqlalchemy import (
    and_,
    create_engine,
    func,
    literal,
    tuple_)

from sqlalchemy.orm import aliased

import sqlalchemy.types

import transaction
import xlsxwriter

# Project imports.

from lingvodoc.cache.caching import (
    initialize_cache,
    TaskStatus)

from lingvodoc.models import (
    Client,
    DBSession,
    DictionaryPerspective as dbPerspective,
    DictionaryPerspectiveToField as dbColumn,
    ENGLISH_LOCALE,
    Entity as dbEntity,
    Field as dbField,
    Language as dbLanguage,
    LexicalEntry as dbLexicalEntry,
    PublishingEntity as dbPublishingEntity,
    TranslationAtom as dbTranslationAtom)

from lingvodoc.queue.celery import celery

from lingvodoc.schema.gql_holders import (
    del_object,
    LingvodocID,
    ObjectVal,
    ResponseError,
    Upload)

from lingvodoc.scripts.save_dictionary import (
    find_group_by_tags)

import lingvodoc.utils as utils

from lingvodoc.utils import ids_to_id_query

from lingvodoc.utils.deletion import real_delete_entity

from lingvodoc.utils.search import (
    find_all_tags,
    find_lexical_entries_by_tags)

import lingvodoc.views.v2.phonology as phonology
from lingvodoc.views.v2.phonology import process_sound_markup

from lingvodoc.views.v2.utils import anonymous_userid


# Setting up logging.
log = logging.getLogger(__name__)


# Mikhail Oslon's analysis functions.

try:

    liboslon = ctypes.CDLL('liboslon.so')

    phonemic_analysis_f = liboslon.PhonemicAnalysis_GetAllOutput

    cognate_analysis_f = liboslon.CognateAnalysis_GetAllOutput
    cognate_acoustic_analysis_f = liboslon.CognateAcousticAnalysis_GetAllOutput
    cognate_distance_analysis_f = liboslon.CognateDistanceAnalysis_GetAllOutput
    cognate_reconstruction_f = liboslon.CognateReconstruct_GetAllOutput
    cognate_reconstruction_multi_f = liboslon.CognateMultiReconstruct_GetAllOutput
    cognate_suggestions_f = liboslon.GuessCognates_GetAllOutput

except:

    log.warning('liboslon.so')

    phonemic_analysis_f = None

    cognate_analysis_f = None
    cognate_acoustic_analysis_f = None
    cognate_distance_analysis_f = None
    cognate_reconstruction_f = None
    cognate_reconstruction_multi_f = None
    cognate_suggestions_f = None


class PhonemicAnalysis(graphene.Mutation):

    class Arguments:

        perspective_id = LingvodocID(required = True)
        transcription_field_id = LingvodocID(required = True)
        translation_field_id = LingvodocID(required = True)
        wrap_flag = graphene.Boolean()

        debug_flag = graphene.Boolean()
        intermediate_flag = graphene.Boolean()

    triumph = graphene.Boolean()
    entity_count = graphene.Int()
    result = graphene.String()

    intermediate_url_list = graphene.List(graphene.String)

    @staticmethod
    def mutate(self, info, **args):
        """
        mutation PhonemicAnalysis {
          phonemic_analysis(
            perspective_id: [70, 5],
            transcription_field_id: [66, 8],
            translation_field_id: [66, 10])
          {
            triumph
            entity_count
            result
          }
        }
        """

        perspective_cid, perspective_oid = args['perspective_id']

        transcription_field_cid, transcription_field_oid = args['transcription_field_id']
        translation_field_cid, translation_field_oid = args['translation_field_id']

        wrap_flag = args.get('wrap_flag', False)

        locale_id = info.context.locale_id

        __debug_flag__ = args.get('debug_flag', False)
        __intermediate_flag__ = args.get('intermediate_flag', False)

        try:

            perspective = DBSession.query(dbPerspective).filter_by(
                client_id = perspective_cid, object_id = perspective_oid).first()

            perspective_name = perspective.get_translation(locale_id)
            dictionary_name = perspective.parent.get_translation(locale_id)

            transcription_rules = (
                '' if not perspective.additional_metadata else
                    perspective.additional_metadata.get('transcription_rules', ''))

            # Showing phonemic analysis info, checking phonemic analysis library presence.

            log.debug(
                '\nphonemic_analysis {0}/{1}:'
                '\n  dictionary: {2}'
                '\n  perspective: {3}'
                '\n  transcription rules: {4}'
                '\n  transcription field: {5}/{6}'
                '\n  translation field: {7}/{8}'
                '\n  wrap_flag: {9}'
                '\n  __debug_flag__: {10}'
                '\n  __intermediate_flag__: {11}'
                '\n  locale_id: {12}'
                '\n  phonemic_analysis_f: {13}'.format(
                    perspective_cid, perspective_oid,
                    repr(dictionary_name.strip()),
                    repr(perspective_name.strip()),
                    repr(transcription_rules),
                    transcription_field_cid, transcription_field_oid,
                    translation_field_cid, translation_field_oid,
                    wrap_flag,
                    __debug_flag__,
                    __intermediate_flag__,
                    locale_id,
                    repr(phonemic_analysis_f)))

            if phonemic_analysis_f is None:

                return ResponseError(message =
                    'Analysis library is absent, please contact system administrator.')

            # Query for non-deleted, published and accepted entities of the specified perspective with the
            # specified field.

            dbTranslation = aliased(dbEntity, name = 'Translation')
            dbPublishingTranslation = aliased(dbPublishingEntity, name = 'PublishingTranslation')

            data_query = (

                DBSession.query(dbEntity).filter(
                    dbLexicalEntry.parent_client_id == perspective_cid,
                    dbLexicalEntry.parent_object_id == perspective_oid,
                    dbLexicalEntry.marked_for_deletion == False,
                    dbEntity.parent_client_id == dbLexicalEntry.client_id,
                    dbEntity.parent_object_id == dbLexicalEntry.object_id,
                    dbEntity.field_client_id == transcription_field_cid,
                    dbEntity.field_object_id == transcription_field_oid,
                    dbEntity.marked_for_deletion == False,
                    dbPublishingEntity.client_id == dbEntity.client_id,
                    dbPublishingEntity.object_id == dbEntity.object_id,
                    dbPublishingEntity.published == True,
                    dbPublishingEntity.accepted == True)

                .outerjoin(dbTranslation, and_(
                    dbTranslation.parent_client_id == dbEntity.parent_client_id,
                    dbTranslation.parent_object_id == dbEntity.parent_object_id,
                    dbTranslation.field_client_id == translation_field_cid,
                    dbTranslation.field_object_id == translation_field_oid,
                    dbTranslation.marked_for_deletion == False))

                .outerjoin(dbPublishingTranslation, and_(
                    dbPublishingTranslation.client_id == dbTranslation.client_id,
                    dbPublishingTranslation.object_id == dbTranslation.object_id,
                    dbPublishingTranslation.published == True,
                    dbPublishingTranslation.accepted == True))

                .add_columns(
                    func.array_agg(dbTranslation.content))

                .group_by(dbEntity))

            # Counting text entities we have. If we haven't got any, we return empty result.

            total_count = data_query.count()

            log.debug(
                'phonemic_analysis {}/{}: {} transcription entities'.format(
                    perspective_cid,
                    perspective_oid,
                    total_count))

            if total_count <= 0:

                return PhonemicAnalysis(
                    triumph = True,
                    entity_count = total_count,
                    result = u'')

            # Otherwise we are going to perform phonemic analysis.

            data_list = [

                (entity.content,
                    translation_list[0] if translation_list else '')

                for entity, translation_list in data_query.all()
                if len(entity.content) > 0]

            if len(data_list) <= 0:

                return PhonemicAnalysis(
                    triumph = True,
                    entity_count = total_count,
                    result = u'No transcription entities with non-empty contents.')

            # Preparing analysis input.

            input = (

                '{0} - {1}\0{2}\0'.format(
                    dictionary_name,
                    perspective_name,
                    transcription_rules) +

                ''.join(

                    '{0}\0{1}\0'.format(
                        transcription, translation)

                    for transcription, translation in data_list))

            log.debug(
                'phonemic_analysis {0}/{1}: '
                'transcription field {2}/{3}, translation field {4}/{5}'
                '\ndata_list:\n{6}'
                '\ninput ({7} rows):\n{8}'.format(
                    perspective_cid, perspective_oid,
                    transcription_field_cid, transcription_field_oid,
                    translation_field_cid, translation_field_oid,
                    pprint.pformat(data_list, width = 108),
                    len(data_list) + 1,
                    pprint.pformat([input[i : i + 256]
                        for i in range(0, len(input), 256)], width = 144)))

            # Saving to a file, if required.

            intermediate_url_list = []

            if __debug_flag__ or __intermediate_flag__:

                perspective = DBSession.query(dbPerspective).filter_by(
                    client_id = perspective_cid, object_id = perspective_oid).first()

                perspective_name = (
                    perspective.get_translation(ENGLISH_LOCALE).strip())

                if len(perspective_name) > 48:
                    perspective_name = perspective_name[:48] + '...'

                dictionary_name = (
                    perspective.parent.get_translation(ENGLISH_LOCALE).strip())

                if len(dictionary_name) > 48:
                    dictionary_name = dictionary_name[:48] + '...'

                phonemic_name_str = (
                    'phonemic {0} {1} {2}'.format(
                    dictionary_name,
                    perspective_name,
                    len(data_list) + 1))

                # Initializing file storage directory, if required.

                if __intermediate_flag__:

                    storage = info.context.request.registry.settings['storage']
                    cur_time = time.time()

                    storage_dir = os.path.join(
                        storage['path'], 'phonemic', str(cur_time))

                for extension, encoding in (
                    ('utf8', 'utf-8'), ('utf16', 'utf-16')):

                    input_file_name = (

                        pathvalidate.sanitize_filename(
                            'input {0}.{1}'.format(
                                phonemic_name_str, extension)))

                    # Saving to the working directory...

                    if __debug_flag__:

                        with open(input_file_name, 'wb') as input_file:
                            input_file.write(input.encode(encoding))

                    # ...and / or to the file storage.

                    if __intermediate_flag__:

                        input_path = os.path.join(
                            storage_dir, input_file_name)

                        os.makedirs(
                            os.path.dirname(input_path),
                            exist_ok = True)

                        with open(input_path, 'wb') as input_file:
                            input_file.write(input.encode(encoding))

                        input_url = ''.join([
                            storage['prefix'],
                            storage['static_route'],
                            'phonemic', '/',
                            str(cur_time), '/',
                            input_file_name])

                        intermediate_url_list.append(input_url)

            # Calling analysis library, starting with getting required output buffer size and continuing
            # with analysis proper.

            output_buffer_size = phonemic_analysis_f(
                None, len(data_list) + 1, None, 0)

            if output_buffer_size <= 0:
                return ResponseError(message = 'Invalid output buffer size')

            log.debug(
                'phonemic_analysis {0}/{1}: output buffer size {2}'.format(
                perspective_cid, perspective_oid,
                output_buffer_size))

            input_buffer = ctypes.create_unicode_buffer(input)
            output_buffer = ctypes.create_unicode_buffer(output_buffer_size + 256)

            result = phonemic_analysis_f(
                input_buffer, len(data_list) + 1, output_buffer, 0)

            # If we don't have a good result, we return an error.

            log.debug(
                'phonemic_analysis {0}/{1}: '
                'transcription field {2}/{3}, translation field {4}/{5}: result {6}'.format(
                    perspective_cid, perspective_oid,
                    transcription_field_cid, transcription_field_oid,
                    translation_field_cid, translation_field_oid,
                    result))

            if result != 1:
                return ResponseError(message =
                    'Phonemic analysis library call error {0}'.format(result))

            output = output_buffer.value

            log.debug(
                'phonemic_analysis {0}/{1}: '
                'transcription field {2}/{3}, translation field {4}/{5}:'
                '\noutput:\n{6}'.format(
                    perspective_cid, perspective_oid,
                    transcription_field_cid, transcription_field_oid,
                    translation_field_cid, translation_field_oid,
                    repr(output)))

            # Reflowing output, if required.

            final_output = output

            if wrap_flag:

                line_list = output.split('\r\n')

                text_wrapper = textwrap.TextWrapper(
                    width = 108, tabsize = 4)

                reflow_list = []

                for line in line_list:
                    reflow_list.extend(text_wrapper.wrap(line))

                wrapped_output = '\n'.join(reflow_list)

                log.debug(
                    'phonemic_analysis {0}/{1}: '
                    'transcription field {2}/{3}, translation field {4}/{5}:'
                    '\nwrapped output:\n{6}'.format(
                        perspective_cid, perspective_oid,
                        transcription_field_cid, transcription_field_oid,
                        translation_field_cid, translation_field_oid,
                        wrapped_output))

                final_output = wrapped_output

            # Returning result.

            return PhonemicAnalysis(

                triumph = True,
                entity_count = total_count,
                result = final_output,

                intermediate_url_list =
                    intermediate_url_list if __intermediate_flag__ else None)

        except Exception as exception:

            traceback_string = ''.join(traceback.format_exception(
                exception, exception, exception.__traceback__))[:-1]

            log.warning('phonemic_analysis: exception')
            log.warning(traceback_string)

            return ResponseError(message =
                'Exception:\n' + traceback_string)


@celery.task
def async_cognate_analysis(
    language_str,
    source_perspective_id,
    base_language_id,
    base_language_name,
    group_field_id,
    perspective_info_list,
    multi_list,
    multi_name_list,
    mode,
    distance_flag,
    reference_perspective_id,
    figure_flag,
    distance_vowel_flag,
    distance_consonant_flag,
    match_translations_value,
    only_orphans_flag,
    locale_id,
    storage,
    task_key,
    cache_kwargs,
    sqlalchemy_url,
    __debug_flag__,
    __intermediate_flag__):
    """
    Sets up and launches cognate analysis in asynchronous mode.
    """

    # NOTE: copied from phonology.
    #
    # This is a no-op with current settings, we use it to enable logging inside celery tasks, because
    # somehow this does it, and otherwise we couldn't set it up.

    logging.debug('async_cognate_analysis')

    # Ok, and now we go on with task execution.

    engine = create_engine(sqlalchemy_url)
    DBSession.configure(bind = engine)
    initialize_cache(cache_kwargs)

    task_status = TaskStatus.get_from_cache(task_key)

    with transaction.manager:

        try:
            CognateAnalysis.perform_cognate_analysis(
                language_str,
                source_perspective_id,
                base_language_id,
                base_language_name,
                group_field_id,
                perspective_info_list,
                multi_list,
                multi_name_list,
                mode,
                None,
                None,
                None,
                None,
                None,
                match_translations_value,
                only_orphans_flag,
                locale_id,
                storage,
                task_status,
                __debug_flag__,
                __intermediate_flag__)

        # Some unknown external exception.

        except Exception as exception:

            traceback_string = ''.join(traceback.format_exception(
                exception, exception, exception.__traceback__))[:-1]

            log.warning(
                'cognate_analysis {0}: exception'.format(
                language_str))

            log.warning(traceback_string)

            if task_status is not None:

                task_status.set(5, 100,
                    'Finished (ERROR), exception:\n' + traceback_string)


class CognateAnalysis(graphene.Mutation):

    class Arguments:

        source_perspective_id = LingvodocID(required = True)
        base_language_id = LingvodocID(required = True)

        group_field_id = LingvodocID(required = True)
        perspective_info_list = graphene.List(graphene.List(LingvodocID), required = True)
        multi_list = graphene.List(ObjectVal)

        mode = graphene.String()

        distance_flag = graphene.Boolean()
        reference_perspective_id = LingvodocID()

        figure_flag = graphene.Boolean()
        distance_vowel_flag = graphene.Boolean()
        distance_consonant_flag = graphene.Boolean()

        match_translations_value = graphene.Int()
        only_orphans_flag = graphene.Boolean()

        debug_flag = graphene.Boolean()
        intermediate_flag = graphene.Boolean()

        synchronous = graphene.Boolean()

    triumph = graphene.Boolean()

    dictionary_count = graphene.Int()
    group_count = graphene.Int()
    not_enough_count = graphene.Int()
    transcription_count = graphene.Int()
    translation_count = graphene.Int()

    result = graphene.String()
    xlsx_url = graphene.String()
    distance_list = graphene.Field(ObjectVal)
    figure_url = graphene.String()

    minimum_spanning_tree = graphene.List(graphene.List(graphene.Int))
    embedding_2d = graphene.List(graphene.List(graphene.Float))
    embedding_3d = graphene.List(graphene.List(graphene.Float))
    perspective_name_list = graphene.List(graphene.String)

    suggestion_list = graphene.List(ObjectVal)
    suggestion_field_id = LingvodocID()

    intermediate_url_list = graphene.List(graphene.String)

    @staticmethod
    def tag_data_std(
        entry_already_set,
        group_list,
        perspective_id,
        field_client_id,
        field_object_id):
        """
        Gets lexical entry grouping data using current standard methods, computes elapsed time.
        """

        start_time = time.time()

        tag_data_list = (DBSession.query(
            dbLexicalEntry, func.count('*'))

            .filter(
                dbLexicalEntry.parent_client_id == perspective_id[0],
                dbLexicalEntry.parent_object_id == perspective_id[1],
                dbLexicalEntry.marked_for_deletion == False,
                dbEntity.parent_client_id == dbLexicalEntry.client_id,
                dbEntity.parent_object_id == dbLexicalEntry.object_id,
                dbEntity.field_client_id == field_client_id,
                dbEntity.field_object_id == field_object_id,
                dbEntity.marked_for_deletion == False,
                dbPublishingEntity.client_id == dbEntity.client_id,
                dbPublishingEntity.object_id == dbEntity.object_id,
                dbPublishingEntity.published == True,
                dbPublishingEntity.accepted == True)

            .group_by(
                dbLexicalEntry.client_id,
                dbLexicalEntry.object_id))

        # Processing each lexical entry with at least one tag.

        for entry, count in tag_data_list.all():

            if (entry.client_id, entry.object_id) in entry_already_set:
                continue

            tag_set = find_all_tags(
                entry, field_client_id, field_object_id, True, True)

            entry_list = find_lexical_entries_by_tags(
                tag_set, field_client_id, field_object_id, True, True)

            entry_id_set = set(
                (tag_entry.client_id, tag_entry.object_id)
                for tag_entry in entry_list)

            entry_already_set.update(entry_id_set)
            group_list.append(entry_id_set)

        return time.time() - start_time

    @staticmethod
    def find_group(
        entry_client_id,
        entry_object_id,
        field_client_id,
        field_object_id):
        """
        Retrieves all lexical entries grouped with a given id-specified entry.
        """

        entry_id_set = set((
            (entry_client_id, entry_object_id),))

        tag_query = (

            DBSession.query(
                dbEntity.content)

            .filter(
                dbEntity.parent_client_id == entry_client_id,
                dbEntity.parent_object_id == entry_object_id,
                dbEntity.field_client_id == field_client_id,
                dbEntity.field_object_id == field_object_id,
                dbEntity.marked_for_deletion == False,
                dbPublishingEntity.client_id == dbEntity.client_id,
                dbPublishingEntity.object_id == dbEntity.object_id,
                dbPublishingEntity.published == True,
                dbPublishingEntity.accepted == True))

        tag_list = list(tag_query.all())
        tag_set = set(tag_list)

        find_group_by_tags(
            DBSession,
            entry_id_set, tag_set, tag_list,
            field_client_id, field_object_id,
            True)

        return entry_id_set

    @staticmethod
    def tag_data_optimized(
        entry_already_set,
        group_list,
        perspective_id,
        field_client_id,
        field_object_id):
        """
        Gets lexical entry grouping data using (hopefully) optimized version of the current standard
        methods, computes elapsed time.
        """

        start_time = time.time()

        tag_data_list = (

            DBSession.query(
                dbLexicalEntry.client_id,
                dbLexicalEntry.object_id,
                func.count('*'))

            .filter(
                dbLexicalEntry.parent_client_id == perspective_id[0],
                dbLexicalEntry.parent_object_id == perspective_id[1],
                dbLexicalEntry.marked_for_deletion == False,
                dbEntity.parent_client_id == dbLexicalEntry.client_id,
                dbEntity.parent_object_id == dbLexicalEntry.object_id,
                dbEntity.field_client_id == field_client_id,
                dbEntity.field_object_id == field_object_id,
                dbEntity.marked_for_deletion == False,
                dbPublishingEntity.client_id == dbEntity.client_id,
                dbPublishingEntity.object_id == dbEntity.object_id,
                dbPublishingEntity.published == True,
                dbPublishingEntity.accepted == True)

            .group_by(
                dbLexicalEntry.client_id,
                dbLexicalEntry.object_id))

        # Processing each lexical entry with at least one tag.

        for entry_client_id, entry_object_id, count in tag_data_list.all():

            if (entry_client_id, entry_object_id) in entry_already_set:
                continue

            entry_id_set = CognateAnalysis.find_group(
                entry_client_id, entry_object_id,
                field_client_id, field_object_id)

            entry_already_set.update(entry_id_set)
            group_list.append(entry_id_set)

        return time.time() - start_time

    @staticmethod
    def tag_data_aggregated(
        perspective_info_list,
        tag_field_id,
        statistics_flag = False,
        optimize_flag = False):
        """
        Gets lexical entry grouping data using aggregated retrieval, computes elapsed time.
        """

        start_time = time.time()

        entry_id_dict = collections.defaultdict(set)
        tag_dict = collections.defaultdict(set)

        tag_set = set()

        # All tags for tagged lexical entries in specified perspectives.

        for perspective_id, transcription_field_id, translation_field_id in perspective_info_list:

            tag_query = (

                DBSession.query(
                    dbEntity.parent_client_id,
                    dbEntity.parent_object_id,
                    dbEntity.content)

                .filter(
                    dbLexicalEntry.parent_client_id == perspective_id[0],
                    dbLexicalEntry.parent_object_id == perspective_id[1],
                    dbLexicalEntry.marked_for_deletion == False,
                    dbEntity.parent_client_id == dbLexicalEntry.client_id,
                    dbEntity.parent_object_id == dbLexicalEntry.object_id,
                    dbEntity.field_client_id == tag_field_id[0],
                    dbEntity.field_object_id == tag_field_id[1],
                    dbEntity.marked_for_deletion == False,
                    dbPublishingEntity.client_id == dbEntity.client_id,
                    dbPublishingEntity.object_id == dbEntity.object_id,
                    dbPublishingEntity.published == True,
                    dbPublishingEntity.accepted == True))

            for entry_client_id, entry_object_id, tag in tag_query.all():

                entry_id_dict[(entry_client_id, entry_object_id)].add(tag)
                tag_set.add(tag)

        tag_list = tag_set

        # While we have tags we don't have all lexical entries for,
        # we get all entries of these tags...

        while tag_list:

            entry_id_query = (

                DBSession.query(
                    dbEntity.parent_client_id,
                    dbEntity.parent_object_id,
                    dbEntity.content)

                .filter(
                    dbLexicalEntry.marked_for_deletion == False,
                    dbEntity.parent_client_id == dbLexicalEntry.client_id,
                    dbEntity.parent_object_id == dbLexicalEntry.object_id,
                    dbEntity.field_client_id == tag_field_id[0],
                    dbEntity.field_object_id == tag_field_id[1],
                    dbEntity.marked_for_deletion == False,
                    dbEntity.content.in_(tag_list),
                    dbPublishingEntity.client_id == dbEntity.client_id,
                    dbPublishingEntity.object_id == dbEntity.object_id,
                    dbPublishingEntity.published == True,
                    dbPublishingEntity.accepted == True))

            entry_id_list = []

            for entry_client_id, entry_object_id, tag in entry_id_query.all():

                entry_id = entry_client_id, entry_object_id
                tag_dict[tag].add(entry_id)

                if entry_id not in entry_id_dict:
                    entry_id_list.append(entry_id)

            # And then get all tags for entries we haven't already done it for.

            tag_list = []

            entry_id_list.sort()
            log.debug('len(entry_id_list): {0}'.format(len(entry_id_list)))

            for i in range((len(entry_id_list) + 16383) // 16384):

                tag_query = (

                    DBSession.query(
                        dbEntity.parent_client_id,
                        dbEntity.parent_object_id,
                        dbEntity.content)

                    # We have to split entries into parts due to the danger of stack overflow in Postgres.

                    .filter(
                        tuple_(dbEntity.parent_client_id, dbEntity.parent_object_id)
                            .in_(entry_id_list[i * 16384 : (i + 1) * 16384]),
                        dbEntity.field_client_id == tag_field_id[0],
                        dbEntity.field_object_id == tag_field_id[1],
                        dbEntity.marked_for_deletion == False,
                        dbPublishingEntity.client_id == dbEntity.client_id,
                        dbPublishingEntity.object_id == dbEntity.object_id,
                        dbPublishingEntity.published == True,
                        dbPublishingEntity.accepted == True))

                for entry_client_id, entry_object_id, tag in tag_query.all():

                    entry_id = entry_client_id, entry_object_id
                    entry_id_dict[entry_id].add(tag)

                    if tag not in tag_dict:
                        tag_list.append(tag)

        # And now grouping lexical entries by tags.

        entry_already_set = set()
        group_list = []

        for entry_id, entry_tag_set in entry_id_dict.items():

            if entry_id in entry_already_set:
                continue

            entry_already_set.add(entry_id)

            group_entry_id_set = set((entry_id,))
            group_tag_set = set(entry_tag_set)

            current_tag_list = entry_tag_set

            # Recursively gathering grouped entries through their tags.

            while current_tag_list:

                tag_list = []

                for current_tag in current_tag_list:
                    for current_entry_id in tag_dict[current_tag]:

                        if current_entry_id in group_entry_id_set:
                            continue

                        group_entry_id_set.add(current_entry_id)

                        for tag in entry_id_dict[current_entry_id]:
                            if tag not in group_tag_set:

                                group_tag_set.add(tag)
                                tag_list.append(tag)

                current_tag_list = tag_list

            # Saving entry group data.

            entry_already_set.update(group_entry_id_set)
            group_list.append(group_entry_id_set)

        # Computing tag statistics for each entry group, if required.

        if statistics_flag:

            redundant_group_count = 0
            redundant_tag_count = 0
            redundant_entity_count = 0

            redundant_group_tag_count = 0
            redundant_group_entry_count = 0
            redundant_group_entity_count = 0

            delta_entity_count = 0

            # Processing each entry group.

            for entry_id_set in group_list:

                tag_set = set.union(
                    *(entry_id_dict[entry_id]
                        for entry_id in entry_id_set))

                tag_list = list(sorted(
                    (tag, len(tag_dict[tag]))
                        for tag in tag_set))

                partial_tag_set = set((tag_list[0][0],))
                partial_entry_id_set = set(tag_dict[tag_list[0][0]])

                index = 1

                # How many partial tag groups we need to cover the entry group completely?

                while (
                    len(partial_entry_id_set) < len(entry_id_set) and
                    index < len(tag_list)):

                    partial_tag_set.add(tag_list[index][0])
                    partial_entry_id_set.update(tag_dict[tag_list[index][0]])

                    index += 1

                group_entity_count = (

                    sum(len(tag_dict[tag])
                        for tag in tag_set))

                # Checking if we have some redundant partial tag groups.

                if index < len(tag_list):

                    redundant_group_count += 1
                    redundant_tag_count += len(tag_list) - index

                    for tag, count in tag_list[index:]:
                        redundant_entity_count += count

                    redundant_group_tag_count += len(tag_list)
                    redundant_group_entry_count += len(entry_id_set)

                    redundant_group_entity_count += group_entity_count

                delta_entity_count += group_entity_count - len(entry_id_set)

            # Showing gathered statistics.

            log.debug(
                '\ngroup_count: {0}/{1}, with {2} tags, {3} entries, {4} entities'
                '\ntag_count: {5}/{6}, {5}/{7} among redundant groups'
                '\nentity_count: {8}/{9}, {8}/{10} among redundant groups'
                '\ndelta(entity): {11}, minimum(entity): {12}'.format(
                    redundant_group_count, len(group_list),
                    redundant_group_tag_count,
                    redundant_group_entry_count,
                    redundant_group_entity_count,
                    redundant_tag_count, len(tag_dict),
                    redundant_group_tag_count,
                    redundant_entity_count,
                    sum(len(entry_id_set)
                        for entry_id_set in tag_dict.values()),
                    redundant_group_entity_count,
                    delta_entity_count,
                    len(entry_id_dict)))

        # If required, optimizing lexical entry groups by ensuring that each group has exactly one tag.

        if optimize_flag:

            redundant_tag_list = []

            for entry_id_set in group_list:

                tag_set = set.union(
                    *(entry_id_dict[entry_id]
                        for entry_id in entry_id_set))

                count, tag = max(
                    (len(tag_dict[tag]), tag)
                    for tag in tag_set)

                # Creating tag entities we need to link current group via selected tag.

                for entry_id in entry_id_set - tag_dict[tag]:

                    tag_entity = dbEntity(
                        client_id = entry_id[0],
                        parent_client_id = entry_id[0],
                        parent_object_id = entry_id[1],
                        field_client_id = tag_field_id[0],
                        field_object_id = tag_field_id[1],
                        content = tag)

                    tag_entity.publishingentity.published = True
                    tag_entity.publishingentity.accepted = True

                tag_set.remove(tag)
                redundant_tag_list.extend(tag_set)

            # Removing tag entities of the redundant tags.

            entity_id_query = (

                dbEntity.__table__
                    .update()
                    .where(and_(
                        dbEntity.marked_for_deletion == False,
                        dbEntity.content.in_(redundant_tag_list),
                        dbPublishingEntity.client_id == dbEntity.client_id,
                        dbPublishingEntity.object_id == dbEntity.object_id,
                        dbPublishingEntity.published == True,
                        dbPublishingEntity.accepted == True))
                    .values(marked_for_deletion = True)
                    .returning(dbEntity.client_id, dbEntity.object_id))

            entity_id_list = list(
                entity_id_query.execute())

            log.debug('entity_id_list: {0} entities\n{1}'.format(
                len(entity_id_list), pprint.pformat(entity_id_list)))

            # Optimization is not fully implemented at the moment.

            raise NotImplementedError

        return entry_already_set, group_list, time.time() - start_time

    @staticmethod
    def tag_data_plpgsql(
        perspective_info_list,
        tag_field_id,
        statistics_flag = False,
        optimize_flag = False):
        """
        Gets lexical entry grouping data using stored PL/pgSQL functions, computes elapsed time.
        """

        __debug_flag__ = False

        start_time = time.time()

        # Getting lexical entries with tag data of the specified tag field from all perspectives.

        perspective_id_list = [
            perspective_id
            for perspective_id, _, _ in perspective_info_list]

        if not perspective_id_list:

            return set(), [], time.time() - start_time

        entry_id_query = (

            DBSession.query(
                dbLexicalEntry.client_id,
                dbLexicalEntry.object_id)

            .filter(

                tuple_(
                    dbLexicalEntry.parent_client_id,
                    dbLexicalEntry.parent_object_id)

                    .in_(
                        ids_to_id_query(
                            perspective_id_list)),

                dbLexicalEntry.marked_for_deletion == False,
                dbEntity.parent_client_id == dbLexicalEntry.client_id,
                dbEntity.parent_object_id == dbLexicalEntry.object_id,
                dbEntity.field_client_id == tag_field_id[0],
                dbEntity.field_object_id == tag_field_id[1],
                dbEntity.marked_for_deletion == False,
                dbPublishingEntity.client_id == dbEntity.client_id,
                dbPublishingEntity.object_id == dbEntity.object_id,
                dbPublishingEntity.published == True,
                dbPublishingEntity.accepted == True)

            .group_by(
                dbLexicalEntry.client_id,
                dbLexicalEntry.object_id))

        entry_id_list = (
            entry_id_query.all())

        if __debug_flag__:

            log.debug(
                '\nentry_id_query:\n' +
                str(entry_id_query.statement.compile(compile_kwargs = {'literal_binds': True})))

            log.debug(
                f'len(entry_id_list): {len(entry_id_list)}')

        # Grouping lexical entries using stored PL/pgSQL function.

        entry_already_set = set()
        group_list = []

        sql_str = '''
            select * from linked_group(
                :field_client_id,
                :field_object_id,
                :client_id,
                :object_id)'''

        for entry_id in entry_id_list:

            if entry_id in entry_already_set:
                continue

            row_list = (

                DBSession.execute(sql_str, {
                    'field_client_id': tag_field_id[0],
                    'field_object_id': tag_field_id[1],
                    'client_id': entry_id[0],
                    'object_id': entry_id[1]})

                .fetchall())

            entry_id_set = set(
                map(tuple, row_list))

            entry_already_set.update(entry_id_set)
            group_list.append(entry_id_set)

        return entry_already_set, group_list, time.time() - start_time

    @staticmethod
    def export_xlsx(
        language_str,
        mode,
        output_str,
        d_output_str,
        perspective_count,
        __debug_flag__ = False,
        cognate_name_str = None):
        """
        Parses results of the cognate analysis and exports them as an XLSX file.
        """

        workbook_stream = io.BytesIO()

        workbook = xlsxwriter.Workbook(
            workbook_stream, {'in_memory': True})

        worksheet_results = (
            workbook.add_worksheet(
                utils.sanitize_worksheet_name('Results')))

        # 20% background yellow, 10% background gray.

        format_yellow = (
            workbook.add_format({'bg_color': '#ffffcc'}))

        format_gray = (
            workbook.add_format({'bg_color': '#e6e6e6'}))

        index = output_str.find('\0')
        size_list = list(map(int, output_str[:index].split(',')))

        log.debug(
            'cognate_analysis {0}: result table size {1}'.format(
            language_str,
            size_list))

        max_width = 0
        row_count = 0

        re_series = r'\s*(\[\S+\]|\?|0)(\s*—\s*(\[\S+\]|\?|0))+\s*'
        re_item_list = r'\s*(\[\S+\]|\?|0)\s*—(\s*—\s*(\[\S+\]|\?|0)\s*—)+\s*'

        def export_table(table_index, table_str, n_col, n_row, source_str):
            """
            Parses from the binary output and exports to the XLSX workbook a table with specified width and
            height.
            """

            nonlocal index

            nonlocal max_width
            nonlocal row_count

            if n_col > max_width:

                max_width = n_col
                worksheet_results.set_column(0, max_width - 1, 16)

            row_list = []

            for i in range(n_row):

                value_list = []

                # Another row of analysis result values.

                for j in range(n_col):

                    index_next = source_str.find('\0', index + 1)
                    value = source_str[index + 1 : index_next]

                    value_list.append(value)
                    index = index_next

                split_list_list = [
                    value.split('|') for value in value_list]

                item_list_count = max(map(len, split_list_list))

                # Checking if we need color formatting.

                cell_format = None

                if (re.match(re_series, value_list[0]) is not None or
                    re.match(re_item_list, '—'.join(value_list))):

                    cell_format = format_yellow

                emphasize_flag_list = [
                    value.startswith('(') and value.endswith(')')
                    for value in value_list[::2]]

                # Some values may actually be sequences, so we check and process them if they are.

                for i in range(item_list_count):

                    item_list = [
                        split_list[i] if i < len(split_list) else ''
                        for split_list in split_list_list]

                    row_list.append(item_list)

                    for j, (x_script, x_lat, emphasize_flag) in (
                        enumerate(zip(item_list[::2], item_list[1::2], emphasize_flag_list))):

                        worksheet_results.write_row(
                            row_count,
                            j * 2,
                            [x_script, x_lat],
                            format_gray
                                if emphasize_flag and (x_script + x_lat).strip() else
                                cell_format)

                    row_count += 1

                # Going on another row of analysis results.

                if source_str[index + 1] != '\0':
                    raise NotImplementedError

                index += 1

            log.debug(
                'cognate_analysis {0}: {1} table {2}:\n{3}'.format(
                language_str,
                table_str, table_index,
                pprint.pformat(row_list, width = 144)))

            # Returning table data.

            return row_list

        # Getting analysis result info, exporting it to the XLSX workbook.

        for table_index, howmany in enumerate(range(len(size_list) // 2)):

            n_col, n_row = size_list[howmany * 2 : howmany * 2 + 2]
            export_table(table_index, 'result', n_col, n_row, output_str)

        # And now we parse formant plot data, if we have any.

        if len(output_str) > index + 1:

            if output_str[index + 1] != '\0':
                raise NotImplementedError

            index += 1
            index_next = output_str.find('\0', index + 1)

            size_list = list(map(int,
                output_str[index + 1 : index_next].split(',')))

            index = index_next

            log.debug(
                'cognate_analysis {0}: plot table size {1}'.format(
                language_str,
                size_list))

            # Getting plot info, exporting it to the XLSX workbook, generating plots.

            worksheet_table_2d = (
                workbook.add_worksheet(
                    utils.sanitize_worksheet_name('F-table')))

            worksheet_chart = (
                workbook.add_worksheet(
                    utils.sanitize_worksheet_name('F-chart')))

            table_2d_row_index = 0
            chart_2d_count = 0

            for table_index, howmany in enumerate(range(len(size_list) // 2)):

                n_col, n_row = size_list[howmany * 2 : howmany * 2 + 2]

                row_list = export_table(
                    table_index, 'plot', n_col, n_row, output_str)

                plot_title = row_list[0][0]

                if not plot_title:
                    continue

                # Getting formant series info.

                series_title_list = []
                series_data_list = []

                for row in row_list[1:]:

                    if not row[0]:
                        continue

                    elif not row[1]:

                        series_title_list.append(row[0])
                        series_data_list.append([])

                    else:

                        series_data_list[-1].append(
                            tuple(map(float, row[1:])))

                log.debug(
                    'cognate_analysis {0}: plot data {1}:\n{2}\n{3}\n{4}'.format(
                    language_str,
                    table_index,
                    repr(plot_title),
                    pprint.pformat(series_title_list, width = 144),
                    pprint.pformat(series_data_list, width = 144)))

                # Proceeding with plot generation only if we have enough data.

                if sum(map(len, series_data_list)) <= 1:
                    continue

                chart_data_2d_list = []

                min_2d_f1, max_2d_f1 = None, None
                min_2d_f2, max_2d_f2 = None, None

                # Generating plot data.

                for series_index, (series_title, series_data) in enumerate(
                    zip(series_title_list, series_data_list)):

                    f_2d_list = list(map(
                        lambda f_tuple: numpy.array(f_tuple[:2]), series_data))

                    f_3d_list = list(map(
                        numpy.array, series_data))

                    if len(f_2d_list) <= 0:
                        continue

                    (filtered_2d_list, outlier_2d_list, mean_2d, ellipse_list,
                        filtered_3d_list, outlier_3d_list, mean_3d, sigma_2d, inverse_3d) = (

                        phonology.chart_data(
                            [(f_2d, ('', '')) for f_2d in f_2d_list],
                            [(f_3d, ('', '')) for f_3d in f_3d_list]))

                    chart_data_2d_list.append((
                        len(filtered_2d_list), len(f_2d_list), series_title,
                        filtered_2d_list, outlier_2d_list, mean_2d, ellipse_list))

                    # Updating F1/F2 maximum/minimum info.

                    f1_list, f2_list = zip(*(x[0] for x in filtered_2d_list))

                    min_f1_list, max_f1_list = min(f1_list), max(f1_list)
                    min_f2_list, max_f2_list = min(f2_list), max(f2_list)

                    if min_2d_f1 == None or min_f1_list < min_2d_f1:
                        min_2d_f1 = min_f1_list

                    if max_2d_f1 == None or max_f1_list > max_2d_f1:
                        max_2d_f1 = max_f1_list

                    if min_2d_f2 == None or min_f2_list < min_2d_f2:
                        min_2d_f2 = min_f2_list

                    if max_2d_f2 == None or max_f2_list > max_2d_f2:
                        max_2d_f2 = max_f2_list

                # Compiling info of the formant scatter chart data series.

                chart_dict_list, table_2d_row_index = (

                    phonology.chart_definition_list(
                        chart_data_2d_list, worksheet_table_2d,
                        min_2d_f1, max_2d_f1, min_2d_f2, max_2d_f2,
                        row_index = table_2d_row_index))

                if chart_dict_list:

                    # Generating the chart, if we have any data.

                    chart = workbook.add_chart({'type': 'scatter'})

                    chart.set_title({
                        'name': plot_title})

                    chart.set_x_axis({
                        'major_gridlines': {'visible': True},
                        'name': 'F2 (Hz)',
                        'reverse': True})

                    chart.set_y_axis({
                        'major_gridlines': {'visible': True},
                        'name': 'F1 (Hz)',
                        'reverse': True})

                    chart.set_legend({
                        'position': 'top'})

                    for chart_dict in chart_dict_list:
                        chart.add_series(chart_dict)

                    chart.set_style(11)
                    chart.set_size({'width': 1024, 'height': 768})

                    worksheet_chart.insert_chart(
                        'A{0}'.format(chart_2d_count * 40 + 1), chart)

                    chart_2d_count += 1

        # If we have distance matrix data, we also parse and export it.

        matrix_info_list = None

        if d_output_str != None:

            index = d_output_str.find('\0')
            size_list = list(map(int, d_output_str[:index].split(',')))

            log.debug(
                'cognate_analysis {0}: distance result table size {1}'.format(
                language_str,
                size_list))

            matrix_info_list = []

            # Parsing and exporting each distance matrix.

            for table_index, howmany in enumerate(range(len(size_list) // 2)):

                n_col, n_row = size_list[howmany * 2 : howmany * 2 + 2]

                row_list = export_table(
                    table_index, 'distance result', n_col, n_row, d_output_str)

                matrix_title = row_list[0][0]

                if not matrix_title:
                    continue

                # Getting distance matrix info.

                matrix_header_list = []
                matrix_data_list = []

                for row in row_list[1:]:

                    if row[0]:

                        matrix_header_list.append(row[0])
                        matrix_data_list.append(row[1 : 1 + perspective_count])

                # Getting distance matrix array, checking if we need to filter its parts.

                matrix_header_array = numpy.array(matrix_header_list)

                matrix_data_array = numpy.array([
                    tuple(map(float, value_list))
                    for value_list in matrix_data_list])

                where = matrix_data_array[:,0] >= 0

                if not all(where):

                    matrix_header_array = matrix_header_array[where]
                    matrix_data_array = matrix_data_array[where, :][:, where]

                    # If we do, we also export filtered version.

                    worksheet_results.write(
                        'A{0}'.format(row_count + 1), 'Filtered:')

                    row_count += 2

                    worksheet_results.write_row(
                        'A{0}'.format(row_count + 1),
                        [''] + list(matrix_header_array))

                    row_count += 1

                    for i, header in enumerate(matrix_header_array):

                        worksheet_results.write_row(
                            'A{0}'.format(row_count + 1),
                            [header] + [round(value) for value in matrix_data_array[i]])

                        row_count += 1

                # Showing info of the matrix we've got.

                matrix_info_list.append((
                    matrix_title,
                    matrix_header_list,
                    matrix_data_list,
                    matrix_header_array,
                    matrix_data_array))

                log.debug(
                    '\ncognate_analysis {0}:'
                    '\ndistance data {1}:'
                    '\n{2}\n{3}\n{4}'
                    '\nmatrix_header_array:\n{5}'
                    '\nmatrix_data_array:\n{6}'.format(
                    language_str,
                    table_index,
                    repr(matrix_title),
                    pprint.pformat(matrix_header_list, width = 144),
                    pprint.pformat(matrix_data_list, width = 144),
                    matrix_header_array,
                    matrix_data_array))

        workbook.close()

        # Saving resulting Excel workbook for debug purposes, if required.

        if __debug_flag__:

            xlsx_file_name = (
                cognate_name_str + '.xlsx')

            workbook_stream.seek(0)

            with open(xlsx_file_name, 'wb') as xlsx_file:
                shutil.copyfileobj(workbook_stream, xlsx_file)

        # Returning exported XLSX data as a binary stream and any parsed distance matrix info.

        return workbook_stream, matrix_info_list

    @staticmethod
    def parse_suggestions(
        language_str,
        output_str,
        perspective_count,
        perspective_source_index,
        entry_id_dict,
        __debug_flag__ = False,
        cognate_name_str = None,
        group_field_id = None):
        """
        Parses cognate suggestions.
        """

        index = -1
        row_list = []

        # Parsing result table rows.

        while index < len(output_str):

            value_list = []

            # Getting required number of values.

            for i in range(
                perspective_count * 2):

                index_next = (
                    output_str.find('\0', index + 1))

                # No more values, meaning we are at the end of the table.

                if index_next == -1:
                    break

                value = (
                    output_str[index + 1 : index_next])

                value_list.append(value)
                index = index_next

            if index_next == -1:
                break

            # Another row of analysis result values.

            log.debug(
                '\n{0} / {1}: {2}'.format(
                    len(row_list),
                    index,
                    repr(value_list)))

            row_list.append(
                value_list)

            if output_str[index + 1] != '\0':
                raise NotImplementedError

            index += 1

        # Parsing cognate suggestions table.

        row_index = 2
        suggestion_list = []

        header_str = row_list[0][0]

        begin_str = 'Предложения для '
        end_str = ': '

        while row_index < len(row_list):

            word_str = row_list[row_index][0]

            assert word_str.startswith(begin_str)

            if word_str.endswith('НЕТ'):

                row_index += 1
                continue

            assert word_str.endswith(end_str)

            word = (
                word_str[
                    len(begin_str) : -len(end_str)])

            word_entry_id = (

                entry_id_dict[(
                    perspective_source_index,
                    word)])

            # Parsing suggestions for the word.

            single_list = []
            group_list = []

            row_index += 1

            if row_list[row_index][perspective_source_index * 2] == header_str:
                row_index += 1

            value_list = row_list[row_index]

            while not (
                row_index >= len(row_list) or
                value_list[0].startswith(begin_str)):

                # Existing groups.

                if value_list[0] == 'Уже имеющиеся ряды: ':

                    row_index += 1
                    value_list = row_list[row_index]

                    while not (
                        value_list[0].startswith(begin_str) or
                        value_list[0] == 'Слова-сироты: '):

                        word_list = []

                        for i in range(perspective_count):

                            if value_list[i * 2] or value_list[i * 2 + 1]:

                                word_list.append(
                                    (i, value_list[i * 2 : i * 2 + 2]))

                        if word_list:
                            group_list.append(word_list)

                        row_index += 1

                        if row_index >= len(row_list):
                            break

                        value_list = row_list[row_index]

                # Single words.

                elif value_list[0] == 'Слова-сироты: ':

                    row_index += 1

                    if row_list[row_index][perspective_source_index * 2] == header_str:
                        row_index += 1

                    value_list = row_list[row_index]

                    while not (
                        value_list[0].startswith(begin_str) or
                        value_list[0] == 'Уже имеющиеся ряды: '):

                        for i in range(perspective_count):

                            if value_list[i * 2] or value_list[i * 2 + 1]:

                                single_list.append(
                                    (i, value_list[i * 2 : i * 2 + 2]))

                        row_index += 1

                        if row_index >= len(row_list):
                            break

                        value_list = row_list[row_index]

                # Something unexpected?

                else:

                    log.debug(value_list)
                    raise NotImplementedError

            # Getting lexical entry identifiers, if required.

            raw_list = single_list

            def f(index, tt_tuple):
                """
                Gets entry id by its perspective index, translations and transcriptions.
                """

                transcription_str, translation_str = tt_tuple

                return (

                    entry_id_dict[(
                        index,
                        transcription_str + (
                            ' ' + translation_str if translation_str else ''))])

            single_list = [
                (index, tt_tuple, f(index, tt_tuple))
                for index, tt_tuple in raw_list]

            # For lexical entry groups we get id of just the first entry.

            raw_list = group_list
            group_list = []

            word_group = None

            for word_list in raw_list:

                entry_id_list = list(f(*w) for w in word_list)

                if word_entry_id in entry_id_list:

                    if word_group is not None:
                        raise NotImplementedError

                    word_group = (word_list, entry_id_list[0])

                else:

                    group_list.append((
                        word_list, entry_id_list[0]))

            # Showing what we've got, saving suggestion if it is non-trivial.

            log.debug('\n' +
                pprint.pformat(
                    (word, word_entry_id, word_group, single_list, group_list),
                    width = 192))

            if single_list or group_list:

                suggestion_list.append((
                    perspective_source_index,
                    word,
                    word_entry_id,
                    word_group,
                    single_list,
                    group_list))

        # Maybe we need to gather suggestions info for debugging?

        if group_field_id is not None:

            data_list = []

            for index, word, word_entry_id, word_group, single_list, group_list in suggestion_list:

                entry_id_list = (
                    [word_entry_id] +
                    [single_info[-1] for single_info in single_list] +
                    [group_info[-1] for group_info in group_list])

                data_list.append(
                    (group_field_id, entry_id_list))

            log.debug(
                '\ndebug data list:\n{}'.format(
                    pprint.pformat(data_list, width = 192)))

        # Showing and returning what we've got.

        log.debug(
            '\nsuggestion_list (length {0}):\n{1}'.format(
                len(suggestion_list),
                pprint.pformat(
                    suggestion_list, width = 192)))

        return suggestion_list

    @staticmethod
    def acoustic_data(
        base_language_id,
        sound_entity_id,
        sound_url,
        markup_entity_id,
        markup_url,
        storage,
        __debug_flag__ = False):
        """
        Extracts acoustic data from a pair of sound recording and its markup, using cache in a manner
        compatible with phonological analysis.
        """

        log_str = (
            'cognate_analysis {0}/{1}: sound entity {2}/{3}, markup entity {4}/{5}'.format(
                base_language_id[0], base_language_id[1],
                sound_entity_id[0], sound_entity_id[1],
                markup_entity_id[0], markup_entity_id[1]))

        textgrid_result_list = (

            process_sound_markup(
                log_str,
                sound_entity_id,
                sound_url,
                markup_entity_id,
                markup_url,
                storage,
                False,
                __debug_flag__))

        if textgrid_result_list is None:
            return None

        # Extracting info of the first vowel from the first vowel-containing tier of the sound/markup
        # analysis results.

        def f():

            for tier_index, tier_name, tier_result_list in textgrid_result_list:

                if (tier_result_list == 'no_vowel' or
                    tier_result_list == 'no_vowel_selected'):
                    continue

                for tier_result in tier_result_list:

                    for (interval_str, interval_r_length, p_mean, i_list, f_list,
                        sign_longest, sign_highest, source_index) in tier_result.interval_data_list:

                        interval_str_list = interval_str.split()

                        return [interval_str_list[0], interval_str_list[-3]] + f_list[:3]

        # Showing what we've finally got and returning it.

        result = f()

        log.debug(
            '{0}: {1}'.format(
            log_str, result))

        return result

    @staticmethod
    def graph_2d_embedding(d_ij, verbose = False):
        """
        Computes 2d embedding of a graph specified by non-negative simmetric distance matrix via stress
        minimization.

        Stress is based on relative strain for non-zero distances and absolute strain for zero distances.

        Let S_ij be source distances, D_ij be 2d distances, then stress is

          Sum[D_ij^2] for S_ij == 0 +
          Sum[D_ij^2 / S_ij^2 + S_ij^2 / D_ij^2] for S_ij > 0.

        Given D_ij^2 = (x_i - x_j)^2 + (y_i - y_j)^2, xy-gradient used for minimization can be computed
        using following:

          d[D_ij^2, x_i] = 2 (x_i - x_j)
          d[D_ij^2, x_j] = -2 (x_i - x_j)
          d[D_ij^2, y_i] = 2 (y_i - y_j)
          d[D_ij^2, y_j] = -2 (y_i - y_j)

        Obviously, d[D_ij^2 / S_ij^2, x_i] = 2 (x_i - x_j) / S_ij^2, and so on.

        And, with checking via WolframAlpha,

          d[S_ij^2 / D_ij^2, x_i] = -2 S_ij^2 (x_i - x_j) / D_ij^4, and so on.

        """

        N = numpy.size(d_ij, 0)

        def f(xy):
            """
            Computes stress given xy-coordinates.
            """

            x = xy[:N]
            y = xy[N:]

            result = 0.0

            for i in range(1, N):
                for j in range(i):

                    dr2 = (x[i] - x[j]) ** 2 + (y[i] - y[j]) ** 2

                    if d_ij[i,j] <= 0:
                        result += 4 * dr2

                    else:
                        d2_ij = d_ij[i,j] ** 2
                        result += dr2 / d2_ij + d2_ij / dr2

            return result

        def df(xy):
            """
            Computes gradient at the given xy-coordinates.
            """

            x = xy[:N]
            y = xy[N:]

            df_x = numpy.zeros(N)
            df_y = numpy.zeros(N)

            for i in range(1, N):
                for j in range(i):

                    dx = x[i] - x[j]
                    dy = y[i] - y[j]

                    dr2 = dx ** 2 + dy ** 2

                    if d_ij[i,j] <= 0:

                        df_x[i] += 4 * dx
                        df_x[j] -= 4 * dx

                        df_y[i] += 4 * dy
                        df_y[j] -= 4 * dy

                    else:

                        d2_ij = d_ij[i,j] ** 2
                        factor = (1 / d2_ij - d2_ij / dr2 ** 2)

                        df_x[i] += dx * factor
                        df_x[j] -= dx * factor

                        df_y[i] += dy * factor
                        df_y[j] -= dy * factor

            return numpy.concatenate((df_x, df_y))

        iter_count = 0

        def f_callback(xy):
            """
            Shows minimization progress, if enabled.
            """

            nonlocal iter_count

            log.debug(
                '\niteration {0}:\nxy:\n{1}\nf:\n{2}\ndf:\n{3}'.format(
                iter_count, xy, f(xy), df(xy)))

            iter_count += 1

        # Performing minization, returning minimization results.
        #
        # To get deterministic results we use the source distance matrix to seed the initial pseudo-random
        # coordinates we start optimization from.

        rng = (

            numpy.random.Generator(
                numpy.random.PCG64(

                    tuple(
                        hash(value)
                        for value in d_ij.flat))))

        result = (

            scipy.optimize.minimize(f,
                rng.random(N * 2),
                jac = df,
                callback = f_callback if verbose else None,
                options = {'disp': verbose}))

        result_x = numpy.stack((result.x[:N], result.x[N:])).T

        return result_x, f(result.x)

    @staticmethod
    def graph_3d_embedding(d_ij, verbose = False):
        """
        Computes 3d embedding of a graph specified by non-negative simmetric distance matrix via stress
        minimization.

        The same as with 2d embedding, see graph_2d_embedding.
        """

        N = numpy.size(d_ij, 0)
        N2 = N * 2

        def f(xyz):
            """
            Computes stress given xyz-coordinates.
            """

            x = xyz[:N]
            y = xyz[N:N2]
            z = xyz[N2:]

            result = 0.0

            for i in range(1, N):
                for j in range(i):

                    dr2 = (x[i] - x[j]) ** 2 + (y[i] - y[j]) ** 2 + (z[i] - z[j]) ** 2

                    if d_ij[i,j] <= 0:
                        result += 4 * dr2

                    else:
                        d2_ij = d_ij[i,j] ** 2
                        result += dr2 / d2_ij + d2_ij / dr2

            return result

        def df(xyz):
            """
            Computes gradient at a given xyz-coordinates.
            """

            x = xyz[:N]
            y = xyz[N:N2]
            z = xyz[N2:]

            df_x = numpy.zeros(N)
            df_y = numpy.zeros(N)
            df_z = numpy.zeros(N)

            for i in range(1, N):
                for j in range(i):

                    dx = x[i] - x[j]
                    dy = y[i] - y[j]
                    dz = z[i] - z[j]

                    dr2 = dx ** 2 + dy ** 2 + dz ** 2

                    if d_ij[i,j] <= 0:

                        df_x[i] += 4 * dx
                        df_x[j] -= 4 * dx

                        df_y[i] += 4 * dy
                        df_y[j] -= 4 * dy

                        df_z[i] += 4 * dz
                        df_z[j] -= 4 * dz

                    else:

                        d2_ij = d_ij[i,j] ** 2
                        factor = (1 / d2_ij - d2_ij / dr2 ** 2)

                        df_x[i] += dx * factor
                        df_x[j] -= dx * factor

                        df_y[i] += dy * factor
                        df_y[j] -= dy * factor

                        df_z[i] += dz * factor
                        df_z[j] -= dz * factor

            return numpy.concatenate((df_x, df_y, df_z))

        iter_count = 0

        def f_callback(xyz):
            """
            Shows minimization progress, if enabled.
            """

            nonlocal iter_count

            log.debug(
                '\niteration {0}:\nxyz:\n{1}\nf:\n{2}\ndf:\n{3}'.format(
                iter_count,
                numpy.stack((xyz[:N], xyz[N:N2], xyz[N2:])).T,
                f(xyz),
                df(xyz)))

            iter_count += 1

        # Performing minization, returning minimization results.
        #
        # To get deterministic results we use the source distance matrix to seed the initial pseudo-random
        # coordinates we start optimization from.

        rng = (

            numpy.random.Generator(
                numpy.random.PCG64(
                    
                    tuple(
                        hash(value)
                        for value in d_ij.flat))))

        result = (

            scipy.optimize.minimize(f,
                rng.random(N * 3),
                jac = df,
                callback = f_callback if verbose else None,
                options = {'disp': verbose}))

        result_x = numpy.stack((result.x[:N], result.x[N:N2], result.x[N2:])).T

        return result_x, f(result.x)

    @staticmethod
    def distance_graph(
            language_str,
            base_language_name,
            distance_data_array,
            distance_header_array,
            mode,
            storage,
            storage_dir,
            analysis_str = 'cognate_analysis',
            __debug_flag__ = False,
            __plot_flag__ = True):

        d_ij = (distance_data_array + distance_data_array.T) / 2

        log.debug(
            f'\n{analysis_str} {language_str}:'
            f'\ndistance_header_array:\n{distance_header_array}'
            f'\ndistance_data_array:\n{distance_data_array}'
            f'\nd_ij:\n{d_ij}')

        # Projecting the graph into a 2d plane via relative distance strain optimization, using PCA to
        # orient it left-right.

        if len(distance_data_array) > 1:

            embedding_2d, strain_2d = (
                CognateAnalysis.graph_2d_embedding(d_ij, verbose = __debug_flag__))

            embedding_2d_pca = (
                sklearn.decomposition.PCA(n_components = 2)
                    .fit_transform(embedding_2d))

            distance_2d = sklearn.metrics.euclidean_distances(embedding_2d)

        else:

            embedding_2d = numpy.zeros((1, 2))
            embedding_2d_pca = numpy.zeros((1, 2))

            strain_2d = 0.0

            distance_2d = numpy.zeros((1, 1))

        log.debug(
            f'\n{analysis_str} {language_str}:'
            f'\nembedding 2d:\n{embedding_2d}'
            f'\nembedding 2d (PCA-oriented):\n{embedding_2d_pca}'
            f'\nstrain 2d:\n{strain_2d}'
            f'\ndistances 2d:\n{distance_2d}')

        # And now the same with 3d embedding.

        if len(distance_data_array) > 1:

            embedding_3d, strain_3d = (
                CognateAnalysis.graph_3d_embedding(d_ij, verbose = __debug_flag__))

            # At least three points, standard PCA-based orientation.

            if len(distance_data_array) >= 3:

                embedding_3d_pca = (
                    sklearn.decomposition.PCA(n_components = 3)
                        .fit_transform(embedding_3d))

            # Only two points, so we take 2d embedding and extend it with zeros.

            else:

                embedding_3d_pca = (

                    numpy.hstack((
                        embedding_2d_pca,
                        numpy.zeros((embedding_2d_pca.shape[0], 1)))))

            # Making 3d embedding actually 3d, if required.

            if embedding_3d_pca.shape[1] <= 2:

                embedding_3d_pca = (

                    numpy.hstack((
                        embedding_3d_pca,
                        numpy.zeros((embedding_3d_pca.shape[0], 1)))))

            distance_3d = (
                sklearn.metrics.euclidean_distances(embedding_3d_pca))

        else:

            embedding_3d = numpy.zeros((1, 3))
            embedding_3d_pca = numpy.zeros((1, 3))

            strain_3d = 0.0

            distance_3d = numpy.zeros((1, 1))

        log.debug(
            f'\n{analysis_str} {language_str}:'
            f'\nembedding 3d:\n{embedding_3d}'
            f'\nembedding 3d (PCA-oriented):\n{embedding_3d_pca}'
            f'\nstrain 3d:\n{strain_3d}'
            f'\ndistances 3d:\n{distance_3d}')

        # Computing minimum spanning tree via standard Jarnik-Prim-Dijkstra algorithm using 2d and 3d
        # embedding distances to break ties.

        if len(distance_data_array) <= 1:
            mst_list = []

        else:

            d_min, d_extra_min, min_i, min_j = min(
                (d_ij[i,j], distance_2d[i,j] + distance_3d[i,j], i, j)
                for i in range(d_ij.shape[0] - 1)
                for j in range(i + 1, d_ij.shape[0]))

            mst_list = [(min_i, min_j)]
            mst_dict = {}

            # MST construction initialization.

            for i in range(d_ij.shape[0]):

                if i == min_i or i == min_j:
                    continue

                d_min_i = (d_ij[i, min_i], distance_2d[i, min_i] + distance_3d[i, min_i])
                d_min_j = (d_ij[i, min_j], distance_2d[i, min_j] + distance_3d[i, min_j])

                mst_dict[i] = (
                    (d_min_i, min_i) if d_min_i <= d_min_j else
                    (d_min_j, min_i))

            # Iterative MST construction.

            while len(mst_dict) > 0:

                (d_min, d_extra_min, i_min, i_from_min) = min(
                    (d, d_extra, i, i_from) for i, ((d, d_extra), i_from) in mst_dict.items())

                log.debug('\n' + pprint.pformat(mst_dict))
                log.debug('\n' + repr((i_from_min, i_min, d_min, d_extra_min)))

                mst_list.append((i_from_min, i_min))
                del mst_dict[i_min]

                # Updating shortest connection info.

                for i_to in mst_dict.keys():

                    d_to = (d_ij[i_min, i_to], distance_2d[i_min, i_to] + distance_3d[i_min, i_to])

                    if d_to < mst_dict[i_to][0]:
                        mst_dict[i_to] = (d_to, i_min)

        log.debug(
            f'\n{analysis_str} {language_str}:'
            f'\nminimum spanning tree:\n{pprint.pformat(mst_list)}')

        # Plotting with matplotlib.
        figure_url = None
        if __plot_flag__:

            figure = pyplot.figure(figsize = (10, 10))
            axes = figure.add_subplot(212)

            axes.set_title(
                'Etymological distance tree (relative distance embedding)',
                fontsize = 14, family = 'Gentium')

            axes.axis('equal')
            axes.axis('off')
            axes.autoscale()

            def f(axes, embedding_pca):
                """
                Plots specified graph embedding on a given axis.
                """

                flag_3d = numpy.size(embedding_pca, 1) > 2

                for index, (position, name) in enumerate(
                    zip(embedding_pca, distance_header_array)):

                    # Checking if any of the previous perspectives are already in this perspective's
                    # position.

                    same_position_index = None

                    for i, p in enumerate(embedding_pca[:index]):
                        if numpy.linalg.norm(position - p) <= 1e-3:

                            same_position_index = i
                            break

                    color = matplotlib.colors.hsv_to_rgb(
                        [(same_position_index or index) * 1.0 / len(distance_header_array), 0.5, 0.75])

                    label_same_str = (
                        '' if same_position_index is None else
                        ' (same as {0})'.format(same_position_index + 1))

                    kwargs = {
                        's': 35,
                        'color': color,
                        'label': '{0}) {1}{2}'.format(index + 1, name, label_same_str)}

                    axes.scatter(*position, **kwargs)

                    # Annotating position with its number, but only if we hadn't already annotated nearby.

                    if same_position_index is None:

                        if flag_3d:

                            axes.text(
                                position[0] + 0.01, position[1], position[2] + 0.01,
                                str(index + 1), None, fontsize = 14)

                        else:

                            axes.annotate(
                                str(index + 1),
                                (position[0] + 0.01, position[1] - 0.005),
                                fontsize = 14)

                # Plotting minimum spanning trees.

                line_list = [
                    (embedding_pca[i], embedding_pca[j])
                    for i, j in mst_list]

                line_collection = (
                    Line3DCollection if flag_3d else LineCollection)(
                        line_list, zorder = 0, color = 'gray')

                axes.add_collection(line_collection)

                pyplot.setp(axes.texts, family = 'Gentium')

            # Plotting our embedding, creating the legend.

            f(axes, embedding_2d_pca)

            pyplot.tight_layout()

            legend = axes.legend(
                scatterpoints = 1,
                loc = 'upper center',
                bbox_to_anchor = (0.5, -0.05),
                frameon = False,
                handlelength = 0.5,
                handletextpad = 0.75,
                fontsize = 14)

            pyplot.setp(legend.texts, family = 'Gentium')
            axes.autoscale_view()

            # Saving generated figure for debug purposes, if required.

            if __debug_flag__:

                figure_file_name = (
                    'figure cognate distance{0}.png'.format(
                    mode_name_str))

                with open(figure_file_name, 'wb') as figure_file:

                    pyplot.savefig(
                        figure_file,
                        bbox_extra_artists = (legend,),
                        bbox_inches = 'tight',
                        pad_inches = 0.25,
                        format = 'png')

                # Also generating 3d embedding figure.

                figure_3d = pyplot.figure()
                figure_3d.set_size_inches(16, 10)

                axes_3d = figure_3d.add_subplot(111, projection = '3d')

                axes_3d.axis('equal')
                axes_3d.view_init(elev = 30, azim = -75)

                f(axes_3d, embedding_3d_pca)

                # Setting up legend.

                axes_3d.set_xlabel('X')
                axes_3d.set_ylabel('Y')
                axes_3d.set_zlabel('Z')

                legend_3d = axes_3d.legend(
                    scatterpoints = 1,
                    loc = 'upper center',
                    bbox_to_anchor = (0.5, -0.05),
                    frameon = False,
                    handlelength = 0.5,
                    handletextpad = 0.75,
                    fontsize = 14)

                pyplot.setp(legend_3d.texts, family = 'Gentium')

                # Fake cubic bounding box to force axis aspect ratios, see
                # https://stackoverflow.com/a/13701747/2016856.

                X = embedding_3d_pca[:,0]
                Y = embedding_3d_pca[:,1]
                Z = embedding_3d_pca[:,2]

                max_range = numpy.array([
                    X.max() - X.min(), Y.max() - Y.min(), Z.max() - Z.min()]).max()

                Xb = (
                    0.5 * max_range * numpy.mgrid[-1:2:2,-1:2:2,-1:2:2][0].flatten() +
                    0.5 * (X.max() + X.min()))

                Yb = (
                    0.5 * max_range * numpy.mgrid[-1:2:2,-1:2:2,-1:2:2][1].flatten() +
                    0.5 * (Y.max() + Y.min()))

                Zb = (
                    0.5 * max_range * numpy.mgrid[-1:2:2,-1:2:2,-1:2:2][2].flatten() +
                    0.5 * (Z.max() + Z.min()))

                for xb, yb, zb in zip(Xb, Yb, Zb):
                   axes_3d.plot([xb], [yb], [zb], 'w')

                axes_3d.autoscale_view()

                # And saving it.

                figure_3d_file_name = (
                    'figure 3d cognate distance{0}.png'.format(
                    mode_name_str))

                with open(figure_3d_file_name, 'wb') as figure_3d_file:

                    figure_3d.savefig(
                        figure_3d_file,
                        bbox_extra_artists = (legend_3d,),
                        bbox_inches = 'tight',
                        pad_inches = 0.25,
                        format = 'png')

            # Storing generated figure as a PNG image.
            current_datetime = datetime.datetime.now(datetime.timezone.utc)
            figure_filename = pathvalidate.sanitize_filename(
                '{0} cognate{1} analysis {2:04d}.{3:02d}.{4:02d}.png'.format(
                    base_language_name[:64],
                    ' ' + mode if mode else '',
                    current_datetime.year,
                    current_datetime.month,
                    current_datetime.day))

            figure_path = os.path.join(storage_dir, figure_filename)
            os.makedirs(os.path.dirname(figure_path), exist_ok = True)

            with open(figure_path, 'wb') as figure_file:

                figure.savefig(
                    figure_file,
                    bbox_extra_artists = (legend,),
                    bbox_inches = 'tight',
                    pad_inches = 0.25,
                    format = 'png')

            cur_time = time.time()
            figure_url = ''.join([
                storage['prefix'], storage['static_route'],
                'cognate', '/', str(cur_time), '/', figure_filename])
        ### Plotting with matplotlib ends

        return (
            figure_url,
            mst_list,
            embedding_2d_pca,
            embedding_3d_pca
        )

    @staticmethod
    def perform_cognate_analysis(
        language_str,
        source_perspective_id,
        base_language_id,
        base_language_name,
        group_field_id,
        perspective_info_list,
        multi_list,
        multi_name_list,
        mode,
        distance_flag,
        reference_perspective_id,
        figure_flag,
        distance_vowel_flag,
        distance_consonant_flag,
        match_translations_value,
        only_orphans_flag,
        locale_id,
        storage,
        task_status = None,
        __debug_flag__ = False,
        __intermediate_flag__ = False):
        """
        Performs cognate analysis in either synchronous or asynchronous mode.
        """

        __result_flag__ = False

        if task_status is not None:
            task_status.set(1, 0, 'Gathering grouping data')

        # Sometimes in debugging mode we should return already computed results.

        if __debug_flag__:

            tag_data_digest = (

                hashlib.md5(

                    repr(list(group_field_id) +
                        [perspective_info[0] for perspective_info in perspective_info_list])

                    .encode('utf-8'))

                .hexdigest())

            result_file_name = (

                '__result_{0}_{1}__.gz'.format(

                    'multi{0}'.format(len(multi_list))
                        if mode == 'multi' else
                        '{0}_{1}'.format(*base_language_id),

                    tag_data_digest))

            if __result_flag__ and os.path.exists(result_file_name):

                with gzip.open(
                    result_file_name, 'rb') as result_file:

                    result_dict = pickle.load(result_file)

                return CognateAnalysis(**result_dict)

        # Gathering entry grouping data.

        perspective_dict = collections.defaultdict(dict)

        entry_already_set = set()
        group_list = []

        text_dict = {}
        entry_id_dict = {}

        if not __debug_flag__:

            entry_already_set, group_list, group_time = (

                CognateAnalysis.tag_data_plpgsql(
                    perspective_info_list, group_field_id))

        else:

            # If we are in debug mode, we try to load existing tag data to reduce debugging time.

            tag_data_file_name = (

                '__tag_data_{0}_{1}__.gz'.format(

                    'multi{0}'.format(len(multi_list))
                        if mode == 'multi' else
                        '{0}_{1}'.format(*base_language_id),

                    tag_data_digest))

            # Checking if we have saved data.

            if os.path.exists(tag_data_file_name):

                with gzip.open(tag_data_file_name, 'rb') as tag_data_file:
                    entry_already_set, group_list, group_time = pickle.load(tag_data_file)

            else:

                # Don't have existing data, so we gather it and then save it for later use.

                entry_already_set, group_list, group_time = (

                    CognateAnalysis.tag_data_plpgsql(
                        perspective_info_list, group_field_id))

                with gzip.open(tag_data_file_name, 'wb') as tag_data_file:
                    pickle.dump((entry_already_set, group_list, group_time), tag_data_file)

        log.debug(
            '\ncognate_analysis {0}:'
            '\n{1} entries, {2} groups, {3:.2f}s elapsed time'.format(
            language_str,
            len(entry_already_set),
            len(group_list),
            group_time))

        if task_status is not None:
            task_status.set(2, 5, 'Gathering analysis source data')

        # Getting text data for each perspective.

        dbSound = aliased(dbEntity, name = 'Sound')
        dbMarkup = aliased(dbEntity, name = 'Markup')

        dbPublishingSound = aliased(dbPublishingEntity, name = 'PublishingSound')
        dbPublishingMarkup = aliased(dbPublishingEntity, name = 'PublishingMarkup')

        phonemic_data_list = []
        suggestions_data_list = []

        sg_total_count = 0
        sg_xcript_count = 0
        sg_xlat_count = 0
        sg_both_count = 0

        source_perspective_index = None
        for index, (perspective_id, transcription_field_id, translation_field_id) in \
            enumerate(perspective_info_list):

            if perspective_id == source_perspective_id:
                source_perspective_index = index

            # Getting and saving perspective info.

            perspective = DBSession.query(dbPerspective).filter_by(
                client_id = perspective_id[0], object_id = perspective_id[1]).first()

            perspective_name = perspective.get_translation(locale_id)
            dictionary_name = perspective.parent.get_translation(locale_id)

            transcription_rules = (
                '' if not perspective.additional_metadata else
                    perspective.additional_metadata.get('transcription_rules', ''))

            perspective_data = perspective_dict[perspective_id]

            perspective_data['perspective_name'] = perspective_name
            perspective_data['dictionary_name'] = dictionary_name
            perspective_data['transcription_rules'] = transcription_rules

            # Preparing to save additional data, if required.

            if mode == 'phonemic':

                phonemic_data_list.append([
                    '{0} - {1}'.format(dictionary_name, perspective_name), ''])

            elif mode == 'suggestions':

                suggestions_data_list.append([])

            log.debug(
                '\ncognate_analysis {0}:'
                '\n  dictionary {1}/{2}: {3}'
                '\n  perspective {4}/{5}: {6}'
                '\n  transcription_rules: {7}'.format(
                language_str,
                perspective.parent_client_id, perspective.parent_object_id,
                repr(dictionary_name.strip()),
                perspective_id[0], perspective_id[1],
                repr(perspective_name.strip()),
                repr(transcription_rules)))

            # Getting text data.

            transcription_query = (

                DBSession.query(
                    dbLexicalEntry.client_id,
                    dbLexicalEntry.object_id).filter(
                        dbLexicalEntry.parent_client_id == perspective_id[0],
                        dbLexicalEntry.parent_object_id == perspective_id[1],
                        dbLexicalEntry.marked_for_deletion == False,
                        dbEntity.parent_client_id == dbLexicalEntry.client_id,
                        dbEntity.parent_object_id == dbLexicalEntry.object_id,
                        dbEntity.field_client_id == transcription_field_id[0],
                        dbEntity.field_object_id == transcription_field_id[1],
                        dbEntity.marked_for_deletion == False,
                        dbPublishingEntity.client_id == dbEntity.client_id,
                        dbPublishingEntity.object_id == dbEntity.object_id,
                        dbPublishingEntity.published == True,
                        dbPublishingEntity.accepted == True)

                .add_columns(
                    func.array_agg(dbEntity.content).label('transcription'))

                .group_by(dbLexicalEntry)).subquery()

            translation_query = (

                DBSession.query(
                    dbLexicalEntry.client_id,
                    dbLexicalEntry.object_id).filter(
                        dbLexicalEntry.parent_client_id == perspective_id[0],
                        dbLexicalEntry.parent_object_id == perspective_id[1],
                        dbLexicalEntry.marked_for_deletion == False,
                        dbEntity.parent_client_id == dbLexicalEntry.client_id,
                        dbEntity.parent_object_id == dbLexicalEntry.object_id,
                        dbEntity.field_client_id == translation_field_id[0],
                        dbEntity.field_object_id == translation_field_id[1],
                        dbEntity.marked_for_deletion == False,
                        dbPublishingEntity.client_id == dbEntity.client_id,
                        dbPublishingEntity.object_id == dbEntity.object_id,
                        dbPublishingEntity.published == True,
                        dbPublishingEntity.accepted == True)

                .add_columns(
                    func.array_agg(dbEntity.content).label('translation'))

                .group_by(dbLexicalEntry)).subquery()

            # Main query for transcription/translation data.

            data_query = (
                DBSession.query(transcription_query)

                .outerjoin(translation_query, and_(
                    transcription_query.c.client_id == translation_query.c.client_id,
                    transcription_query.c.object_id == translation_query.c.object_id))

                .add_columns(
                    translation_query.c.translation))

            # If we need to do an acoustic analysis, we also get sound/markup data.

            if mode == 'acoustic':

                sound_markup_query = (

                    DBSession.query(
                        dbLexicalEntry.client_id,
                        dbLexicalEntry.object_id).filter(
                            dbLexicalEntry.parent_client_id == perspective_id[0],
                            dbLexicalEntry.parent_object_id == perspective_id[1],
                            dbLexicalEntry.marked_for_deletion == False,
                            dbMarkup.parent_client_id == dbLexicalEntry.client_id,
                            dbMarkup.parent_object_id == dbLexicalEntry.object_id,
                            dbMarkup.marked_for_deletion == False,
                            dbMarkup.additional_metadata.contains({'data_type': 'praat markup'}),
                            dbPublishingMarkup.client_id == dbMarkup.client_id,
                            dbPublishingMarkup.object_id == dbMarkup.object_id,
                            dbPublishingMarkup.published == True,
                            dbPublishingMarkup.accepted == True,
                            dbSound.client_id == dbMarkup.self_client_id,
                            dbSound.object_id == dbMarkup.self_object_id,
                            dbSound.marked_for_deletion == False,
                            dbPublishingSound.client_id == dbSound.client_id,
                            dbPublishingSound.object_id == dbSound.object_id,
                            dbPublishingSound.published == True,
                            dbPublishingSound.accepted == True)

                    .add_columns(

                        func.jsonb_agg(func.jsonb_build_array(
                            dbSound.client_id, dbSound.object_id, dbSound.content,
                            dbMarkup.client_id, dbMarkup.object_id, dbMarkup.content))

                        .label('sound_markup'))

                    .group_by(dbLexicalEntry)).subquery()

                # Adding sound/markup retrieval to the main query.

                data_query = (
                    data_query

                    .outerjoin(sound_markup_query, and_(
                        transcription_query.c.client_id == sound_markup_query.c.client_id,
                        transcription_query.c.object_id == sound_markup_query.c.object_id))

                    .add_columns(
                        sound_markup_query.c.sound_markup))

            # If we are in asynchronous mode, we need to look up how many data rows we need
            # to process for this perspective.

            if task_status is not None:

                row_count = data_query.count()

                log.debug(
                    'cognate_analysis {0}: perspective {1}/{2}: {3} data rows'.format(
                    language_str,
                    perspective_id[0], perspective_id[1],
                    row_count))

            # Grouping transcriptions and translations by lexical entries.

            for row_index, row in enumerate(data_query.all()):

                entry_id = tuple(row[:2])
                transcription_list, translation_list = row[2:4]

                transcription_list = (
                    [] if not transcription_list else [
                        transcription.strip()
                        for transcription in transcription_list
                        if transcription.strip()])

                # If we have no trascriptions for this lexical entry, we skip it altogether.

                if not transcription_list:
                    continue

                translation_list = (
                    [] if not translation_list else [
                        translation.strip()
                        for translation in translation_list
                        if translation.strip()])

                # Saving transcription / translation data.

                translation_str = (
                    translation_list[0] if translation_list else '')

                if mode == 'phonemic':

                    for transcription in transcription_list:
                        phonemic_data_list[-1].extend([transcription, translation_str])

                elif mode == 'suggestions' and entry_id not in entry_already_set:

                    suggestions_data_list[-1].append([
                        '|'.join(transcription_list),
                        '|'.join(translation_list)])

                    sg_total_count += 1

                    # Counting how many instances of more than one transcription and / or translation
                    # we have.

                    if len(transcription_list) > 1:
                        sg_xcript_count += 1

                    if len(translation_list) > 1:
                        sg_xlat_count += 1

                    if len(transcription_list) > 1 and len(translation_list) > 1:
                        sg_both_count += 1

                # If we are fetching additional acoustic data, it's possible we have to process
                # sound recordings and markup this lexical entry has.

                if len(row) > 4 and row[4]:

                    row_list = row[4][0]

                    result = (
                        CognateAnalysis.acoustic_data(
                            base_language_id,
                            tuple(row_list[0:2]), row_list[2],
                            tuple(row_list[3:5]), row_list[5],
                            storage,
                            __debug_flag__))

                    # Updating task progress, if required.

                    if task_status is not None:

                        percent = int(math.floor(90.0 *
                            (index + float(row_index + 1) / row_count) /
                            len(perspective_info_list)))

                        task_status.set(2, 5 + percent, 'Gathering analysis source data')

                    entry_data_list = (index,
                        transcription_list,
                        translation_list,
                        result)

                # No additional acoustic data.

                else:
                    entry_data_list = (index, transcription_list, translation_list)

                text_dict[entry_id] = entry_data_list

                entry_id_key = (

                    index,
                    '|'.join(transcription_list) + (
                        ' ʽ' + '|'.join(translation_list) + 'ʼ' if translation_list else ''))

                entry_id_dict[entry_id_key] = entry_id

        # Showing some info on non-grouped entries, if required.

        if mode == 'suggestions':

            log.debug(
                '\ncognate_analysis {0}:'
                '\n{1} non-grouped entries'
                '\n{2} with multiple transcriptions'
                '\n{3} with multiple translations'
                '\n{4} with multiple transcriptions and translations'.format(
                language_str,
                sg_total_count,
                sg_xcript_count,
                sg_xlat_count,
                sg_both_count))

            # Also, if we are computing cognate suggestions, we should have a valid source perspective, it's
            # an error otherwise.

            if source_perspective_index is None:

                return ResponseError(message =
                    'Cognate suggestions require that the source perspective '
                    'is among the ones being analyzed.')

        if task_status is not None:
            task_status.set(3, 95, 'Performing analysis')

        # Ok, and now we form the source data for analysis.

        result_list = [[]]

        perspective_id_list = []
        perspective_name_list = []

        for perspective_id, transcription_field_id, translation_field_id in perspective_info_list:

            perspective_id_list.append(perspective_id)
            perspective_data = perspective_dict[perspective_id]

            perspective_str = '{0} - {1}'.format(
                perspective_data['dictionary_name'],
                perspective_data['perspective_name'])

            perspective_name_list.append(perspective_str)

            # Also going to use transcription transformation rules.

            result_list[0].extend([
                perspective_str,
                perspective_data['transcription_rules']])

        log.debug(
            '\ncognate_analysis {0}:'
            '\nsource_perspective_index: {1}'
            '\nperspective_list:\n{2}'
            '\nheader_list:\n{3}'.format(
            language_str,
            source_perspective_index,
            pprint.pformat(perspective_name_list, width = 108),
            pprint.pformat(result_list[0], width = 108)))

        # Each group of lexical entries.

        not_enough_count = 0

        total_transcription_count = 0
        total_translation_count = 0

        not_suggestions = mode != 'suggestions'

        for entry_id_set in group_list:

            group_entry_id_list = [[]
                for i in range(len(perspective_info_list))]

            group_transcription_list = [[]
                for i in range(len(perspective_info_list))]

            group_translation_list = [[]
                for i in range(len(perspective_info_list))]

            group_acoustic_list = [None
                for i in range(len(perspective_info_list))]

            transcription_count = 0
            translation_count = 0

            for entry_id in entry_id_set:

                if entry_id not in text_dict:
                    continue

                # Processing text data of each entry of the group.

                entry_data_list = text_dict[entry_id]

                (index,
                    transcription_list,
                    translation_list) = (

                    entry_data_list[:3])

                group_entry_id_list[index].append(entry_id)

                group_transcription_list[index].extend(transcription_list)
                group_translation_list[index].extend(translation_list)

                transcription_count += len(transcription_list)
                translation_count += len(translation_list)

                if (len(entry_data_list) > 3 and
                    entry_data_list[3] and
                    group_acoustic_list[index] is None):

                    group_acoustic_list[index] = entry_data_list[3]

            # Dropping groups with transcriptions from no more than a single dictionary, if required.

            if (not_suggestions and
                sum(min(1, len(transcription_list))
                    for transcription_list in group_transcription_list) <= 1):

                not_enough_count += 1
                continue

            total_transcription_count += transcription_count
            total_translation_count += translation_count

            result_list.append([])

            group_zipper = zip(
                group_entry_id_list,
                group_transcription_list,
                group_translation_list,
                group_acoustic_list)

            # Forming row of the source data table based on the entry group.

            for (
                index, (
                    entry_id_list,
                    transcription_list,
                    translation_list,
                    acoustic_list)) in (

                enumerate(group_zipper)):

                transcription_str = '|'.join(transcription_list)
                translation_str = '|'.join(translation_list)

                result_list[-1].append(transcription_str)
                result_list[-1].append(translation_str)

                if mode == 'acoustic':
                    result_list[-1].extend(acoustic_list or ['', '', '', '', ''])

                # Saving mapping from the translation / transcription info string to an id of one entry of
                # the group.

                if transcription_list or translation_list:

                    entry_id_key = (

                        index,
                        transcription_str + (
                            ' ʽ' + translation_str + 'ʼ' if translation_str else ''))

                    entry_id_dict[entry_id_key] = entry_id_list[0]

        # Showing what we've gathered.

        log.debug(
            '\ncognate_analysis {0}:'
            '\n  len(group_list): {1}'
            '\n  len(result_list): {2}'
            '\n  not_enough_count: {3}'
            '\n  transcription_count: {4}'
            '\n  translation_count: {5}'
            '\n  result_list:\n{6}'.format(
                language_str,
                len(group_list),
                len(result_list),
                not_enough_count,
                total_transcription_count,
                total_translation_count,
                pprint.pformat(result_list, width = 108)))

        # If we have no data at all, we return empty result.

        if len(result_list) <= 1 and not_suggestions:

            return CognateAnalysis(
                triumph = True,
                dictionary_count = len(perspective_info_list),
                group_count = len(group_list),
                not_enough_count = not_enough_count,
                transcription_count = total_transcription_count,
                translation_count = total_translation_count,
                result = '',
                xlsx_url = '',
                distance_list = [],
                figure_url = '',
                intermediate_url_list = None)

        analysis_f = (
            cognate_acoustic_analysis_f if mode == 'acoustic' else
            cognate_reconstruction_f if mode == 'reconstruction' else
            cognate_reconstruction_multi_f if mode == 'multi' else
            cognate_suggestions_f if mode == 'suggestions' else
            cognate_analysis_f)

        # Preparing analysis input.

        phonemic_input_list = [
            ''.join(text + '\0' for text in text_list)
            for text_list in phonemic_data_list]

        suggestions_result_list = []

        for tt_list in itertools.zip_longest(
            *suggestions_data_list, fillvalue = ['', '']):

            suggestions_result_list.append([])

            for tt in tt_list:
                suggestions_result_list[-1].extend(tt)

        if mode == 'suggestions':

            # Showing additional ungrouped input data, if required.

            log.debug(
                '\ncognate_analysis {0}:'
                '\nsuggestions_result_list:\n{1}'.format(
                    language_str,
                    pprint.pformat(suggestions_result_list, width = 144)))

        result_input = (

            ''.join(
                ''.join(text + '\0' for text in text_list)

                for text_list in (
                    result_list + suggestions_result_list)))

        input = '\0'.join(phonemic_input_list + [result_input])

        log.debug(
            '\ncognate_analysis {0}:'
            '\nanalysis_f: {1}'
            '\ninput ({2} columns, {3} rows{4}):\n{5}'.format(
                language_str,
                repr(analysis_f),
                len(perspective_info_list),
                len(result_list),
                '' if mode != 'suggestions' else
                    ', {0} ungrouped rows'.format(len(suggestions_result_list)),
                pprint.pformat([input[i : i + 256]
                    for i in range(0, len(input), 256)], width = 144)))

        # Saving input to a file, if required.

        storage_dir = None
        intermediate_url_list = []

        if __debug_flag__ or __intermediate_flag__:

            language_name_str = (
                ' '.join(multi_name_list) if mode == 'multi' else
                base_language_name.strip())

            if len(language_name_str) > 64:
                language_name_str = language_name_str[:64] + '...'

            mode_name_str = (

                '{0} {1} {2} {3}{4}'.format(

                    ' multi{0}'.format(len(multi_list))
                        if mode == 'multi' else
                        (' ' + mode if mode else ''),

                    language_name_str,

                    ' '.join(str(count) for id, count in multi_list)
                        if mode == 'multi' else
                        len(perspective_info_list),

                    len(result_list),

                    '' if not_suggestions else
                        ' {} {} {} {}'.format(
                            len(suggestions_result_list),
                            source_perspective_index,
                            match_translations_value,
                            int(only_orphans_flag))))

            cognate_name_str = (
                'cognate' + mode_name_str)

            # Initializing file storage directory, if required.

            if __intermediate_flag__ and storage_dir is None:

                cur_time = time.time()

                storage_dir = os.path.join(
                    storage['path'], 'cognate', str(cur_time))

            for extension, encoding in ('utf8', 'utf-8'), ('utf16', 'utf-16'):

                input_file_name = (

                    pathvalidate.sanitize_filename(
                        'input {0}.{1}'.format(
                            cognate_name_str, extension)))

                # Saving to the working directory...

                if __debug_flag__:

                    with open(input_file_name, 'wb') as input_file:
                        input_file.write(input.encode(encoding))

                # ...and / or to the file storage.

                if __intermediate_flag__:

                    input_path = os.path.join(
                        storage_dir, input_file_name)

                    os.makedirs(
                        os.path.dirname(input_path),
                        exist_ok = True)

                    with open(input_path, 'wb') as input_file:
                        input_file.write(input.encode(encoding))

                    input_url = ''.join([
                        storage['prefix'],
                        storage['static_route'],
                        'cognate', '/',
                        str(cur_time), '/',
                        input_file_name])

                    intermediate_url_list.append(input_url)

        # Calling analysis library, starting with getting required output buffer size and continuing
        # with analysis proper.

        if mode == 'multi':

            multi_count_list = [
                perspective_count
                for language_id, perspective_count in multi_list]

            perspective_count_array = (
                ctypes.c_int * len(multi_list))(*multi_count_list)

            # int CognateMultiReconstruct_GetAllOutput(
            #   LPTSTR bufIn, int* pnCols, int nGroups, int nRows, LPTSTR bufOut, int flags)

            output_buffer_size = analysis_f(
                None,
                perspective_count_array,
                len(multi_list),
                len(result_list),
                None,
                1)

        elif mode == 'suggestions':

            # int GuessCognates_GetAllOutput(
            #   LPTSTR bufIn, int nCols, int nRowsCorresp, int nRowsRest, int iDictThis, int lookMeaning,
            #   int onlyOrphans, LPTSTR bufOut, int flags)

            output_buffer_size = analysis_f(
                None,
                len(perspective_info_list),
                len(result_list),
                len(suggestions_result_list),
                source_perspective_index,
                match_translations_value,
                int(only_orphans_flag),
                None,
                1)

        else:

            # int CognateAnalysis_GetAllOutput(
            #   LPTSTR bufIn, int nCols, int nRows, LPTSTR bufOut, int flags)

            output_buffer_size = analysis_f(
                None,
                len(perspective_info_list),
                len(result_list),
                None,
                1)

        log.debug(
            '\ncognate_analysis {0}: output buffer size {1}'.format(
            language_str,
            output_buffer_size))

        input_buffer = ctypes.create_unicode_buffer(input)

        # Saving input buffer to a file, if required.

        if __debug_flag__:

            input_file_name = (
                'input {0}.buffer'.format(
                    cognate_name_str))

            with open(input_file_name, 'wb') as input_file:
                input_file.write(bytes(input_buffer))

        output_buffer = ctypes.create_unicode_buffer(output_buffer_size + 256)

        if mode == 'multi':

            result = analysis_f(
                input_buffer,
                perspective_count_array,
                len(multi_list),
                len(result_list),
                output_buffer,
                1)

        elif mode == 'suggestions':

            result = analysis_f(
                input_buffer,
                len(perspective_info_list),
                len(result_list),
                len(suggestions_result_list),
                source_perspective_index,
                match_translations_value,
                int(only_orphans_flag),
                output_buffer,
                1)

        else:

            result = analysis_f(
                input_buffer,
                len(perspective_info_list),
                len(result_list),
                output_buffer,
                1)

        log.debug(
            '\ncognate_analysis {0}: result {1}'.format(
            language_str,
            result))

        # If we don't have a good result, we return an error.

        if result <= 0:

            if task_status is not None:

                task_status.set(5, 100,
                    'Finished (ERROR): library call error {0}'.format(result))

            return ResponseError(message =
                'Cognate analysis library call error {0}'.format(result))

        output = output_buffer.value

        log.debug(
            '\ncognate_analysis {}:\noutput ({}):\n{}'.format(
            language_str,
            len(output),
            pprint.pformat([output[i : i + 256]
                for i in range(0, len(output), 256)], width = 144)))

        # Saving output buffer and output to files, if required.

        if __debug_flag__:

            output_file_name = (
                'output {0}.buffer'.format(
                    cognate_name_str))

            with open(output_file_name, 'wb') as output_file:
                output_file.write(bytes(output_buffer))

            for extension, encoding in ('utf8', 'utf-8'), ('utf16', 'utf-16'):

                output_file_name = (
                    'output {0}.{1}'.format(
                        cognate_name_str,
                        extension))

                with open(output_file_name, 'wb') as output_file:
                    output_file.write(output.encode(encoding))

        # Reflowing output.

        line_list = output.split('\r\n')

        text_wrapper = textwrap.TextWrapper(
            width = max(196, len(perspective_info_list) * 40), tabsize = 20)

        reflow_list = []

        for line in line_list:
            reflow_list.extend(text_wrapper.wrap(line))

        wrapped_output = '\n'.join(reflow_list)

        log.debug(
            'cognate_analysis {0}:\nwrapped output:\n{1}'.format(
            language_str,
            wrapped_output))

        # Getting binary output for parsing and exporting.

        if mode == 'multi':

            result_binary = analysis_f(
                input_buffer,
                perspective_count_array,
                len(multi_list),
                len(result_list),
                output_buffer,
                2)

        # If we are in the suggestions mode, we currently just return the output.

        elif mode == 'suggestions':

            result_binary = analysis_f(
                input_buffer,
                len(perspective_info_list),
                len(result_list),
                len(suggestions_result_list),
                source_perspective_index,
                match_translations_value,
                int(only_orphans_flag),
                output_buffer,
                2)

        else:

            result_binary = analysis_f(
                input_buffer,
                len(perspective_info_list),
                len(result_list),
                output_buffer,
                2)

        log.debug(
            'cognate_analysis {0}: result_binary {1}'.format(
            language_str,
            result_binary))

        if result_binary <= 0:

            if task_status is not None:

                task_status.set(5, 100,
                    'Finished (ERROR): library call (binary) error {0}'.format(result_binary))

            return ResponseError(message =
                'Cognate analysis library call (binary) error {0}'.format(result_binary))

        # Showing what we've got from the binary output call.

        output_binary = output_buffer[:result_binary]

        output_binary_list = [
            output_binary[i : i + 256]
            for i in range(0, len(output_binary), 256)]

        log.debug(
            '\ncognate_analysis {0}:'
            '\noutput_binary:\n{1}'.format(
            language_str,
            pprint.pformat(
                output_binary_list, width = 144)))

        # Saving binary output buffer and binary output to files, if required.

        if __debug_flag__:

            output_file_name = (
                'output binary {0}.buffer'.format(
                    cognate_name_str))

            with open(
                output_file_name, 'wb') as output_file:

                output_file.write(
                    bytes(output_buffer))

            for extension, encoding in ('utf8', 'utf-8'), ('utf16', 'utf-16'):

                output_file_name = (
                    'output binary {0}.{1}'.format(
                        cognate_name_str, extension))

                with open(
                    output_file_name, 'wb') as output_file:

                    output_file.write(
                        output_binary.encode(encoding))

        # For cognate suggestions we just parse and return suggestions.

        if mode == 'suggestions':

            suggestion_list = (

                CognateAnalysis.parse_suggestions(
                    language_str,
                    output_binary,
                    len(perspective_info_list),
                    source_perspective_index,
                    entry_id_dict,
                    __debug_flag__,
                    cognate_name_str if __debug_flag__ else None,
                    group_field_id if __debug_flag__ else None))

            result_dict = (

                dict(

                    triumph = True,

                    dictionary_count = len(perspective_info_list),
                    group_count = len(group_list),
                    not_enough_count = not_enough_count,
                    transcription_count = total_transcription_count,
                    translation_count = total_translation_count,

                    result = output,

                    perspective_name_list = perspective_name_list,

                    suggestion_list = suggestion_list,
                    suggestion_field_id = group_field_id,

                    intermediate_url_list =
                        intermediate_url_list if __intermediate_flag__ else None))

            if __debug_flag__ and __result_flag__:

                with gzip.open(
                    result_file_name, 'wb') as result_file:

                    pickle.dump(result_dict, result_file)

            return CognateAnalysis(**result_dict)

        # Performing etymological distance analysis, if required.

        d_output = None
        d_output_binary = None

        if distance_flag or figure_flag:

            d_output_buffer_size = cognate_distance_analysis_f(
                None, len(perspective_info_list), len(result_list), None, 1)

            log.debug(
                'cognate_analysis {0}: distance output buffer size {1}'.format(
                language_str,
                d_output_buffer_size))

            d_output_buffer = ctypes.create_unicode_buffer(d_output_buffer_size + 256)

            d_result = cognate_distance_analysis_f(
                input_buffer, len(perspective_info_list), len(result_list), d_output_buffer, 1)

            # If we don't have a good result, we return an error.

            log.debug(
                'cognate_analysis {0}: distance result {1}'.format(
                language_str,
                d_result))

            if d_result <= 0:

                if task_status is not None:

                    task_status.set(5, 100,
                        'Finished (ERROR): library call error {0}'.format(d_result))

                return ResponseError(message =
                    'Cognate analysis library call error {0}'.format(d_result))

            # Showing what we've got.

            d_output = d_output_buffer.value

            distance_output_list = [
                d_output[i : i + 256]
                for i in range(0, len(d_output), 256)]

            log.debug(
                'cognate_analysis {0}:\ndistance output:\n{1}'.format(
                language_str,
                pprint.pformat(
                    distance_output_list, width = 144)))

            # Saving distance output buffer and distance output to files, if required.

            if __debug_flag__:

                d_output_file_name = (
                    'output {0}.buffer'.format(
                        cognate_name_str))

                with open(
                    d_output_file_name, 'wb') as d_output_file:

                    d_output_file.write(
                        bytes(d_output_buffer))

                for extension, encoding in ('utf8', 'utf-8'), ('utf16', 'utf-16'):

                    d_output_file_name = (
                        'output {0}.{1}'.format(
                            cognate_name_str, extension))

                    with open(
                        d_output_file_name, 'wb') as d_output_file:

                        d_output_file.write(
                            d_output.encode(encoding))

            # Getting binary output for parsing and exporting.

            d_result_binary = cognate_distance_analysis_f(
                input_buffer, len(perspective_info_list), len(result_list), d_output_buffer, 2)

            log.debug(
                'cognate_analysis {0}: distance result_binary {1}'.format(
                language_str,
                d_result_binary))

            if d_result_binary <= 0:

                if task_status is not None:

                    task_status.set(5, 100,
                        'Finished (ERROR): library call (binary) error {0}'.format(d_result_binary))

                return ResponseError(message =
                    'Cognate analysis library call (binary) error {0}'.format(d_result_binary))

            # Showing what we've got from the binary output call.

            d_output_binary = d_output_buffer[:d_result_binary]

            d_output_binary_list = [
                d_output_binary[i : i + 256]
                for i in range(0, len(d_output_binary), 256)]

            log.debug(
                '\ncognate_analysis {0}:'
                '\ndistance output_binary:\n{1}'.format(
                language_str,
                pprint.pformat(
                    d_output_binary_list, width = 144)))

        # Indicating task's final stage, if required.

        if task_status is not None:
            task_status.set(4, 99, 'Exporting analysis results to XLSX')

        # Parsing analysis results and exporting them as an Excel file.

        workbook_stream, distance_matrix_list = (

            CognateAnalysis.export_xlsx(
                language_str,
                mode,
                output_binary,
                d_output_binary,
                len(perspective_info_list),
                __debug_flag__,
                cognate_name_str if __debug_flag__ else None))

        current_datetime = datetime.datetime.now(datetime.timezone.utc)

        xlsx_filename = pathvalidate.sanitize_filename(
            '{0} cognate{1} analysis {2:04d}.{3:02d}.{4:02d}.xlsx'.format(
                base_language_name[:64],
                ' ' + mode if mode else '',
                current_datetime.year,
                current_datetime.month,
                current_datetime.day))

        if storage_dir is None:

            cur_time = time.time()
            storage_dir = os.path.join(storage['path'], 'cognate', str(cur_time))

        # Storing Excel file with the results.

        xlsx_path = os.path.join(storage_dir, xlsx_filename)
        os.makedirs(os.path.dirname(xlsx_path), exist_ok = True)

        workbook_stream.seek(0)

        with open(xlsx_path, 'wb') as xlsx_file:
            shutil.copyfileobj(workbook_stream, xlsx_file)

        xlsx_url = ''.join([
            storage['prefix'], storage['static_route'],
            'cognate', '/', str(cur_time), '/', xlsx_filename])

        # Selecting one of the distance matrices, if we have any.

        distance_header_array = None

        if distance_matrix_list is not None:

            distance_matrix = distance_matrix_list[-1]

            if distance_vowel_flag and distance_consonant_flag:
                pass

            elif distance_vowel_flag:
                distance_matrix = distance_matrix_list[0]

            elif distance_consonant_flag:
                distance_matrix = distance_matrix_list[1]

            (distance_title,
                distance_header_list,
                distance_data_list,
                distance_header_array,
                distance_data_array) = distance_matrix

        # Generating list of etymological distances to the reference perspective, if required.

        distance_list = None

        if distance_flag and reference_perspective_id is not None:

            reference_index = None

            for index, perspective_id in enumerate(perspective_id_list):
                if perspective_id == reference_perspective_id:

                    reference_index = index
                    break

            if reference_index is not None:

                distance_value_list = list(map(
                    float, distance_data_list[reference_index]))

                max_distance = float(max(distance_value_list))

                # Compiling and showing relative distance list.

                if max_distance > 0:
                    distance_list = [
                        (perspective_id, distance / max_distance)

                        for perspective_id, distance in zip(
                            perspective_id_list, distance_value_list)]

                else:

                    distance_list = distance_value_list

                log.debug(
                    '\ncognate_analysis {0}:'
                    '\n  perspective_id_list: {1}'
                    '\n  perspective_name_list:\n{2}'
                    '\n  reference_perspective_id: {3}'
                    '\n  reference_index: {4}'
                    '\n  distance_value_list: {5}'
                    '\n  max_distance: {6}'
                    '\n  distance_list: {7}'.format(
                    language_str,
                    perspective_id_list,
                    pprint.pformat(perspective_name_list, width = 144),
                    reference_perspective_id,
                    reference_index,
                    distance_value_list,
                    max_distance,
                    distance_list))

        # Generating distance graph, if required.
        figure_url = None
        mst_list = None
        embedding_2d_pca = None
        embedding_3d_pca = None

        if figure_flag:
            figure_url, mst_list, embedding_2d_pca, embedding_3d_pca = \
                CognateAnalysis.distance_graph(
                    language_str,
                    base_language_name,
                    distance_data_array,
                    distance_header_array,
                    mode,
                    storage,
                    storage_dir,
                    __debug_flag__
                )

        # Finalizing task status, if required, returning result.
        if task_status is not None:

            result_link_list = (
                [xlsx_url] +
                ([] if figure_url is None else [figure_url]) +
                (intermediate_url_list if __intermediate_flag__ else []))

            task_status.set(5, 100, 'Finished',
                result_link_list = result_link_list)

        result_dict = (

            dict(

                triumph = True,

                dictionary_count = len(perspective_info_list),
                group_count = len(group_list),
                not_enough_count = not_enough_count,
                transcription_count = total_transcription_count,
                translation_count = total_translation_count,

                result = wrapped_output,
                xlsx_url = xlsx_url,
                distance_list = distance_list,
                figure_url = figure_url,

                minimum_spanning_tree = mst_list,
                embedding_2d = embedding_2d_pca,
                embedding_3d = embedding_3d_pca,
                perspective_name_list = distance_header_array,

                intermediate_url_list =
                    intermediate_url_list if __intermediate_flag__ else None))

        if __debug_flag__ and __result_flag__:

            with gzip.open(
                result_file_name, 'wb') as result_file:

                pickle.dump(result_dict, result_file)

        return CognateAnalysis(**result_dict)

    @staticmethod
    def mutate(
        self,
        info,

        source_perspective_id,
        base_language_id,

        group_field_id,
        perspective_info_list,
        multi_list = None,

        mode = None,

        distance_flag = None,
        reference_perspective_id = None,

        figure_flag = None,
        distance_vowel_flag = None,
        distance_consonant_flag = None,

        match_translations_value = 1,
        only_orphans_flag = True,

        debug_flag = False,
        intermediate_flag = False,
        synchronous = False):
        """
        mutation CognateAnalysis {
          cognate_analysis(
            base_language_id: [508, 41],
            group_field_id: [66, 25],
            perspective_info_list: [
              [[425, 4], [66, 8], [66, 10]],
              [[1552, 1759], [66, 8], [66, 10]],
              [[418, 4], [66, 8], [66, 10]]])
          {
            triumph
            entity_count
            dictionary_count
            group_count
            not_enough_count
            text_count
            result
          }
        }
        """

        # Administrator / perspective author / editing permission check.

        error_str = (
            'Only administrator, perspective author and users with perspective editing permissions '
            'can perform cognate analysis.')

        client_id = info.context.client_id

        if not client_id:
            return ResponseError(error_str)

        user = Client.get_user_by_client_id(client_id)

        author_client_id_set = (

            set(
                client_id
                for (client_id, _), _, _ in perspective_info_list))

        author_id_check = (

            DBSession

                .query(

                    DBSession
                        .query(literal(1))
                        .filter(
                            Client.id.in_(author_client_id_set),
                            Client.user_id == user.id)
                        .exists())

                .scalar())

        if (user.id != 1 and
            not author_id_check and
            not info.context.acl_check_if('edit', 'perspective', source_perspective_id)):

            return ResponseError(error_str)

        # Debug mode check.

        if debug_flag and user.id != 1:

            return (

                ResponseError(
                    message = 'Only administrator can use debug mode.'))

        language_str = (
            '{0}/{1}, language {2}/{3}'.format(
                source_perspective_id[0], source_perspective_id[1],
                base_language_id[0], base_language_id[1]))

        try:

            # Getting base language info.

            locale_id = info.context.locale_id

            base_language = DBSession.query(dbLanguage).filter_by(
                client_id = base_language_id[0], object_id = base_language_id[1]).first()

            base_language_name = base_language.get_translation(locale_id)

            request = info.context.request
            storage = request.registry.settings['storage']

            # Getting multi-language info, if required.

            if multi_list is None:
                multi_list = []

            multi_name_list = []

            for language_id, perspective_count in multi_list:

                language = DBSession.query(dbLanguage).filter_by(
                    client_id = language_id[0], object_id = language_id[1]).first()

                multi_name_list.append(
                    language.get_translation(locale_id))

            # Language tag.

            if mode == 'multi':

                multi_str = ', '.join(
                    '{0}/{1}'.format(*id)
                    for id, count in multi_list)

                language_str = (
                    '{0}/{1}, languages {2}'.format(
                        source_perspective_id[0], source_perspective_id[1],
                        multi_str))

            # Showing cognate analysis info, checking cognate analysis library presence.

            log.debug(
                 '\ncognate_analysis {}:'
                 '\n  base language: {}'
                 '\n  group field: {}/{}'
                 '\n  perspectives and transcription/translation fields: {}'
                 '\n  multi_list: {}'
                 '\n  multi_name_list: {}'
                 '\n  mode: {}'
                 '\n  distance_flag: {}'
                 '\n  reference_perspective_id: {}'
                 '\n  figure_flag: {}'
                 '\n  distance_vowel_flag: {}'
                 '\n  distance_consonant_flag: {}'
                 '\n  match_translations_value: {}'
                 '\n  only_orphans_flag: {} ({})'
                 '\n  debug_flag: {}'
                 '\n  intermediate_flag: {}'
                 '\n  cognate_analysis_f: {}'
                 '\n  cognate_acoustic_analysis_f: {}'
                 '\n  cognate_distance_analysis_f: {}'
                 '\n  cognate_reconstruction_f: {}'
                 '\n  cognate_reconstruction_multi_f: {}'
                 '\n  cognate_suggestions_f: {}'.format(
                    language_str,
                    repr(base_language_name.strip()),
                    group_field_id[0], group_field_id[1],
                    perspective_info_list,
                    multi_list,
                    multi_name_list,
                    repr(mode),
                    distance_flag,
                    reference_perspective_id,
                    figure_flag,
                    distance_vowel_flag,
                    distance_consonant_flag,
                    match_translations_value,
                    only_orphans_flag, int(only_orphans_flag),
                    debug_flag,
                    intermediate_flag,
                    repr(cognate_analysis_f),
                    repr(cognate_acoustic_analysis_f),
                    repr(cognate_distance_analysis_f),
                    repr(cognate_reconstruction_f),
                    repr(cognate_reconstruction_multi_f),
                    repr(cognate_suggestions_f)))

            # Checking if we have analysis function ready.

            analysis_f = (
                cognate_acoustic_analysis_f if mode == 'acoustic' else
                cognate_reconstruction_f if mode == 'reconstruction' else
                cognate_reconstruction_multi_f if mode == 'multi' else
                cognate_suggestions_f if mode == 'suggestions' else
                cognate_analysis_f)

            if analysis_f is None:

                return ResponseError(message =
                    'Analysis library fuction \'{0}()\' is absent, '
                    'please contact system administrator.'.format(
                        'CognateAcousticAnalysis_GetAllOutput' if mode == 'acoustic' else
                        'CognateReconstruct_GetAllOutput' if mode == 'reconstruction' else
                        'CognateMultiReconstruct_GetAllOutput' if mode == 'multi' else
                        'GuessCognates_GetAllOutput' if mode == 'suggestions' else
                        'CognateAnalysis_GetAllOutput'))

            # Transforming client/object pair ids from lists to 2-tuples.

            source_perspective_id = tuple(source_perspective_id)
            base_language_id = tuple(base_language_id)
            group_field_id = tuple(group_field_id)

            perspective_info_list = [

                (tuple(perspective_id),
                    tuple(transcription_field_id),
                    tuple(translation_field_id))

                for perspective_id,
                    transcription_field_id,
                    translation_field_id in perspective_info_list]

            multi_list = [
                [tuple(language_id), perspective_count]
                for language_id, perspective_count in multi_list]

            if reference_perspective_id is not None:
                reference_perspective_id = tuple(reference_perspective_id)

            # If we are to use acoustic data, we will launch cognate analysis in asynchronous mode.

            if mode == 'acoustic':

                client_id = info.context.client_id

                user_id = (
                    Client.get_user_by_client_id(client_id).id
                        if client_id else anonymous_userid(request))

                task_status = TaskStatus(
                    user_id, 'Cognate acoustic analysis', base_language_name, 5)

                # Launching cognate acoustic analysis asynchronously.

                request.response.status = HTTPOk.code

                if synchronous:
                    CognateAnalysis.perform_cognate_analysis(
                        language_str,
                        source_perspective_id,
                        base_language_id,
                        base_language_name,
                        group_field_id,
                        perspective_info_list,
                        multi_list,
                        multi_name_list,
                        mode,
                        None,
                        None,
                        None,
                        None,
                        None,
                        match_translations_value,
                        only_orphans_flag,
                        locale_id,
                        storage,
                        task_status,
                        debug_flag,
                        intermediate_flag)

                else:

                    async_cognate_analysis.delay(
                        language_str,
                        source_perspective_id,
                        base_language_id,
                        base_language_name,
                        group_field_id,
                        perspective_info_list,
                        multi_list,
                        multi_name_list,
                        mode,
                        distance_flag,
                        reference_perspective_id,
                        figure_flag,
                        distance_vowel_flag,
                        distance_consonant_flag,
                        match_translations_value,
                        only_orphans_flag,
                        locale_id,
                        storage,
                        task_status.key,
                        request.registry.settings['cache_kwargs'],
                        request.registry.settings['sqlalchemy.url'],
                        debug_flag,
                        intermediate_flag)

                # Signifying that we've successfully launched asynchronous cognate acoustic analysis.

                return CognateAnalysis(triumph = True)

            # We do not use acoustic data, so we perform cognate analysis synchronously.
            else:

                return CognateAnalysis.perform_cognate_analysis(
                    language_str,
                    source_perspective_id,
                    base_language_id,
                    base_language_name,
                    group_field_id,
                    perspective_info_list,
                    multi_list,
                    multi_name_list,
                    mode,
                    distance_flag,
                    reference_perspective_id,
                    figure_flag,
                    distance_vowel_flag,
                    distance_consonant_flag,
                    match_translations_value,
                    only_orphans_flag,
                    locale_id,
                    storage,
                    None,
                    debug_flag,
                    intermediate_flag)

        # Exception occured while we tried to perform cognate analysis.

        except Exception as exception:

            traceback_string = ''.join(traceback.format_exception(
                exception, exception, exception.__traceback__))[:-1]

            log.warning(
                'cognate_analysis {0}: exception'.format(
                language_str))

            log.warning(traceback_string)

            return ResponseError(message =
                'Exception:\n' + traceback_string)


class SwadeshAnalysis(graphene.Mutation):
    class Arguments:

        source_perspective_id = LingvodocID(required = True)
        base_language_id = LingvodocID(required = True)

        group_field_id = LingvodocID(required = True)
        perspective_info_list = graphene.List(graphene.List(LingvodocID), required = True)

        debug_flag = graphene.Boolean()

    triumph = graphene.Boolean()

    result = graphene.String()
    xlsx_url = graphene.String()
    minimum_spanning_tree = graphene.List(graphene.List(graphene.Int))
    embedding_2d = graphene.List(graphene.List(graphene.Float))
    embedding_3d = graphene.List(graphene.List(graphene.Float))
    perspective_name_list = graphene.List(graphene.String)

    @staticmethod
    def get_entry_text(entry):
        return f"{entry['swadesh']} [ {entry['transcription']} ] {entry['translation']}"

    @staticmethod
    def export_dataframe(result_pool, distance_data_array, bundles, get_entry_text):
        '''
        Keys:
        result_pool[perspective_id][entry_id]
        Fields:
        'group': group_index,
        'borrowed': bool,
        'swadesh': swadesh_lex,
        'transcription': transcription_list[0],
        'translation': translation_lex
        '''

        distances = pd.DataFrame(distance_data_array,
                                 columns=[perspective['name'] for perspective in result_pool.values()])
        # Start index for distances from 1 to match with dictionaries numbers
        distances.index += 1

        groups = pd.DataFrame()
        # Insert 'lines' column as the first one
        groups['lines'] = 0

        borrowed = pd.DataFrame()
        singles = pd.DataFrame()

        singles_index = 0
        borrowed_index = 0
        # re-group by group number and add joined values
        for perspective in result_pool.values():
            dict_name = perspective['name']
            for entry in perspective.values():
                # 'entry' iterator may present string value of 'name' field
                # but not a dictionary for one of entries. Continue in this case.
                if not isinstance(entry, dict):
                    continue
                group_num = entry['group']
                entry_text = get_entry_text(entry)
                if group_num is not None and group_num in bundles:
                    # Concatinate existing value if is and a new one, store the result to 'groups' dataframe
                    value = ""
                    if dict_name in groups:
                        cell = groups[dict_name].get(group_num)
                        value = cell if pd.notnull(cell) else value
                    value = f"{value}\n{entry_text}".strip()
                    groups.loc[group_num, dict_name] = value

                    # Count result lines to set rows height in xlsx after
                    lines = value.count('\n') + 1
                    cell = groups.loc[group_num].get('lines')
                    if pd.isnull(cell) or cell < lines:
                        groups.loc[group_num, 'lines'] = lines
                elif entry.get('borrowed'):
                    borrowed.loc[borrowed_index, dict_name] = entry_text
                    borrowed_index += 1
                else:
                    singles.loc[singles_index, dict_name] = entry_text
                    singles_index += 1

        return {
            'Cognates': groups if len(groups) < 2 else groups.sort_values(groups.columns[1]),
            'Singles': singles.sort_index(),
            'Borrowed': borrowed.sort_index(),
            'Distances': distances.sort_index()
        }

    @staticmethod
    def export_xlsx(
            result,
            base_language_name,
            storage
    ):
        # Exporting analysis results as an Excel file.

        current_datetime = datetime.datetime.now(datetime.timezone.utc)
        xlsx_filename = pathvalidate.sanitize_filename(
            '{0} {1} {2:04d}.{3:02d}.{4:02d}.xlsx'.format(
                base_language_name[:64],
                'glottochronology',
                current_datetime.year,
                current_datetime.month,
                current_datetime.day))

        cur_time = time.time()
        storage_dir = os.path.join(storage['path'], 'glottochronology', str(cur_time))

        # Storing Excel file with the results.

        xlsx_path = os.path.join(storage_dir, xlsx_filename)
        os.makedirs(os.path.dirname(xlsx_path), exist_ok=True)

        with pd.ExcelWriter(xlsx_path, engine='xlsxwriter') as writer:
            header_format = writer.book.add_format({'bold': True,
                                                    'text_wrap': True,
                                                    'valign': 'top',
                                                    'fg_color': '#D7E4BC',
                                                    'border': 1})
            for sheet_name, df in result.items():
                index = (sheet_name == 'Distances')
                startcol = int(index)
                # Exclude 'lines' column
                columns = df.columns[int(sheet_name == 'Cognates'):]
                # Check if the table is empty
                if columns.empty:
                    continue

                df.to_excel(writer,
                            sheet_name=sheet_name,
                            index=index,
                            startrow=1,
                            columns=columns,
                            header=False)

                worksheet = writer.sheets[sheet_name]
                worksheet.set_row(0, 70)
                worksheet.set_column(startcol, len(columns) - 1 + startcol, 30)
                # Write the column headers with the defined format.
                for col_num, value in enumerate(columns):
                    worksheet.write(0, col_num + startcol, value, header_format)
                # Set rows specific height
                if sheet_name == 'Cognates':
                    for row_num, coeff in enumerate(df['lines']):
                        if coeff > 1:
                            worksheet.set_row(row_num + 1, 14 * coeff)

        xlsx_url = ''.join([
            storage['prefix'], storage['static_route'],
            'glottochronology', '/', str(cur_time), '/', xlsx_filename])

        return xlsx_url

    @staticmethod
    def export_html(result, tiny_dicts=None, huge_size=1048576):
        result_tables = (
            build_table(result['Distances'], 'orange_light', width="300px", index=True),
            build_table(result['Cognates'], 'blue_light', width="300px").replace("\\n","<br>"),
            build_table(result['Singles'], 'green_light', width="300px"),
            build_table(result['Borrowed'], 'yellow_light', width="300px"))

        # Control output size
        spl = "<pre>\n\n</pre>"
        html_result = f"{result_tables[0]}" \
                      f"{spl}" \
                      f"{result_tables[1]}" \
                      f"{spl}" \
                      f"{result_tables[2]}" \
                      f"{spl}" \
                      f"{result_tables[3]}"

        if len(html_result) > huge_size:
            html_result = f"{result_tables[0]}" \
                          f"{spl}" \
                          f"{result_tables[1]}" \
                          f"<pre>\n\nNote: The table with single words is not shown due to huge summary size</pre>"

        if len(html_result) > huge_size:
            html_result = f"{result_tables[0]}" \
                          f"<pre>\n\nNote: The result tables with words are not shown due to huge summary size</pre>"

        html_result += ("<pre>Note: The following dictionaries contain too less words and were not processed: \n\n" +
                        '\n'.join(tiny_dicts) + "</pre>") if tiny_dicts else ""

        return html_result

    @staticmethod
    def swadesh_statistics(
            language_str,
            base_language_id,
            base_language_name,
            group_field_id,
            perspective_info_list,
            locale_id,
            storage,
            debug_flag = False):

        swadesh_list = ['я','ты','мы','этот, это','тот, то','кто','что','не','все','много','один','два','большой',
                        'долгий','маленький','женщина','мужчина','человек','рыба','птица','собака','вошь','дерево',
                        'семя','лист','корень','кора','кожа','мясо','кровь','кость','жир','яйцо','рог','хвост','перо',
                        'волосы','голова','ухо','глаз','нос','рот','зуб','язык (орган)','ноготь','нога (стопа)','колено',
                        'рука (кисть)','живот','горло','грудь','сердце','печень','пить','есть (кушать)','кусать','видеть',
                        'слышать','знать','спать','умирать','убивать','плавать','летать','гулять','приходить','лежать',
                        'сидеть','стоять','дать','сказать','солнце','луна','звезда','вода','дождь','камень','песок',
                        'земля','облако','дым','огонь','пепел','гореть','дорога,тропа','гора','красный','зелёный',
                        'жёлтый','белый','чёрный','ночь','тёплый','холодный','полный','новый','хороший','круглый',
                        'сухой','имя']

        def compare_translations(swadesh_lex, dictionary_lex):
            def split_lex(lex):
                # Split by commas and open brackets to separate
                # various forms of lexeme and extra note if is
                lex = ' '.join(lex.lower().split()) # reduce multi spaces
                if "убрать из стословника" in lex:
                    return set()

                return set(form.replace('ё', 'е').strip()
                           for form in lex.replace('(', ',').split(',')
                           if form.strip() and ')' not in form)  # exclude notes

            # return true if the intersection is not empty
            return bool(split_lex(swadesh_lex) & split_lex(dictionary_lex))


        # Gathering entry grouping data.

        if not debug_flag:

            _, group_list, _ = (
                CognateAnalysis.tag_data_plpgsql(
                    perspective_info_list, group_field_id))

        else:

            # If we are in debug mode, we try to load existing tag data to reduce debugging time.

            tag_data_digest = (

                hashlib.md5(

                    repr(list(group_field_id) +
                        [perspective_info[0] for perspective_info in perspective_info_list])

                    .encode('utf-8'))

                .hexdigest())

            tag_data_file_name = (
                f'__tag_data_{base_language_id[0]}_{base_language_id[1]}_{tag_data_digest}__.gz')

            # Checking if we have saved data.

            if os.path.exists(tag_data_file_name):

                with gzip.open(tag_data_file_name, 'rb') as tag_data_file:
                    _, group_list, _ = pickle.load(tag_data_file)

            else:

                # Don't have existing data, so we gather it and then save it for later use.

                r1, group_list, r3 = (

                    CognateAnalysis.tag_data_plpgsql(
                        perspective_info_list, group_field_id))

                with gzip.open(tag_data_file_name, 'wb') as tag_data_file:
                    pickle.dump((r1, group_list, r3), tag_data_file)


        # Getting text data for each perspective.
        # entries_set gathers entry_id(s) of words met in Swadesh' list
        # swadesh_total gathers numbers of words within Swadesh' list
        entries_set = {}
        swadesh_total = {}
        result_pool = {}
        tiny_dicts = set()
        for index, (perspective_id, transcription_field_id, translation_field_id) in \
                enumerate(perspective_info_list):

            # Getting and saving perspective info.
            perspective = (
                DBSession
                    .query(dbPerspective)
                    .filter_by(client_id=perspective_id[0], object_id=perspective_id[1])
                    .first()
            )
            dictionary_name = perspective.parent.get_translation(locale_id)

            # GC
            del perspective

            # Getting text data.
            transcription_query = (
                DBSession
                    .query(
                        dbLexicalEntry.client_id,
                        dbLexicalEntry.object_id)
                    .filter(
                        dbLexicalEntry.parent_client_id == perspective_id[0],
                        dbLexicalEntry.parent_object_id == perspective_id[1],
                        dbLexicalEntry.marked_for_deletion == False,
                        dbEntity.parent_client_id == dbLexicalEntry.client_id,
                        dbEntity.parent_object_id == dbLexicalEntry.object_id,
                        dbEntity.field_client_id == transcription_field_id[0],
                        dbEntity.field_object_id == transcription_field_id[1],
                        dbEntity.marked_for_deletion == False,
                        dbPublishingEntity.client_id == dbEntity.client_id,
                        dbPublishingEntity.object_id == dbEntity.object_id,
                        dbPublishingEntity.published == True,
                        dbPublishingEntity.accepted == True)
                    .add_columns(
                        func.array_agg(dbEntity.content).label('transcription'))
                    .group_by(dbLexicalEntry)
                    .subquery())

            translation_query = (
                DBSession
                    .query(
                        dbLexicalEntry.client_id,
                        dbLexicalEntry.object_id)
                    .filter(
                        dbLexicalEntry.parent_client_id == perspective_id[0],
                        dbLexicalEntry.parent_object_id == perspective_id[1],
                        dbLexicalEntry.marked_for_deletion == False,
                        dbEntity.parent_client_id == dbLexicalEntry.client_id,
                        dbEntity.parent_object_id == dbLexicalEntry.object_id,
                        dbEntity.field_client_id == translation_field_id[0],
                        dbEntity.field_object_id == translation_field_id[1],
                        dbEntity.marked_for_deletion == False,
                        dbPublishingEntity.client_id == dbEntity.client_id,
                        dbPublishingEntity.object_id == dbEntity.object_id,
                        dbPublishingEntity.published == True,
                        dbPublishingEntity.accepted == True)
                    .add_columns(
                        func.array_agg(dbEntity.content).label('translation'))
                    .group_by(dbLexicalEntry)
                    .subquery())

            # Main query for transcription/translation data.
            data_query = (
                DBSession
                    .query(transcription_query)
                    .outerjoin(translation_query, and_(
                        transcription_query.c.client_id == translation_query.c.client_id,
                        transcription_query.c.object_id == translation_query.c.object_id))
                    .add_columns(
                        translation_query.c.translation)
                    .all())

            # GC
            del transcription_query
            del translation_query

            # Grouping translations by lexical entries.
            entries_set[perspective_id] = set()
            swadesh_total[perspective_id] = set()
            result_pool[perspective_id] = {'name': dictionary_name}
            for row_index, row in enumerate(data_query):
                entry_id = tuple(row[:2])
                transcription_list, translation_list = row[2:4]

                # If we have no transcriptions for this lexical entry, we skip it altogether.
                if not transcription_list:
                    continue

                translation_list = (
                    [] if not translation_list else [
                        translation.strip()
                        for translation in translation_list
                        if translation.strip()])

                # Parsing translations and matching with Swadesh's words
                transcription_lex = ', '.join(transcription_list)
                for swadesh_num, swadesh_lex in enumerate(swadesh_list):
                    for translation_lex in translation_list:
                        if compare_translations(swadesh_lex, translation_lex):
                            # Store the entry's content in human-readable format
                            result_pool[perspective_id][entry_id] = {
                                'group': None,
                                'borrowed': (" заим." in f" {transcription_lex} {translation_lex}"),
                                'swadesh': swadesh_lex,
                                'transcription': transcription_lex,
                                'translation': translation_lex
                            }
                            # Store entry_id and number of the lex within Swadesh's list
                            entries_set[perspective_id].add(entry_id)
                            if not result_pool[perspective_id][entry_id]['borrowed']:
                                # Total list of Swadesh's words in the perspective,
                                # they can have not any etymological links
                                swadesh_total[perspective_id].add(swadesh_num)

            # Forget the dictionary if it contains less than 50 Swadesh words
            if len(swadesh_total[perspective_id]) < 50:
                del entries_set[perspective_id]
                del swadesh_total[perspective_id]
                del result_pool[perspective_id]
                tiny_dicts.add(dictionary_name)

            # GC
            del data_query

        # Checking if found entries have links
        means = collections.OrderedDict()
        for perspective_id, entries in entries_set.items():
            means[perspective_id] = collections.defaultdict(set)
            for group_index, group in enumerate(group_list):
                # Select etymologically linked entries
                linked = entries & group
                for entry_id in linked:
                    result_pool[perspective_id][entry_id]['group'] = group_index
                    swadesh = result_pool[perspective_id][entry_id]['swadesh']
                    # Store the correspondence: perspective { meanings(1/2/3) { etymological_groups(1.1/1.2/2.1/3.1)
                    if not result_pool[perspective_id][entry_id]['borrowed']:
                        means[perspective_id][swadesh].add(group_index)

        dictionary_count = len(means)
        distance_data_array = numpy.full((dictionary_count, dictionary_count), 50, dtype='float')
        complex_data_array = numpy.full((dictionary_count, dictionary_count), "n/a", dtype='object')
        distance_header_array = numpy.full(dictionary_count, "<noname>", dtype='object')

        # Calculate intersection between lists of linked meanings (Swadesh matching)
        # So length of this intersection is the similarity of corresponding perspectives
        # means_total is amount of Swadesh's lexemes met in the both perspectives
        bundles = set()
        # Calculate each-to-each distances, exclude self-to-self
        for n1, (perspective1, means1) in enumerate(means.items()):
            # Numerate dictionaries
            result_pool[perspective1]['name'] = f"{n1 + 1}. {result_pool[perspective1]['name']}"
            distance_header_array[n1] = result_pool[perspective1]['name']
            for n2, (perspective2, means2) in enumerate(means.items()):
                if n1 == n2:
                    distance_data_array[n1][n2] = 0
                    complex_data_array[n1][n2] = "n/a"
                else:
                    # Common meanings of entries which have etymological links
                    # but this links may be not mutual
                    means_common = means1.keys() & means2.keys()
                    means_linked = 0
                    # Checking if the found meanings have common links
                    for swadesh in means_common:
                        links_common = means1[swadesh] & means2[swadesh]
                        if links_common:
                            # Bundles are links with two or more entries in the result table
                            bundles.update(links_common)
                            means_linked += 1

                    means_total = len(swadesh_total[perspective1] & swadesh_total[perspective2])

                    if n2 > n1 and means_linked >= means_total:
                        log.debug(f"{n1+1},{n2+1} : "
                                  f"{len(means_common)} but {means_linked} of {means_total} : "
                                  f"{', '.join(sorted(means_common))}")

                    c = means_linked / means_total if means_total > 0 else 0
                    distance = math.sqrt( math.log(c) / -0.1 / math.sqrt(c) ) if c > 0 else 25
                    percent = means_linked * 100 // means_total if means_total > 0 else 0
                    distance_data_array[n1][n2] = round(distance, 2)
                    complex_data_array[n1][n2] = f"{distance_data_array[n1][n2]:.2f} ({percent}%)"

        result = SwadeshAnalysis.export_dataframe(result_pool, complex_data_array, bundles, SwadeshAnalysis.get_entry_text)

        # GC
        del result_pool

        xlsx_url = SwadeshAnalysis.export_xlsx(result, base_language_name, storage)

        # 'lines' field is not needed any more
        del result['Cognates']['lines']

        html_result = SwadeshAnalysis.export_html(result, tiny_dicts)

        _, mst_list, embedding_2d_pca, embedding_3d_pca = \
            CognateAnalysis.distance_graph(
                language_str,
                base_language_name,
                distance_data_array,
                distance_header_array,
                None,
                None,
                None,
                analysis_str = 'swadesh_analysis',
                __debug_flag__ = debug_flag,
                __plot_flag__ = False
            )

        result_dict = (
            dict(
                triumph = True,

                result = html_result,
                xlsx_url = xlsx_url,
                minimum_spanning_tree = mst_list,
                embedding_2d = embedding_2d_pca,
                embedding_3d = embedding_3d_pca,
                perspective_name_list = distance_header_array))

        return SwadeshAnalysis(**result_dict)

    @staticmethod
    def mutate(
        self,
        info,
        source_perspective_id,
        base_language_id,
        group_field_id,
        perspective_info_list,
        debug_flag = False):
        """
        mutation SwadeshAnalysis {
          swadesh_analysis(
            base_language_id: [508, 41],
            group_field_id: [66, 25],
            perspective_info_list: [
              [[425, 4], [66, 8], [66, 10]],
              [[1552, 1759], [66, 8], [66, 10]],
              [[418, 4], [66, 8], [66, 10]]])
          {
            triumph          }
        }
        """

        # Administrator / perspective author / editing permission check.
        error_str = (
            'Only administrator, perspective author and users with perspective editing permissions '
            'can perform Swadesh analysis.')

        client_id = info.context.client_id

        if not client_id:
            return ResponseError(error_str)

        user = Client.get_user_by_client_id(client_id)

        author_client_id_set = (

            set(
                client_id
                for (client_id, _), _, _ in perspective_info_list))

        author_id_check = (

            DBSession

                .query(

                    DBSession
                        .query(literal(1))
                        .filter(
                            Client.id.in_(author_client_id_set),
                            Client.user_id == user.id)
                        .exists())

                .scalar())

        if (user.id != 1 and
            not author_id_check and
            not info.context.acl_check_if('edit', 'perspective', source_perspective_id)):

            return ResponseError(error_str)

        # Debug mode check.

        if debug_flag and user.id != 1:

            return (

                ResponseError(
                    message = 'Only administrator can use debug mode.'))

        language_str = (
            '{0}/{1}, language {2}/{3}'.format(
                source_perspective_id[0], source_perspective_id[1],
                base_language_id[0], base_language_id[1]))

        try:

            # Getting base language info.

            locale_id = info.context.locale_id

            base_language = DBSession.query(dbLanguage).filter_by(
                client_id = base_language_id[0], object_id = base_language_id[1]).first()

            base_language_name = base_language.get_translation(locale_id)

            request = info.context.request
            storage = request.registry.settings['storage']

            # Transforming client/object pair ids from lists to 2-tuples.

            source_perspective_id = tuple(source_perspective_id)
            base_language_id = tuple(base_language_id)
            group_field_id = tuple(group_field_id)

            perspective_info_list = [

                (tuple(perspective_id),
                    tuple(transcription_field_id),
                    tuple(translation_field_id))

                for perspective_id,
                    transcription_field_id,
                    translation_field_id in perspective_info_list]

            return SwadeshAnalysis.swadesh_statistics(
                language_str,
                base_language_id,
                base_language_name,
                group_field_id,
                perspective_info_list,
                locale_id,
                storage,
                debug_flag)

        # Exception occured while we tried to perform swadesh analysis.
        except Exception as exception:

            traceback_string = ''.join(traceback.format_exception(
                exception, exception, exception.__traceback__))[:-1]

            log.warning(
                'swadesh_analysis {0}: exception'.format(
                language_str))

            log.warning(traceback_string)

            return ResponseError(message =
                'Exception:\n' + traceback_string)


class MorphCognateAnalysis(graphene.Mutation):
    class Arguments:

        source_perspective_id = LingvodocID(required = True)
        base_language_id = LingvodocID(required = True)

        group_field_id = LingvodocID(required = True)
        perspective_info_list = graphene.List(graphene.List(LingvodocID), required = True)

        debug_flag = graphene.Boolean()

    triumph = graphene.Boolean()

    result = graphene.String()
    xlsx_url = graphene.String()
    minimum_spanning_tree = graphene.List(graphene.List(graphene.Int))
    embedding_2d = graphene.List(graphene.List(graphene.Float))
    embedding_3d = graphene.List(graphene.List(graphene.Float))
    perspective_name_list = graphene.List(graphene.String)

    @staticmethod
    def get_entry_text(entry):
        return f"{'; '.join(entry['affix'])} ( {'; '.join(entry['meaning'])} )"

    @staticmethod
    def morph_cognate_statistics(
            language_str,
            base_language_id,
            base_language_name,
            group_field_id,
            perspective_info_list,
            locale_id,
            storage,
            debug_flag = False):

        # Gathering entry grouping data.

        if not debug_flag:

            _, group_list, _ = (
                CognateAnalysis.tag_data_plpgsql(
                    perspective_info_list, group_field_id))

        else:

            # If we are in debug mode, we try to load existing tag data to reduce debugging time.

            tag_data_digest = (
                hashlib.md5(
                    repr(list(group_field_id) +
                        [perspective_info[0] for perspective_info in perspective_info_list])
                    .encode('utf-8'))
                .hexdigest())

            tag_data_file_name = (
                f'__tag_data_{base_language_id[0]}_{base_language_id[1]}_{tag_data_digest}__.gz')

            # Checking if we have saved data.

            if os.path.exists(tag_data_file_name):

                with gzip.open(tag_data_file_name, 'rb') as tag_data_file:
                    _, group_list, _ = pickle.load(tag_data_file)

            else:

                # Don't have existing data, so we gather it and then save it for later use.

                r1, group_list, r3 = (

                    CognateAnalysis.tag_data_plpgsql(
                        perspective_info_list, group_field_id))

                with gzip.open(tag_data_file_name, 'wb') as tag_data_file:
                    pickle.dump((r1, group_list, r3), tag_data_file)

        # Getting text data for each perspective.
        to_canon_meaning = collections.defaultdict(dict)
        meaning_to_links = {}
        result_pool = {}
        tiny_dicts = set()
        meaning_re = re.compile('[.\dA-Z]+')
        meaning_with_comment_re = re.compile('[.\dA-Z]+ *\([.,:;\d\w ]+\)')

        for index, (perspective_id, affix_field_id, meaning_field_id) in \
                enumerate(perspective_info_list):

            # Getting and saving perspective info.
            perspective = (
                DBSession
                    .query(dbPerspective)
                    .filter_by(client_id=perspective_id[0], object_id=perspective_id[1])
                    .first()
            )
            dictionary_name = perspective.parent.get_translation(locale_id)

            # GC
            del perspective

            # Getting text data.
            affix_query = (
                DBSession
                    .query(
                        dbLexicalEntry.client_id,
                        dbLexicalEntry.object_id)
                    .filter(
                        dbLexicalEntry.parent_client_id == perspective_id[0],
                        dbLexicalEntry.parent_object_id == perspective_id[1],
                        dbLexicalEntry.marked_for_deletion == False,
                        dbEntity.parent_client_id == dbLexicalEntry.client_id,
                        dbEntity.parent_object_id == dbLexicalEntry.object_id,
                        dbEntity.field_client_id == affix_field_id[0],
                        dbEntity.field_object_id == affix_field_id[1],
                        dbEntity.marked_for_deletion == False,
                        dbPublishingEntity.client_id == dbEntity.client_id,
                        dbPublishingEntity.object_id == dbEntity.object_id,
                        dbPublishingEntity.published == True,
                        dbPublishingEntity.accepted == True)
                    .add_columns(
                        func.array_agg(dbEntity.content).label('affix'))
                    .group_by(dbLexicalEntry)
                    .subquery())

            meaning_query = (
                DBSession
                    .query(
                        dbLexicalEntry.client_id,
                        dbLexicalEntry.object_id)
                    .filter(
                        dbLexicalEntry.parent_client_id == perspective_id[0],
                        dbLexicalEntry.parent_object_id == perspective_id[1],
                        dbLexicalEntry.marked_for_deletion == False,
                        dbEntity.parent_client_id == dbLexicalEntry.client_id,
                        dbEntity.parent_object_id == dbLexicalEntry.object_id,
                        dbEntity.field_client_id == meaning_field_id[0],
                        dbEntity.field_object_id == meaning_field_id[1],
                        dbEntity.marked_for_deletion == False,
                        dbPublishingEntity.client_id == dbEntity.client_id,
                        dbPublishingEntity.object_id == dbEntity.object_id,
                        dbPublishingEntity.published == True,
                        dbPublishingEntity.accepted == True)
                    .add_columns(
                        func.array_agg(dbEntity.content).label('meaning'))
                    .group_by(dbLexicalEntry)
                    .subquery())

            # Main query for transcription/translation data.
            data_query = (
                DBSession
                    .query(affix_query)
                    .outerjoin(meaning_query, and_(
                        affix_query.c.client_id == meaning_query.c.client_id,
                        affix_query.c.object_id == meaning_query.c.object_id))
                    .add_columns(
                        meaning_query.c.meaning)
                    .all())

            # GC
            del affix_query
            del meaning_query

            meaning_to_links[perspective_id] = {}
            result_pool[perspective_id] = {'name': dictionary_name}

            for row in data_query:
                entry_id = tuple(row[:2])

                if not (affix_list := row[2]) or not (meaning_list := row[3]):
                    continue

                affix = list(map(lambda a: a.strip(), affix_list))
                meaning = []
                for m in meaning_list:
                    if ((meaning_search := re.search(meaning_with_comment_re, m)) or
                            (meaning_search := re.search(meaning_re, m))):
                        meaning_value = meaning_search.group(0)
                        meaning.append(" ".join(meaning_value.split()))

                if not meaning:
                    continue

                # Compounding a dictionary to convert every meaning to the first one within each row
                # Initialize group of links for every sub_meaning to have a full set of sub_meanings
                for sub_meaning in meaning:
                    to_canon_meaning[perspective_id][sub_meaning] = meaning[0]
                    meaning_to_links[perspective_id][sub_meaning] = set()

                # Grouping affixes and meanings by lexical entries.
                result_pool[perspective_id][entry_id] = {
                    'group': None,
                    'affix': affix,
                    'meaning': meaning
                }

            # Forget the dictionary if it contains less than 30 sub-meanings
            if len(to_canon_meaning[perspective_id]) < 30:
                del meaning_to_links[perspective_id]
                del result_pool[perspective_id]
                tiny_dicts.add(dictionary_name)

            # GC
            del data_query

        # Checking if found entries have links
        for perspective_id, entries in result_pool.items():
            for group_index, group in enumerate(group_list):
                # Select etymologically linked entries
                linked = entries.keys() & group
                for entry_id in linked:
                    result_pool[perspective_id][entry_id]['group'] = group_index
                    meaning = result_pool[perspective_id][entry_id]['meaning']
                    for sub_meaning in meaning:
                        meaning_to_links[perspective_id][sub_meaning].add(group_index)

        dictionary_count = len(result_pool)
        distance_data_array = numpy.full((dictionary_count, dictionary_count), 50, dtype='float')
        complex_data_array = numpy.full((dictionary_count, dictionary_count), "n/a", dtype='object')
        distance_header_array = numpy.full(dictionary_count, "<noname>", dtype='object')

        bundles = set()
        # Calculate each-to-each distances, exclude self-to-self
        for n1, (perspective1, meaning_to_links1) in enumerate(meaning_to_links.items()):
            # Numerate dictionaries
            result_pool[perspective1]['name'] = f"{n1 + 1}. {result_pool[perspective1]['name']}"
            distance_header_array[n1] = result_pool[perspective1]['name']

            to_canon_meaning1 = to_canon_meaning[perspective1]
            canon_meanings1_set = set(to_canon_meaning1.values())

            for n2, (perspective2, meaning_to_links2) in enumerate(meaning_to_links.items()):
                if n1 == n2:
                    distance_data_array[n1][n2] = 0
                    complex_data_array[n1][n2] = "n/a"
                else:
                    # Compile new meaning_to_links2 using canon_meanings instead of sub_meanings
                    canon_meaning_to_links2 = collections.defaultdict(set)
                    for sub_meaning, links in meaning_to_links2.items():
                        if canon_meaning := to_canon_meaning1.get(sub_meaning):
                            canon_meaning_to_links2[canon_meaning].update(links)

                    # Common canonical meanings of perspective1 and perspective2
                    meanings_common = canon_meanings1_set & canon_meaning_to_links2.keys()
                    meanings_total = len(meanings_common)

                    meanings_linked = 0
                    # Checking if the found meanings have common links
                    for meaning in meanings_common:
                        links_common = meaning_to_links1[meaning] & canon_meaning_to_links2[meaning]
                        if links_common:
                            # Bundles are links with two or more entries in the result table
                            bundles.update(links_common)
                            meanings_linked += 1

                    '''
                    if debug_flag and n2 > n1 and meanings_linked >= meanings_total:
                        log.debug(f"{n1+1},{n2+1} : "
                                  f"{len(meanings_common)} but {meanings_linked} of {meanings_total} : "
                                  f"{', '.join(sorted(meanings_common))}")
                    '''

                    # meanings_linked > 0 meanings that meanings_total > 0 even more so
                    distance = math.log(meanings_linked / meanings_total) / -0.14 if meanings_linked > 0 else 50
                    percent = meanings_linked * 100 // meanings_total if meanings_total > 0 else 0
                    distance_data_array[n1][n2] = round(distance, 2)
                    complex_data_array[n1][n2] = f"{distance_data_array[n1][n2]:.2f} ({percent}%)"

        result = SwadeshAnalysis.export_dataframe(result_pool, complex_data_array, bundles, MorphCognateAnalysis.get_entry_text)

        # GC
        del result_pool

        xlsx_url = SwadeshAnalysis.export_xlsx(result, base_language_name, storage)

        # 'lines' field is not needed any more
        del result['Cognates']['lines']

        html_result = SwadeshAnalysis.export_html(result, tiny_dicts)

        _, mst_list, embedding_2d_pca, embedding_3d_pca = \
            CognateAnalysis.distance_graph(
                language_str,
                base_language_name,
                distance_data_array,
                distance_header_array,
                None,
                None,
                None,
                analysis_str = 'morph_cognate_analysis',
                __debug_flag__ = debug_flag,
                __plot_flag__ = False
            )

        result_dict = (
            dict(
                triumph = True,

                result = html_result,
                xlsx_url = xlsx_url,
                minimum_spanning_tree = mst_list,
                embedding_2d = embedding_2d_pca,
                embedding_3d = embedding_3d_pca,
                perspective_name_list = distance_header_array))

        return MorphCognateAnalysis(**result_dict)

    @staticmethod
    def mutate(
        self,
        info,
        source_perspective_id,
        base_language_id,
        group_field_id,
        perspective_info_list,
        debug_flag = False):
        """
        mutation MorphCognateAnalysis {
          morph_cognate_analysis(
            base_language_id: [508, 41],
            group_field_id: [66, 25],
            perspective_info_list: [
              [[425, 4], [66, 8], [66, 10]],
              [[1552, 1759], [66, 8], [66, 10]],
              [[418, 4], [66, 8], [66, 10]]])
          {
            triumph          }
        }
        """

        # Administrator / perspective author / editing permission check.
        error_str = (
            'Only administrator, perspective author and users with perspective editing permissions '
            'can perform Swadesh analysis.')

        client_id = info.context.client_id

        if not client_id:
            return ResponseError(error_str)

        user = Client.get_user_by_client_id(client_id)

        author_client_id_set = (

            set(
                client_id
                for (client_id, _), _, _ in perspective_info_list))

        author_id_check = (

            DBSession

                .query(

                    DBSession
                        .query(literal(1))
                        .filter(
                            Client.id.in_(author_client_id_set),
                            Client.user_id == user.id)
                        .exists())

                .scalar())

        if (user.id != 1 and
            not author_id_check and
            not info.context.acl_check_if('edit', 'perspective', source_perspective_id)):

            return ResponseError(error_str)

        # Debug mode check.

        if debug_flag and user.id != 1:

            return (

                ResponseError(
                    message = 'Only administrator can use debug mode.'))

        language_str = (
            '{0}/{1}, language {2}/{3}'.format(
                source_perspective_id[0], source_perspective_id[1],
                base_language_id[0], base_language_id[1]))

        try:

            # Getting base language info.

            locale_id = info.context.locale_id

            base_language = DBSession.query(dbLanguage).filter_by(
                client_id = base_language_id[0], object_id = base_language_id[1]).first()

            base_language_name = base_language.get_translation(locale_id)

            request = info.context.request
            storage = request.registry.settings['storage']

            # Transforming client/object pair ids from lists to 2-tuples.

            source_perspective_id = tuple(source_perspective_id)
            base_language_id = tuple(base_language_id)
            group_field_id = tuple(group_field_id)

            perspective_info_list = [

                (tuple(perspective_id),
                    tuple(affix_field_id),
                    tuple(meaning_field_id))

                for perspective_id,
                    affix_field_id,
                    meaning_field_id in perspective_info_list]

            return MorphCognateAnalysis.morph_cognate_statistics(
                language_str,
                base_language_id,
                base_language_name,
                group_field_id,
                perspective_info_list,
                locale_id,
                storage,
                debug_flag)

        # Exception occured while we tried to perform swadesh analysis.
        except Exception as exception:

            traceback_string = ''.join(traceback.format_exception(
                exception, exception, exception.__traceback__))[:-1]

            log.warning(
                'morph_cognate_analysis {0}: exception'.format(
                language_str))

            log.warning(traceback_string)

            return ResponseError(message =
                'Exception:\n' + traceback_string)


class XlsxBulkDisconnect(graphene.Mutation):
    """
    Parses uploaded XLSX file, disconnects highlighted cognates.

    Example:

      curl 'http://localhost:6543/graphql'
        -H 'Content-Type: multipart/form-data'
        -H 'Cookie: locale_id=2; auth_tkt=$TOKEN; client_id=$ID'
        -F operations='{
          "query":
            "mutation xlsxBulkDisconnect($xlsxFile: Upload) {
              xlsx_bulk_disconnect(
                xlsx_file: $xlsxFile,
                debug_flag: false)
              {
                triumph
                entry_info_count
                skip_count
                group_count
                disconnect_count
              }
            }",
          "variables": {
            "xlsxFile": null
          }}'
        -F map='{ "1": ["variables.xlsx_file"] }'
        -F 1=@"./Словарь_среднечепецкого_диалекта_северного_наречия_удмуртского_я.xlsx"
  
      Set $TOKEN and $ID to valid admin user authentication info.
  
      To use in a shell, join into a single line or add escaping backslashes at the end of the lines.
    """

    class Arguments:
        xlsx_file = Upload()
        debug_flag = graphene.Boolean()

    entry_info_count = graphene.Int()
    skip_count = graphene.Int()

    group_count = graphene.Int()
    disconnect_count = graphene.Int()

    triumph = graphene.Boolean()

    sql_search_str = ('''

        select
        L.client_id,
        L.object_id

        from
        dictionary D,
        dictionaryperspective P,
        dictionaryperspectivetofield Fc,
        dictionaryperspectivetofield Fl,
        lexicalentry L

        where

        (D.translation_gist_client_id, D.translation_gist_object_id) in (

          select distinct
          parent_client_id,
          parent_object_id

          from
          translationatom

          where
          content = :d_name and
          marked_for_deletion = false) and

        D.marked_for_deletion = false and

        P.parent_client_id = D.client_id and
        P.parent_object_id = D.object_id and
        P.marked_for_deletion = false and

        exists (

          select 1

          from
          translationatom A

          where
          A.parent_client_id = P.translation_gist_client_id and
          A.parent_object_id = P.translation_gist_object_id and
          A.marked_for_deletion = false and
          A.content = :p_name) and

        Fc.parent_client_id = P.client_id and
        Fc.parent_object_id = P.object_id and
        Fc.marked_for_deletion = false and

        Fl.parent_client_id = P.client_id and
        Fl.parent_object_id = P.object_id and
        Fl.marked_for_deletion = false and

        (
          Fc.field_client_id = 66 and Fc.field_object_id = 8 and
          Fl.field_client_id = 66 and Fl.field_object_id = 10 or

          exists (

            select 1

            from
            field F,
            translationatom A

            where
            F.client_id = Fc.field_client_id and
            F.object_id = Fc.field_object_id and
            A.parent_client_id = F.translation_gist_client_id and
            A.parent_object_id = F.translation_gist_object_id and
            A.marked_for_deletion = false and
            A.locale_id = 2 and
            lower(regexp_replace(A.content, '\W+', '', 'g')) =
              'phonemictranscription') and

          exists (

            select 1

            from
            field F,
            translationatom A

            where
            F.client_id = Fl.field_client_id and
            F.object_id = Fl.field_object_id and
            A.parent_client_id = F.translation_gist_client_id and
            A.parent_object_id = F.translation_gist_object_id and
            A.marked_for_deletion = false and
            A.locale_id = 2 and
            lower(regexp_replace(A.content, '\W+', '', 'g')) =
              'meaning') or

          exists (

            select 1

            from
            field F,
            translationatom A

            where
            F.client_id = Fc.field_client_id and
            F.object_id = Fc.field_object_id and
            A.parent_client_id = F.translation_gist_client_id and
            A.parent_object_id = F.translation_gist_object_id and
            A.marked_for_deletion = false and
            A.locale_id = 1 and
            lower(regexp_replace(A.content, '\W+', '', 'g')) =
              'фонологическаятранскрипция') and

          exists (

            select 1

            from
            field F,
            translationatom A

            where
            F.client_id = Fl.field_client_id and
            F.object_id = Fl.field_object_id and
            A.parent_client_id = F.translation_gist_client_id and
            A.parent_object_id = F.translation_gist_object_id and
            A.marked_for_deletion = false and
            A.locale_id = 1 and
            lower(regexp_replace(A.content, '\W+', '', 'g')) =
              'значение') or

          exists (

            select 1

            from
            field F,
            translationatom A

            where
            F.client_id = Fc.field_client_id and
            F.object_id = Fc.field_object_id and
            A.parent_client_id = F.translation_gist_client_id and
            A.parent_object_id = F.translation_gist_object_id and
            A.marked_for_deletion = false and
            A.locale_id = 2 and
            lower(regexp_replace(A.content, '\W+', '', 'g')) =
              'transcriptionofparadigmaticforms') and

          exists (

            select 1

            from
            field F,
            translationatom A

            where
            F.client_id = Fl.field_client_id and
            F.object_id = Fl.field_object_id and
            A.parent_client_id = F.translation_gist_client_id and
            A.parent_object_id = F.translation_gist_object_id and
            A.marked_for_deletion = false and
            A.locale_id = 2 and
            lower(regexp_replace(A.content, '\W+', '', 'g')) =
              'translationofparadigmaticforms') or

          exists (

            select 1

            from
            field F,
            translationatom A

            where
            F.client_id = Fc.field_client_id and
            F.object_id = Fc.field_object_id and
            A.parent_client_id = F.translation_gist_client_id and
            A.parent_object_id = F.translation_gist_object_id and
            A.marked_for_deletion = false and
            A.locale_id = 1 and
            lower(regexp_replace(A.content, '\W+', '', 'g')) =
              'транскрипцияпарадигматическихформ') and

          exists (

            select 1

            from
            field F,
            translationatom A

            where
            F.client_id = Fl.field_client_id and
            F.object_id = Fl.field_object_id and
            A.parent_client_id = F.translation_gist_client_id and
            A.parent_object_id = F.translation_gist_object_id and
            A.marked_for_deletion = false and
            A.locale_id = 1 and
            lower(regexp_replace(A.content, '\W+', '', 'g')) =
              'переводпарадигматическихформ') or

          :p_name ~* '.*starling.*' and

          (
            exists (

              select 1

              from
              field F,
              translationatom A

              where
              F.client_id = Fc.field_client_id and
              F.object_id = Fc.field_object_id and
              A.parent_client_id = F.translation_gist_client_id and
              A.parent_object_id = F.translation_gist_object_id and
              A.marked_for_deletion = false and
              A.locale_id = 2 and
              lower(regexp_replace(A.content, '\W+', '', 'g')) =
                'protoform') and

            exists (

              select 1

              from
              field F,
              translationatom A

              where
              F.client_id = Fl.field_client_id and
              F.object_id = Fl.field_object_id and
              A.parent_client_id = F.translation_gist_client_id and
              A.parent_object_id = F.translation_gist_object_id and
              A.marked_for_deletion = false and
              A.locale_id = 2 and
              lower(regexp_replace(A.content, '\W+', '', 'g')) =
                'protoformmeaning') or

            exists (

              select 1

              from
              field F,
              translationatom A

              where
              F.client_id = Fc.field_client_id and
              F.object_id = Fc.field_object_id and
              A.parent_client_id = F.translation_gist_client_id and
              A.parent_object_id = F.translation_gist_object_id and
              A.marked_for_deletion = false and
              A.locale_id = 1 and
              lower(regexp_replace(A.content, '\W+', '', 'g')) =
                'праформа') and

            exists (

              select 1

              from
              field F,
              translationatom A

              where
              F.client_id = Fl.field_client_id and
              F.object_id = Fl.field_object_id and
              A.parent_client_id = F.translation_gist_client_id and
              A.parent_object_id = F.translation_gist_object_id and
              A.marked_for_deletion = false and
              A.locale_id = 1 and
              lower(regexp_replace(A.content, '\W+', '', 'g')) =
                'значениепраформы'))) and

        L.parent_client_id = P.client_id and
        L.parent_object_id = P.object_id and
        L.marked_for_deletion = false and

        exists (

          select 1

          from
          public.entity E

          where
          E.parent_client_id = L.client_id and
          E.parent_object_id = L.object_id and
          E.field_client_id = 66 and
          E.field_object_id = 25 and
          E.marked_for_deletion = false and
          E.content in {}){}{};

        ''')

    sql_tag_str = ('''

        (

          select distinct
          E.content

          from
          lexicalentry L,
          public.entity E

          where

          L.parent_client_id = {} and
          L.parent_object_id = {} and
          L.marked_for_deletion = false{}

          and

          E.parent_client_id = L.client_id and
          E.parent_object_id = L.object_id and
          E.field_client_id = 66 and
          E.field_object_id = 25 and
          E.marked_for_deletion = false)

        ''')

    sql_xcript_str = ('''

        and

        (
          select count(*)

          from
          public.entity E

          where
          E.parent_client_id = L.client_id and
          E.parent_object_id = L.object_id and
          E.field_client_id = Fc.field_client_id and
          E.field_object_id = Fc.field_object_id and
          E.marked_for_deletion = false and
          E.content ~ :xc_regexp
        )
        = {}

        ''')

    sql_xlat_str = ('''

        and

        (
          select count(*)

          from
          public.entity E

          where
          E.parent_client_id = L.client_id and
          E.parent_object_id = L.object_id and
          E.field_client_id = Fl.field_client_id and
          E.field_object_id = Fl.field_object_id and
          E.marked_for_deletion = false and
          E.content ~ :xl_regexp
        )
        = {}

        ''')

    def escape(string):
        """
        Escapes special regexp characters in literal strings for PostgreSQL regexps, see
        https://stackoverflow.com/questions/4202538/escape-regex-special-characters-in-a-python-string/12012114.
        """

        return re.sub(r'([!$()*+.:<=>?[\\\]^{|}-])', r'\\\1', string)

    @staticmethod
    def get_entry_id(
        perspective_id,
        entry_info,
        row_index_set,
        tag_table_name):
        """
        Tries to get id of a cognate entry.
        """

        (content_info,
            dp_name,
            xcript_tuple,
            xlat_tuple) = (

            entry_info)

        # We have to have at least some entry info.

        if not xcript_tuple and not xlat_tuple and not content_info:
            return None

        d_name, p_name = (
            dp_name.split(' › '))

        param_dict = {
            'd_name': d_name,
            'p_name': p_name}

        if xcript_tuple:

            param_dict.update({

                'xc_regexp':

                    r'^\s*(' +

                    '|'.join(
                        XlsxBulkDisconnect.escape(xcript)
                        for xcript in xcript_tuple) +

                    r')\s*$'})

        if xlat_tuple:

            param_dict.update({

                'xl_regexp':

                    r'^\s*(' +

                    '|'.join(
                        XlsxBulkDisconnect.escape(xlat)
                        for xlat in xlat_tuple) +

                    r')\s*$'})

        # Trying to find the entry.

        if content_info:

            sql_str_a_list = []

            for i, (field_id_set, content_str) in enumerate(content_info):

                sql_str_a_list.append(

                    '''
                    and

                    exists (

                      select 1

                      from
                      public.entity E

                      where
                      E.parent_client_id = L.client_id and
                      E.parent_object_id = L.object_id and
                      E.marked_for_deletion = false and
                      (E.field_client_id, E.field_object_id) in ({}) and
                      E.content = :content_{})
                    '''

                    .format(
                        ', '.join(map(str, field_id_set)),
                        i))

                param_dict[
                    'content_{}'.format(i)] = content_str

            sql_str_a = (

                XlsxBulkDisconnect.sql_tag_str.format(
                    perspective_id[0],
                    perspective_id[1],
                    ''.join(sql_str_a_list)))

        else:

            sql_str_a = (

                '(select * from {})'.format(
                    tag_table_name))

        sql_str_b = (

            XlsxBulkDisconnect.sql_xcript_str.format(len(xcript_tuple))
                if xcript_tuple else '')

        sql_str_c = (

            XlsxBulkDisconnect.sql_xlat_str.format(len(xlat_tuple))
                if xlat_tuple else '')

        sql_str = (

            XlsxBulkDisconnect.sql_search_str.format(
                sql_str_a,
                sql_str_b,
                sql_str_c))

        result_list = (

            DBSession

                .execute(
                    sql_str,
                    param_dict)

                .fetchall())

        result_list = [

            (entry_cid, entry_oid)
            for entry_cid, entry_oid in result_list]

        log.debug(
            '\nresult_list: {}'.format(
                result_list))

        # If we haven't found anything, no problem, just going on ahead.

        if not result_list:
            return None

        # We shouldn't have any duplicate results.

        result_set = set(result_list)

        if len(result_set) < len(result_list):

            log.warning(

                '\n' +

                str(
                    sqlalchemy
                        .text(sql_str)
                        .bindparams(**param_dict)
                        .compile(compile_kwargs = {'literal_binds': True})) +

                '\nresult_list: {}'
                '\nresult_set: {}'.format(
                    result_list,
                    result_set))

            result_list = list(result_set)

        # If we've got the unambiguous entry info, ok, cool, otherwise no problem, skipping this and going
        # ahead.

        if len(result_list) <= len(row_index_set):
            return result_list

        return None

    @staticmethod
    def mutate(root, info, **args):

        __debug_flag__ = args.get('debug_flag', False)

        try:

            client_id = info.context.client_id
            client = DBSession.query(Client).filter_by(id = client_id).first()

            if not client or client.user_id != 1:
                return ResponseError('Only administrator can bulk disconnect.')

            request = info.context.request

            if '1' not in request.POST:
                return ResponseError('XLSX file is required.')

            multipart = request.POST.pop('1')

            xlsx_file_name = multipart.filename
            xlsx_file = multipart.file

            log.debug(
                '\n{}\n{}'.format(
                    xlsx_file_name,
                    type(xlsx_file)))

            settings = (
                request.registry.settings)

            # Processing XLSX workbook, assuming each worksheet has data of a single perspective.

            workbook = (
                openpyxl.load_workbook(xlsx_file))

            entry_info_count = 0
            skip_count = 0

            group_count = 0
            disconnect_count = 0

            tag_table_name = None

            for sheet_name in workbook.sheetnames:

                worksheet = workbook[sheet_name]

                # Assuming the first row has field names.

                field_name_list = []
                cognates_index = None

                for i in itertools.count(1):

                    cell = worksheet.cell(1, i)

                    if cell.value:

                        field_name_list.append(cell.value)
                        cognates_index = i

                    else:
                        break

                # Trying to parse perspective's fields.

                (perspective_name,
                    perspective_cid,
                    perspective_oid) = (

                    re.match(
                        r'^(.*)_(\d+)_(\d+)$',
                        sheet_name)

                        .groups())

                perspective_id = (
                    perspective_cid, perspective_oid)

                log.debug(
                    '\nperspective: \'{}\' {}/{}'
                    '\nfield_name_list:\n{}'.format(
                        perspective_name,
                        perspective_cid,
                        perspective_oid,
                        field_name_list))

                field_id_set_list = []

                for field_name in field_name_list[:-1]:

                    result_list = (

                        DBSession

                            .query(
                                dbField.client_id,
                                dbField.object_id)

                            .filter(
                                dbColumn.parent_client_id == perspective_cid,
                                dbColumn.parent_object_id == perspective_oid,
                                dbColumn.marked_for_deletion == False,

                                tuple_(
                                    dbColumn.field_client_id,
                                    dbColumn.field_object_id)

                                    .in_(
                                        sqlalchemy.text('select * from text_field_id_view')),

                                dbField.client_id == dbColumn.field_client_id,
                                dbField.object_id == dbColumn.field_object_id,
                                dbField.marked_for_deletion == False,
                                dbTranslationAtom.parent_client_id == dbField.translation_gist_client_id,
                                dbTranslationAtom.parent_object_id == dbField.translation_gist_object_id,
                                dbTranslationAtom.marked_for_deletion == False,
                                dbTranslationAtom.content == field_name)

                            .distinct()

                            .all())

                    field_id_set_list.append(

                        tuple(set(
                            tuple(field_id) for field_id in result_list)

                            if result_list else None))

                # Getting all possible cognate tags of the entries of the perspective.

                if tag_table_name is None:

                    tag_table_name = (

                        'tag_table_' +
                        str(uuid.uuid4()).replace('-', '_'))

                    DBSession.execute('''

                        create temporary table

                        {} (
                          tag TEXT,
                          primary key (tag))

                        on commit drop;

                        '''.format(
                            tag_table_name))

                else:

                    DBSession.execute(
                        'truncate table {};'.format(
                            tag_table_name))

                DBSession.execute('''

                    insert into {}

                    select
                    E.content

                    from
                    lexicalentry L,
                    public.entity E

                    where
                    L.parent_client_id = {} and
                    L.parent_object_id = {} and
                    L.marked_for_deletion = false and
                    E.parent_client_id = L.client_id and
                    E.parent_object_id = L.object_id and
                    E.field_client_id = 66 and
                    E.field_object_id = 25 and
                    E.marked_for_deletion = false

                    on conflict do nothing;

                    '''.format(
                        tag_table_name,
                        perspective_cid,
                        perspective_oid))

                # Processing cognate groups.

                color_set = set()

                content_info = None

                entry_info_list = []
                entry_info_dict = collections.defaultdict(set)

                entry_content_info = None
                entry_dp_name = None

                for i in range(2, worksheet.max_row):

                    # Getting text field info if we have any.

                    previous_content_info = content_info

                    content_info = []
                    content_flag = False

                    for j, field_id_set in enumerate(field_id_set_list):

                        field_cell = worksheet.cell(i, j + 1)

                        if field_cell.value:

                            content_flag = True

                            if field_id_set is not None:

                                content_info.append(
                                    (field_id_set, field_cell.value))

                    if content_flag:

                        content_info = (
                            tuple(content_info) if content_info else None)

                    else:

                        content_info = (
                            previous_content_info)

                    # Do we have a beginning of another cognate entry info?

                    dp_cell = worksheet.cell(i, cognates_index)

                    if dp_cell.value:

                        if (entry_dp_name and
                            entry_highlight_list and
                            len(entry_highlight_list) >= len(entry_xcript_list)):

                            entry_info = (
                                entry_content_info,
                                entry_dp_name,
                                tuple(xcript for xcript in entry_xcript_list if xcript),
                                tuple(xlat[1:-1] for xlat in entry_xlat_list))

                            if entry_info not in entry_info_dict:
                                entry_info_list.append(entry_info)

                            entry_info_dict[entry_info].add(entry_row_index)

                        entry_dp_name = dp_cell.value
                        entry_content_info = content_info

                        entry_row_index = i

                        entry_xcript_list = []
                        entry_xlat_list = []

                        entry_highlight_list = []

                    # Do we have transcription and / or translation, is transcription highlighted?

                    xc_cell = worksheet.cell(i, cognates_index + 1)

                    if xc_cell.value or dp_cell.value:

                        entry_xcript_list.append(xc_cell.value)

                        color = xc_cell.fill.fgColor.rgb

                        if color != '00000000':
                            color_set.add(color)

                        if color == 'FFFFFF00':
                            entry_highlight_list.append(xc_cell.value)

                    xl_cell = worksheet.cell(i, cognates_index + 2)

                    if xl_cell.value:
                        entry_xlat_list.append(xl_cell.value)

                log.debug(
                    '\n' +
                    pprint.pformat(
                        entry_info_list, width = 192))

                # Processing highlighted entries.

                entry_id_list = []
                skip_list = []

                for entry_info in entry_info_list:

                    id_list = (

                        XlsxBulkDisconnect.get_entry_id(
                            perspective_id,
                            entry_info,
                            entry_info_dict[entry_info],
                            tag_table_name))

                    if id_list:

                        entry_id_list.extend(
                            id_list)

                    else:

                        skip_list.append((
                            entry_info,
                            entry_info_dict[entry_info]))

                log.debug(
                    '\nentry_id_list:\n{}'
                    '\nskip_list:\n{}'
                    '\nlen(entry_info_list): {}'
                    '\nlen(entry_id_list): {}'
                    '\nlen(skip_list): {}'.format(

                        pprint.pformat(
                            entry_id_list, width = 192),

                        pprint.pformat(
                            skip_list, width = 192),

                        len(entry_info_list),
                        len(entry_id_list),
                        len(skip_list)))

                # Performing disconnects.

                entry_id_set = set(entry_id_list)
                already_set = set()

                perspective_group_count = 0
                perspective_disconnect_count = 0

                for entry_id in entry_id_list:

                    if entry_id in already_set:
                        continue

                    result_list = (

                        DBSession

                            .execute(
                                'select * from linked_group(66, 25, {}, {})'.format(
                                    *entry_id))

                            .fetchall())

                    cognate_id_set = (

                        set(
                            (entry_cid, entry_oid)
                            for entry_cid, entry_oid in result_list))

                    disconnect_set = (
                        cognate_id_set & entry_id_set)

                    leave_set = (
                        cognate_id_set - disconnect_set)

                    log.debug(
                        '\ncognate_id_set ({}):\n{}'
                        '\ndisconnect_set ({}):\n{}'
                        '\nleave_set ({}):\n{}'.format(
                        len(cognate_id_set),
                        cognate_id_set,
                        len(disconnect_set),
                        disconnect_set,
                        len(leave_set),
                        leave_set))

                    # Disconnecting highlighted entries, see `class DeleteGroupingTags()`.

                    entity_list = (

                        DBSession

                            .query(dbEntity)

                            .filter(

                                tuple_(
                                    dbEntity.parent_client_id,
                                    dbEntity.parent_object_id)
                                    .in_(disconnect_set),

                                dbEntity.field_client_id == 66,
                                dbEntity.field_object_id == 25,
                                dbEntity.marked_for_deletion == False)

                            .all())

                    for entity in entity_list:

                        if 'desktop' in settings:

                            real_delete_entity(
                                entity,
                                settings)

                        else:

                            del_object(
                                entity,
                                'xlsx_bulk_disconnect',
                                client_id)

                    # Connecting disconnected entries together, if there is more than one, see
                    # `class ConnectLexicalEntries()`.

                    n = 10

                    rnd = (
                        random.SystemRandom())

                    choice_str = (
                        string.digits + string.ascii_letters)

                    tag_str = (

                        time.asctime(time.gmtime()) +

                        ''.join(
                            rnd.choice(choice_str)
                            for c in range(n)))

                    for entry_id in disconnect_set:

                        dbEntity(
                            client_id = client_id,
                            parent_client_id = entry_id[0],
                            parent_object_id = entry_id[1],
                            field_client_id = 66,
                            field_object_id = 25,
                            content = tag_str,
                            published = True,
                            accepted = True)

                    already_set.update(disconnect_set)

                    perspective_group_count += 1
                    perspective_disconnect_count += len(disconnect_set)

                # Finished this perspective.

                log.debug(
                    '\n\'{}\' {}/{}:'
                    '\nperspective_group_count: {}'
                    '\nperspective_disconnect_count: {}'.format(
                        perspective_name,
                        perspective_cid,
                        perspective_oid,
                        perspective_group_count,
                        perspective_disconnect_count))

                entry_info_count += len(entry_info_list)
                skip_count += len(skip_list)

                group_count += perspective_group_count
                disconnect_count += perspective_disconnect_count

            # Finished bulk disconnects.

            log.debug(
                '\n{} perspectives'
                '\nentry_info_count: {}'
                '\nskip_count: {}'
                '\ngroup_count: {}'
                '\ndisconnect_count: {}'.format(
                    len(workbook.sheetnames),
                    entry_info_count,
                    skip_count,
                    group_count,
                    disconnect_count))

            return (

                XlsxBulkDisconnect(
                    entry_info_count = entry_info_count,
                    skip_count = skip_count,
                    group_count = group_count,
                    disconnect_count = disconnect_count,
                    triumph = True))

        except Exception as exception:

            traceback_string = (

                ''.join(
                    traceback.format_exception(
                        exception, exception, exception.__traceback__))[:-1])

            log.warning('xlsx_bulk_disconnect: exception')
            log.warning(traceback_string)

            return (

                ResponseError(
                    'Exception:\n' + traceback_string))


if __name__ == '__main__':

    # Cognate analysis debugging script.

    import pdb
    import sys

    dictionary_count, line_count = map(
        int, re.findall(r'\d+', sys.argv[1])[:2])

    print(dictionary_count, line_count)

    with open(sys.argv[1], 'rb') as input_file:
        input = input_file.read().decode('utf-16')

    input_buffer = ctypes.create_unicode_buffer(input)

    output_buffer_size = cognate_analysis_f(
        None, dictionary_count, line_count, None, 1)

    print(output_buffer_size)

    output_buffer = ctypes.create_unicode_buffer(output_buffer_size + 256)

    result = cognate_analysis_f(
        input_buffer, dictionary_count, line_count, output_buffer, 1)

    print(result)

