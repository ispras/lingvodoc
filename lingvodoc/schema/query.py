import base64
import collections
import copy
import ctypes
import datetime
import gzip
import hashlib
import io
import itertools
import logging
import math
import os.path
import pickle
import pprint
import re
import shutil
import textwrap
import time
import traceback
import types
import unicodedata
import urllib.parse
import uuid
import zipfile

import graphene
import graphene.types
from graphql.language.ast import BooleanValue, IntValue, ListValue

import lingvodoc.utils as utils
from lingvodoc.utils.deletion import real_delete_entity
from lingvodoc.utils.elan_functions import tgt_to_eaf
import requests
from lingvodoc.schema.gql_entity import (
    Entity,
    CreateEntity,
    UpdateEntity,
    DeleteEntity,
    UpdateEntityContent,
    BulkCreateEntity,
    ApproveAllForUser,
    BulkUpdateEntityContent,
    is_subject_for_parsing)
from lingvodoc.schema.gql_column import (
    Column,
    CreateColumn,
    UpdateColumn,
    DeleteColumn
)
from lingvodoc.schema.gql_basegroup import (
    BaseGroup,
    CreateBasegroup,
    AddUserToBasegroup)
from lingvodoc.schema.gql_group import (
    Group
)
from lingvodoc.schema.gql_organization import (
    Organization,
    CreateOrganization,
    UpdateOrganization,
    DeleteOrganization
)
# from lingvodoc.schema.gql_publishingentity import (
#     PublishingEntity
# )
from lingvodoc.schema.gql_translationatom import (
    TranslationAtom,
    CreateTranslationAtom,
    UpdateTranslationAtom,
    DeleteTranslationAtom)
from lingvodoc.schema.gql_translationgist import (
    TranslationGist,
    CreateTranslationGist,
    DeleteTranslationGist
)
from lingvodoc.schema.gql_userblobs import (
    UserBlobs,
    CreateUserBlob,
    DeleteUserBlob
)
from lingvodoc.schema.gql_field import (
    Field,
    CreateField,
    # UpdateField,
    # DeleteField
)

from lingvodoc.schema.gql_dictionary import (
    Dictionary,
    CreateDictionary,
    UpdateDictionary,
    UpdateDictionaryStatus,
    AddDictionaryRoles,
    DeleteDictionaryRoles,
    DeleteDictionary,
    UndeleteDictionary,
    UpdateDictionaryAtom)

from lingvodoc.schema.gql_search import (
    AdvancedSearch,
    AdvancedSearchSimple,
    EafSearch)

from lingvodoc.schema.gql_lexicalentry import (
    LexicalEntry,
    CreateLexicalEntry,
    DeleteLexicalEntry,
    BulkDeleteLexicalEntry,
    BulkUndeleteLexicalEntry,
    BulkCreateLexicalEntry,
    ConnectLexicalEntries,
    DeleteGroupingTags,
)

from lingvodoc.schema.gql_language import (
    Language,
    CreateLanguage,
    UpdateLanguage,
    DeleteLanguage,
    MoveLanguage,
    UpdateLanguageAtom)
from lingvodoc.schema.gql_merge import MergeBulk
from lingvodoc.schema.gql_dictionaryperspective import (
    DictionaryPerspective as Perspective,
    CreateDictionaryPerspective,
    UpdateDictionaryPerspective,
    UpdatePerspectiveStatus,
    AddPerspectiveRoles,
    DeletePerspectiveRoles,
    DeleteDictionaryPerspective,
    UndeleteDictionaryPerspective,
    UpdatePerspectiveAtom)
from lingvodoc.schema.gql_user import (
    User,
    CreateUser,
    UpdateUser,
    ActivateDeactivateUser
)
from lingvodoc.schema.gql_grant import (
    Grant,
    CreateGrant,
    UpdateGrant,
    # DeleteGrant
)
from lingvodoc.schema.gql_sync import (
    DownloadDictionary,
    DownloadDictionaries,
    Synchronize
)

from lingvodoc.schema.gql_holders import (
    AdditionalMetadata,
    client_id_check,
    CreatedAt,
    gql_none_value,
    LingvodocID,
    ObjectVal,
    PermissionException,
    get_published_translation_gist_id_cte_query,
    get_published_translation_gist_id_subquery_query,
    ResponseError,
    UnstructuredData,
    Upload,
)

from lingvodoc.schema.gql_userrequest import (
    UserRequest,
    CreateGrantPermission,
    AddDictionaryToGrant,
    AdministrateOrg,
    ParticipateOrg,
    AddDictionaryToOrganization,
    AcceptUserRequest,
    # DeleteUserRequest
)

from lingvodoc.schema.gql_parser import Parser
from lingvodoc.schema.gql_parserresult import DeleteParserResult, UpdateParserResult, ParserResult

import lingvodoc.acl as acl
import time
import random
import string

import lingvodoc.models as models
from lingvodoc.models import (
    BaseGroup as dbBaseGroup,
    Client,
    DBSession,
    Dictionary as dbDictionary,
    DictionaryPerspective as dbPerspective,
    DictionaryPerspectiveToField as dbColumn,
    ENGLISH_LOCALE,
    Email as dbEmail,
    Entity as dbEntity,
    Field as dbField,
    Grant as dbGrant,
    Group as dbGroup,
    Language as dbLanguage,
    LexicalEntry as dbLexicalEntry,
    Locale as dbLocale,
    Organization as dbOrganization,
    Parser as dbParser,
    ParserResult as dbParserResult,
    PublishingEntity as dbPublishingEntity,
    RUSSIAN_LOCALE,
    SLBigInteger,
    TranslationAtom as dbTranslationAtom,
    TranslationGist as dbTranslationGist,
    UnstructuredData as dbUnstructuredData,
    User as dbUser,
    UserBlobs as dbUserBlobs,
    UserRequest as dbUserRequest,
    ValencyAnnotationData as dbValencyAnnotationData,
    ValencyEafData as dbValencyEafData,
    ValencyInstanceData as dbValencyInstanceData,
    ValencyMergeData as dbValencyMergeData,
    ValencyMergeIdSequence as dbValencyMergeIdSequence,
    ValencyParserData as dbValencyParserData,
    ValencySentenceData as dbValencySentenceData,
    ValencySourceData as dbValencySourceData,
    user_to_group_association,
)
from pyramid.request import Request

from lingvodoc.utils.proxy import try_proxy, ProxyPass

import sqlalchemy

from sqlalchemy import (
    func,
    and_,
    or_,
    tuple_,
    create_engine,
    literal,
    union,
    cast,
    Boolean,
)

import sqlalchemy.dialects.postgresql as postgresql

from sqlalchemy.sql.elements import ColumnElement

import sqlalchemy.types

from zope.sqlalchemy import mark_changed
from lingvodoc.views.v2.utils import (
    view_field_from_object,
    storage_file,
    as_storage_file
)
from sqlalchemy.orm import aliased, joinedload

from sqlalchemy.sql.expression import Grouping
from sqlalchemy.sql.functions import coalesce
from lingvodoc.schema.gql_tasks import Task, DeleteTask
from lingvodoc.schema.gql_convert_dictionary import ConvertDictionary, ConvertFiveTiers
from pyramid.security import authenticated_userid

from lingvodoc.utils.phonology import (
    gql_phonology as utils_phonology,
    gql_phonology_skip_list as utils_phonology_skip_list,
    gql_phonology_tier_list as utils_phonology_tier_list,
    gql_phonology_link_perspective_data,
    gql_sound_and_markup)

from lingvodoc.utils import starling_converter, render_statement, ids_to_id_cte, ids_to_id_query

from lingvodoc.utils.search import (
    translation_gist_search,
    recursive_sort,
    eaf_words,
    find_all_tags,
    find_lexical_entries_by_tags,
    get_id_to_field_dict)

import lingvodoc.cache.caching as caching
from lingvodoc.cache.caching import initialize_cache, TaskStatus

import lingvodoc.views.v2.phonology as phonology
from lingvodoc.views.v2.phonology import (
    AudioPraatLike,
    format_textgrid_result,
    get_vowel_class,
    Phonology_Parameters,
    process_sound,
    process_sound_markup,
    process_textgrid)

from lingvodoc.views.v2.utils import anonymous_userid

from sqlite3 import connect
from lingvodoc.utils.merge import merge_suggestions
import tempfile

import lingvodoc.scripts.export_parser_result as export_parser_result
import lingvodoc.scripts.valency as valency
import lingvodoc.scripts.valency_verb_cases as valency_verb_cases

from lingvodoc.scripts.save_dictionary import (
    find_group_by_tags,
    save_dictionary as sync_save_dictionary)

from lingvodoc.views.v2.save_dictionary.core import async_save_dictionary
import json

from pyramid.httpexceptions import (
    HTTPError,
    HTTPOk
)

from lingvodoc.scripts import elan_parser
from lingvodoc.utils.creation import create_entity, edit_role

from lingvodoc.queue.celery import celery
from lingvodoc.schema.gql_holders import del_object

# So that matplotlib does not require display stuff, in particular, tkinter. See e.g. https://
# stackoverflow.com/questions/4931376/generating-matplotlib-graphs-without-a-running-x-server.
import matplotlib
matplotlib.use('Agg', warn = False)

from matplotlib.collections import LineCollection
from matplotlib import pyplot

import minio

from mpl_toolkits.mplot3d import Axes3D
from mpl_toolkits.mplot3d import proj3d

from mpl_toolkits.mplot3d.art3d import Line3DCollection

import numpy
import openpyxl
import pathvalidate
import psycopg2.errors
import pydub
import pylab
import pympi

import scipy.optimize
import scipy.sparse.csgraph

import sklearn.decomposition
import sklearn.manifold
import sklearn.metrics
import sklearn.mixture

import transaction
import xlsxwriter

from lingvodoc.schema.gql_copy_field import CopySingleField, CopySoundMarkupFields

import lingvodoc.version

from lingvodoc.schema.gql_parserresult import ExecuteParser

from lingvodoc.cache.caching import CACHE

import lingvodoc.scripts.docx_import as docx_import

import pandas as pd
from pretty_html_table import build_table


# Setting up logging.
log = logging.getLogger(__name__)


# Mikhail Oslon's analysis functions.

try:

    liboslon = ctypes.CDLL('liboslon.so')

    phonemic_analysis_f = liboslon.PhonemicAnalysis_GetAllOutput

    cognate_analysis_f = liboslon.CognateAnalysis_GetAllOutput
    cognate_acoustic_analysis_f = liboslon.CognateAcousticAnalysis_GetAllOutput
    cognate_distance_analysis_f = liboslon.CognateDistanceAnalysis_GetAllOutput
    cognate_reconstruction_f = liboslon.CognateReconstruct_GetAllOutput
    cognate_reconstruction_multi_f = liboslon.CognateMultiReconstruct_GetAllOutput
    cognate_suggestions_f = liboslon.GuessCognates_GetAllOutput

except:

    log.warning('liboslon.so')

    phonemic_analysis_f = None

    cognate_analysis_f = None
    cognate_acoustic_analysis_f = None
    cognate_distance_analysis_f = None
    cognate_reconstruction_f = None
    cognate_reconstruction_multi_f = None
    cognate_suggestions_f = None


class TierList(graphene.ObjectType):
    tier_count = graphene.Field(ObjectVal)
    total_count = graphene.Int()


class SkipList(graphene.ObjectType):
    markup_count = graphene.Int()
    neighbour_list = graphene.Field(ObjectVal)
    skip_list = graphene.Field(ObjectVal)
    total_neighbour_count = graphene.Int()
    total_skip_count = graphene.Int()


class Link_Perspective_Data(graphene.ObjectType):
    field_data_list = graphene.List(ObjectVal)
    perspective_id_list = graphene.List(LingvodocID)


class LexicalEntriesAndEntities(graphene.ObjectType):
    entities = graphene.List(Entity)
    lexical_entries = graphene.List(LexicalEntry)


class Permissions(graphene.ObjectType):
    edit = graphene.List(Perspective)
    view = graphene.List(Perspective)
    publish = graphene.List(Perspective)
    limited = graphene.List(Perspective)


class StarlingField(graphene.InputObjectType):
    starling_name = graphene.String(required=True)
    starling_type = graphene.Int(required=True)
    field_id = LingvodocID(required=True)
    fake_id = graphene.String()
    link_fake_id = LingvodocID()  # graphene.String()


class DialeqtInfo(graphene.ObjectType):
    dictionary_name = graphene.String()
    dialeqt_id = graphene.String()


class MergeSuggestions(graphene.ObjectType):
    match_result = graphene.List(ObjectVal)
    user_has_permissions = graphene.Boolean()


class LanguageTree(graphene.ObjectType):
    tree = ObjectVal()
    languages = graphene.List(Language)


def get_dict_attributes(sqconn):
    dict_trav = sqconn.cursor()
    dict_trav.execute("""SELECT
                        dict_name,
                        dict_identificator,
                        dict_description
                        FROM
                        dict_attributes
                        WHERE
                        id = 1;""")
    req = dict()
    for dictionary in dict_trav:
        req['dictionary_name'] = dictionary[0]
        req['dialeqt_id'] = dictionary[1]
    return req


class Resolver_Selection(object):
    """
    Stores a set of selected fields together with columns required to resolve them.
    """

    def __init__(self, field_column_dict):

        self.field_column_dict = field_column_dict

        self.field_set = set()
        self.column_set = set()

        self.column_list = []

        self.object_flag = False

    def __call__(self, *args):

        self.add(*args)

    def __contains__(self, field_str):

        return (
            field_str in self.field_column_dict)

    def add(self, *args):

        for value in args:

            if isinstance(value, str):

                if value in self.field_set:
                    return

                self.field_set.add(value)

                for column in (
                    self.field_column_dict.get(value, ())):

                    if column not in self.column_set:

                        self.column_set.add(column)
                        self.column_list.append(column)

            elif value not in self.column_set:

                self.column_set.add(value)
                self.column_list.append(value)

    def remove(self, *args):

        for arg in args:

            self.field_set.remove(arg)


def language_resolver_args(
    id_list = None,
    only_in_toc = False,
    only_with_dictionaries_recursive = False,
    dictionary_category = None,
    dictionary_published = None,
    language_id = None,
    by_grants = False,
    grant_id = None,
    by_organizations = False,
    organization_id = None,
    in_tree_order = False):

    return (

        types.SimpleNamespace(
            id_list = id_list,
            only_in_toc = only_in_toc,
            only_with_dictionaries_recursive = only_with_dictionaries_recursive,
            dictionary_category = dictionary_category,
            dictionary_published = dictionary_published,
            language_id = language_id,
            by_grants = by_grants,
            grant_id = grant_id,
            by_organizations = by_organizations,
            organization_id = organization_id,
            in_tree_order = in_tree_order))


def language_toc_sql(self):

    toc_table_name = (

        'toc_language_ids_' +
         str(uuid.uuid4()).replace('-', '_'))

    count_table_name = (

        'dictionary_counts_' +
         str(uuid.uuid4()).replace('-', '_'))

    result_table_name = (

        'recursive_counts_' +
         str(uuid.uuid4()).replace('-', '_'))

    sql_str = f'''

        create temporary table
          {toc_table_name}
          on commit drop as

          select
            client_id,
            object_id,
            parent_client_id,
            parent_object_id

          from
            language

          where
            marked_for_deletion = false and
            (additional_metadata -> 'toc_mark') :: boolean;

        create temporary table
          {count_table_name}
          on commit drop as

          with recursive

          underlying_language_ids as (

            select *
            from {toc_table_name}

            union

            select
              L.client_id,
              L.object_id,
              L.parent_client_id,
              L.parent_object_id

            from
              language L,
              underlying_language_ids U

            where
              L.marked_for_deletion = false and
              L.parent_client_id = U.client_id and
              L.parent_object_id = U.object_id
          )

          select
            L.*, count(D.*)

          from
            underlying_language_ids L

          left outer join
            dictionary D

          on
            D.parent_client_id = L.client_id and
            D.parent_object_id = L.object_id and
            D.marked_for_deletion = false and
            D.category = 0

          group by
            L.client_id,
            L.object_id,
            L.parent_client_id,
            L.parent_object_id;

        create temporary table

          {result_table_name} (
            client_id BIGINT,
            object_id BIGINT,
            count BIGINT
          )

        on commit drop;

        do $$

        begin

        while exists (
          select 1 from {count_table_name}) loop

          with

          iteration_counts as (

            select C1.*

              from
                {count_table_name} C1

              left outer join
                {count_table_name} C2

              on
                C1.client_id = C2.parent_client_id and
                C1.object_id = C2.parent_object_id

              where
                C2 is null
          ),

          delete_cte as (

            delete

            from
              {count_table_name} D

            using
              iteration_counts I

            where
              D.client_id = I.client_id and
              D.object_id = I.object_id
          ),

          update_cte as (

            update
              {count_table_name} D

            set
              count = D.count + S.sum

            from (

              select
                parent_client_id,
                parent_object_id,
                sum(count)

              from
                iteration_counts

              group by
                parent_client_id,
                parent_object_id

            ) S

            where
              D.client_id = S.parent_client_id and
              D.object_id = S.parent_object_id
          )

          insert into
            {result_table_name}

          select
            client_id,
            object_id,
            count

          from
            iteration_counts;

        end loop;

        end

        $$;

        select
          R.*

        from
          {result_table_name} R,
          {toc_table_name} T

        where
          R.client_id = T.client_id and
          R.object_id = T.object_id;

    '''

    return (

        DBSession
            .execute(sql_str)
            .fetchall())

def language_toc_python(self):

    # Getting base counts.

    toc_table_name = (

        'toc_language_ids_' +
         str(uuid.uuid4()).replace('-', '_'))

    sql_str = f'''

        create temporary table
          {toc_table_name}
          on commit drop as

          select
            client_id,
            object_id,
            parent_client_id,
            parent_object_id

          from
            language

          where
            marked_for_deletion = false and
            (additional_metadata -> 'toc_mark') :: boolean;

        with recursive

        underlying_language_ids as (

          select *
          from {toc_table_name}

          union

          select
            L.client_id,
            L.object_id,
            L.parent_client_id,
            L.parent_object_id

          from
            language L,
            underlying_language_ids U

          where
            L.marked_for_deletion = false and
            L.parent_client_id = U.client_id and
            L.parent_object_id = U.object_id
        )

        select
          U.*,
          count(D.*)
            filter (where D.client_id is not null)

        from
          underlying_language_ids U

        left outer join
          dictionary D

        on
          D.parent_client_id = U.client_id and
          D.parent_object_id = U.object_id and
          D.marked_for_deletion = false and
          D.category = 0

        group by
          U.client_id,
          U.object_id,
          U.parent_client_id,
          U.parent_object_id;

    '''

    count_list = (

        DBSession
            .execute(sql_str)
            .fetchall())

    from_to_dict = collections.defaultdict(list)

    count_dict = {}

    for client_id, object_id, parent_client_id, parent_object_id, count in count_list:

        id = (client_id, object_id)
        parent_id = (parent_client_id, parent_object_id)

        from_to_dict[parent_id].append(id)

        count_dict[id] = count

    # Getting language TOC ids, computing language TOC total counts.

    toc_id_list = [

        tuple(id) for id in

            DBSession
                .execute(f'select client_id, object_id from {toc_table_name}')
                .fetchall()]

    total_count_dict = {}

    def f(id):

        total_count = total_count_dict.get(id)

        if total_count is not None:
            return total_count

        total_count = (

            count_dict[id] +

            sum(f(to_id)
                for to_id in from_to_dict[id]))

        total_count_dict[id] = total_count

        return total_count

    return [
        (id[0], id[1], f(id))
        for id in toc_id_list]


def toc_translations_cache(self, toc_list):

    gist_id_list = (

        DBSession

            .query(
                dbLanguage.client_id,
                dbLanguage.object_id,
                dbLanguage.translation_gist_client_id,
                dbLanguage.translation_gist_object_id)

            .filter(

                tuple_(
                    dbLanguage.client_id,
                    dbLanguage.object_id)

                    .in_(
                        ids_to_id_query(
                            (client_id, object_id)
                            for client_id, object_id, count in toc_list)))

            .all())

    return {

        (client_id, object_id):
            models.get_translations(gist_client_id, gist_object_id)

        for client_id, object_id, gist_client_id, gist_object_id in gist_id_list}

def toc_translations_db(self, toc_list):

    translation_list = (

        DBSession

            .query(
                dbLanguage.client_id,
                dbLanguage.object_id,
                dbTranslationAtom.locale_id,
                dbTranslationAtom.content)

            .filter(

                tuple_(
                    dbLanguage.client_id,
                    dbLanguage.object_id)

                    .in_(
                        ids_to_id_query(
                            (client_id, object_id)
                            for client_id, object_id, count in toc_list)),

                dbTranslationAtom.parent_client_id == dbLanguage.translation_gist_client_id,
                dbTranslationAtom.parent_object_id == dbLanguage.translation_gist_object_id,
                dbTranslationAtom.marked_for_deletion == False)

            .all())

    content_dict = collections.defaultdict(list)

    for client_id, object_id, locale_id, content in translation_list:
        content_dict[(client_id, object_id)].append((locale_id, content))

    result_dict = {}

    for client_id, object_id, count in toc_list:

        id = (client_id, object_id)
        content_list = content_dict.get(id)

        result_dict[id] = (
            None if content_list is None else
                {str(key): value for key, value in content_list})

    return result_dict


class Language_Resolver(object):
    """
    Resolves list of languages, possibly with underlying dictionaries and their perspectives, for
    'languages' and 'language_tree' queries.
    """

    def __init__(
        self,
        info,
        field_asts,
        args,
        debug_flag = False):

        self.info = info

        self.field_asts = field_asts
        self.variable_values = info.variable_values

        self.args = args

        self.debug_flag = debug_flag

        self.dictionary_selection = None
        self.perspective_selection = None
        self.column_selection = None

        self.published_query = None
        self.published_condition_count = 0

    def argument_value(self, argument):

        try:

            return argument.value.value

        except AttributeError:

            return (

                self.variable_values.get(
                    argument.value.name.value, None))

    def parse_languages(self, field_asts):

        ls = (
            self.language_selection)

        already_set = set()

        if self.debug_flag:

            log.debug(
                f'\n field_asts (languages):\n {field_asts}')

        for field in field_asts:

            name_str = field.name.value

            if name_str in ls:

                ls(name_str)

            elif name_str == 'dictionary_count':

                if 'dictionary_count' in already_set:

                    ls.dictionary_count = None
                    continue

                already_set.add('dictionary_count')

                dc = (

                    types.SimpleNamespace(
                        recursive = None,
                        category = None,
                        published = None))

                ls.dictionary_count = dc

                for argument in field.arguments:

                    name_str = argument.name.value

                    if name_str == 'recursive':

                        dc.recursive = (
                            self.argument_value(argument))

                    elif name_str == 'category':

                        dc.category = (
                            self.argument_value(argument))

                    elif name_str == 'published':

                        dc.published = (
                            self.argument_value(argument))

                    else:

                        ls.dictionary_count = None
                        break

            elif name_str == 'dictionaries':

                if 'dictionaries' in already_set:

                    ls.dictionaries = None
                    continue

                already_set.add('dictionaries')

                d = (

                    types.SimpleNamespace(
                        deleted = None,
                        category = None,
                        published = None))

                ls.dictionaries = d

                for argument in field.arguments:

                    name_str = argument.name.value

                    if name_str == 'category':

                        d.category = (
                            self.argument_value(argument))

                    elif name_str == 'deleted':

                        d.deleted = (
                            self.argument_value(argument))

                    elif name_str == 'published':

                        d.published = (
                            self.argument_value(argument))

                    else:

                        ls.dictionaries = None
                        break

                if ls.dictionaries:

                    ds = (

                        Resolver_Selection({

                            'additional_metadata': (
                                dbDictionary.additional_metadata,),

                            'category': (
                                dbDictionary.category,),

                            'created_at': (
                                dbDictionary.created_at,),

                            'domain': (
                                dbDictionary.domain,),

                            'id': (
                                dbDictionary.client_id,
                                dbDictionary.object_id),

                            'marked_for_deletion': (
                                dbDictionary.marked_for_deletion,),

                            'parent_id': (
                                dbDictionary.parent_client_id,
                                dbDictionary.parent_object_id),

                            'state_translation_gist_id': (
                                dbDictionary.state_translation_gist_client_id,
                                dbDictionary.state_translation_gist_object_id),

                            'status_translations': (
                                dbDictionary.state_translation_gist_client_id,
                                dbDictionary.state_translation_gist_object_id),

                            'translation_gist_id': (
                                dbDictionary.translation_gist_client_id,
                                dbDictionary.translation_gist_object_id),

                            'translations': (
                                dbDictionary.translation_gist_client_id,
                                dbDictionary.translation_gist_object_id)}))

                    self.dictionary_selection = ds

                    self.parse_dictionaries(
                        field.selection_set.selections)

                    # Would need creation time to order dictionaries in a standard way, from newest to
                    # oldest.

                    ds('id')
                    ds('created_at')

                    ls('id')
                    ds('parent_id')

            elif name_str == 'translation':

                if 'translation' in already_set:

                    ls.translation = None
                    continue

                already_set.add('translation')

                t = (

                    types.SimpleNamespace(
                        locale_id = None))

                ls.translation = t

                for argument in field.arguments:

                    name_str = argument.name.value

                    if name_str == 'locale_id':

                        t.locale_id = (
                            self.argument_value(argument))

                    else:

                        ls.translation = None
                        break

                if ls.translation:

                    ls('translation_gist_id')
                    ls('translation')

            elif name_str != '__typename':

                ls.object_flag = True

        if (ls.dictionaries and
            ls.dictionaries.published is not None):

            self.published_condition_count += 1

    def parse_dictionaries(self, field_asts):

        ds = (
            self.dictionary_selection)

        ds.perspectives = None
        ds.status = None

        already_set = set()

        for field in field_asts:

            name_str = field.name.value

            if name_str in ds:

                ds(name_str)

            elif name_str == 'perspectives':

                if 'perspectives' in already_set:

                    ds.perspectives = None
                    continue

                already_set.add('perspectives')

                p = (

                    types.SimpleNamespace(
                        with_verb_data = None))

                ds.perspectives = p

                for argument in field.arguments:

                    name_str = argument.name.value

                    if name_str == 'with_verb_data':

                        p.with_verb_data = (
                            self.argument_value(argument))

                    else:

                        ds.perspectives = None
                        break

                if ds.perspectives:

                    ps = (

                        Resolver_Selection({

                            'additional_metadata': (
                                dbPerspective.additional_metadata,),

                            'created_at': (
                                dbPerspective.created_at,),

                            'id': (
                                dbPerspective.client_id,
                                dbPerspective.object_id),

                            'import_hash': (
                                dbPerspective.import_hash,),

                            'import_source': (
                                dbPerspective.import_source,),

                            'marked_for_deletion': (
                                dbPerspective.marked_for_deletion,),

                            'parent_id': (
                                dbPerspective.parent_client_id,
                                dbPerspective.parent_object_id),

                            'state_translation_gist_id': (
                                dbPerspective.state_translation_gist_client_id,
                                dbPerspective.state_translation_gist_object_id),

                            'status': (
                                dbPerspective.state_translation_gist_client_id,
                                dbPerspective.state_translation_gist_object_id),

                            'status_translations': (
                                dbPerspective.state_translation_gist_client_id,
                                dbPerspective.state_translation_gist_object_id),

                            'translation_gist_id': (
                                dbPerspective.translation_gist_client_id,
                                dbPerspective.translation_gist_object_id),

                            'translations': (
                                dbPerspective.translation_gist_client_id,
                                dbPerspective.translation_gist_object_id)}))

                    self.perspective_selection = ps

                    self.parse_perspectives(
                        field.selection_set.selections)

                    # Would need creation time to order perspectives from older to newer.

                    ps('id')
                    ps('created_at')

                    ds('id')
                    ps('parent_id')

            elif name_str == 'status':

                if 'status' in already_set:

                    ds.status = None
                    continue

                already_set.add('status')

                s = (

                    types.SimpleNamespace(
                        locale_id = None))

                ds.status = s

                for argument in field.arguments:

                    name_str = argument.name.value

                    if name_str == 'locale_id':

                        s.locale_id = (
                            self.argument_value(argument))

                    else:

                        ds.status = None
                        break

                if ds.status:

                    ds('state_translation_gist_id')
                    ds('status')

            elif name_str != '__typename':

                ds.object_flag = True

        ds.translations_flag = (
            'translations' in ds.field_set)

        ds.status_flag = (
            ds.status is not None)

        ds.status_translations_flag = (
            'status_translations' in ds.field_set)

    def parse_perspectives(self, field_asts):

        ps = (
            self.perspective_selection)

        ps.columns = None
        ps.status = None

        already_set = set()

        for field in field_asts:

            name_str = field.name.value

            if name_str in ps:

                ps(name_str)

            elif name_str == 'columns':

                if 'columns' in already_set:

                    ps.columns = None
                    continue

                already_set.add('columns')

                ps.columns = (
                    types.SimpleNamespace())

                for argument in field.arguments:

                    ps.columns = None
                    break

                if ps.columns:

                    cs = (

                        Resolver_Selection({

                            'created_at': (
                                dbColumn.created_at,),

                            'field_id': (
                                dbColumn.field_client_id,
                                dbColumn.field_object_id),

                            'id': (
                                dbColumn.client_id,
                                dbColumn.object_id),

                            'link_id': (
                                dbColumn.link_client_id,
                                dbColumn.link_object_id),

                            'marked_for_deletion': (
                                dbColumn.marked_for_deletion,),

                            'parent_id': (
                                dbColumn.parent_client_id,
                                dbColumn.parent_object_id),

                            'position': (
                                dbColumn.position,),

                            'self_id': (
                                dbColumn.self_client_id,
                                dbColumn.self_object_id)}))

                    self.column_selection = cs

                    self.parse_columns(
                        field.selection_set.selections)

                    cs('id')

                    ps('id')
                    cs('parent_id')

            elif name_str == 'status':

                if 'status' in already_set:

                    ps.status = None
                    continue

                already_set.add('status')

                s = (

                    types.SimpleNamespace(
                        locale_id = None))

                ps.status = s

                for argument in field.arguments:

                    name_str = argument.name.value

                    if name_str == 'locale_id':

                        s.locale_id = (
                            self.argument_value(argument))

                    else:

                        ps.status = None
                        break

                if ps.status:

                    ps('state_translation_gist_id')
                    ps('status')

            elif name_str != '__typename':

                ps.object_flag = True

        ps.translations_flag = (
            'translations' in ps.field_set)

        ps.status_flag = (
            ps.status is not None)

        ps.status_translations_flag = (
            'status_translations' in ps.field_set)

    def parse_columns(self, field_asts):

        cs = (
            self.column_selection)

        for field in field_asts:

            name_str = field.name.value

            if name_str in cs:

                cs(name_str)

            elif name_str != '__typename':

                cs.object_flag = True

    def published_condition(self, published):

        state_id_tuple = (

            tuple_(
                dbDictionary.state_translation_gist_client_id,
                dbDictionary.state_translation_gist_object_id))

        if self.published_query is None:

            # Query or CTE.
            #
            # NOTE:
            #
            # Need to use a subquery query in case we join with translationgist / translationatom
            # outside of it.

            self.published_query = (

                get_published_translation_gist_id_subquery_query()
                    if self.published_condition_count <= 1 else
                    get_published_translation_gist_id_cte_query())

        return (

            (state_id_tuple.in_
                if published else
                state_id_tuple.notin_)

                (self.published_query))

    def translations_join(
        self,
        selection,
        gist_client_id_str,
        gist_object_id_str,
        name):

        if selection.cte is not None:

            translation_query = (

                DBSession

                    .query(
                        selection.c.client_id,
                        selection.c.object_id)

                    .outerjoin(
                        dbTranslationAtom,

                        and_(
                            dbTranslationAtom.parent_client_id == getattr(selection.c, gist_client_id_str),
                            dbTranslationAtom.parent_object_id == getattr(selection.c, gist_object_id_str),
                            dbTranslationAtom.marked_for_deletion == False))

                    .add_columns(

                        func.jsonb_object_agg(
                            dbTranslationAtom.locale_id,
                            dbTranslationAtom.content)

                            .filter(dbTranslationAtom.locale_id != None)
                            .label(name))

                    .group_by(
                        selection.c.client_id,
                        selection.c.object_id)

                    .subquery())

            selection.query = (

                selection.query

                    .join(
                        translation_query,

                        and_(
                            translation_query.c.client_id == selection.c.client_id,
                            translation_query.c.object_id == selection.c.object_id))

                    .add_columns(
                        translation_query.c[name]))

        else:

            selection.query = (

                selection.query

                    .outerjoin(
                        dbTranslationAtom,

                        and_(
                            dbTranslationAtom.parent_client_id == getattr(selection.c, gist_client_id_str),
                            dbTranslationAtom.parent_object_id == getattr(selection.c, gist_object_id_str),
                            dbTranslationAtom.marked_for_deletion == False))

                    .add_columns(

                        func.jsonb_object_agg(
                            dbTranslationAtom.locale_id,
                            dbTranslationAtom.content)

                            .filter(dbTranslationAtom.locale_id != None)
                            .label(name))

                    .group_by(
                        selection.c.client_id,
                        selection.c.object_id))

    def translation_join(
        self,
        selection,
        locale_id,
        gist_client_id_str,
        gist_object_id_str,
        name):

        if selection.cte is not None:

            translation_query = (

                DBSession

                    .query(
                        selection.c.client_id,
                        selection.c.object_id)

                    .outerjoin(
                        dbTranslationAtom,

                        and_(
                            dbTranslationAtom.parent_client_id == getattr(selection.c, gist_client_id_str),
                            dbTranslationAtom.parent_object_id == getattr(selection.c, gist_object_id_str),
                            dbTranslationAtom.marked_for_deletion == False,
                            dbTranslationAtom.locale_id == locale_id))

                    .add_columns(

                        dbTranslationAtom.content
                            .label(name))

                    .distinct(
                        selection.c.client_id,
                        selection.c.object_id,
                        dbTranslationAtom.locale_id)

                    .order_by(
                        selection.c.client_id,
                        selection.c.object_id,
                        dbTranslationAtom.locale_id,
                        dbTranslationAtom.created_at.desc())

                    .subquery())

            selection.query = (

                selection.query

                    .join(
                        translation_query,

                        and_(
                            translation_query.c.client_id == selection.c.client_id,
                            translation_query.c.object_id == selection.c.object_id))

                    .add_columns(
                        translation_query.c[name]))

        else:

            selection.query = (

                selection.query

                    .outerjoin(
                        dbTranslationAtom,

                        and_(
                            dbTranslationAtom.parent_client_id == getattr(selection.c, gist_client_id_str),
                            dbTranslationAtom.parent_object_id == getattr(selection.c, gist_object_id_str),
                            dbTranslationAtom.marked_for_deletion == False,
                            dbTranslationAtom.locale_id == locale_id))

                    .add_columns(

                        dbTranslationAtom.content
                            .label(name))

                    .distinct(
                        selection.c.client_id,
                        selection.c.object_id,
                        dbTranslationAtom.locale_id)

                    .order_by(
                        selection.c.client_id,
                        selection.c.object_id,
                        dbTranslationAtom.locale_id,
                        dbTranslationAtom.created_at.desc()))

    def dictionary_condition_list(
        self, category, deleted, published):

        condition_list = []

        if category is not None:

            condition_list.append(
                dbDictionary.category == category)

        if deleted is not None:

            condition_list.append(
                dbDictionary.marked_for_deletion == deleted)

        if published is not None:

            condition_list.append(
                self.published_condition(published))

        if self.dictionary_id_c is not None:

            condition_list.append(

                tuple_(
                    dbDictionary.client_id,
                    dbDictionary.object_id)

                    .in_(
                        DBSession.query(
                            self.dictionary_id_c.client_id,
                            self.dictionary_id_c.object_id)))

        return condition_list

    def dictionary_count_join(
        self, name, category, deleted, published):

        ls = (
            self.language_selection)

        # Using a base language-dictionary CTE, if we have one.

        if self.ld_base_cte_args == (category, deleted, published):

            ld_base_c = self.ld_base_cte.c

            # Assuming that we have the language CTE due to at least two count joins, so a join to a
            # separate count query.

            count_query = (

                DBSession

                    .query(
                        ld_base_c.client_id,
                        ld_base_c.object_id,

                        func.count()
                            .filter(ld_base_c.dictionary_client_id != None)
                            .label(name))

                    .group_by(
                        ld_base_c.client_id,
                        ld_base_c.object_id)

                    .subquery())

            ls.query = (

                ls.query

                    .join(
                        count_query,

                        and_(
                            count_query.c.client_id == ls.c.client_id,
                            count_query.c.object_id == ls.c.object_id))

                    .add_columns(
                        count_query.c[name]))

            return

        # No base CTE, using join to the dictionaries.

        condition_list = (

            self.dictionary_condition_list(
                category,
                deleted,
                published))

        if ls.cte is not None:

            count_query = (

                DBSession

                    .query(
                        ls.c.client_id,
                        ls.c.object_id)

                    .outerjoin(
                        dbDictionary,

                        and_(
                            dbDictionary.parent_client_id == ls.c.client_id,
                            dbDictionary.parent_object_id == ls.c.object_id,
                            *condition_list))

                    .add_columns(

                        func.count()
                            .filter(dbDictionary.client_id != None)
                            .label(name))

                    .group_by(
                        ls.c.client_id,
                        ls.c.object_id)

                    .subquery())

            ls.query = (

                ls.query

                    .join(
                        count_query,

                        and_(
                            count_query.c.client_id == ls.c.client_id,
                            count_query.c.object_id == ls.c.object_id))

                    .add_columns(
                        count_query.c[name]))

        else:

            ls.query = (

                ls.query

                    .outerjoin(
                        dbDictionary,

                        and_(
                            dbDictionary.parent_client_id == ls.c.client_id,
                            dbDictionary.parent_object_id == ls.c.object_id,
                            *condition_list))

                    .add_columns(

                        func.count()
                            .filter(dbDictionary.client_id != None)
                            .label(name))

                    .group_by(
                        ls.c.client_id,
                        ls.c.object_id))

    def run(self):

        # Analyzing query.

        if self.debug_flag:

            log.debug(
                f'\n field_asts:\n {self.field_asts}')

        ls = (

            Resolver_Selection({

                'additional_metadata': (
                    dbLanguage.additional_metadata,),

                'created_at': (
                    dbLanguage.created_at,),

                'id': (
                    dbLanguage.client_id,
                    dbLanguage.object_id),

                'in_toc': (
                    dbLanguage.additional_metadata,),

                'marked_for_deletion': (
                    dbLanguage.marked_for_deletion,),

                'parent_id': (
                    dbLanguage.parent_client_id,
                    dbLanguage.parent_object_id),

                'translation_gist_id': (
                    dbLanguage.translation_gist_client_id,
                    dbLanguage.translation_gist_object_id),

                'translations': (
                    dbLanguage.translation_gist_client_id,
                    dbLanguage.translation_gist_object_id)}))

        self.language_selection = ls

        ls.dictionary_count = None
        ls.dictionaries = None
        ls.translation = None

        for field in self.field_asts:

            if field.name.value != 'languages':
                continue

            self.parse_languages(
                field.selection_set.selections)

            break

        ls('id')

        ls.in_toc_flag = (
            'in_toc' in ls.field_set)

        ls.translation_flag = (
            ls.translation is not None)

        ls.translations_flag = (
            'translations' in ls.field_set)

        def recursive_count_f(count_dict):
            """
            Creates recursive count computing function.
            """

            recursive_count_dict = {}

            def f(id):

                recursive_count = recursive_count_dict.get(id)

                if recursive_count is not None:
                    return recursive_count

                recursive_count = (

                    count_dict[id] +

                    sum(f(to_id)
                        for to_id in self.from_to_dict[id]))

                recursive_count_dict[id] = recursive_count

                return recursive_count

            return f

        def simple_count_f(count_dict):
            """
            Creates simple count computing function.
            """

            def f(id):

                return count_dict.get(id, 0)

            return f

        # If we are both filtering by dictionary counts and sorting languages by grants / organizations,
        # we'll be using a base language-dictionary CTE for both.

        self.grant_or_organization_id = (

            self.args.grant_id is not None or
            self.args.organization_id is not None)

        self.grant_or_organization = (

            self.args.by_grants and
                self.args.grant_id is None or

            self.args.by_organizations and
                self.args.organization_id is None)

        self.ld_base_cte = None
        self.ld_base_cte_args = None

        if self.grant_or_organization:

            if self.args.only_with_dictionaries_recursive:

                self.ld_base_cte_args = (

                    self.args.dictionary_category,
                    False,
                    self.args.dictionary_published)

            # No common base CTE, it would be a separate join, we check if we'll be needing a published
            # condition check.

            elif self.args.dictionary_published is not None:

                self.published_condition_count += 1

            # For aggregate count representation.

            empty_count_dict = {}
            self.aggregate_count_dict = {}

        #
        # Determining if we'll need any dictionary count and other joins.
        #
        # NOTE:
        #
        # We have two ways to compute recursive dictionary counts for languages, fully on SQL side and
        # partially on Python side.
        #
        # For example, ToC languages.
        #
        # Full SQL computation required 3 temporary tables and a big SQL script, partially Python
        # computation has only 1 temporary table, will receive raw dictionary counts for 400+ related
        # languages with their ids and parent ids, and transfer around 100 less lines of SQL to the DB.
        #
        # Testing (with log level WARN instead of DEBUG):
        #
        #   t0 = time.time()
        #
        #   for i in range(256):
        #       language_toc_sql(self)
        #
        #   t1 = time.time()
        #
        #   result_1 = (
        #       sorted(language_toc_sql(self)))
        #
        #   print(len(result_1))
        #   pprint.pprint(result_1, width = 192)
        #   print(hashlib.md5(repr(result_1).encode('utf-8')).hexdigest())
        #
        #   t2 = time.time()
        #
        #   for i in range(256):
        #       language_toc_python(self)
        #
        #   t3 = time.time()
        #
        #   result_2 = (
        #       sorted(language_toc_python(self)))
        #
        #   print(len(result_2))
        #   pprint.pprint(result_2, width = 192)
        #   print(hashlib.md5(repr(result_2).encode('utf-8')).hexdigest())
        #
        #   log.debug(
        #       f'\nt1 - t0: {t1 - t0:.6f}s'
        #       f'\nt3 - t2: {t3 - t2:.6f}s')
        #
        # Looks like language_toc_python() is about 20% faster.
        #
        # So we should compute languages' recursive dictionary count via Python.
        #

        dictionary_count_dict = {}

        dictionary_filter_f = None
        dictionary_count_f = None

        ls.recursive_count_flag = False

        if (self.args.only_with_dictionaries_recursive and
            (dc := ls.dictionary_count) and
            dc.category == self.args.dictionary_category and
            dc.published == self.args.dictionary_published):

            # Single shared count.

            dc_args = (
                dc.category,
                False,
                dc.published)

            count_dict = {}

            dictionary_count_dict[dc_args] = (
                ('dictionary_count', count_dict))

            ls.recursive_count_flag = True

            if self.args.dictionary_published is not None:

                self.published_condition_count += 1

            dictionary_filter_f = (
                recursive_count_f(count_dict))

            dictionary_count_f = (

                dictionary_filter_f
                    if dc.recursive else
                    simple_count_f(count_dict))

        else:

            # Only one or two separate counts.

            if self.args.only_with_dictionaries_recursive:

                dc_args = (
                    self.args.dictionary_category,
                    False,
                    self.args.dictionary_published)

                count_dict = {}

                dictionary_count_dict[dc_args] = (
                    ('dictionary_count_0', count_dict))

                ls.recursive_count_flag = True

                if self.args.dictionary_published is not None:

                    self.published_condition_count += 1

                dictionary_filter_f = (
                    recursive_count_f(count_dict))

            if (dc := ls.dictionary_count):

                dc_args = (
                    dc.category,
                    False,
                    dc.published)

                count_dict = {}

                dictionary_count_dict[dc_args] = (
                    ('dictionary_count_1', count_dict))

                if dc.recursive:

                    ls.recursive_count_flag = True

                if dc.published is not None:

                    self.published_condition_count += 1

                dictionary_count_f = (

                    recursive_count_f(count_dict)
                        if dc.recursive else
                        simple_count_f(count_dict))

        join_count = (
            len(dictionary_count_dict))

        if self.grant_or_organization:
            join_count += 1

        if ls.translation_flag:
            join_count += 1

        if ls.translations_flag:
            join_count += 1

        # If we are returning languages in tree order and at the same time getting a translation, we have to
        # use a CTE because otherwise order by of translation's distinct on messes up tree order preliminary
        # order by.
        #
        # If we are constructing language trees by grants or organizations, we'll need to use a CTE due to
        # an involved multi-stage per grant / per organization dictionary count retrieval requiring it.

        ls.cte_flag = (
            self.args.in_tree_order and ls.translation_flag or
            self.grant_or_organization or
            join_count > 1)

        ls.join_flag = (
            join_count >= 1)

        if ls.object_flag:

            ls.column_list = [dbLanguage]

        else:

            # For ToC inclusion checking, if required.

            if ls.in_toc_flag:

                ls('additional_metadata')

            # For standard ordering.

            if self.args.in_tree_order:

                ls('parent_id')

                if ls.cte_flag:

                    ls('additional_metadata')

            # For computing recursive dictionary counts and/or getting languages bottom-up from
            # dictionaries.

            if (ls.recursive_count_flag or
                self.grant_or_organization_id):

                ls('parent_id')

        # Base language query.

        self.dictionary_id_c = None

        if (language_id := self.args.language_id) is not None:

            # Language subtree.

            if ls.dictionaries:

                column_list = (

                    dbLanguage.client_id,
                    dbLanguage.object_id)

            else:

                column_list = (

                    ls.column_list)

            base_cte = (

                DBSession

                    .query(
                        *column_list)

                    .filter(
                        dbLanguage.client_id == language_id[0],
                        dbLanguage.object_id == language_id[1],
                        dbLanguage.marked_for_deletion == False)

                    .cte(recursive = True))

            recursive_query = (

                DBSession

                    .query(
                        *column_list)

                    .filter(
                        dbLanguage.parent_client_id == base_cte.c.client_id,
                        dbLanguage.parent_object_id == base_cte.c.object_id,
                        dbLanguage.marked_for_deletion == False))

            language_cte = (
                base_cte.union(recursive_query))

            ls.c = language_cte.c

            if (ls.object_flag and
                not ls.dictionaries):

                language_cte = (

                    aliased(
                        dbLanguage,
                        language_cte,
                        adapt_on_names = True))

                ls.c = language_cte

            ls.query = (
                DBSession.query(language_cte))

        elif self.grant_or_organization_id:

            # Languages of a grant or of an organization.
            #
            # We'll have to get dictionary ids through raw SQL as SQLAlchemy is bad with PostgreSQL's
            # jsonb_to_recordset.

            sql_str = f'''

                select P.*

                from
                  {'public.grant'
                    if self.args.grant_id is not None else
                    'organization'} S

                cross join
                  jsonb_to_recordset(S.additional_metadata -> 'participant')
                    P (client_id bigint, object_id bigint)

                where S.id = {
                  self.args.grant_id
                    if self.args.grant_id is not None else
                    self.args.organization_id};

                '''

            dictionary_id_query = (

                sqlalchemy

                    .text(sql_str)

                    .columns(
                        client_id = SLBigInteger,
                        object_id = SLBigInteger)

                    .alias())

            self.dictionary_id_c = (
                dictionary_id_query.c)

            if ls.dictionaries:

                # If we are going to get dictionaries, we save dictionary ids in a temporary table to be
                # used for filtering later.

                dictionary_table_name = (

                    'dictionary_' +
                     str(uuid.uuid4()).replace('-', '_'))

                dictionary_id_table = (

                    sqlalchemy.Table(
                        dictionary_table_name,
                        models.Base.metadata,
                        sqlalchemy.Column('client_id', SLBigInteger),
                        sqlalchemy.Column('object_id', SLBigInteger),
                        prefixes = ['temporary'],
                        postgresql_on_commit = 'drop'))

                dictionary_id_table.create(
                    DBSession.connection())

                DBSession.execute(

                    dictionary_id_table

                        .insert()

                        .from_select(
                            (self.dictionary_id_c.client_id, self.dictionary_id_c.object_id),
                            dictionary_id_query))

                self.dictionary_id_c = (
                    dictionary_id_table.c)

            elif dictionary_count_dict:

                # Othwerwise, if we'll need to filter based on dictionary ids when getting dictionary
                # counts, we turn dictionary id query into a CTE.

                dictionary_id_cte = (
                    dictionary_id_query.cte())

                self.dictionary_id_c = (
                    dictionary_id_cte.c)

            # Getting languages bottom-up from dictionaries through recursive CTE.

            base_cte = (

                DBSession

                    .query(
                        *ls.column_list)

                    .filter(
                        dbDictionary.client_id == self.dictionary_id_c.client_id,
                        dbDictionary.object_id == self.dictionary_id_c.object_id,
                        dbLanguage.client_id == dbDictionary.parent_client_id,
                        dbLanguage.object_id == dbDictionary.parent_object_id,
                        dbLanguage.marked_for_deletion == False)

                    .group_by(
                        dbLanguage.client_id,
                        dbLanguage.object_id)

                    .cte(recursive = True))

            recursive_query = (

                DBSession

                    .query(
                        *ls.column_list)

                    .filter(
                        dbLanguage.client_id == base_cte.c.parent_client_id,
                        dbLanguage.object_id == base_cte.c.parent_object_id,
                        dbLanguage.marked_for_deletion == False)

                    .group_by(
                        dbLanguage.client_id,
                        dbLanguage.object_id))

            language_cte = (
                base_cte.union(recursive_query))

            ls.c = language_cte.c

            if ls.object_flag:

                language_cte = (

                    aliased(
                        dbLanguage,
                        language_cte,
                        adapt_on_names = True))

                ls.c = language_cte

            ls.query = (
                DBSession.query(language_cte))

        else:

            # Just all languages.

            ls.query = (

                DBSession

                    .query(
                        *ls.column_list)

                    .filter_by(
                        marked_for_deletion = False))

            ls.c = dbLanguage

        # Filtering by ids, if required.

        if self.args.id_list is not None:

            ls.query = (

                ls.query.filter(

                    tuple_(
                        ls.c.client_id,
                        ls.c.object_id)

                        .in_(
                            ids_to_id_query(
                                self.args.id_list))))

        # Filtering by language table of contents status, if required.

        if self.args.only_in_toc:

            ls.query = (

                ls.query.filter(

                    cast(
                        ls.c.additional_metadata['toc_mark'],
                        Boolean)))

        # If we are going to get dictionaries, we save language info in a temporary table, unless we've got
        # dictionary ids directly from grant or organization.
        #
        # We also might need language info to trace language tree paths up to the root to enable proper
        # ordering of languages in tree order or to get language / dictionary / perspective tree path info.

        if (ls.dictionaries and
            not self.grant_or_organization_id

            or

            self.args.in_tree_order and
            (self.args.id_list or self.args.only_in_toc)):

            self.language_table_name = (

                'language_' +
                 str(uuid.uuid4()).replace('-', '_'))

            # Testing shows that only id temporary table is slightly faster than the full one, like 7.55 to
            # 7.88 relative times.
            #
            # If we could, we should have used insert CTE, but as of now SQLAlchemy does not support adding
            # unrelated CTEs to queries.

            if True:

                language_table = (

                    sqlalchemy.Table(
                        self.language_table_name,
                        models.Base.metadata,
                        sqlalchemy.Column('client_id', SLBigInteger),
                        sqlalchemy.Column('object_id', SLBigInteger),
                        prefixes = ['temporary'],
                        postgresql_on_commit = 'drop'))

                language_table.create(
                    DBSession.connection())

                DBSession.execute(

                    language_table

                        .insert()

                        .from_select(
                            (ls.c.client_id, ls.c.object_id),
                            ls.query.with_entities(
                                ls.c.client_id, ls.c.object_id)))

                ls.query = (

                    DBSession

                        .query(
                            *ls.column_list)

                        .filter(

                            tuple_(
                                dbLanguage.client_id,
                                dbLanguage.object_id)

                                .in_(
                                    DBSession.query(language_table))))

                ls.c = dbLanguage

            else:

                language_table = (

                    sqlalchemy.Table(
                        self.language_table_name,
                        models.Base.metadata,
                        *(sqlalchemy.Column(column.name, column.type)
                            for column in ls.column_list),
                        prefixes = ['temporary'],
                        postgresql_on_commit = 'drop'))

                language_table.create(
                    DBSession.connection())

                DBSession.execute(

                    language_table

                        .insert()

                        .from_select(
                            ls.column_list,
                            ls.query))

                ls.query = (

                    DBSession.query(
                        language_table))

                ls.c = language_table.c

        # If we are going to query both translations and dictionary counts or other non-singular sets of
        # derived attributes, we'll need to use separate joins to a base CTE.

        ls.cte = None

        if ls.cte_flag:

            ls.cte = ls.query.cte()
            ls.c = ls.cte.c

            if ls.object_flag:

                ls.cte = (

                    aliased(
                        dbLanguage,
                        ls.cte,
                        adapt_on_names = True))

                ls.c = ls.cte

            ls.query = (
                DBSession.query(ls.cte))

        #
        # Getting translations through a join, if required.
        #
        # NOTE:
        #
        # Testing getting translations from cache / from DB with example of ToC languages
        # (with log level WARN instead of DEBUG):
        #
        #   t0 = time.time()
        #
        #   for i in range(256):
        #       toc_translations_cache(self, toc_list)
        #
        #   t1 = time.time()
        #
        #   result_1 = toc_translations_cache(self, toc_list)
        #
        #   print(len(result_1))
        #
        #   result_1_sorted = (
        #
        #       sorted(
        #           (key, None if value is None else sorted(value.items()))
        #           for key, value in result_1.items()))
        #
        #   pprint.pprint(result_1_sorted, width = 192)
        #
        #   print(
        #       hashlib.md5(
        #           repr(result_1_sorted).encode('utf-8'))
        #           .hexdigest())
        #
        #   t2 = time.time()
        #
        #   for i in range(256):
        #       toc_translations_db(self, toc_list)
        #
        #   t3 = time.time()
        #
        #   result_2 = toc_translations_db(self, toc_list)
        #
        #   print(len(result_2))
        #
        #   result_2_sorted = (
        #
        #       sorted(
        #           (key, None if value is None else sorted(value.items()))
        #           for key, value in result_2.items()))
        #
        #   pprint.pprint(result_2_sorted, width = 192)
        #
        #   print(
        #       hashlib.md5(
        #           repr(result_2_sorted).encode('utf-8'))
        #           .hexdigest())
        #
        #   log.debug(
        #       f'\nt1 - t0: {t1 - t0:.6f}s'
        #       f'\nt3 - t2: {t3 - t2:.6f}s')
        #
        # Looks like toc_translations_db() is about 2-3 times faster.
        #
        # So if possible, we should get translations from DB.
        #

        if ls.translation_flag:

            self.translation_join(
                ls,
                ls.translation.locale_id or self.info.context.get('locale_id'),
                'translation_gist_client_id',
                'translation_gist_object_id',
                'translation')

        if ls.translations_flag:

            self.translations_join(
                ls,
                'translation_gist_client_id',
                'translation_gist_object_id',
                'translations')

        # Creating a base language-dictionary CTE, if required.

        if self.ld_base_cte_args:

            condition_list = (

                self.dictionary_condition_list(
                    *self.ld_base_cte_args))

            # Assuming that we have the language CTE due to at least two count joins.

            if ls.cte is None:
                raise NotImplementedError

            self.ld_base_cte = (

                DBSession

                    .query(
                        ls.c.client_id,
                        ls.c.object_id)

                    .outerjoin(
                        dbDictionary,

                        and_(
                            dbDictionary.parent_client_id == ls.c.client_id,
                            dbDictionary.parent_object_id == ls.c.object_id,
                            *condition_list))

                    .add_columns(
                        dbDictionary.client_id.label('dictionary_client_id'),
                        dbDictionary.object_id.label('dictionary_object_id'))

                    .cte('ld_base'))

        # Getting dictionary counts through a join, if required.

        for dc_args, dc_name_count in dictionary_count_dict.items():

            self.dictionary_count_join(
                dc_name_count[0], *dc_args)

        # Getting aggregate grant/organization counts, if required.
        #
        # In that case we'll have the language CTE due to an involved multi-stage per grant / per
        # organization dictionary count retrieval requiring it.

        if self.grant_or_organization:

            sql_str = f'''

                select
                  P.client_id,
                  P.object_id,
                  G.id

                from
                  {'public.grant'
                    if self.args.by_grants else
                    'organization'} G

                cross join
                  jsonb_to_recordset(G.additional_metadata -> 'participant')
                    P (client_id bigint, object_id bigint)

                {''
                  if self.args.by_grants else
                  'where G.marked_for_deletion = false'}

                '''

            participant_query = (

                sqlalchemy

                    .text(sql_str)

                    .columns(
                        client_id = SLBigInteger,
                        object_id = SLBigInteger,
                        id = SLBigInteger)

                    .alias('participant'))

            # Using a base language-dictionary CTE, if we have one.

            if self.ld_base_cte is not None:

                ld_base_c = self.ld_base_cte.c

                joint_query = (

                    DBSession

                        .query(
                            ld_base_c.client_id,
                            ld_base_c.object_id)

                        .outerjoin(
                            participant_query,

                            and_(
                                ld_base_c.dictionary_client_id == participant_query.c.client_id,
                                ld_base_c.dictionary_object_id == participant_query.c.object_id))

                        .add_columns(

                            sqlalchemy

                                .literal_column(f'''
                                    (case when ld_base.dictionary_client_id is null then null else
                                        coalesce(participant.id :: text, '') end)
                                    ''')

                                .label('group_id'))

                        .subquery())

            # No base CTE, using join to the dictionaries.

            else:

                condition_list = (

                    self.dictionary_condition_list(
                        self.args.dictionary_category,
                        False,
                        self.args.dictionary_published))

                joint_query = (

                    DBSession

                        .query(
                            ls.c.client_id,
                            ls.c.object_id)

                        .outerjoin(
                            dbDictionary,

                            and_(
                                dbDictionary.parent_client_id == ls.c.client_id,
                                dbDictionary.parent_object_id == ls.c.object_id,
                                *condition_list))

                        .add_columns(
                            dbDictionary.client_id.label('dictionary_client_id'),
                            dbDictionary.object_id.label('dictionary_object_id'))

                        .outerjoin(
                            participant_query,

                            and_(
                                dbDictionary.client_id == participant_query.client_id,
                                dbDictionary.object_id == participant_query.object_id))

                        .add_columns(

                            sqlalchemy

                                .literal_column(f'''
                                    (case when dictionary.client_id is null then null else
                                        coalesce(participant.id :: text, '') end)
                                    ''')

                                .label('group_id'))

                        .subquery())

            base_count_query = (

                DBSession

                    .query(
                        joint_query.c.client_id,
                        joint_query.c.object_id,
                        joint_query.c.group_id,

                        func.count()
                            .filter(joint_query.c.group_id != None)
                            .label('base_count'))

                    .group_by(
                        joint_query.c.client_id,
                        joint_query.c.object_id,
                        joint_query.c.group_id)

                    .subquery())

            aggregate_count_query = (

                DBSession

                    .query(
                        base_count_query.c.client_id,
                        base_count_query.c.object_id,

                        func
                            .jsonb_object_agg(
                                base_count_query.c.group_id,
                                base_count_query.c.base_count)

                            .filter(
                                base_count_query.c.group_id != None)

                            .label('aggregate_count'))

                    .group_by(
                        base_count_query.c.client_id,
                        base_count_query.c.object_id)

                    .subquery())

            ls.query = (

                ls.query

                    .join(
                        aggregate_count_query,

                        and_(
                            aggregate_count_query.c.client_id == ls.c.client_id,
                            aggregate_count_query.c.object_id == ls.c.object_id))

                    .add_columns(
                        aggregate_count_query.c.aggregate_count))

        # If we require languages in the language tree order, we establish preliminary ordering.

        if self.args.in_tree_order:

            ls.query = (

                ls.query

                    .order_by(
                        ls.c.additional_metadata['younger_siblings'],
                        ls.c.client_id.desc(),
                        ls.c.object_id.desc()))

        # Getting language data.

        result_list = ls.query.all()

        if self.debug_flag:

            log.debug(
                f'\n language_query ({len(result_list)} languages):\n ' +
                render_statement(ls.query.statement))

        # Generating language tree data and dictionary counts, if required.

        self.arbitrary_root_set = (

            self.args.in_tree_order and
            (self.args.id_list or self.args.only_in_toc))

        if (self.args.in_tree_order or
            ls.recursive_count_flag or
            self.grant_or_organization):

            if not self.arbitrary_root_set:

                self.from_to_dict = collections.defaultdict(list)

            for result in result_list:

                if ls.object_flag:

                    language = (
                        result[0] if ls.join_flag else result)

                    id = language.id
                    parent_id = language.parent_id

                else:

                    id = (
                        result.client_id, result.object_id)

                    parent_id = (
                        result.parent_client_id, result.parent_object_id)

                # Processing language tree data, unless we don't have full paths to the language hierarchy
                # root in our language set and we'll have to get fully traceable language set later and
                # build language tree data from it.

                if not self.arbitrary_root_set:

                    self.from_to_dict[parent_id].append(id)

                for name, count_dict in dictionary_count_dict.values():

                    count_dict[id] = (
                        getattr(result, name))

                if self.grant_or_organization:

                    self.aggregate_count_dict[id] = (
                        result.aggregate_count or empty_count_dict)

        if self.debug_flag:

            for name, count_dict in dictionary_count_dict.values():

                log.debug(

                    f'\n count_dict({name}):\n' +

                    pprint.pformat(
                        count_dict, width = 144))

            if self.grant_or_organization:

                log.debug(

                    f'\n aggregate_count_dict:\n' +

                    pprint.pformat(
                        self.aggregate_count_dict, width = 144))

        gql_language_list = []

        gql_language_dict_flag = (

            self.args.in_tree_order or
            ls.dictionaries)

        if gql_language_dict_flag:

            gql_language_dict = {}

        self.filtered_out_id_set = set()

        if ls.in_toc_flag:

            ls.field_set.discard('in_toc')
            ls.field_set.discard('additional_metadata')

        if ls.object_flag:

            # We are getting full ORM dbLanguage objects.

            attribute_set = ls.field_set.copy()

            attribute_set.discard('translation')
            attribute_set.discard('translations')

            for result in result_list:

                language = (
                    result[0] if ls.join_flag else result)

                language_id = language.id

                # Filtering based on a dictionary count if required. 

                if (dictionary_filter_f and
                    dictionary_filter_f(language_id) <= 0):

                    self.filtered_out_id_set.add(language_id)
                    continue

                gql_language = (
                    Language(id = language_id))

                gql_language.dbObject = language

                # Computed attributes.

                if ls.in_toc_flag:

                    metadata = (
                        result.additional_metadata)

                    gql_language.in_toc = (

                        metadata is not None and
                        metadata.get('toc_mark', False))

                    gql_language.additional_metadata = (
                        AdditionalMetadata.from_object(metadata))

                if ls.translation_flag:

                    translation = (
                        result.translation)

                    gql_language.translation = (
                        translation if translation is not None else gql_none_value)

                if ls.translations_flag:

                    translations = (
                        result.translations)

                    gql_language.translations = (
                        translations if translations is not None else gql_none_value)

                if dictionary_count_f:

                    gql_language.dictionary_count = (
                        dictionary_count_f(language_id))

                # Standard attributes.

                for attribute in attribute_set:

                    value = (
                        getattr(language, attribute))

                    if attribute == 'additional_metadata':

                        value = AdditionalMetadata.from_object(value)

                    elif attribute == 'created_at':

                        value = CreatedAt.from_timestamp(value)

                    setattr(
                        gql_language,
                        attribute,
                        value)

                if not self.args.in_tree_order:

                    gql_language_list.append(gql_language)

                if gql_language_dict_flag:

                    gql_language_dict[language_id] = gql_language

        else:

            # We are getting attribute values as they are.

            for result in result_list:

                language_id = (
                    result.client_id, result.object_id)

                # Filtering based on a dictionary count if required. 

                if (dictionary_filter_f and
                    dictionary_filter_f(language_id) <= 0):

                    self.filtered_out_id_set.add(language_id)
                    continue

                gql_language = (
                    Language(id = language_id))

                if ls.in_toc_flag:

                    metadata = (
                        result.additional_metadata)

                    gql_language.in_toc = (

                        metadata is not None and
                        metadata.get('toc_mark', False))

                    gql_language.additional_metadata = (
                        AdditionalMetadata.from_object(metadata))

                if dictionary_count_f:

                    gql_language.dictionary_count = (
                        dictionary_count_f(language_id))

                for field_str in ls.field_set:

                    if field_str == 'id':

                        gql_language.id = language_id

                        continue

                    elif field_str == 'parent_id':

                        parent_client_id = (
                            result.parent_client_id)

                        gql_language.parent_id = (

                            gql_none_value

                            if parent_client_id is None else

                            (parent_client_id,
                                result.parent_object_id))

                        continue

                    elif field_str == 'translation_gist_id':

                        gql_language.translation_gist_id = (

                            result.translation_gist_client_id,
                            result.translation_gist_object_id)

                        continue

                    value = getattr(result, field_str)

                    if field_str == 'additional_metadata':

                        value = AdditionalMetadata.from_object(value)

                    elif field_str == 'created_at':

                        value = CreatedAt.from_timestamp(value)

                    setattr(
                        gql_language,
                        field_str,
                        value if value is not None else gql_none_value)

                if not self.args.in_tree_order:

                    gql_language_list.append(gql_language)

                if gql_language_dict_flag:

                    gql_language_dict[language_id] = gql_language

        # If we need to return languages in the standard language tree order, we order languages by
        # recursively traversing the tree we build earlier, the order is guaranteed by the preliminary
        # ordering we received languages from the DB in.

        if self.args.in_tree_order:

            def f(id):

                gql_language = (
                    gql_language_dict.get(id))

                if gql_language:

                    gql_language_list.append(
                        gql_language)

                for to_id in self.from_to_dict[id]:

                    f(to_id)

            # Filtering languages explicitly or based on ToC status, an arbitrary number of tree roots, and
            # we'll have to find them and traverse them in the right order, so we'll have to compile full
            # language ordering info.

            if (self.args.id_list or
                self.args.only_in_toc):

                base_cte = (

                    DBSession

                        .query(
                            dbLanguage.client_id,
                            dbLanguage.object_id,
                            dbLanguage.parent_client_id,
                            dbLanguage.parent_object_id,
                            dbLanguage.additional_metadata)

                        .filter(
                            dbLanguage.client_id == language_table.c.client_id,
                            dbLanguage.object_id == language_table.c.object_id)

                        .cte(recursive = True))

                recursive_query = (

                    DBSession

                        .query(
                            dbLanguage.client_id,
                            dbLanguage.object_id,
                            dbLanguage.parent_client_id,
                            dbLanguage.parent_object_id,
                            dbLanguage.additional_metadata)

                        .filter(
                            dbLanguage.client_id == base_cte.c.parent_client_id,
                            dbLanguage.object_id == base_cte.c.parent_object_id))

                language_cte = (
                    base_cte.union(recursive_query))

                language_id_id_list = (

                    DBSession

                        .query(
                            language_cte.c.client_id,
                            language_cte.c.object_id,
                            language_cte.c.parent_client_id,
                            language_cte.c.parent_object_id)

                        .order_by(
                            language_cte.c.additional_metadata['younger_siblings'],
                            language_cte.c.client_id.desc(),
                            language_cte.c.object_id.desc())

                        .all())

                self.from_to_dict = collections.defaultdict(list)

                for language_id_id in language_id_id_list:

                    id = (
                        language_id_id[0], language_id_id[1])

                    parent_id = (
                        language_id_id[2], language_id_id[3])

                    self.from_to_dict[parent_id].append(id)

            # Traversing language tree we have in the standard order, compiling final list of languages.

            root_id = (

                tuple(self.args.language_id) if self.args.language_id else
                (None, None))

            f(root_id)

            # Checking compatibility with recursive_sort(), if required.

            if self.debug_flag:

                id_list = [

                    gql_language.id
                    for gql_language in gql_language_list]

                id_set = set(id_list)

                reference_list = [

                    (client_id, object_id)
                    for _, client_id, object_id, _ in

                        recursive_sort(

                            DBSession

                                .query(dbLanguage)

                                .filter_by(
                                    marked_for_deletion = False)

                                .order_by(
                                    dbLanguage.parent_client_id,
                                    dbLanguage.parent_object_id,
                                    dbLanguage.additional_metadata['younger_siblings'],
                                    dbLanguage.client_id.desc(),
                                    dbLanguage.object_id.desc())

                                .all())

                    if (client_id, object_id) in id_set]

                if id_list != reference_list:

                    log.warning(
                        f'\nid_list:\n{pprint.pformat(id_list, width = 192)}'
                        f'\nreference_list:\n{pprint.pformat(reference_list, width = 192)}')

                    raise NotImplementedError

        if self.debug_flag:

            log.debug(
                f'\nlen(gql_language_list): {len(gql_language_list)}')

        # Getting dictionaries, if required.

        if (d := ls.dictionaries):

            condition_list = (

                self.dictionary_condition_list(
                    d.category,
                    d.deleted,
                    d.published))

            if self.dictionary_id_c is None:

                condition_list.extend((
                    dbDictionary.parent_client_id == language_table.c.client_id,
                    dbDictionary.parent_object_id == language_table.c.object_id))

            if self.filtered_out_id_set:

                condition_list.append(

                    tuple_(
                        dbDictionary.parent_client_id,
                        dbDictionary.parent_object_id)

                        .notin_(

                            ids_to_id_query(
                                self.filtered_out_id_set)))

            ds = self.dictionary_selection

            if ds.object_flag:

                ds.column_list = [dbDictionary]

            ds.query = (

                DBSession

                    .query(
                        *ds.column_list)

                    .filter(
                        *condition_list))

            ds.c = dbDictionary

            # If we are going to get perspectives, we save dictionary info in a temporary table.

            if ds.perspectives:

                dictionary_table_name = (

                    'dictionary_' +
                     str(uuid.uuid4()).replace('-', '_'))

                dictionary_table = (

                    sqlalchemy.Table(
                        dictionary_table_name,
                        models.Base.metadata,
                        sqlalchemy.Column('client_id', SLBigInteger),
                        sqlalchemy.Column('object_id', SLBigInteger),
                        prefixes = ['temporary'],
                        postgresql_on_commit = 'drop'))

                dictionary_table.create(
                    DBSession.connection())

                DBSession.execute(

                    dictionary_table

                        .insert()

                        .from_select(
                            (ds.c.client_id, ds.c.object_id),
                            ds.query.with_entities(
                                ds.c.client_id, ds.c.object_id)))

                ds.query = (

                    DBSession

                        .query(
                            *ds.column_list)

                        .filter(

                            tuple_(
                                dbDictionary.client_id,
                                dbDictionary.object_id)

                                .in_(
                                    DBSession.query(dictionary_table))))

                ds.c = dbDictionary

            # Checking for joins.

            join_count = 0

            if ds.translations_flag:
                join_count += 1

            if ds.status_flag:
                join_count += 1

            if ds.status_translations_flag:
                join_count += 1

            ds.cte_flag = (
                join_count > 1)

            ds.join_flag = (
                join_count >= 1)

            ds.cte = None

            if ds.cte_flag:

                ds.cte = ds.query.cte()
                ds.c = ds.cte.c

                if ds.object_flag:

                    ds.cte = (

                        aliased(
                            dbDictionary,
                            ds.cte,
                            adapt_on_names = True))

                    ds.c = ds.cte

                ds.query = (
                    DBSession.query(ds.cte))

            if ds.translations_flag:

                self.translations_join(
                    ds,
                    'translation_gist_client_id',
                    'translation_gist_object_id',
                    'translations')

            if ds.status_flag:

                self.translation_join(
                    ds,
                    ds.status.locale_id or self.info.context.get('locale_id'),
                    'state_translation_gist_client_id',
                    'state_translation_gist_object_id',
                    'status')

            if ds.status_translations_flag:

                self.translations_join(
                    ds,
                    'state_translation_gist_client_id',
                    'state_translation_gist_object_id',
                    'status_translations')

            # Dictionaries go in standard order, from newest to oldest.

            ds.query = (

                ds.query

                    .order_by(
                        ds.c.created_at.desc(),
                        ds.c.client_id.desc(),
                        ds.c.object_id.desc()))

            # Getting and processing dictionary data.

            result_list = ds.query.all()

            if self.debug_flag:

                log.debug(
                    '\n dictionary_query:\n ' +
                    render_statement(ds.query.statement))

            for gql_language in gql_language_dict.values():

                gql_language.dictionaries = []

            gql_dictionary_dict_flag = (

                ds.perspectives)

            if gql_dictionary_dict_flag:

                gql_dictionary_dict = {}

            if ds.object_flag:

                # We are getting full ORM dbDictionary objects.

                attribute_set = ds.field_set.copy()

                attribute_set.discard('status')
                attribute_set.discard('status_translations')
                attribute_set.discard('translations')

                for result in result_list:

                    dictionary = (
                        result[0] if ds.join_flag else result)

                    dictionary_id = dictionary.id

                    gql_dictionary = (
                        Dictionary(id = dictionary_id))

                    gql_dictionary.dbObject = dictionary

                    # Computed attributes.

                    if ds.status_flag:

                        translation = (
                            result.status)

                        gql_dictionary.status = (
                            translation if translation is not None else gql_none_value)

                    if ds.status_translations_flag:

                        translations = (
                            result.status_translations)

                        gql_dictionary.translations = (
                            translations if translations is not None else gql_none_value)

                    if ds.translations_flag:

                        translations = (
                            result.translations)

                        gql_dictionary.translations = (
                            translations if translations is not None else gql_none_value)

                    # Standard attributes.

                    for attribute in attribute_set:

                        value = (
                            getattr(dictionary, attribute))

                        if attribute == 'additional_metadata':

                            value = AdditionalMetadata.from_object(value)

                        elif attribute == 'created_at':

                            value = CreatedAt.from_timestamp(value)

                        setattr(
                            gql_dictionary,
                            attribute,
                            value)

                    (gql_language_dict[
                        dictionary.parent_id]

                        .dictionaries
                        .append(gql_dictionary))

                    if gql_dictionary_dict_flag:

                        gql_dictionary_dict[dictionary_id] = gql_dictionary

            else:

                # We are getting attribute values as they are.

                for result in result_list:

                    dictionary_id = (
                        result.client_id, result.object_id)

                    dictionary_parent_id = (
                        result.parent_client_id, result.parent_object_id)

                    gql_dictionary = (
                        Dictionary(id = dictionary_id))

                    for field_str in ds.field_set:

                        if field_str == 'id':

                            gql_dictionary.id = dictionary_id

                            continue

                        elif field_str == 'parent_id':

                            gql_dictionary.parent_id = dictionary_parent_id

                            continue

                        elif field_str == 'state_translation_gist_id':

                            gql_dictionary.state_translation_gist_id = (

                                result.state_translation_gist_client_id,
                                result.state_translation_gist_object_id)

                            continue

                        elif field_str == 'translation_gist_id':

                            gql_dictionary.translation_gist_id = (

                                result.translation_gist_client_id,
                                result.translation_gist_object_id)

                            continue

                        value = getattr(result, field_str)

                        if field_str == 'additional_metadata':

                            value = AdditionalMetadata.from_object(value)

                        elif field_str == 'created_at':

                            value = CreatedAt.from_timestamp(value)

                        setattr(
                            gql_dictionary,
                            field_str,
                            value if value is not None else gql_none_value)

                    (gql_language_dict[
                        dictionary_parent_id]

                        .dictionaries
                        .append(gql_dictionary))

                    if gql_dictionary_dict_flag:

                        gql_dictionary_dict[dictionary_id] = gql_dictionary

        # Getting perspectives, if required.

        if (ps := self.perspective_selection):

            p = ds.perspectives

            parent_id_tuple = (

                tuple_(
                    dbPerspective.parent_client_id,
                    dbPerspective.parent_object_id))

            condition_list = [

                parent_id_tuple.in_(

                    DBSession.query(
                        dictionary_table.c.client_id,
                        dictionary_table.c.object_id)),

                dbPerspective.marked_for_deletion == False]

            if p.with_verb_data is not None:

                with_accepted_cte = (

                    DBSession

                        .query(
                            dbValencySourceData.perspective_client_id,
                            dbValencySourceData.perspective_object_id)

                        .filter(
                            dbValencySourceData.id == dbValencySentenceData.source_id,
                            dbValencySentenceData.id == dbValencyInstanceData.sentence_id,
                            dbValencyInstanceData.id == dbValencyAnnotationData.instance_id,
                            dbValencyAnnotationData.accepted == True)

                        .group_by(
                            dbValencySourceData.perspective_client_id,
                            dbValencySourceData.perspective_object_id)

                        .cte())

                id_tuple = (

                    tuple_(
                        dbPerspective.client_id,
                        dbPerspective.object_id))

                condition_list.append(

                    (id_tuple.in_ if p.with_verb_data else
                        id_tuple.notin_)(

                        DBSession.query(
                            with_accepted_cte)))

            if ps.object_flag:

                ps.column_list = [dbPerspective]

            ps.query = (

                DBSession

                    .query(
                        *ps.column_list)

                    .filter(
                        *condition_list))

            ps.c = dbPerspective

            # If we are going to get columns, we save perspective info in a temporary table.

            if ps.columns:

                self.perspective_table_name = (

                    'perspective_' +
                     str(uuid.uuid4()).replace('-', '_'))

                perspective_table = (

                    sqlalchemy.Table(
                        self.perspective_table_name,
                        models.Base.metadata,
                        sqlalchemy.Column('client_id', SLBigInteger),
                        sqlalchemy.Column('object_id', SLBigInteger),
                        prefixes = ['temporary'],
                        postgresql_on_commit = 'drop'))

                perspective_table.create(
                    DBSession.connection())

                DBSession.execute(

                    perspective_table

                        .insert()

                        .from_select(
                            (ps.c.client_id, ps.c.object_id),
                            ps.query.with_entities(
                                ps.c.client_id, ps.c.object_id)))

                ps.query = (

                    DBSession

                        .query(
                            *ps.column_list)

                        .filter(

                            tuple_(
                                dbPerspective.client_id,
                                dbPerspective.object_id)

                                .in_(
                                    DBSession.query(perspective_table))))

                ps.c = dbPerspective

            # Checking for joins.

            join_count = 0

            if ps.translations_flag:
                join_count += 1

            if ps.status_flag:
                join_count += 1

            if ps.status_translations_flag:
                join_count += 1

            ps.cte_flag = (
                join_count > 1)

            ps.join_flag = (
                join_count >= 1)

            ps.cte = None

            if ps.cte_flag:

                ps.cte = ps.query.cte()
                ps.c = ps.cte.c

                if ps.object_flag:

                    ps.cte = (

                        aliased(
                            dbPerspective,
                            ps.cte,
                            adapt_on_names = True))

                    ps.c = ps.cte

                ps.query = (
                    DBSession.query(ps.cte))

            if ps.translations_flag:

                self.translations_join(
                    ps,
                    'translation_gist_client_id',
                    'translation_gist_object_id',
                    'translations')

            if ps.status_flag:

                self.translation_join(
                    ps,
                    ps.status.locale_id or self.info.context.get('locale_id'),
                    'state_translation_gist_client_id',
                    'state_translation_gist_object_id',
                    'status')

            if ps.status_translations_flag:

                self.translations_join(
                    ps,
                    'state_translation_gist_client_id',
                    'state_translation_gist_object_id',
                    'status_translations')

            # Perspectives go from older to newer.

            ps.query = (

                ps.query

                    .order_by(
                        ps.c.created_at,
                        ps.c.client_id,
                        ps.c.object_id))

            # Getting and processing perspective data.

            result_list = ps.query.all()

            if self.debug_flag:

                log.debug(
                    '\n perspective_query:\n ' +
                    render_statement(ps.query.statement))

            for gql_dictionary in gql_dictionary_dict.values():

                gql_dictionary.perspectives = []

            gql_perspective_dict_flag = (

                ps.columns)

            if gql_perspective_dict_flag:

                gql_perspective_dict = {}

            if ps.object_flag:

                # We are getting full ORM dbPerspective objects.

                attribute_set = ps.field_set.copy()

                attribute_set.discard('status')
                attribute_set.discard('status_translations')
                attribute_set.discard('translations')

                for result in result_list:

                    perspective = (
                        result[0] if ps.join_flag else result)

                    gql_perspective = (
                        Perspective(id = perspective.id))

                    gql_perspective.dbObject = perspective

                    # Computed attributes.

                    if ps.status_flag:

                        translation = (
                            result.status)

                        gql_perspective.status = (
                            translation if translation is not None else gql_none_value)

                    if ps.status_translations_flag:

                        translations = (
                            result.status_translations)

                        gql_perspective.translations = (
                            translations if translations is not None else gql_none_value)

                    if ps.translations_flag:

                        translations = (
                            result.translations)

                        gql_perspective.translations = (
                            translations if translations is not None else gql_none_value)

                    # Standard attributes.

                    for attribute in attribute_set:

                        value = (
                            getattr(perspective, attribute))

                        if attribute == 'additional_metadata':

                            value = AdditionalMetadata.from_object(value)

                        elif attribute == 'created_at':

                            value = CreatedAt.from_timestamp(value)

                        setattr(
                            gql_perspective,
                            attribute,
                            value)

                    (gql_dictionary_dict[
                        perspective.parent_id]

                        .perspectives
                        .append(gql_perspective))

                    if gql_perspective_dict_flag:

                        gql_perspective_dict[perspective_id] = gql_perspective

            else:

                # We are getting attribute values as they are.

                for result in result_list:

                    perspective_id = (
                        result.client_id, result.object_id)

                    perspective_parent_id = (
                        result.parent_client_id, result.parent_object_id)

                    gql_perspective = (
                        Perspective(id = perspective_id))

                    for field_str in ps.field_set:

                        if field_str == 'id':

                            gql_perspective.id = perspective_id

                            continue

                        elif field_str == 'parent_id':

                            gql_perspective.parent_id = perspective_parent_id

                            continue

                        elif field_str == 'state_translation_gist_id':

                            gql_perspective.state_translation_gist_id = (

                                result.state_translation_gist_client_id,
                                result.state_translation_gist_object_id)

                            continue

                        elif field_str == 'translation_gist_id':

                            gql_perspective.translation_gist_id = (

                                result.translation_gist_client_id,
                                result.translation_gist_object_id)

                            continue

                        value = getattr(result, field_str)

                        if field_str == 'additional_metadata':

                            value = AdditionalMetadata.from_object(value)

                        elif field_str == 'created_at':

                            value = CreatedAt.from_timestamp(value)

                        setattr(
                            gql_perspective,
                            field_str,
                            value if value is not None else gql_none_value)

                    (gql_dictionary_dict[
                        perspective_parent_id]

                        .perspectives
                        .append(gql_perspective))

                    if gql_perspective_dict_flag:

                        gql_perspective_dict[perspective_id] = gql_perspective

        # Getting columns, if required.

        if (cs := self.column_selection):

            c = ps.columns

            parent_id_tuple = (

                tuple_(
                    dbColumn.parent_client_id,
                    dbColumn.parent_object_id))

            if cs.object_flag:

                cs.column_list = [dbColumn]

            cs.query = (

                DBSession

                    .query(
                        *cs.column_list)

                    .filter(

                        parent_id_tuple.in_(

                            DBSession.query(
                                perspective_table.c.client_id,
                                perspective_table.c.object_id))))

            cs.c = dbColumn

            # Getting and processing column data.

            result_list = cs.query.all()

            if self.debug_flag:

                log.debug(
                    '\n column_query:\n ' +
                    render_statement(cs.query.statement))

            for gql_perspective in gql_perspective_dict.values():

                gql_perspective.columns = []

            if cs.object_flag:

                # We are getting full ORM dbColumn objects.

                for result in result_list:

                    column = result

                    gql_column = (
                        Column(id = column.id))

                    gql_column.dbObject = column

                    # Standard attributes.

                    for attribute in cs.field_set:

                        value = (
                            getattr(column, attribute))

                        if attribute == 'created_at':

                            value = CreatedAt.from_timestamp(value)

                        setattr(
                            gql_column,
                            attribute,
                            value)

                    (gql_perspective_dict[
                        column.parent_id]

                        .columns
                        .append(gql_column))

            else:

                # We are getting attribute values as they are.

                for result in result_list:

                    column_id = (
                        result.client_id, result.object_id)

                    column_parent_id = (
                        result.parent_client_id, result.parent_object_id)

                    gql_column = (
                        Column(id = column_id))

                    for field_str in cs.field_set:

                        if field_str == 'field_id':

                            gql_column.field_id = (

                                result.field_client_id,
                                result.field_object_id)

                            continue

                        elif field_str == 'id':

                            gql_column.id = column_id

                            continue

                        elif field_str == 'link_id':

                            link_client_id = (
                                result.link_client_id)

                            gql_column.link_id = (

                                gql_none_value

                                if link_client_id is None else

                                (link_client_id,
                                    result.link_object_id))

                            continue

                        elif field_str == 'parent_id':

                            gql_column.parent_id = column_parent_id

                            continue

                        elif field_str == 'self_id':

                            self_client_id = (
                                result.self_client_id)

                            gql_column.self_id = (

                                gql_none_value

                                if self_client_id is None else

                                (self_client_id,
                                    result.self_object_id))

                            continue

                        value = getattr(result, field_str)

                        if field_str == 'created_at':

                            value = CreatedAt.from_timestamp(value)

                        setattr(
                            gql_column,
                            field_str,
                            value if value is not None else gql_none_value)

                    (gql_perspective_dict[
                        column_parent_id]

                        .columns
                        .append(gql_column))

        return gql_language_list


class Query(graphene.ObjectType):
    client = graphene.String()
    dictionaries = graphene.List(Dictionary, published=graphene.Boolean(),
                                 mode=graphene.Int(),
                                 category=graphene.Int(),
                                 proxy=graphene.Boolean())
    dictionary = graphene.Field(Dictionary, id=LingvodocID())
    perspectives = graphene.List(Perspective,
        published=graphene.Boolean(),
        with_phonology_data=graphene.Boolean(),
        with_valency_data=graphene.Boolean())
    perspective = graphene.Field(Perspective, id=LingvodocID())
    entity = graphene.Field(Entity, id=LingvodocID())
    language = graphene.Field(Language, id=LingvodocID())

    languages = (

        graphene.List(
            Language,
            id_list = graphene.List(LingvodocID),
            only_in_toc = graphene.Boolean(),
            only_with_dictionaries_recursive = graphene.Boolean(),
            dictionary_category = graphene.Int(),
            dictionary_published = graphene.Boolean(),
            in_tree_order = graphene.Boolean()))

    user = graphene.Field(User, id=graphene.Int())
    users = graphene.List(User, search=graphene.String())
    field = graphene.Field(Field, id=LingvodocID())
    translationgist = graphene.Field(TranslationGist, id=LingvodocID())
    userblob = graphene.Field(UserBlobs, id=LingvodocID())
    translationatom = graphene.Field(TranslationAtom, id=LingvodocID())
    organization = graphene.Field(Organization, id=LingvodocID())

    organizations = (

        graphene.List(
            Organization,
            has_participant = graphene.Boolean(),
            participant_deleted = graphene.Boolean(),
            participant_category = graphene.Int(),
            participant_published = graphene.Boolean()))

    lexicalentry = graphene.Field(LexicalEntry, id=LingvodocID())
    basic_search = graphene.Field(LexicalEntriesAndEntities, searchstring=graphene.String(),
                                  can_add_tags=graphene.Boolean(),
                                  perspective_id=LingvodocID(), field_id=LingvodocID(),
                                  search_in_published=graphene.Boolean())
    advanced_lexicalentries = graphene.List(LexicalEntry, searchstrings=graphene.List(ObjectVal),
                                            perspectives=LingvodocID(),
                                            adopted=graphene.Boolean(),
                                            adopted_type=LingvodocID(),
                                            with_entimology=graphene.Boolean())
    translationgists = graphene.List(TranslationGist, gists_type=graphene.String())

    translation_search = (
        graphene.List(
            TranslationGist,
            searchstring = graphene.String(),
            search_case_insensitive = graphene.Boolean(),
            search_regular_expression = graphene.Boolean(),
            translation_type = graphene.String(),
            deleted = graphene.Boolean(),
            order_by_type = graphene.Boolean(),
            no_result_error_flag = graphene.Boolean()))

    translation_service_search = graphene.Field(TranslationGist, searchstring=graphene.String())
    advanced_translation_search = graphene.List(TranslationGist, searchstrings=graphene.List(graphene.String))
    optimized_translation_search = graphene.List(graphene.String, searchstrings=graphene.List(graphene.String))
    all_locales = graphene.List(ObjectVal)
    user_blobs = graphene.List(UserBlobs, data_type=graphene.String(), is_global=graphene.Boolean())
    userrequest = graphene.Field(UserRequest, id=graphene.Int())
    userrequests = graphene.List(UserRequest)
    all_basegroups = graphene.List(BaseGroup)
    all_data_types = graphene.List(TranslationGist)
    all_fields = graphene.List(Field, common=graphene.Boolean())
    common_fields = graphene.List(Field)
    all_statuses = graphene.List(TranslationGist)
    template_fields = graphene.List(Field, mode=graphene.String())
    template_modes = graphene.List(graphene.String)
    grant = graphene.Field(Grant, id=graphene.Int())

    grants = (

        graphene.List(
            Grant,
            has_participant = graphene.Boolean(),
            participant_deleted = graphene.Boolean(),
            participant_category = graphene.Int(),
            participant_published = graphene.Boolean()))

    column = graphene.Field(Column, id=LingvodocID())

    phonology_tier_list = graphene.Field(TierList, perspective_id=LingvodocID(required=True))
    phonology_skip_list = graphene.Field(SkipList, perspective_id=LingvodocID(required=True))

    phonology_link_perspective_data = graphene.Field(
        Link_Perspective_Data,
        perspective_id = LingvodocID(required = True),
        field_id_list = graphene.List(LingvodocID, required = True))

    connected_words = graphene.Field(LexicalEntriesAndEntities, id=LingvodocID(required=True),
                                     field_id=LingvodocID(required=True), mode=graphene.String())

    advanced_search = (

        graphene.Field(
            AdvancedSearch,
            languages = graphene.List(LingvodocID),
            dicts_to_filter = graphene.List(LingvodocID),
            tag_list = graphene.List(graphene.String),
            category = graphene.Int(),
            adopted = graphene.Boolean(),
            etymology = graphene.Boolean(),
            diacritics = graphene.String(),
            search_strings = graphene.List(graphene.List(ObjectVal), required = True),
            mode = graphene.String(),
            search_metadata = ObjectVal(),
            simple = graphene.Boolean(),
            xlsx_export = graphene.Boolean(),
            cognates_flag = graphene.Boolean(),
            load_entities = graphene.Boolean(),
            debug_flag = graphene.Boolean()))

    advanced_search_simple = (

        graphene.Field(
            AdvancedSearchSimple,
            languages = graphene.List(LingvodocID),
            dicts_to_filter = graphene.List(LingvodocID),
            tag_list = graphene.List(graphene.String),
            category = graphene.Int(),
            adopted = graphene.Boolean(),
            etymology = graphene.Boolean(),
            diacritics = graphene.String(),
            search_strings = graphene.List(graphene.List(ObjectVal), required = True),
            mode = graphene.String()))

    convert_markup = graphene.Field(
        graphene.String, id=LingvodocID(required=True))

    eaf_wordlist = graphene.Field(
        graphene.List(graphene.String), id=LingvodocID(required=True))

    permission_lists = graphene.Field(Permissions, proxy=graphene.Boolean(required=True))
    tasks = graphene.List(Task)
    is_authenticated = graphene.Boolean()
    dictionary_dialeqt_get_info = graphene.Field(DialeqtInfo, blob_id=LingvodocID(required=True))

    convert_five_tiers_validate = (

        graphene.Field(
            graphene.List(ObjectVal),
            markup_id_list = graphene.List(LingvodocID, required=True)))

    merge_suggestions = graphene.Field(MergeSuggestions, perspective_id=LingvodocID(required=True),
                                       algorithm=graphene.String(required=True),
                                       entity_type_primary=graphene.String(),
                                       entity_type_secondary=graphene.String(),
                                       levenshtein=graphene.Int(),
                                       threshold=graphene.Float(),
                                       field_selection_list=graphene.List(ObjectVal), )
    select_tags_metadata = ObjectVal()
    get_user_groups = graphene.Field(
        graphene.List(graphene.String), id=graphene.Int(required=False))
    perspectives_fields_intersection = graphene.Field(
        graphene.List(Field), perspectives=graphene.List(LingvodocID),)

    eaf_search = (
        graphene.Field(EafSearch,
            perspective_id = LingvodocID(),
            search_query = graphene.Argument(ObjectVal),
            debug_flag = graphene.Boolean()))

    version = graphene.String()
    version_uniparser = ObjectVal()

    client_list = (
        graphene.Field(
            graphene.List(graphene.List(graphene.Int)),
            client_id_list = graphene.List(graphene.Int, required = True)))
    parser_results = graphene.Field((graphene.List(ParserResult)),
                                    entity_id = LingvodocID(), parser_id=LingvodocID())
    parser_result = graphene.Field(ParserResult, id=LingvodocID())
    parsers = graphene.Field(graphene.List(Parser))

    unstructured_data = (

        graphene.Field(
            UnstructuredData,
            id = graphene.String(required = True)))

    valency_data = (

        graphene.Field(
            ObjectVal,
            perspective_id = LingvodocID(required = True),
            offset = graphene.Int(),
            limit = graphene.Int(),
            verb_prefix = graphene.String(),
            case_flag = graphene.Boolean(),
            accept_value = graphene.Boolean(),
            sort_order_list = graphene.List(graphene.String),
            debug_flag = graphene.Boolean()))

    language_toc = graphene.List(Language)

    language_tree = (

        graphene.Field(
            LanguageTree,
            dictionary_category = graphene.Int(),
            dictionary_published = graphene.Boolean(),
            language_id = LingvodocID(),
            by_grants = graphene.Boolean(),
            grant_id = graphene.Int(),
            by_organizations = graphene.Boolean(),
            organization_id = graphene.Int(),
            debug_flag = graphene.Boolean()))

    fill_logs = graphene.String(worker = graphene.Int())

    def resolve_fill_logs(self, info, worker=1):
        # Check if the current user is administrator
        client_id = info.context.get('client_id')
        user_id = DBSession.query(Client.user_id).filter_by(id=client_id).scalar()
        if user_id != 1:
            return ResponseError("Only administrator can fill in logs using request.")

        period = 300
        step = 0.25
        times = int(period // step)
        for _ in range(times):
            log.debug(f"Lingvodoc({worker}) " * 500)
            time.sleep(step)
        return "Done"

    def resolve_language_tree(
        self,
        info,
        dictionary_category = None,
        dictionary_published = None,
        language_id = None,
        by_grants = False,
        grant_id = None,
        by_organizations = False,
        organization_id = None,
        debug_flag = False):

        try:

            language_field_asts = []

            for field in info.field_asts:

                if field.name.value != 'language_tree':
                    continue

                language_field_asts = (
                    field.selection_set.selections)

            resolver = (

                Language_Resolver(

                    info,
                    language_field_asts,

                    language_resolver_args(
                        only_with_dictionaries_recursive = True,
                        dictionary_category = dictionary_category,
                        dictionary_published = dictionary_published,
                        language_id = language_id,
                        by_grants = by_grants,
                        grant_id = grant_id,
                        by_organizations = by_organizations,
                        organization_id = organization_id,
                        in_tree_order = True),

                    debug_flag = debug_flag))

            gql_language_list = (
                resolver.run())

            from_to_dict = (
                resolver.from_to_dict)

            filtered_out_id_set = (
                resolver.filtered_out_id_set)

            tree_object = None

            # Constructing per grant / per organization language trees.

            if resolver.grant_or_organization:

                count_dict = (
                    resolver.aggregate_count_dict)

                count_dict[(None, None)] = {}

                recursive_count_dict = {}

                def f(language_id):

                    group_count_dict = (
                        recursive_count_dict.get(language_id))

                    if group_count_dict is not None:
                        return group_count_dict

                    group_count_dict = (

                        collections.Counter(
                            count_dict[language_id]))

                    for to_id in from_to_dict[language_id]:

                        group_count_dict += f(to_id)

                    recursive_count_dict[language_id] = group_count_dict

                    return group_count_dict

                # Getting properly ordered list of grants / organizations with dictionaries.

                id_str_set = (

                    f((None, None)).keys())

                id_empty_flag = (

                    '' in id_str_set)

                id_list = [

                    int(id_str)
                    for id_str in id_str_set
                    if id_str]

                if by_grants:

                    id_row_list = (

                        DBSession

                            .query(
                                cast(dbGrant.id, models.UnicodeText))

                            .filter(
                                dbGrant.id.in_(
                                    utils.values_query(
                                        id_list, models.SLBigInteger)))

                            .order_by(
                                dbGrant.grant_number,
                                dbGrant.id)

                            .all())

                else:

                    id_row_list = (

                        DBSession

                            .query(
                                cast(dbOrganization.id, models.UnicodeText))

                            .filter(
                                dbOrganization.id.in_(
                                    utils.values_query(
                                        id_list, models.SLBigInteger)))

                            .order_by(
                                dbOrganization.id)

                            .all())

                id_str_list = [

                    id_row[0]
                    for id_row in id_row_list]

                if id_empty_flag:

                    id_str_list.append('')

                if debug_flag:

                    log.debug(
                        f'\n id_list:\n{id_list}')

                tree_object_list = []

                def g(
                    language_id,
                    group_id_str,
                    dictionary_count = 0):

                    node_list = [

                        language_id if language_id != (None, None) else
                        int(group_id_str) if group_id_str else
                        None]

                    item_list = []

                    for to_id in from_to_dict[language_id]:

                        count = (
                            f(to_id).get(group_id_str, 0))

                        item, count = (
                            g(to_id, group_id_str, count))

                        if count > 0:

                            item_list.append(item)
                            dictionary_count += count

                    if item_list:

                        node_list.append(item_list)

                    return (
                        node_list, dictionary_count)

                for id_str in id_str_list:

                    tree_object, _ = (
                        g((None, None), id_str))

                    tree_object_list.append(
                        tree_object)

                tree_object = [
                    None, tree_object_list]

            # Constructing standard language tree, if we actually have language data.

            elif gql_language_list:

                root_id = (

                    tuple(language_id) if language_id else
                    (None, None))

                def f(language_id):

                    node_list = [

                        None if language_id == (None, None) else
                        language_id]

                    item_list = [

                        f(to_id)
                        for to_id in from_to_dict[language_id]
                        if to_id not in filtered_out_id_set]

                    if item_list:

                        node_list.append(item_list)

                    return node_list
                        
                tree_object = f(root_id)

            if debug_flag:

                log.debug(

                    '\n tree_object:\n' +

                    pprint.pformat(
                        tree_object, width = 144))

            return (

                LanguageTree(
                    tree = tree_object,
                    languages = gql_language_list))

        except Exception as exception:

            traceback_string = (

                ''.join(
                    traceback.format_exception(
                        exception, exception, exception.__traceback__))[:-1])

            log.warning('language_tree: exception')
            log.warning(traceback_string)

            return (
                ResponseError(
                    'Exception:\n' + traceback_string))

    def resolve_valency_data(
        self,
        info,
        perspective_id,
        offset = 0,
        limit = 25,
        verb_prefix = None,
        case_flag = False,
        accept_value = None,
        sort_order_list = None,
        debug_flag = False,
        **args):

        log.debug(
            f'\nperspective_id: {perspective_id}'
            f'\noffset: {offset}'
            f'\nlimit: {limit}'
            f'\nverb_prefix: {repr(verb_prefix)}'
            f'\ncase_flag: {case_flag}'
            f'\naccept_value: {accept_value}'
            f'\nsort_order_list: {sort_order_list}'
            f'\ndebug_flag: {debug_flag}')

        if sort_order_list is None:
            sort_order_list = ['verb', 'case', 'accept']

        verb_flag = verb_prefix is not None
        accept_flag = accept_value is not None

        # We'll need source deletion info and acceptance info to filter out not accepted instances from
        # deleted sources.

        source_id_cte = (

            DBSession

                .query(
                    dbValencySourceData.id)

                .filter(
                    dbValencySourceData.perspective_client_id == perspective_id[0],
                    dbValencySourceData.perspective_object_id == perspective_id[1])

                .cte())

        source_parser_query = (

            DBSession

                .query(

                    source_id_cte.c.id
                        .label('id'),

                    (dbParserResult.marked_for_deletion |
                        dbEntity.marked_for_deletion |
                        dbLexicalEntry.marked_for_deletion)

                        .label('marked_for_deletion'))

                .filter(
                    source_id_cte.c.id == dbValencyParserData.id,
                    dbParserResult.client_id == dbValencyParserData.parser_result_client_id,
                    dbParserResult.object_id == dbValencyParserData.parser_result_object_id,
                    dbEntity.client_id == dbParserResult.entity_client_id,
                    dbEntity.object_id == dbParserResult.entity_object_id,
                    dbLexicalEntry.client_id == dbEntity.parent_client_id,
                    dbLexicalEntry.object_id == dbEntity.parent_object_id))

        source_eaf_query = (

            DBSession

                .query(

                    source_id_cte.c.id
                        .label('id'),

                    (dbEntity.marked_for_deletion |
                        dbLexicalEntry.marked_for_deletion)

                        .label('marked_for_deletion'))

                .filter(
                    source_id_cte.c.id == dbValencyEafData.id,
                    dbEntity.client_id == dbValencyEafData.entity_client_id,
                    dbEntity.object_id == dbValencyEafData.entity_object_id,
                    dbLexicalEntry.client_id == dbEntity.parent_client_id,
                    dbLexicalEntry.object_id == dbEntity.parent_object_id))

        source_subquery = (

            source_parser_query
                .union(source_eaf_query)
                .subquery())

        accept_subquery = (

            DBSession

                .query(
                    dbValencyAnnotationData.instance_id,

                    func.bool_or(dbValencyAnnotationData.accepted)
                        .label('accept_value'))

                .group_by(
                    dbValencyAnnotationData.instance_id)

                .subquery())

        instance_query = (

            DBSession

                .query(
                    dbValencyInstanceData)

                .filter(
                    dbValencySentenceData.source_id == source_subquery.c.id,
                    dbValencyInstanceData.sentence_id == dbValencySentenceData.id)

                .outerjoin(
                    accept_subquery,
                    dbValencyInstanceData.id == accept_subquery.c.instance_id)

                .filter(
                    or_(
                        ~source_subquery.c.marked_for_deletion,
                        func.coalesce(
                            accept_subquery.c.accept_value, False))))

        # We'll need that for getting verb merge groupings.

        merge_filter_list = [
            dbValencyMergeData.perspective_client_id == perspective_id[0],
            dbValencyMergeData.perspective_object_id == perspective_id[1]]

        if verb_prefix:

            verb_prefix_filter_str = (
                verb_prefix.replace('%', '\\%') + '%')

            merge_filter_list.append(

                dbValencyMergeData.verb_lex.ilike(
                    verb_prefix_filter_str))

        lex_id_cte = (

            DBSession

                .query(
                    dbValencyMergeData.verb_lex,
                    dbValencyMergeData.merge_id)

                .filter(
                    *merge_filter_list)

                .cte())

        id_lex_list_cte = (

            DBSession

                .query(
                    dbValencyMergeData.merge_id,

                    func.array_agg(
                        postgresql.aggregate_order_by(
                            dbValencyMergeData.verb_lex,
                            dbValencyMergeData.verb_lex))

                        .label('verb_lex_list'))

                .filter(
                    dbValencyMergeData.perspective_client_id == perspective_id[0],
                    dbValencyMergeData.perspective_object_id == perspective_id[1],

                    dbValencyMergeData.merge_id.in_(
                        DBSession.query(lex_id_cte.c.merge_id)))

                .group_by(
                    dbValencyMergeData.merge_id)

                .cte())

        # Filtering by verb prefix, if required.

        if verb_prefix:

            verb_lex_subquery = (

                DBSession

                    .query(
                        dbValencyMergeData.verb_lex)

                    .filter(
                        dbValencyMergeData.perspective_client_id == perspective_id[0],
                        dbValencyMergeData.perspective_object_id == perspective_id[1],

                        dbValencyMergeData.merge_id.in_(
                            DBSession.query(lex_id_cte.c.merge_id)))

                    .subquery())

            instance_query = (

                instance_query.filter(

                    or_(
                        dbValencyInstanceData.verb_lex.ilike(
                            verb_prefix_filter_str),

                        dbValencyInstanceData.verb_lex.in_(
                            verb_lex_subquery))))

        instance_count = (
            instance_query.count())

        if debug_flag:

            log.debug(
                '\ninstance_query:\n' +
                str(instance_query.statement.compile(compile_kwargs = {'literal_binds': True})))

            log.debug(
                f'\ninstance_count: {instance_count}')

        # Getting ready to sort, if required.

        order_by_list = []

        for sort_type in sort_order_list:

            if sort_type == 'verb':

                if verb_flag:

                    lex_list_subquery = (

                        DBSession

                            .query(
                                dbValencyMergeData.verb_lex,
                                id_lex_list_cte.c.verb_lex_list)

                            .filter(
                                dbValencyMergeData.perspective_client_id == perspective_id[0],
                                dbValencyMergeData.perspective_object_id == perspective_id[1],
                                dbValencyMergeData.merge_id == id_lex_list_cte.c.merge_id)

                            .subquery())

                    instance_query = (

                        instance_query

                            .outerjoin(
                                lex_list_subquery,
                                dbValencyInstanceData.verb_lex == lex_list_subquery.c.verb_lex))

                    order_by_list.extend((

                        func.coalesce(
                            lex_list_subquery.c.verb_lex_list,
                            postgresql.array(
                                [dbValencyInstanceData.verb_lex])),

                        dbValencyInstanceData.verb_lex))

            elif sort_type == 'case':

                if case_flag:

                    # Getting case ordering mapping as a temporary table.

                    case_table_name = (

                        'case_table_' +
                        str(uuid.uuid4()).replace('-', '_'))

                    case_value_str = (

                        ', '.join(
                            f'(\'{case_str}\', {index})'
                            for index, case_str in (
                                enumerate(CreateValencyData.case_list))))

                    DBSession.execute(f'''

                        create temporary table

                        {case_table_name} (
                          case_str TEXT PRIMARY KEY,
                          order_value INT NOT NULL)

                        on commit drop;

                        insert into {case_table_name}
                        values {case_value_str};

                        ''')

                    class tmpCaseOrder(models.Base):

                        __tablename__ = case_table_name

                        case_str = (
                            sqlalchemy.Column(sqlalchemy.types.UnicodeText, primary_key = True))

                        order_value = (
                            sqlalchemy.Column(sqlalchemy.types.Integer, nullable = False))

                    # Ordering by cases.

                    instance_query = (

                        instance_query.outerjoin(
                            tmpCaseOrder,
                            dbValencyInstanceData.case_str == tmpCaseOrder.case_str))

                    order_by_list.append(
                        tmpCaseOrder.order_value)

            elif sort_type == 'accept':

                if accept_flag:

                    order_by_list.append(
                        func.coalesce(accept_subquery.c.accept_value, False) != accept_value)

        order_by_list.append(
            dbValencyInstanceData.id)

        # Getting annotation instances and related info.

        instance_query = (

            instance_query
                .order_by(*order_by_list)
                .offset(offset)
                .limit(limit))

        instance_list = instance_query.all()

        if debug_flag:

            log.debug(
                f'\ninstance_query ({len(instance_list)}):\n' +
                str(instance_query.statement.compile(compile_kwargs = {'literal_binds': True})))

        instance_id_set = (
            set(instance.id for instance in instance_list))

        instance_verb_lex_set = (
            set(instance.verb_lex for instance in instance_list))

        sentence_id_set = (
            set(instance.sentence_id for instance in instance_list))

        log.debug(
            '\ninstance_id_set: {}'
            '\nsentence_id_set: {}'.format(
                instance_id_set,
                sentence_id_set))

        sentence_list = []

        if sentence_id_set:

            sentence_list = (

                DBSession

                    .query(
                        dbValencySentenceData)

                    .filter(
                        dbValencySentenceData.id.in_(

                            utils.values_query(
                                sentence_id_set, models.SLBigInteger)))

                    .all())

        annotation_list = []

        if instance_id_set:

            annotation_list = (

                DBSession

                    .query(
                        dbValencyAnnotationData.instance_id,

                        func.jsonb_agg(
                            func.jsonb_build_array(
                                dbValencyAnnotationData.user_id,
                                dbValencyAnnotationData.accepted)))

                    .filter(
                        dbValencyAnnotationData.instance_id.in_(

                            utils.values_query(
                                instance_id_set, models.SLBigInteger)))

                    .group_by(
                        dbValencyAnnotationData.instance_id)

                    .all())

        user_id_set = (

            set(user_id
                for _, user_annotation_list in annotation_list
                for user_id, _ in user_annotation_list))

        user_list = []

        if user_id_set:

            user_list = (

                DBSession

                    .query(
                        dbUser.id, dbUser.name)

                    .filter(
                        dbUser.id.in_(

                            utils.values_query(
                                user_id_set, models.SLBigInteger)))

                    .all())

        instance_list = [

            {'id': instance.id,
                'sentence_id': instance.sentence_id,
                'index': instance.index,
                'verb_lex': instance.verb_lex,
                'case_str': instance.case_str}

            for instance in instance_list]

        sentence_list = [
            dict(sentence.data, id = sentence.id)
            for sentence in sentence_list]

        merge_list = []

        if instance_verb_lex_set:

            merge_id_subquery = (

                DBSession

                    .query(
                        lex_id_cte.c.merge_id)

                    .filter(
                        lex_id_cte.c.verb_lex.in_(

                            utils.values_query(
                                instance_verb_lex_set, models.String)))

                    .subquery())

            merge_list = (

                DBSession

                    .query(
                        id_lex_list_cte.c.verb_lex_list)

                    .filter(
                        id_lex_list_cte.c.merge_id.in_(
                            merge_id_subquery))

                    .all())

            merge_list = [
                item[0] for item in merge_list]

        log.debug(

            '\ninstance_list ({}):\n{}'
            '\nmerge_list ({}):\n{}'
            '\nsentence_list ({}):\n{}'
            '\nannotation_list ({}):\n{}'
            '\nuser_list ({}):\n{}'.format(

                len(instance_list),
                pprint.pformat(instance_list, width = 192),
                len(merge_list),
                pprint.pformat(merge_list, width = 192),
                len(sentence_list),
                pprint.pformat(sentence_list, width = 192),
                len(annotation_list),
                pprint.pformat(annotation_list, width = 192),
                len(user_list),
                pprint.pformat(user_list, width = 192)))

        result_dict = {
            'instance_count': instance_count,
            'instance_list': instance_list,
            'merge_list': merge_list,
            'sentence_list': sentence_list,
            'annotation_list': annotation_list,
            'user_list': user_list}

        # Getting all verbs, without filtering, if required.

        if verb_flag:

            if verb_prefix:

                verb_query = (

                    DBSession

                        .query(
                            dbValencyInstanceData.verb_lex,
                            dbValencyInstanceData.verb_lex.ilike(verb_prefix_filter_str)))

            else:

                verb_query = (

                    DBSession

                        .query(
                            dbValencyInstanceData.verb_lex))

            verb_list = (

                verb_query

                    .filter(
                        dbValencySourceData.perspective_client_id == perspective_id[0],
                        dbValencySourceData.perspective_object_id == perspective_id[1],
                        dbValencySentenceData.source_id == dbValencySourceData.id,
                        dbValencyInstanceData.sentence_id == dbValencySentenceData.id)

                    .distinct()

                    .order_by(dbValencyInstanceData.verb_lex)

                    .all())

            if verb_prefix:

                result_dict['verb_list'] = (
                    [list(row) for row in verb_list])

            else:

                result_dict['verb_list'] = (
                    [[row[0], True] for row in verb_list])

        return result_dict

    def resolve_unstructured_data(self, info, id):
        return UnstructuredData(id = id)

    def resolve_client_list(
        self,
        info,
        client_id_list):

        client_list = (

            DBSession

                .query(
                    Client.id,
                    Client.user_id)

                .filter(
                    Client.id.in_(set(client_id_list)))

                .distinct())

        return client_list

    def resolve_version(self, info):
        return lingvodoc.version.__version__

    def resolve_version_uniparser(self, info):
        return lingvodoc.version.uniparser_version_dict

    def resolve_eaf_search(
        self,
        info,
        perspective_id = None,
        search_query = None,
        debug_flag = False):

        return EafSearch.constructor(
            info, perspective_id, search_query, debug_flag)

    def resolve_perspectives_fields_intersection(self, info, perspectives=None):
        """
        query qwe($perspective_list: [LingvodocID] ){
        perspective_fields_intersection(perspectives: $perspective_list){
            id
            data_type
            translation
        }
        }


        {
            "perspective_list": [
                [
                    504,
                    5
                ],
                [
                    743,
                    127939
                ]
            ]
        }
        """

        field_ids_cte = DBSession.query(
            dbColumn.field_client_id, dbColumn.field_object_id).filter(
            tuple_(dbColumn.parent_client_id, dbColumn.parent_object_id).in_(perspectives)).distinct().cte()
        field_objects = DBSession.query(dbField).join(field_ids_cte).filter(dbField.client_id==field_ids_cte.c.field_client_id,
                                                                            dbField.object_id==field_ids_cte.c.field_object_id,
                                                                            ).distinct()


        gql_fields = list()
        for db_field in field_objects:
            if db_field.data_type == "Text":
                gql_field = Field()
                gql_field.dbObject = db_field
                gql_fields.append(gql_field)

        return gql_fields

    def resolve_get_user_groups(self, info, id=None):
        if id:
            user = DBSession.query(dbUser).filter(dbUser.id == id).first()
        else:
            client_id = info.context.get('client_id')
            user = Client.get_user_by_client_id(client_id)
        if not user:
            return None
        all_basegroups = DBSession.query(dbBaseGroup).all()
        user_groups = list()
        for base in all_basegroups:
            group = DBSession.query(dbGroup).filter(dbGroup.base_group_id==base.id).first()
            if group:
                if user in group.users:
                    user_groups.append(base.name)
        return user_groups

    def resolve_select_tags_metadata(self, info):

        def get_sorted_metadata_keys(metadata_name):

            all_values = DBSession.query(dbDictionary.additional_metadata[metadata_name]) \
                .filter(dbDictionary.additional_metadata[metadata_name] != None,
                        dbDictionary.marked_for_deletion==False)

            value_set = set()

            for value, in all_values:

                value_set.update(
                    (value_str.strip() for value_str in re.split(r'\s*,\s*', value))
                        if isinstance(value, str) else
                        value)

            return sorted(value_set)

        menu_json_data = {}
        authors_list = get_sorted_metadata_keys("authors")
        menu_json_data["hasAudio"] = [False, True]
        menu_json_data["authors"] = authors_list
        menu_json_data["humanSettlement"] = get_sorted_metadata_keys("humanSettlement")
        menu_json_data["years"] = get_sorted_metadata_keys("years")
        menu_json_data["kind"] = ["Expedition", "Archive"]
        menu_json_data["nativeSpeakersCount"] = ["Vulnerable",
                                                 "Definitely endangered",
                                                 "Critically endangered",
                                                 "Extinct",
                                                 "Severely endangered",
                                                 "Safe"]
        return menu_json_data



    def resolve_merge_suggestions(self, info, perspective_id, algorithm, threshold=0.1,
                                  entity_type_primary='Transcription',
                                  entity_type_secondary='Translation',
                                  levenshtein=1,
                                  field_selection_list=None
                                  ):
        request = info.context.request
        locale_id = info.context.get('locale_id')

        old_field_selection_list = copy.deepcopy(field_selection_list)

        for field_selection in old_field_selection_list:
            field_selection['client_id'] = field_selection['field_id'][0]
            field_selection['object_id'] = field_selection['field_id'][1]

        result = merge_suggestions(request=request,
                                   perspective_client_id=perspective_id[0],
                                   perspective_object_id=perspective_id[1],
                                   algorithm=algorithm,
                                   entity_type_primary=entity_type_primary,
                                   entity_type_secondary=entity_type_secondary,
                                   threshold=threshold, levenshtein=levenshtein,
                                   field_selection_list=old_field_selection_list,
                                   locale_id=locale_id)

        return MergeSuggestions(user_has_permissions=result['user_has_permissions'],
                                match_result=result['match_result'])

    def resolve_convert_five_tiers_validate(self, info, markup_id_list):

        result_list = []

        for index, markup_id in enumerate(markup_id_list):

            if not markup_id:
                result_list.append(False)
                continue

            client_id, object_id = markup_id
            entity = DBSession.query(dbEntity).filter_by(client_id=client_id, object_id=object_id).first()

            if not entity:
                return ResponseError(f'No entity {client_id} / {object_id}.')

            try:

                storage = (
                    info.context.request.registry.settings['storage'])

                with storage_file(
                    storage, entity.content) as content_stream:

                    content = content_stream.read()

            except:
                return ResponseError(f'Cannot access file \'{entity.content}\'.')

            fd, filename = tempfile.mkstemp()
            with open(filename, 'wb') as temp:
                markup = tgt_to_eaf(content, entity.additional_metadata)
                temp.write(markup.encode("utf-8"))
                temp.flush()

                elan_check = elan_parser.ElanCheck(filename)
                elan_check.parse()
                is_valid = elan_check.check()

                if index == 0 and is_valid:
                    elan_reader = elan_parser.Elan(filename)
                    elan_reader.parse()
                    result_list.append(
                        elan_reader.preview())
                else:
                    result_list.append(is_valid)

            os.close(fd)
            os.remove(filename)

        return result_list

    def resolve_dictionary_dialeqt_get_info(self, info, blob_id):  # TODO: test
        blob_client_id, blob_object_id = blob_id
        blob = DBSession.query(dbUserBlobs).filter_by(client_id=blob_client_id, object_id=blob_object_id).first()
        if blob:
            filename = blob.real_storage_path
            sqconn = connect(filename)
            try:
                dict_attributes = get_dict_attributes(sqconn)
            except:
                raise ResponseError(message="database disk image is malformed")
            return DialeqtInfo(dictionary_name=dict_attributes['dictionary_name'], dialeqt_id=dict_attributes["dialeqt_id"])
        raise ResponseError(message="No such blob in the system")

    def resolve_tasks(self, info):
        request = info.context.request
        client_id = authenticated_userid(request)
        if not client_id:
            tasks_dicts = TaskStatus.get_user_tasks(anonymous_userid(request), clear_out=True)
            tasks = [Task(**task_dict) for task_dict in tasks_dicts]
            return tasks
        user = Client.get_user_by_client_id(authenticated_userid(request))
        tasks_dicts = TaskStatus.get_user_tasks(user.id, clear_out=True)
        tasks = [Task(**task_dict) for task_dict in tasks_dicts]
        return tasks

    def resolve_permission_lists(self, info, proxy):
        request = info.context.request
        if proxy:
            try_proxy(request)
        client_id = authenticated_userid(request)

        subreq = Request.blank('/translation_service_search')
        subreq.method = 'POST'
        subreq.headers = request.headers
        headers = dict()
        if request.headers.get('Cookie'):
            headers = {'Cookie': request.headers['Cookie']}
        subreq.headers = headers
        subreq.json = {'searchstring': 'Published'}
        resp = request.invoke_subrequest(subreq)

        if 'error' not in resp.json:
            published_gist_object_id, published_gist_client_id = resp.json['object_id'], resp.json['client_id']
        else:
            raise KeyError("Something wrong with the base", resp.json['error'])

        subreq = Request.blank('/translation_service_search')
        subreq.method = 'POST'
        subreq.headers = request.headers
        headers = dict()
        if request.headers.get('Cookie'):
            headers = {'Cookie': request.headers['Cookie']}
        subreq.headers = headers
        subreq.json = {'searchstring': 'Limited access'}  # todo: fix
        resp = request.invoke_subrequest(subreq)

        if 'error' not in resp.json:
            limited_gist_object_id, limited_gist_client_id = resp.json['object_id'], resp.json['client_id']
        else:
            raise KeyError("Something wrong with the base", resp.json['error'])


        dblimited = DBSession.query(dbPerspective).filter(
            and_(dbPerspective.state_translation_gist_client_id == limited_gist_client_id,
                 dbPerspective.state_translation_gist_object_id == limited_gist_object_id)
        )

        # limited_perms = [("limited", True), ("read", False), ("write", False), ("publish", False)]
        limited = list()
        for dbperspective in dblimited.all():
            perspective = Perspective(id=[dbperspective.client_id, dbperspective.object_id])
            perspective.dbObject = dbperspective
            perspective.list_name='limited'
            limited.append(perspective)
            # fulfill_permissions_on_perspectives(intermediate, pers, limited_perms)


        dbpublished = DBSession.query(dbPerspective).filter(
            and_(dbPerspective.state_translation_gist_client_id == published_gist_client_id,
                 dbPerspective.state_translation_gist_object_id == published_gist_object_id)
        )
        existing = list()
        view = list()
        for dbperspective in dbpublished.all():
            perspective = Perspective(id=[dbperspective.client_id, dbperspective.object_id])
            perspective.dbObject = dbperspective
            perspective.list_name='view'
            view.append(perspective)
            existing.append([dbperspective.client_id, dbperspective.object_id])

        if not client_id:
            return Permissions(limited=limited, view=view, edit=list(), publish=list())

        user = DBSession.query(Client).filter(client_id == Client.id).first()
        if not user:
            return None
        user_id = user.user_id
        editor_basegroup = DBSession.query(dbBaseGroup).filter(
            and_(dbBaseGroup.subject == "lexical_entries_and_entities", dbBaseGroup.action == "create")).first()
        editable_perspectives = DBSession.query(dbPerspective).join(dbGroup, and_(
            dbPerspective.client_id == dbGroup.subject_client_id,
            dbPerspective.object_id == dbGroup.subject_object_id)).join(dbGroup.users).filter(
            and_(dbUser.id == user_id,
                 dbGroup.base_group_id == editor_basegroup.id,
                 dbPerspective.marked_for_deletion == False)).all()
        edit = list()
        for dbperspective in editable_perspectives:
            perspective = Perspective(id=[dbperspective.client_id, dbperspective.object_id])
            perspective.dbObject = dbperspective
            perspective.list_name='edit'
            edit.append(perspective)

        reader_basegroup = DBSession.query(dbBaseGroup).filter(
            and_(dbBaseGroup.subject == "approve_entities", dbBaseGroup.action == "view")).first()
        readable_perspectives = DBSession.query(dbPerspective).join(dbGroup, and_(
            dbPerspective.client_id == dbGroup.subject_client_id,
            dbPerspective.object_id == dbGroup.subject_object_id)).join(dbGroup.users).filter(
            and_(dbUser.id == user_id, dbGroup.base_group_id == reader_basegroup.id)).all()

        view = list()
        for dbperspective in readable_perspectives:
            if [dbperspective.client_id, dbperspective.object_id] not in existing:
                perspective = Perspective(id=[dbperspective.client_id, dbperspective.object_id])
                perspective.dbObject = dbperspective
                perspective.list_name='view'
                view.append(perspective)

        publisher_basegroup = DBSession.query(dbBaseGroup).filter(
            and_(dbBaseGroup.subject == "approve_entities", dbBaseGroup.action == "create")).first()

        approvable_perspectives = DBSession.query(dbPerspective).join(dbGroup, and_(
            dbPerspective.client_id == dbGroup.subject_client_id,
            dbPerspective.object_id == dbGroup.subject_object_id)).join(dbGroup.users).filter(
            and_(dbUser.id == user_id, dbGroup.base_group_id == publisher_basegroup.id)).all()
        publish = list()
        for dbperspective in approvable_perspectives:
            perspective = Perspective(id=[dbperspective.client_id, dbperspective.object_id])
            perspective.dbObject = dbperspective
            perspective.list_name='publish'
            publish.append(perspective)
        return Permissions(limited=limited, view=view, edit=edit, publish=publish)


    def resolve_advanced_search(
        self,
        info,
        search_strings,
        languages = None,
        dicts_to_filter = None,
        tag_list = None,
        category = None,
        adopted = None,
        etymology = None,
        diacritics = None,
        search_metadata = None,
        mode = 'published',
        simple = True,
        xlsx_export = False,
        cognates_flag = True,
        load_entities = True,
        debug_flag = False):

        if mode == 'all':
            publish = None
            accept = True
        elif mode == 'published':
            publish = True
            accept = True
        elif mode == 'not_accepted':
            publish = None
            accept = False
        elif mode == 'deleted':
            publish = None
            accept = None
        elif mode == 'all_with_deleted':
            publish = None
            accept = None
        else:
            return ResponseError('mode: <all|published|not_accepted>')

        if simple:

            return (

                AdvancedSearchSimple().constructor(
                    info,
                    languages,
                    dicts_to_filter,
                    tag_list,
                    category,
                    adopted,
                    etymology,
                    diacritics,
                    search_strings,
                    publish,
                    accept,
                    xlsx_export,
                    cognates_flag,
                    debug_flag))

        return (

            AdvancedSearch().constructor(
                info,
                languages,
                dicts_to_filter,
                tag_list,
                category,
                adopted,
                etymology,
                diacritics,
                search_strings,
                publish,
                accept,
                search_metadata,
                xlsx_export,
                cognates_flag,
                load_entities,
                debug_flag))

    def resolve_advanced_search_simple(
        self,
        info,
        search_strings,
        languages = None,
        dicts_to_filter = None,
        tag_list = None,
        category = None,
        adopted = None,
        etymology = None,
        diacritics = None,
        mode = 'published'):

        if mode == 'all':
            publish = None
            accept = True
        elif mode == 'published':
            publish = True
            accept = True
        elif mode == 'not_accepted':
            publish = None
            accept = False
        elif mode == 'deleted':
            publish = None
            accept = None
        elif mode == 'all_with_deleted':
            publish = None
            accept = None
        else:
            return ResponseError(message="mode: <all|published|not_accepted>")

        return (

            AdvancedSearchSimple().constructor(
                info,
                languages,
                dicts_to_filter,
                tag_list,
                category,
                adopted,
                etymology,
                diacritics,
                search_strings,
                publish,
                accept))

    def resolve_template_modes(self, info):
        return ['corpora']

    def resolve_template_fields(self, info, mode=None):
        response = list()
        request = info.context.request
        if mode == 'corpora':
            data_type_query = DBSession.query(dbField) \
                .join(dbTranslationGist,
                      and_(dbField.translation_gist_object_id == dbTranslationGist.object_id,
                           dbField.translation_gist_client_id == dbTranslationGist.client_id)) \
                .join(dbTranslationGist.translationatom)
            sound_field = data_type_query.filter(dbTranslationAtom.locale_id == 2,
                                                 dbTranslationAtom.content == 'Sound').one()  # todo: a way to find this fields if wwe cannot use one
            markup_field = data_type_query.filter(dbTranslationAtom.locale_id == 2,
                                                  dbTranslationAtom.content == 'Markup').one()
            comment_field = data_type_query.filter(dbTranslationAtom.locale_id == 2,
                                                    dbTranslationAtom.content == 'Comment').one()
            sound_field =view_field_from_object(request=request, field=sound_field)
            markup_field = view_field_from_object(request=request, field=markup_field)
            comment_field = view_field_from_object(request=request, field=comment_field)
            fake_id_1 = '6f355d7a-e68d-44ab-9cf6-36f78e8f1b34'  # chosen by fair dice roll
            fake_id_2 = '51fbe0b6-2cea-4d40-a994-f6bb6f501d48'  # guaranteed to be random
            f = Field(id=[sound_field["client_id"], sound_field["object_id"]], fake_id = fake_id_1)
            f2 = Field(id=[markup_field["client_id"], markup_field["object_id"]], fake_id = fake_id_2, self_fake_id = fake_id_1)
            f.dbObject = DBSession.query(dbField).filter_by(client_id=sound_field["client_id"], object_id=sound_field["object_id"]).first()
            f2.dbObject = DBSession.query(dbField).filter_by(client_id=markup_field["client_id"], object_id=markup_field["object_id"]).first()
            response.append(f)
            response.append(f2)

            f3 = Field(id=[comment_field["client_id"], comment_field["object_id"]])
            f3.dbObject = DBSession.query(dbField).filter_by(client_id=comment_field["client_id"]).first()
            response.append(f3)
            # response[0]['contains'] = [view_field_from_object(request=request, field=markup_field)]
            # response.append(view_field_from_object(request=request, field=markup_field))
            # response.append(view_field_from_object(request=request, field=comment_field))
            #
            # return response
            #
            # response.append(TranslationGist(id=[sound_field.translation_gist_client_id, sound_field.data_type_translation_gist_object_id]))
            # response.append(TranslationGist(id=[markup_field.data_type_translation_gist_client_id, markup_field.data_type_translation_gist_object_id]))
            # response.append(TranslationGist(id=[comment_field.data_type_translation_gist_client_id, comment_field.data_type_translation_gist_object_id]))
            return response
        else:
            raise ResponseError(message='no such mode')
    def resolve_all_statuses(self, info):
        request = info.context.request
        gql_statuses = list()
        for status in ['WiP', 'Published', 'Limited access', 'Hidden']:
            db_tr_gist = translation_gist_search(status)
            gql_tr_gist = TranslationGist(id=[db_tr_gist.client_id, db_tr_gist.object_id ])
            gql_tr_gist.dbObject = db_tr_gist
            gql_statuses.append(gql_tr_gist)
        return gql_statuses

    def resolve_all_fields(self, info, common=False):
        fields = DBSession.query(dbField).filter_by(marked_for_deletion=False).all()
        if common:
            field_to_psersp_dict = collections.defaultdict(list)
            p_to_field = DBSession.query(dbColumn.parent_client_id,
                                         dbColumn.parent_object_id,
                                         dbColumn.field_client_id,
                                         dbColumn.field_object_id).filter(
                dbColumn.marked_for_deletion == False
            ).all()

            for perspective_client_id, perspective_object_id, field_client_id, field_object_id in p_to_field:
                field_to_psersp_dict[(field_client_id, field_object_id)].append((perspective_client_id,
                                                                                 perspective_object_id))
        gql_fields = list()
        for db_field in fields:
            if common:
                if len(field_to_psersp_dict[(db_field.client_id, db_field.object_id)]) <= 3:
                    continue
                if db_field.data_type != "Text":
                    continue
            gql_field = Field(id=[db_field.client_id, db_field.object_id])
            gql_field.dbObject = db_field
            gql_fields.append(gql_field)

        return gql_fields




    def resolve_all_data_types(self, info):
        response = list()
        for data_type in ['Text', 'Image', 'Sound', 'Markup', 'Link', 'Grouping Tag']:
            db_tr_gist = translation_gist_search(data_type)
            gql_tr_gist = TranslationGist(id=[db_tr_gist.client_id, db_tr_gist.object_id ])
            gql_tr_gist.dbObject = db_tr_gist
            response.append(gql_tr_gist)
        return response

    def resolve_dictionaries(self, info, published=None, mode=None, category=None, proxy=False):
        """
        example:

        query DictionaryList {
            dictionaries(published: true) {
                id
                translation
                parent_id
                translation_gist_id
                state_translation_gist_id
                category
                domain
            }
        }
        """
        request = info.context.request
        if proxy:
            try_proxy(request)

        client_id = info.context.get('client_id')
        client = DBSession.query(Client).filter_by(id=client_id).first()

        dbdicts = (

            DBSession
                .query(dbDictionary)
                .filter_by(marked_for_deletion = False))

        published_cte_query = (
            get_published_translation_gist_id_cte_query())

        if published:

            dbdicts = (

                dbdicts

                    .filter(
                        tuple_(
                            dbDictionary.state_translation_gist_client_id,
                            dbDictionary.state_translation_gist_object_id)

                            .in_(published_cte_query))

                    .join(dbPerspective)

                    .filter(
                        dbPerspective.marked_for_deletion == False,

                        tuple_(
                            dbPerspective.state_translation_gist_client_id,
                            dbPerspective.state_translation_gist_object_id)

                            .in_(published_cte_query))

                    .group_by(dbDictionary))

        if category is not None:
            if category:
                dbdicts = dbdicts.filter(dbDictionary.category == 1)
            else:
                dbdicts = dbdicts.filter(dbDictionary.category == 0)
        dbdicts = dbdicts.order_by(dbDictionary.created_at.desc())
        if mode is not None and client:
            user = DBSession.query(dbUser).filter_by(id=client.user_id).first()

            if not mode:
                # my dictionaries

                client_query = (

                    DBSession
                        .query(Client.id)
                        .filter(Client.user_id == user.id)
                        .subquery()) # user,id?

                dbdicts = dbdicts.filter(dbDictionary.client_id.in_(client_query))

            else:
                # available dictionaries

                dictstemp_set = set()
                group_tuples = []
                isadmin = False
                for group in user.groups: # todo: LOOK AT ME this is really bad. rewrite me from group point of view
                    subject_id = (group.subject_client_id, group.subject_object_id)
                    if group.parent.dictionary_default:
                        if group.subject_override:
                            isadmin = True
                            break
                        dictstemp_set.add(subject_id)
                    if group.parent.perspective_default:
                        if group.subject_override:
                            isadmin = True
                            break
                    group_tuples.append(subject_id)

                if not isadmin:

                    for i in range(0, len(group_tuples), 1000):

                        dictstemp_set.update(

                            DBSession
                                .query(dbDictionary.client_id, dbDictionary.object_id)
                                .join(dbPerspective)
                                .filter(
                                    tuple_(
                                        dbPerspective.client_id,
                                        dbPerspective.object_id)
                                        .in_(group_tuples[i : i + 1000]))
                                .all())

                    dbdicts = [o for o in dbdicts if (o.client_id, o.object_id) in dictstemp_set]

        dictionaries_list = list()
        for dbdict in dbdicts:
            gql_dict = Dictionary(id=[dbdict.client_id, dbdict.object_id])
            gql_dict.dbObject = dbdict
            dictionaries_list.append(gql_dict)
        return dictionaries_list

    def resolve_dictionary(self, info, id):
        return Dictionary(id=id)

    def resolve_perspectives(
        self,
        info,
        published = None,
        with_phonology_data = None,
        with_valency_data = None):
        """
        example:

        query LanguagesList {
            perspectives(published: true) {
                id
                translation
                parent_id
                translation_gist_id
                state_translation_gist_id
                import_source
                import_hash
            }
        }
        """

        perspective_query = (

            DBSession

                .query(
                    dbPerspective)

                .filter(
                    dbPerspective.marked_for_deletion == False))

        if published:

            db_published_gist = translation_gist_search('Published')
            state_translation_gist_client_id = db_published_gist.client_id
            state_translation_gist_object_id = db_published_gist.object_id
            db_la_gist = translation_gist_search('Limited access')
            limited_client_id, limited_object_id = db_la_gist.client_id, db_la_gist.object_id

            """
            atom_perspective_name_alias = aliased(dbTranslationAtom, name="PerspectiveName")
            atom_perspective_name_fallback_alias = aliased(dbTranslationAtom, name="PerspectiveNameFallback")
            persps = DBSession.query(dbPerspective,
                                     dbTranslationAtom,
                                     coalesce(atom_perspective_name_alias.content,
                                              atom_perspective_name_fallback_alias.content,
                                              "No translation for your locale available").label("Translation")
                                     ).filter(dbPerspective.marked_for_deletion == False)
            """

            perspective_query = perspective_query.filter(
                or_(and_(dbPerspective.state_translation_gist_object_id == state_translation_gist_object_id,
                         dbPerspective.state_translation_gist_client_id == state_translation_gist_client_id),
                    and_(dbPerspective.state_translation_gist_object_id == limited_object_id,
                         dbPerspective.state_translation_gist_client_id == limited_client_id)))

        # If required, filtering out pespectives without phonology data.
        #
        # Experiments have shown that filtering is faster through id in select group by than through exists
        # subquery. For previous exists-based filterting see the file's history.

        if with_phonology_data is not None:

            dbMarkup = aliased(dbEntity, name = 'Markup')
            dbSound = aliased(dbEntity, name = 'Sound')

            dbPublishingMarkup = aliased(dbPublishingEntity, name = 'PublishingMarkup')
            dbPublishingSound = aliased(dbPublishingEntity, name = 'PublishingSound')

            phonology_query = (

                DBSession

                    .query(
                        dbLexicalEntry.parent_client_id,
                        dbLexicalEntry.parent_object_id)

                    .filter(
                        dbLexicalEntry.marked_for_deletion == False,
                        dbMarkup.parent_client_id == dbLexicalEntry.client_id,
                        dbMarkup.parent_object_id == dbLexicalEntry.object_id,
                        dbMarkup.marked_for_deletion == False,
                        dbMarkup.additional_metadata.contains({'data_type': 'praat markup'}),
                        dbPublishingMarkup.client_id == dbMarkup.client_id,
                        dbPublishingMarkup.object_id == dbMarkup.object_id,
                        dbPublishingMarkup.published == True,
                        dbPublishingMarkup.accepted == True,
                        dbSound.client_id == dbMarkup.self_client_id,
                        dbSound.object_id == dbMarkup.self_object_id,
                        dbSound.marked_for_deletion == False,
                        dbPublishingSound.client_id == dbSound.client_id,
                        dbPublishingSound.object_id == dbSound.object_id,
                        dbPublishingSound.published == True,
                        dbPublishingSound.accepted == True)

                    .group_by(
                        dbLexicalEntry.parent_client_id,
                        dbLexicalEntry.parent_object_id))

            id_tuple = (

                tuple_(
                    dbPerspective.client_id,
                    dbPerspective.object_id))

            perspective_query = (

                perspective_query.filter(

                    (id_tuple.in_ if with_phonology_data else
                        id_tuple.notin_)(

                        DBSession.query(
                            phonology_query.cte()))))

        # If required, filtering out perspectives without valency data.
        #
        # NOTE: We explicitly need a union, if we try to use an or condition, due to something or other in
        # PostgreSQL's planner query execution time jumps from 2 to 180 seconds, like what?

        if with_valency_data is not None:

            parser_result_query = (

                DBSession

                    .query(
                        dbLexicalEntry.parent_client_id,
                        dbLexicalEntry.parent_object_id)

                    .filter(
                        dbLexicalEntry.marked_for_deletion == False,
                        dbEntity.parent_client_id == dbLexicalEntry.client_id,
                        dbEntity.parent_object_id == dbLexicalEntry.object_id,
                        dbEntity.marked_for_deletion == False,
                        dbEntity.content.op('~*')('.*\.(doc|docx|odt)'),
                        dbPublishingEntity.client_id == dbEntity.client_id,
                        dbPublishingEntity.object_id == dbEntity.object_id,
                        dbPublishingEntity.published == True,
                        dbPublishingEntity.accepted == True,
                        dbParserResult.entity_client_id == dbEntity.client_id,
                        dbParserResult.entity_object_id == dbEntity.object_id,
                        dbParserResult.marked_for_deletion == False)

                    .group_by(
                        dbLexicalEntry.parent_client_id,
                        dbLexicalEntry.parent_object_id))

            eaf_corpus_query = (

                DBSession

                    .query(
                        dbLexicalEntry.parent_client_id,
                        dbLexicalEntry.parent_object_id)

                    .filter(
                        dbLexicalEntry.marked_for_deletion == False,
                        dbEntity.parent_client_id == dbLexicalEntry.client_id,
                        dbEntity.parent_object_id == dbLexicalEntry.object_id,
                        dbEntity.marked_for_deletion == False,
                        dbEntity.content.ilike('%.eaf'),
                        dbEntity.additional_metadata.contains({'data_type': 'elan markup'}),
                        dbPublishingEntity.client_id == dbEntity.client_id,
                        dbPublishingEntity.object_id == dbEntity.object_id,
                        dbPublishingEntity.published == True,
                        dbPublishingEntity.accepted == True)

                    .group_by(
                        dbLexicalEntry.parent_client_id,
                        dbLexicalEntry.parent_object_id))

            union_list = [
                DBSession.query(parser_result_query.cte()),
                DBSession.query(eaf_corpus_query.cte()),
            ]

            id_tuple = (

                tuple_(
                    dbPerspective.client_id,
                    dbPerspective.object_id))

            perspective_query = (

                perspective_query.filter(

                    (id_tuple.in_ if with_valency_data else
                        id_tuple.notin_)(

                        union(*union_list))))

        log.debug(
            '\nperspective_query:\n' +
            render_statement(perspective_query.statement))

        perspectives_list = []

        for db_persp in perspective_query.all():

            gql_persp = Perspective(id=[db_persp.client_id, db_persp.object_id])
            gql_persp.dbObject = db_persp
            perspectives_list.append(gql_persp)

        return perspectives_list


    def resolve_perspective(self, info, id):
        return Perspective(id=id)

    def resolve_language(self, info, id):
        return Language(id=id)

    def resolve_languages(
        self,
        info,
        id_list = None,
        only_in_toc = False,
        only_with_dictionaries_recursive = False,
        dictionary_category = None,
        dictionary_published = None,
        in_tree_order = False):
        """
        example:

        query LanguagesList {
            languages {
                id
                translation
                parent_id
                translation_gist_id
            }
        }
        """

        try:

            __debug_flag__ = False

            resolver = (

                Language_Resolver(

                    info,
                    info.field_asts,

                    language_resolver_args(
                        id_list = id_list,
                        only_in_toc = only_in_toc,
                        only_with_dictionaries_recursive = only_with_dictionaries_recursive,
                        dictionary_category = dictionary_category,
                        dictionary_published = dictionary_published,
                        in_tree_order = in_tree_order),

                    debug_flag = __debug_flag__))

            return (
                resolver.run())

        except Exception as exception:

            traceback_string = (

                ''.join(
                    traceback.format_exception(
                        exception, exception, exception.__traceback__))[:-1])

            log.warning('languages: exception')
            log.warning(traceback_string)

            return (
                ResponseError(
                    'Exception:\n' + traceback_string))

    def resolve_entity(self, info, id):
        return Entity(id=id)

    def resolve_user(self, info, id=None):
        if id is None:
            client_id = info.context.get('client_id')
            client = DBSession.query(Client).filter_by(id=client_id).first()
            if not client:
                return None
            id = client.user_id
        return User(id=id)

    def resolve_is_authenticated(self, info):
        client_id = info.context.get('client_id')
        client = DBSession.query(Client).filter_by(id=client_id).first()
        if not client:
            return False
        return True

    def resolve_users(self, info, search=None):
        """
        example:

        query UsersList {
            users(search: "modis") {
                login
                name
                intl_name
            }
        }
        """
        users = DBSession.query(dbUser).join(dbUser.email)
        if search:
            name = search + '%'
            users = users.filter(or_(
                dbUser.name.startswith(name),
                dbUser.login.startswith(name),
                dbUser.intl_name.startswith(name),
                dbEmail.email.startswith(name)
            ))
        users_list = list()
        for db_user in users:
            gql_user = User(name=db_user.name)
            gql_user.dbObject = db_user
            users_list.append(gql_user)

        return users_list


    # def resolve_datetime(self, args, context, info):
    #     id = args.get('id')
    #     return DateTime(id=id)

    def resolve_basegroup(self, info, id):
        return BaseGroup(id=id)

    def resolve_client(self, info):
        context = info.context
        return context.get('client')

    def resolve_column(self, info, id):
        return Column(id=id)

    def resolve_group(self, info, id):
        return Group(id=id)

    def resolve_organization(self, info, id):
        return Organization(id=id)

    def resolve_organizations(
        self,
        info,
        has_participant = None,
        participant_deleted = None,
        participant_published = None,
        participant_category = None):

        __debug_flag__ = False

        # Analyzing query.

        object_flag = False

        selection_dict = {

            'additional_metadata': (
                dbOrganization.additional_metadata,),

            'created_at': (
                dbOrganization.created_at,),

            'id': (
                dbOrganization.id,),

            'marked_for_deletion': (
                dbOrganization.marked_for_deletion,),

            'about_translations': (
                dbOrganization.about_translation_gist_client_id,
                dbOrganization.about_translation_gist_object_id),

            'translations': (
                dbOrganization.translation_gist_client_id,
                dbOrganization.translation_gist_object_id)}

        selection_set = set()
        selection_list = []

        for field in info.field_asts:

            if field.name.value != 'organizations':
                continue

            for subfield in field.selection_set.selections:

                name_str = subfield.name.value

                if name_str in selection_dict:

                    selection_set.add(name_str)
                    selection_list.extend(selection_dict[name_str])

                elif name_str != '__typename':

                    object_flag = True

        if object_flag:

            selection_list = [dbOrganization]

        elif 'id' not in selection_set:

            # For standard organization ordering.

            selection_set.add('id')
            selection_list.extend(selection_dict['id'])

        # If we are going to query both the usual and about translations, we'll have to use separate joins
        # to a CTE-based queries to avoid joining to actually a cross product of two translations.

        translations_flag = 'translations' in selection_set
        about_translations_flag = 'about_translations' in selection_set

        cte_flag = (
            translations_flag and
            about_translations_flag)

        organization_query = None
        organization_cte = None

        # No participant filtering, getting every organization.

        if has_participant is None:

            organization_query = (

                DBSession

                    .query(
                        *selection_list)

                    .filter_by(
                        marked_for_deletion = False))

            organization_c = dbOrganization

        # Simple participant count filter.

        elif (
            participant_deleted is None and
            participant_category is None and
            participant_published is None):

            participant_count = (

                func.jsonb_array_length(
                    dbOrganization.additional_metadata['participant']))

            organization_query = (

                DBSession

                    .query(
                        *selection_list)

                    .filter(
                        dbOrganization.marked_for_deletion == False,

                        participant_count > 0 if has_participant else
                        participant_count <= 0))

            organization_c = dbOrganization

        # Additional conditions on participants, we'll have to check them through a join.
        #
        # We have to use raw SQL due to SQLAlchemy being bad with PostgreSQL's jsonb_to_recordset.

        else:

            dictionary_condition_list = []

            if participant_deleted is not None:

                dictionary_condition_list.append(
                    '\n and D.marked_for_deletion = true' if participant_deleted else
                    '\n and D.marked_for_deletion = false')

            if participant_category is not None:

                dictionary_condition_list.append(
                    f'\n and D.category = {participant_category}')

            if participant_published is not None:

                if not participant_published:
                    raise NotImplementedError

                dictionary_condition_list.append(f'''

                    and (
                      D.state_translation_gist_client_id,
                      D.state_translation_gist_object_id) in (

                      select
                        T.client_id,
                        T.object_id

                      from
                        translationgist T,
                        translationatom A

                      where
                        T.marked_for_deletion = false and
                        T.type = 'Service' and
                        A.parent_client_id = T.client_id and
                        A.parent_object_id = T.object_id and
                        A.locale_id = 2 and
                        A.marked_for_deletion = false and (
                          A.content = 'Published' or
                          A.content = 'Limited access'))

                    ''')

            dictionary_condition_str = (
                ''.join(dictionary_condition_list))

            if object_flag:

                selection_str = 'O.*'

            else:

                selection_str = (

                    ', '.join(
                        f'O.{selection.name}'
                        for selection in selection_list))

            sql_text = (

                sqlalchemy.text(f'''

                    select
                      {selection_str}

                    from
                      organization O

                    cross join
                      jsonb_to_recordset(O.additional_metadata -> 'participant')
                        P (client_id bigint, object_id bigint)

                    join
                      dictionary D

                    on
                      D.client_id = P.client_id and
                      D.object_id = P.object_id {dictionary_condition_str}

                    where
                      O.marked_for_deletion = false

                    group by
                      O.id

                    '''))

            if object_flag:

                sql_text = (

                    aliased(
                        dbOrganization,

                        sql_text

                            .columns(
                                **{column.name: column.type
                                    for column in dbOrganization.__table__.c})

                            .alias(),

                        adapt_on_names = True))

                organization_c = (
                    sql_text)

            else:

                sql_text = (

                    sql_text

                        .columns(
                            *selection_list)

                        .alias())

                organization_c = (
                    sql_text.c)

            organization_query = (
                DBSession.query(sql_text))

            if not has_participant:
                raise NotImplementedError

        # Establishing a CTE if we'll need it.

        if cte_flag:

            organization_cte = organization_query.cte()
            organization_c = organization_cte.c

            if object_flag:

                organization_cte = (

                    aliased(
                        dbOrganization,
                        organization_cte,
                        adapt_on_names = True))

                organization_c = organization_cte

            organization_query = (
                DBSession.query(organization_cte))

        # Getting translations through a join, if required.

        if translations_flag:

            if organization_cte is not None:

                translation_query = (

                    DBSession

                        .query(
                            organization_c.id)

                        .outerjoin(
                            dbTranslationAtom,

                            and_(
                                dbTranslationAtom.parent_client_id == organization_c.translation_gist_client_id,
                                dbTranslationAtom.parent_object_id == organization_c.translation_gist_object_id,
                                dbTranslationAtom.marked_for_deletion == False))

                        .add_columns(

                            func.jsonb_object_agg(
                                dbTranslationAtom.locale_id,
                                dbTranslationAtom.content)

                                .filter(dbTranslationAtom.locale_id != None)
                                .label('translations'))

                        .group_by(
                            organization_c.id)

                        .subquery())

                organization_query = (

                    organization_query

                        .join(
                            translation_query,
                            translation_query.c.id == organization_c.id)

                        .add_columns(
                            translation_query.c.translations))

            else:

                organization_query = (

                    organization_query

                        .outerjoin(
                            dbTranslationAtom,

                            and_(
                                dbTranslationAtom.parent_client_id == organization_c.translation_gist_client_id,
                                dbTranslationAtom.parent_object_id == organization_c.translation_gist_object_id,
                                dbTranslationAtom.marked_for_deletion == False))

                        .add_columns(

                            func.jsonb_object_agg(
                                dbTranslationAtom.locale_id,
                                dbTranslationAtom.content)

                                .filter(dbTranslationAtom.locale_id != None)
                                .label('translations'))

                        .group_by(
                            *([organization_c] if object_flag else organization_c)))

        # Getting about translations through a join, if required.

        if about_translations_flag:

            if organization_cte is not None:

                about_translation_query = (

                    DBSession

                        .query(
                            organization_c.id)

                        .outerjoin(
                            dbTranslationAtom,

                            and_(
                                dbTranslationAtom.parent_client_id == organization_c.about_translation_gist_client_id,
                                dbTranslationAtom.parent_object_id == organization_c.about_translation_gist_object_id,
                                dbTranslationAtom.marked_for_deletion == False))

                        .add_columns(

                            func.jsonb_object_agg(
                                dbTranslationAtom.locale_id,
                                dbTranslationAtom.content)

                                .filter(dbTranslationAtom.locale_id != None)
                                .label('about_translations'))

                        .group_by(
                            organization_c.id)

                        .subquery())

                organization_query = (

                    organization_query

                        .join(
                            about_translation_query,
                            about_translation_query.c.id == organization_c.id)

                        .add_columns(
                            about_translation_query.c.about_translations))

            else:

                organization_query = (

                    organization_query

                        .outerjoin(
                            dbTranslationAtom,

                            and_(
                                dbTranslationAtom.parent_client_id == organization_c.about_translation_gist_client_id,
                                dbTranslationAtom.parent_object_id == organization_c.about_translation_gist_object_id,
                                dbTranslationAtom.marked_for_deletion == False))

                        .add_columns(

                            func.jsonb_object_agg(
                                dbTranslationAtom.locale_id,
                                dbTranslationAtom.content)

                                .filter(dbTranslationAtom.locale_id != None)
                                .label('about_translations'))

                        .group_by(
                            *([organization_c] if object_flag else organization_c)))

        # Getting organization info, preparing it for GraphQL.

        organization_query = (

            organization_query
                .order_by(organization_c.id))

        result_list = organization_query.all()

        if __debug_flag__:

            log.debug(
                '\n organization_query:\n ' +
                render_statement(organization_query.statement))

        gql_organization_list = []

        if object_flag:

            # We are getting full ORM dbOrganization objects.

            attribute_set = selection_set.copy()

            attribute_set.discard('translations')
            attribute_set.discard('about_translations')

            for result in result_list:

                organization = (
                    result[0] if translations_flag or about_translations_flag else
                    result)

                gql_organization = (
                    Organization(id = organization.id))

                gql_organization.dbObject = organization

                if translations_flag:

                    translations = (
                        result.translations)

                    gql_organization.translations = (
                        translations if translations is not None else gql_none_value)

                if about_translations_flag:

                    translations = (
                        result.about_translations)

                    gql_organization.about_translations = (
                        translations if translations is not None else gql_none_value)

                for attribute in attribute_set:

                    value = getattr(organization, attribute)

                    if attribute == 'additional_metadata':

                        value = AdditionalMetadata.from_object(value)

                    elif attribute == 'created_at':

                        value = CreatedAt.from_timestamp(value)

                    setattr(
                        gql_organization,
                        attribute,
                        value)

                gql_organization_list.append(gql_organization)

        else:

            # We are getting attribute values as they are.

            for result in result_list:

                gql_organization = (
                    Organization(id = result.id))

                for selection in selection_set:

                    value = getattr(result, selection)

                    if selection == 'additional_metadata':

                        value = AdditionalMetadata.from_object(value)

                    elif selection == 'created_at':

                        value = CreatedAt.from_timestamp(value)

                    setattr(
                        gql_organization,
                        selection,
                        value if value is not None else gql_none_value)

                gql_organization_list.append(gql_organization)

        return gql_organization_list

    # def resolve_passhash(self, args, context, info):
    #     id = args.get('id')
    #     return Passhash(id=id)

    # def resolve_objecttoc(self, args, context, info):
    #     id = args.get('id')
    #     return ObjectTOC(id=id)


    def resolve_translationatom(self, info, id):
        return TranslationAtom(id=id)

    def resolve_translationgist(self, info, id):
        return TranslationGist(id=id)

    def resolve_translationgists(self, info, gists_type=None):
        """
        example:
        query GistsList {
            translationgists {
                id
                type
            }
        }
        """

        gists_query = DBSession.query(dbTranslationGist).filter_by(marked_for_deletion=False)
        if gists_type:
            gists_query = gists_query.filter(dbTranslationGist.type == gists_type)
        else:
            gists_query = gists_query.order_by(dbTranslationGist.type)
        gists = gists_query.all()
        gists_list = list()
        for db_gist in gists:
            gql_gist = TranslationGist(id=[db_gist.client_id, db_gist.object_id])
            gql_gist.dbObject = db_gist
            gists_list.append(gql_gist)

        return gists_list

    def resolve_translation_search(
        self,
        info,
        searchstring = None,
        search_case_insensitive = False,
        search_regular_expression = False,
        translation_type = None,
        deleted = None,
        order_by_type = False,
        no_result_error_flag = True):
        """
        query TranslationsList {
            translation_search(searchstring: "словарь") {
                id
                type
                translationatoms {
                     id
                     content
                }
            }
        }
        """

        # Analyzing query.

        atoms_flag = False
        atoms_deleted = None

        def f(argument):

            try:

                return argument.value.value

            except AttributeError:

                return (
                    info.variable_values.get(
                        argument.value.name.value, None))

        for field in info.field_asts:

            if field.name.value != 'translation_search':
                continue

            for subfield in field.selection_set.selections:

                if subfield.name.value != 'translationatoms':
                    continue

                atoms_flag = True

                for argument in subfield.arguments:

                    if argument.name.value != 'deleted':

                        atoms_flag = False
                        break

                    atoms_deleted = f(argument)

        # Getting ready to get gists.

        gist_query = (

            DBSession.query(
                dbTranslationGist))

        if deleted is not None:

            gist_query = (

                gist_query.filter(
                    dbTranslationGist.marked_for_deletion == deleted))

        if searchstring:

            if search_regular_expression:

                search_filter = (

                    dbTranslationAtom.content.op('~*') if search_case_insensitive else
                    dbTranslationAtom.content.op('~'))(

                        searchstring)

            else:

                search_filter = (

                    dbTranslationAtom.content.ilike if search_case_insensitive else
                    dbTranslationAtom.content.like)(

                        '%' + searchstring + '%')

            gist_id_query = (

                DBSession

                    .query(
                        dbTranslationAtom.parent_client_id,
                        dbTranslationAtom.parent_object_id)

                    .filter(search_filter))

            gist_query = (

                gist_query.filter(

                    tuple_(
                        dbTranslationGist.client_id,
                        dbTranslationGist.object_id)

                        .in_(gist_id_query)))

        if translation_type:

            gist_query = (

                gist_query
                    .filter(dbTranslationGist.type == translation_type))

        # If we need to get atoms, we'll use the gist query as subquery before we add ordering to it.

        if atoms_flag:

            gist_subquery = (
                gist_query.subquery())

            atom_query = (

                DBSession

                    .query(
                        dbTranslationAtom)

                    .filter(

                        tuple_(
                            dbTranslationAtom.parent_client_id,
                            dbTranslationAtom.parent_object_id)

                            .in_(
                                DBSession.query(
                                    gist_subquery.c.client_id,
                                    gist_subquery.c.object_id))))

            if atoms_deleted is not None:

                atom_query = (

                    atom_query.filter(
                        dbTranslationAtom.marked_for_deletion == atoms_deleted))

        if order_by_type and not translation_type:

            gist_query = (

                gist_query
                    .order_by(dbTranslationGist.type))

        try:

            gist_list = gist_query.all()

        except sqlalchemy.exc.DataError as data_error:

            if isinstance(data_error.orig, psycopg2.errors.InvalidRegularExpression):
                return ResponseError('InvalidRegularExpression')

            raise

        if gist_list or not no_result_error_flag:

            gql_gist_list = []

            if atoms_flag:
                gql_gist_dict = {}

            for gist in gist_list:

                id = gist.id

                gql_gist = TranslationGist(id = id)
                gql_gist.dbObject = gist

                gql_gist_list.append(gql_gist)

                if atoms_flag:

                    gql_gist.translationatoms = []
                    gql_gist_dict[id] = gql_gist

            # Getting atoms info if required.

            if atoms_flag:

                for atom in atom_query.all():

                    gql_atom = TranslationAtom(id = atom.id)
                    gql_atom.dbObject = atom

                    gql_gist_dict[atom.parent_id].translationatoms.append(gql_atom)

            return gql_gist_list

        raise ResponseError(message="Error: no result")

    def resolve_translation_service_search(self, info, searchstring):
        """
        query TranslationsList {
            translation_service_search(searchstring: "Converting 80%") {
                id
                type
                translationatoms {
                     id
                     content
                }
            }
        }
        """
        db_translationgist = translation_gist_search(searchstring)
        if not db_translationgist:
            raise ResponseError(message="Error: no result")
        # translationatoms_list = list()
        # for db_translationatom in db_translationgist.translationatom:
        #     translationatom_object = TranslationAtom(id=[db_translationatom.client_id, db_translationatom.object_id])
        #     translationatom_object.dbObject = db_translationatom
        #     translationatoms_list.append(translationatom_object)
        gql_translationgist = TranslationGist(id=[db_translationgist.client_id, db_translationgist.object_id])
        gql_translationgist.dbObject = db_translationgist
        return gql_translationgist


    def resolve_advanced_translation_search(self, info, searchstrings):
        """
        query TranslationsList {
            advanced_translation_search(searchstrings: ["Converting 80%", "Available dictionaries"]) {
                id
                type
                translationatoms {
                     id
                     content
                }
            }
        }
        """

        debug_flag = False

        if not searchstrings:
            raise ResponseError(message = "Error: no search strings")

        search_table_name = (

            'search_table_' +
            str(uuid.uuid4()).replace('-', '_'))

        DBSession.execute(f'''

            create temporary table

            {search_table_name} (
              index INT PRIMARY KEY,
              search_string TEXT NOT NULL)

            on commit drop;

            ''')

        class tmpSearchString(models.Base):

            __tablename__ = search_table_name

            index = (
                sqlalchemy.Column(sqlalchemy.types.Integer, primary_key = True))

            search_string = (
                sqlalchemy.Column(sqlalchemy.types.UnicodeText, nullable = False))

        insert_list = [
            {'index': index, 'search_string': search_string}
            for index, search_string in enumerate(searchstrings)]

        insert_query = (

            tmpSearchString.__table__
                .insert()
                .values(insert_list))

        if debug_flag:

            log.debug(
                '\ninsert_query:\n' +
                str(insert_query.compile(compile_kwargs = {'literal_binds': True})))

        DBSession.execute(insert_query)

        gist_query = (

            DBSession

                .query(
                    tmpSearchString.index,
                    dbTranslationGist)

                .filter(
                    dbTranslationGist.type == 'Service',
                    dbTranslationGist.marked_for_deletion == False,
                    dbTranslationAtom.parent_client_id == dbTranslationGist.client_id,
                    dbTranslationAtom.parent_object_id == dbTranslationGist.object_id,
                    dbTranslationAtom.locale_id == 2,
                    dbTranslationAtom.marked_for_deletion == False,
                    dbTranslationAtom.content == tmpSearchString.search_string)

                .distinct(
                    tmpSearchString.index)

                .order_by(
                    tmpSearchString.index,
                    dbTranslationGist.client_id,
                    dbTranslationGist.object_id))

        if debug_flag:

            log.debug(
                '\ngist_query:\n' +
                str(gist_query.statement.compile(compile_kwargs = {'literal_binds': True})))

        gist_list = gist_query.all()

        result_list = [None] * len(searchstrings)

        for index, gist in gist_list:

            gql_gist = (

                TranslationGist(
                    id = [gist.client_id, gist.object_id],
                    translation = searchstrings[index]))

            gql_gist.dbObject = gist

            result_list[index] = gql_gist

        return result_list

    def resolve_optimized_translation_search(self, info, searchstrings):

        if not searchstrings:
            raise ResponseError(message = "Error: no search strings")

        search_table_name = (

            'search_table_' +
            str(uuid.uuid4()).replace('-', '_'))

        DBSession.execute(f'''

            create temporary table

            {search_table_name} (
              index INT PRIMARY KEY,
              search_string TEXT NOT NULL)

            on commit drop;

            ''')

        insert_sql_str_list = [

            f'insert into {search_table_name} values ',
            '(0, \'{}\')'.format(searchstrings[0].replace('\'', '\'\''))]

        for index, search_string in enumerate(searchstrings[1:], 1):

            insert_sql_str_list.append(
                ', ({}, \'{}\')'.format(index, search_string.replace('\'', '\'\'')))

        insert_sql_str_list.append(';')

        DBSession.execute(
            ''.join(insert_sql_str_list))

        locale_id = (
            info.context.get('locale_id'))

        row_list = (

            DBSession

                .execute(f'''

                    select
                    distinct on (S.index)
                    S.index,
                    A2.content

                    from
                    {search_table_name} S

                    left outer join
                    translationatom A1
                    on
                    A1.content = S.search_string

                    left outer join
                    translationgist G
                    on
                    A1.parent_client_id = G.client_id and
                    A1.parent_object_id = G.object_id

                    left outer join
                    translationatom A2
                    on
                    A2.parent_client_id = G.client_id and
                    A2.parent_object_id = G.object_id

                    where
                    G.type = 'Service' and
                    G.marked_for_deletion = false and
                    A1.locale_id = 2 and
                    A1.marked_for_deletion = false and
                    A2.locale_id = {locale_id} and
                    A2.marked_for_deletion = false

                    order by
                    S.index, G.client_id, G.object_id, A2.client_id, A2.object_id;

                    ''')

                .fetchall())

        result_list = [None] * len(searchstrings)

        for index, result in row_list:
            result_list[index] = result

        return result_list

    def resolve_userblob(self, info, id):
        return UserBlobs(id=id)

    def resolve_field(self, info, id):
        return Field(id=id)

    def resolve_lexicalentry(self, info, id):
        return LexicalEntry(id=id)

    def resolve_all_locales(self, info):
        response = list()
        locales = DBSession.query(dbLocale).all()
        for locale in locales:
            locale_json = dict()
            locale_json['shortcut'] = locale.shortcut
            locale_json['intl_name'] = locale.intl_name
            locale_json['created_at'] = locale.created_at
            locale_json['id'] = locale.id
            response.append(locale_json)
        return response


    def resolve_basic_search(self, info, searchstring, search_in_published, field_id=None, perspective_id=None, can_add_tags=None): #basic_search() function
        """
        query EntriesList {
            basic_search(searchstring: "следить", search_in_published: true) {
                id
                entities {
                     id
                     content
                }
            }
        }

        """

        if (not searchstring or
            len(searchstring) < 1):

            raise ResponseError(message="Bad string")

        field = None
        if field_id:
            field_client_id, field_object_id = field_id[0], field_id[1]
            field = DBSession.query(dbField).filter_by(client_id=field_client_id, object_id=field_object_id).first()

        client_id = info.context.get('client_id')

        group = DBSession.query(dbGroup).filter(dbGroup.subject_override == True).join(dbBaseGroup) \
            .filter(dbBaseGroup.subject == 'lexical_entries_and_entities', dbBaseGroup.action == 'view') \
            .join(dbUser, dbGroup.users).join(Client) \
            .filter(Client.id == client_id).first()

        # See get_hidden() in models.py.

        hidden_id = (

            DBSession

                .query(
                    dbTranslationGist.client_id,
                    dbTranslationGist.object_id)

                .join(dbTranslationAtom)

                .filter(
                    dbTranslationGist.type == 'Service',
                    dbTranslationAtom.content == 'Hidden',
                    dbTranslationAtom.locale_id == 2)

                .first())

        # NOTE: due to no-op nature of publishing entity checks (see note below), removing joins
        # with PublishingEntity.

        results_cursor = (

            DBSession

                .query(dbLexicalEntry)

                .join(dbEntity)
                .join(dbPerspective)
                .join(dbDictionary)

                .filter(
                    dbEntity.content.like('%' + searchstring + '%'),
                    dbEntity.marked_for_deletion == False,
                    dbLexicalEntry.marked_for_deletion == False,
                    dbPerspective.marked_for_deletion == False,
                    dbDictionary.marked_for_deletion == False,
                    or_(dbDictionary.state_translation_gist_client_id != hidden_id[0],
                        dbDictionary.state_translation_gist_object_id != hidden_id[1]),
                    or_(dbPerspective.state_translation_gist_client_id != hidden_id[0],
                        dbPerspective.state_translation_gist_object_id != hidden_id[1])))

        if perspective_id:

            results_cursor = (

                results_cursor.filter(
                    dbPerspective.client_id == perspective_id[0],
                    dbPerspective.object_id == perspective_id[1]))

        # NOTE:
        #
        # Well, intention is clear, but this does not work and is actually a no-op.
        #
        # To work, should be:
        #
        #   results_cursor = results_cursor.filter(...)
        #
        # So, broken or at least not working as intended for almost 2.5 years.
        #
        # But this GraphQL api is currently used in only a single place, searching words for grouping, so
        # apparently all right, no problem.
        #
        # Commenting out for optimization.

#       if search_in_published is not None:
#           results_cursor.filter(dbPublishingEntity.published == search_in_published)

#       results_cursor.filter(dbPublishingEntity.accepted == True)

        if not group:

            # We do not have a single group giving us all necessary permissions.
            #
            # So, we look in either published perspectives or perspectives we have nesessary permissions
            # for.

            db_published_gist = translation_gist_search('Published')

            published_id = (
                db_published_gist.client_id,
                db_published_gist.object_id)

            group_query = (

                DBSession

                    .query(dbBaseGroup.action)
                    .join(dbGroup)

                    .filter(
                        dbGroup.subject_client_id == dbPerspective.client_id,
                        dbGroup.subject_object_id == dbPerspective.object_id,
                        dbGroup.id == user_to_group_association.c.group_id,
                        user_to_group_association.c.user_id == Client.user_id,
                        Client.id == client_id,
                        dbBaseGroup.subject == 'lexical_entries_and_entities'))

            # Do we need both view and create permissions?

            if can_add_tags:

                group_query = (

                    group_query

                        .filter(or_(
                            dbBaseGroup.action == 'create',
                            dbBaseGroup.action == 'view'))

                        .group_by(dbBaseGroup.action)
                        .subquery())

                group_count_query = (

                    DBSession
                        .query(func.count(group_query.c.action))
                        .as_scalar())

                group_condition = (
                    group_count_query == 2)

            # Only view persmissions.

            else:

                group_query = (
                    group_query.filter(dbBaseGroup.action == 'view'))

                group_condition = group_query.exists()

            results_cursor = (

                results_cursor.filter(

                    or_(

                        and_(
                            dbPerspective.state_translation_gist_client_id == published_id[0],
                            dbPerspective.state_translation_gist_object_id == published_id[1]),

                        group_condition)))

        if field:
            results_cursor = results_cursor.join(dbPerspective.dictionaryperspectivetofield).filter(
                dbColumn.field == field)

        lexes = results_cursor.distinct().all()

        lexes_composite_list = [
            (lex.client_id, lex.object_id, lex.parent_client_id, lex.parent_object_id)
            for lex in lexes]

        entities = (

            dbLexicalEntry.graphene_track_multiple(
                lexes_composite_list,
                publish = search_in_published,
                accept = True,
                check_perspective = False))

        def graphene_entity(cur_entity, cur_publishing):
            ent = Entity(id = (cur_entity.client_id, cur_entity.object_id))
            ent.dbObject = cur_entity
            ent.publishingentity = cur_publishing
            return ent

        def graphene_obj(dbobj, cur_cls):
            obj = cur_cls(id=(dbobj.client_id, dbobj.object_id))
            obj.dbObject = dbobj
            return obj

        entities = [graphene_entity(entity[0], entity[1]) for entity in entities]
        lexical_entries = [graphene_obj(lex, LexicalEntry) for lex in lexes]
        return LexicalEntriesAndEntities(entities=entities, lexical_entries=lexical_entries)

    def resolve_advanced_lexicalentries(self, info, searchstrings, perspectives=None, adopted=None,
                                        adopted_type=None, with_etimology=None): #advanced_search() function

        """
        query EntriesList {
            advanced_lexicalentries(searchstrings: [{searchstring: "смотреть следить"}]) {
                id
                entities {
                     id
                     content
                }
            }
        }

        """
        request = info.context.get('request')

        if not perspectives:
            db_published_gist = translation_gist_search('Published')
            state_translation_gist_client_id = db_published_gist.client_id
            state_translation_gist_object_id = db_published_gist.object_id
            db_la_gist = translation_gist_search('Limited access')
            limited_client_id, limited_object_id = db_la_gist.client_id, db_la_gist.object_id

            perspectives = [(persp.client_id, persp.object_id) for persp in DBSession.query(dbPerspective).filter(
                dbPerspective.marked_for_deletion == False,
                or_(and_(dbPerspective.state_translation_gist_client_id == state_translation_gist_client_id,
                         dbPerspective.state_translation_gist_object_id == state_translation_gist_object_id),
                    and_(dbPerspective.state_translation_gist_client_id == limited_client_id,
                         dbPerspective.state_translation_gist_object_id == limited_object_id))).all()]

        def make_query(searchstring, perspectives):
            results_cursor = DBSession.query(dbLexicalEntry).join(dbEntity.parent) \
                .join(dbEntity.field).join(dbTranslationAtom,
                                         and_(dbField.translation_gist_client_id == dbTranslationAtom.parent_client_id,
                                              dbField.translation_gist_object_id == dbTranslationAtom.parent_object_id,
                                              dbField.marked_for_deletion == False)) \
                .distinct(dbEntity.parent_client_id, dbEntity.parent_object_id)
            if perspectives:
                results_cursor = results_cursor.filter(
                    tuple_(dbLexicalEntry.parent_client_id, dbLexicalEntry.parent_object_id).in_(perspectives))
            if not searchstring["searchstring"]:
                raise ResponseError(message="Error: bad argument 'searchstring'.")
            search_parts = searchstring["searchstring"].split()
            search_expression = dbEntity.content.like('%' + search_parts[0] + '%')
            to_do_or = searchstring.get('search_by_or', True)

            for part in search_parts[1:]:
                search_expression = or_(search_expression, dbEntity.content.like('%' + part + '%'))
            if 'entity_type' in searchstring and searchstring['entity_type']:
                search_expression = and_(search_expression, dbField.client_id == searchstring['entity_type'][0],
                                         dbField.object_id == searchstring['entity_type'][1])

            results_cursor = results_cursor.filter(search_expression)
            return results_cursor, to_do_or

        if not searchstrings[0]:
            raise ResponseError(message="Error: bad argument 'searchstrings'")

        results_cursor, to_do_or = make_query(searchstrings[0], perspectives)

        pre_results = set(results_cursor.all())
        if adopted:
            results_cursor = DBSession.query(dbLexicalEntry).join(dbEntity.parent).filter(
                dbEntity.content.like('%заим.%'))
            if adopted_type:
                results_cursor = results_cursor.join(dbEntity.field) \
                    .join(dbTranslationAtom,
                          and_(dbField.translation_gist_client_id == dbTranslationAtom.parent_client_id,
                               dbField.translation_gist_object_id == dbTranslationAtom.parent_object_id,
                               dbField.marked_for_deletion == False)) \
                    .filter(dbTranslationAtom.content == adopted_type,
                            dbTranslationAtom.locale_id == 2)
            pre_results = pre_results & set(results_cursor.all())
        if with_etimology:
            results_cursor = DBSession.query(dbLexicalEntry).join(dbEntity.parent).join(dbEntity.field) \
                .join(dbTranslationAtom,
                      and_(dbField.data_type_translation_gist_client_id == dbTranslationAtom.parent_client_id,
                           dbField.data_type_translation_gist_object_id == dbTranslationAtom.parent_object_id,
                           dbField.marked_for_deletion == False)) \
                .filter(dbTranslationAtom.content == 'Grouping Tag',
                        dbTranslationAtom.locale_id == 2)

        pre_results = pre_results & set(results_cursor.all())

        for search_string in searchstrings[1:]:
            results_cursor, to_do_or_new = make_query(search_string, perspectives)
            if to_do_or:
                pre_results = pre_results | set(results_cursor.all())
            else:
                pre_results = pre_results & set(results_cursor.all())
            to_do_or = to_do_or_new

        lexes_composite_list = [(lex.created_at,
                                 lex.client_id, lex.object_id, lex.parent_client_id, lex.parent_object_id,
                                 lex.marked_for_deletion, lex.additional_metadata,
                                 lex.additional_metadata.get('came_from')
                                 if lex.additional_metadata and 'came_from' in lex.additional_metadata else None)
                                for lex in pre_results]

        lexical_entries = dbLexicalEntry.track_multiple(lexes_composite_list, int(request.cookies.get('locale_id') or 2),
                                              publish=True, accept=True)

        lexical_entries_list = list()
        for entry in lexical_entries:
            entities = []
            for ent in entry['contains']:
                del ent["contains"]
                del ent["level"]
                del ent["accepted"]
                del ent["published"]
                if "link_client_id" in ent and "link_object_id" in ent:
                    ent["link_id"] = (ent["link_client_id"], ent["link_object_id"])
                else:
                    ent["link_id"] = None
                ent["field_id"] = (ent["field_client_id"], ent["field_object_id"])
                if "self_client_id" in ent and "self_object_id" in ent:
                    ent["self_id"] = (ent["self_client_id"], ent["self_object_id"])
                else:
                    ent["self_id"] = None
                if "content" not in ent:
                    ent["content"] = None
                if "additional_metadata" in ent:
                    ent["additional_metadata_string"] = ent["additional_metadata"]
                    del ent["additional_metadata"]
                if 'entity_type' in ent:
                    del ent['entity_type']

                gr_entity_object = Entity(id=[ent['client_id'],
                                              ent['object_id']],
                                          # link_id = (ent["link_client_id"], ent["link_object_id"]),
                                          parent_id=(ent["parent_client_id"], ent["parent_object_id"]),
                                          **ent  # all other args from sub_result
                                          )
                entities.append(gr_entity_object)

            del entry["published"]
            del entry["contains"]
            del entry["level"]
            gr_lexicalentry_object = LexicalEntry(id=[entry['client_id'],
                                                      entry['object_id']],
                                                  entities=entities, **entry)

            lexical_entries_list.append(gr_lexicalentry_object)
        return lexical_entries_list

    # @client_id_check()
    def resolve_user_blobs(self, info, data_type=None, is_global=None):
        allowed_global_types = ["sociolinguistics", "pdf"]
        client_id = info.context.get('client_id')
        client = DBSession.query(Client).filter_by(id=client_id).first()
        user_blobs = list()
        if not data_type and is_global:
            raise ResponseError("Error: cannot list globally without data_type")
        if data_type and is_global:
            if data_type in allowed_global_types:
                user_blobs = DBSession.query(dbUserBlobs).filter_by(marked_for_deletion=False, data_type=data_type).all()
            else:
                raise ResponseError(message="Error: you can not list that data type globally.")
        elif not client:
            raise ResponseError('not authenticated')
        if data_type:
            if not is_global:
                user_blobs = DBSession.query(dbUserBlobs).filter_by(marked_for_deletion=False, user_id=client.user_id, data_type=data_type).all()
        else:
            user_blobs = DBSession.query(dbUserBlobs).filter_by(marked_for_deletion=False, user_id=client.user_id).all()
        user_blobs_list = list()
        for db_blob in user_blobs:
            gql_blob = UserBlobs(id=[db_blob.client_id, db_blob.object_id])
            gql_blob.dbObject = db_blob
            user_blobs_list.append(gql_blob)
        return user_blobs_list

    def resolve_userrequest(self, info, id):
        """
        query myQuery {
          userrequest(id: 6) {
                id
           }
        }
        """
        return UserRequest(id=id)

    #@client_id_check()
    def resolve_userrequests(self, info):
        """
        query myQuery {
          userrequests {
                id
                sender_id
                type
           }
        }
        """
        client_id = info.context.get('client_id')

        client = DBSession.query(Client).filter_by(id=client_id).first()
        user = DBSession.query(dbUser).filter_by(id=client.user_id).first()
        if not user:
            raise ResponseError(message="This client id is orphaned. Try to logout and then login once more.")

        userrequests = DBSession.query(dbUserRequest).filter(dbUserRequest.recipient_id == user.id).order_by(
            dbUserRequest.created_at).all()
        userrequests_list = list()
        for db_userrequest in userrequests:
            gql_userrequest = UserRequest(id=db_userrequest.id)
            userrequests_list.append(gql_userrequest)
        return userrequests_list

    def resolve_all_basegroups(self, info):  # tested
        basegroups = list()
        for basegroup_object in DBSession.query(dbBaseGroup).all():
            basegroup = BaseGroup(id=basegroup_object.id)
            basegroup.dbObject = basegroup_object
            basegroups.append(basegroup)
        return basegroups

    def resolve_grant(self, info, id):
        return Grant(id=id)

    def resolve_grants(
        self,
        info,
        has_participant = None,
        participant_deleted = None,
        participant_published = None,
        participant_category = None):
        """
        query myQuery {
          grants {
                id
           }
        }
        """

        __debug_flag__ = False

        # Analyzing query.

        object_flag = False

        selection_dict = {

            'additional_metadata': (
                dbGrant.additional_metadata,),

            'begin': (
                dbGrant.begin,),

            'created_at': (
                dbGrant.created_at,),

            'end': (
                dbGrant.end,),

            'grant_number': (
                dbGrant.grant_number,),

            'grant_url': (
                dbGrant.grant_url,),

            'id': (
                dbGrant.id,),

            'issuer_translations': (
                dbGrant.issuer_translation_gist_client_id,
                dbGrant.issuer_translation_gist_object_id),

            'issuer_url': (
                dbGrant.issuer_url,),

            'owners': (
                dbGrant.owners,),

            'translations': (
                dbGrant.translation_gist_client_id,
                dbGrant.translation_gist_object_id,)}

        selection_set = set()
        selection_list = []

        owners_flag = False
        email_flag = False

        for field in info.field_asts:

            if field.name.value != 'grants':
                continue

            for subfield in field.selection_set.selections:

                name_str = subfield.name.value

                if name_str in selection_dict:

                    selection_set.add(name_str)
                    selection_list.extend(selection_dict[name_str])

                    if name_str == 'owners':

                        owners_flag = True

                        for owners_field in subfield.selection_set.selections:

                            if owners_field.name.value == 'email':

                                email_flag = True
                                break

                elif name_str != '__typename':

                    object_flag = True

        if object_flag:

            selection_list = [dbGrant]

        elif 'grant_number' not in selection_set:

            # For standard grant ordering.

            selection_set.add('grant_number')
            selection_list.extend(selection_dict['grant_number'])

        # If we are going to query both the usual and issuer translations, we'll have to use separate joins
        # to a CTE-based queries to avoid joining to actually a cross product of two translations.

        translations_flag = (
            'translations' in selection_set)

        issuer_translations_flag = (
            'issuer_translations' in selection_set)

        cte_flag = (
            translations_flag and
            issuer_translations_flag)

        grant_query = None
        grant_cte = None

        # No participant filtering, getting every grant.

        if has_participant is None:

            grant_query = (
                DBSession.query(*selection_list))

            grant_c = dbGrant

        # Simple participant count filter.

        elif (
            participant_deleted is None and
            participant_category is None and
            participant_published is None):

            participant_count = (

                func.jsonb_array_length(
                    dbGrant.additional_metadata['participant']))

            grant_query = (

                DBSession

                    .query(
                        *selection_list)

                    .filter(
                        participant_count > 0 if has_participant else
                        participant_count <= 0))

            grant_c = dbGrant

        # Additional conditions on participants, we'll have to check them through a join.
        #
        # We have to use raw SQL due to SQLAlchemy being bad with PostgreSQL's jsonb_to_recordset.

        else:

            dictionary_condition_list = []

            if participant_deleted is not None:

                dictionary_condition_list.append(
                    '\n and D.marked_for_deletion = true' if participant_deleted else
                    '\n and D.marked_for_deletion = false')

            if participant_category is not None:

                dictionary_condition_list.append(
                    f'\n and D.category = {participant_category}')

            if participant_published is not None:

                if not participant_published:
                    raise NotImplementedError

                dictionary_condition_list.append(f'''

                    and (
                      D.state_translation_gist_client_id,
                      D.state_translation_gist_object_id) in (

                      select
                        T.client_id,
                        T.object_id

                      from
                        translationgist T,
                        translationatom A

                      where
                        T.marked_for_deletion = false and
                        T.type = 'Service' and
                        A.parent_client_id = T.client_id and
                        A.parent_object_id = T.object_id and
                        A.locale_id = 2 and
                        A.marked_for_deletion = false and (
                          A.content = 'Published' or
                          A.content = 'Limited access'))

                    ''')

            dictionary_condition_str = (
                ''.join(dictionary_condition_list))

            if object_flag:

                selection_str = 'G.*'

            else:

                selection_str = (

                    ', '.join(
                        f'G.{selection.name}'
                        for selection in selection_list))

            sql_text = (

                sqlalchemy.text(f'''

                    select
                      {selection_str}

                    from
                      public.grant G

                    cross join
                      jsonb_to_recordset(G.additional_metadata -> 'participant')
                        P (client_id bigint, object_id bigint)

                    join
                      dictionary D

                    on
                      D.client_id = P.client_id and
                      D.object_id = P.object_id {dictionary_condition_str}

                    group by
                      G.id

                    '''))

            if object_flag:

                sql_text = (

                    aliased(
                        dbGrant,

                        sql_text

                            .columns(
                                **{column.name: column.type
                                    for column in dbGrant.__table__.c})

                            .alias(),

                        adapt_on_names = True))

                grant_c = (
                    sql_text)

            else:

                sql_text = (

                    sql_text

                        .columns(
                            *selection_list)

                        .alias())

                grant_c = (
                    sql_text.c)

            grant_query = (
                DBSession.query(sql_text))

            if not has_participant:
                raise NotImplementedError

        # Establishing a CTE if we'll need it.

        if cte_flag:

            grant_cte = grant_query.cte()
            grant_c = grant_cte.c

            if object_flag:

                grant_cte = (

                    aliased(
                        dbGrant,
                        grant_cte,
                        adapt_on_names = True))

                grant_c = grant_cte

            grant_query = (
                DBSession.query(grant_cte))

        # Getting translations through a join, if required.

        if translations_flag:

            if grant_cte is not None:

                translation_query = (

                    DBSession

                        .query(
                            grant_c.id)

                        .outerjoin(
                            dbTranslationAtom,

                            and_(
                                dbTranslationAtom.parent_client_id == grant_c.translation_gist_client_id,
                                dbTranslationAtom.parent_object_id == grant_c.translation_gist_object_id,
                                dbTranslationAtom.marked_for_deletion == False))

                        .add_columns(

                            func.jsonb_object_agg(
                                dbTranslationAtom.locale_id,
                                dbTranslationAtom.content)

                                .filter(dbTranslationAtom.locale_id != None)
                                .label('translations'))

                        .group_by(
                            grant_c.id)

                        .subquery())

                grant_query = (

                    grant_query

                        .join(
                            translation_query,
                            translation_query.c.id == grant_c.id)

                        .add_columns(
                            translation_query.c.translations))

            else:

                grant_query = (

                    grant_query

                        .outerjoin(
                            dbTranslationAtom,

                            and_(
                                dbTranslationAtom.parent_client_id == grant_c.translation_gist_client_id,
                                dbTranslationAtom.parent_object_id == grant_c.translation_gist_object_id,
                                dbTranslationAtom.marked_for_deletion == False))

                        .add_columns(

                            func.jsonb_object_agg(
                                dbTranslationAtom.locale_id,
                                dbTranslationAtom.content)

                                .filter(dbTranslationAtom.locale_id != None)
                                .label('translations'))

                        .group_by(
                            *([grant_c] if object_flag else grant_c)))

        # Getting issuer translations through a join, if required.

        if issuer_translations_flag:

            if grant_cte is not None:

                issuer_translation_query = (

                    DBSession

                        .query(
                            grant_c.id)

                        .outerjoin(
                            dbTranslationAtom,

                            and_(
                                dbTranslationAtom.parent_client_id == grant_c.issuer_translation_gist_client_id,
                                dbTranslationAtom.parent_object_id == grant_c.issuer_translation_gist_object_id,
                                dbTranslationAtom.marked_for_deletion == False))

                        .add_columns(

                            func.jsonb_object_agg(
                                dbTranslationAtom.locale_id,
                                dbTranslationAtom.content)

                                .filter(dbTranslationAtom.locale_id != None)
                                .label('issuer_translations'))

                        .group_by(
                            grant_c.id)

                        .subquery())

                grant_query = (

                    grant_query

                        .join(
                            issuer_translation_query,
                            issuer_translation_query.c.id == grant_c.id)

                        .add_columns(
                            issuer_translation_query.c.issuer_translations))

            else:

                grant_query = (

                    grant_query

                        .outerjoin(
                            dbTranslationAtom,

                            and_(
                                dbTranslationAtom.parent_client_id == grant_c.issuer_translation_gist_client_id,
                                dbTranslationAtom.parent_object_id == grant_c.issuer_translation_gist_object_id,
                                dbTranslationAtom.marked_for_deletion == False))

                        .add_columns(

                            func.jsonb_object_agg(
                                dbTranslationAtom.locale_id,
                                dbTranslationAtom.content)

                                .filter(dbTranslationAtom.locale_id != None)
                                .label('issuer_translations'))

                        .group_by(
                            *([grant_c] if object_flag else grant_c)))

        # Getting grant info, preparing it for GraphQL.

        grant_query = (

            grant_query
                .order_by(grant_c.grant_number))

        result_list = grant_query.all()

        if __debug_flag__:

            log.debug(
                '\n grant_query:\n ' +
                render_statement(grant_query.statement))

        gql_grant_list = []

        if object_flag:

            # We are getting full ORM dbGrant objects.

            attribute_set = selection_set.copy()

            attribute_set.discard('translations')
            attribute_set.discard('issuer_translations')

            for result in result_list:

                grant = (
                    result[0] if translations_flag or issuer_translations_flag else
                    result)

                gql_grant = (
                    Grant(id = grant.id))

                gql_grant.dbObject = grant

                if translations_flag:

                    translations = (
                        result.translations)

                    gql_grant.translations = (
                        translations if translations is not None else gql_none_value)

                if issuer_translations_flag:

                    translations = (
                        result.issuer_translations)

                    gql_grant.issuer_translations = (
                        translations if translations is not None else gql_none_value)

                for attribute in attribute_set:

                    value = getattr(grant, attribute)

                    if attribute == 'additional_metadata':

                        value = AdditionalMetadata.from_object(value)

                    elif attribute == 'begin' or attribute == 'end':

                        value = Grant.from_date(value)

                    elif attribute == 'created_at':

                        value = CreatedAt.from_timestamp(value)

                    setattr(
                        gql_grant,
                        attribute,
                        value)

                gql_grant_list.append(gql_grant)

        else:

            # We are getting attribute values as they are.

            for result in result_list:

                gql_grant = (
                    Grant(id = result.id))

                for selection in selection_set:

                    value = getattr(result, selection)

                    if selection == 'additional_metadata':

                        value = AdditionalMetadata.from_object(value)

                    elif selection == 'begin' or selection == 'end':

                        value = Grant.from_date(value)

                    elif selection == 'created_at':

                        value = CreatedAt.from_timestamp(value)

                    setattr(
                        gql_grant,
                        selection,
                        value if value is not None else gql_none_value)

                gql_grant_list.append(gql_grant)

        # Loading owners if required.

        if (gql_grant_list and
            owners_flag):

            owner_id_set = set()

            owner_id_set.update(

                *(gql_grant.owners
                    for gql_grant in gql_grant_list))

            user_query = (

                DBSession
                    .query(dbUser)

                    .filter(

                        dbUser.id.in_(
                            utils.values_query(
                                owner_id_set, models.SLBigInteger))))

            if email_flag:

                user_query = (

                    user_query.options(
                        joinedload(dbUser.email)))

            gql_user_dict = {}

            for user in user_query:

                user_id = user.id

                gql_user = User(id = user_id)
                gql_user.dbObject = user

                gql_user_dict[user_id] = gql_user

            for gql_grant in gql_grant_list:

                gql_grant.owners = [

                    gql_user_dict[owner_id]
                    for owner_id in gql_grant.owners]

        return gql_grant_list

    def resolve_phonology_tier_list(self, info, perspective_id):
        """
        query MyQuery {
          phonology_tier_list(perspective_id: [330, 4]) {
            tier_count
            total_count
          }
        }
        """

        answer = utils_phonology_tier_list(*perspective_id)
        return TierList(**answer)

    def resolve_phonology_skip_list(self, info, perspective_id):
        """
        query MyQuery {
          phonology_skip_list(perspective_id: [1251, 14]) {
            markup_count
            neighbour_list
            skip_list
            total_neighbour_count
            total_skip_count
          }
        }
        """

        answer = utils_phonology_skip_list(*perspective_id)
        return SkipList(**answer)

    def resolve_phonology_link_perspective_data(self, info, perspective_id, field_id_list):
        """
        query MyQuery {
          phonology_link_perspective_data(
            perspective_id: [657, 4],
            field_id_list: [[1, 213]])
          {
            perspective_id_list
          }
        }
        """

        answer = gql_phonology_link_perspective_data(perspective_id, field_id_list)
        return Link_Perspective_Data(**answer)

    # def resolve_convert_starling(self, info, starling_dictionaries):
    #     """
    #     query myQuery {
    #         convert_starling(parent_id:[1,1] blob_id:[1,1] translation_atoms:[], add_etymology:true , field_map:{min_created_at:1})
    #     }
    #     """
    #     cache_kwargs = info.context["request"].registry.settings["cache_kwargs"]
    #     sqlalchemy_url = info.context["request"].registry.settings["sqlalchemy.url"]
    #     task_names = []
    #     for st_dict in starling_dictionaries:
    #         # TODO: fix
    #         task_names.append(st_dict.get("translation_atoms")[0].get("content"))
    #     name = ",".join(task_names)
    #     user_id = Client.get_user_by_client_id(info.context["client_id"]).id
    #     task = TaskStatus(user_id, "Starling dictionary conversion", name, 10)
    #     starling_converter.convert(info, starling_dictionaries, cache_kwargs, sqlalchemy_url, task.key)
    #     return True

    def resolve_connected_words(self, info, id, field_id, mode=None):

        client_id = id[0]
        object_id = id[1]
        field_client_id = field_id[0]
        field_object_id = field_id[1]
        if mode == 'all':
            publish = None
            accept = True
        elif mode == 'published':
            publish = True
            accept = True
        elif mode == 'not_accepted':
            publish = None
            accept = False

        # NOTE: modes 'deleted' and 'all_with_deleted' are currently not implemented.

        elif mode == 'deleted':
            publish = None
            accept = None
        elif mode == 'all_with_deleted':
            publish = None
            accept = None

        else:
            raise ResponseError(message="mode: <all|published|not_accepted>")

        # Getting lexical entry group info.

        marked_for_deletion = (

            DBSession
                .query(dbLexicalEntry.marked_for_deletion)

                .filter_by(
                    client_id = client_id,
                    object_id = object_id)

                .scalar())

        if (marked_for_deletion is None or
            marked_for_deletion):

            raise ResponseError(message = 'No such lexical entry in the system')

        if publish is None and accept is None:

            sql_str = '''

                select * from linked_group_no_publishing(
                    :field_client_id,
                    :field_object_id,
                    :client_id,
                    :object_id)

                '''

        else:

            sql_str = '''

                select * from linked_group(
                    :field_client_id,
                    :field_object_id,
                    :client_id,
                    :object_id,
                    :publish,
                    :accept)

                '''

        entry_query = (

            DBSession
                .query(dbLexicalEntry)
                .filter(

                    tuple_(
                        dbLexicalEntry.client_id,
                        dbLexicalEntry.object_id)

                    .in_(sqlalchemy.text(sql_str)))

                .params({
                    'field_client_id': field_client_id,
                    'field_object_id': field_object_id,
                    'client_id': client_id,
                    'object_id': object_id,
                    'publish': publish,
                    'accept': accept}))

        lexes = entry_query.all()

        lexes_composite_list = [
            (entry.client_id, entry.object_id, entry.parent_client_id, entry.parent_object_id)
            for entry in lexes]

        entities = dbLexicalEntry.graphene_track_multiple(lexes_composite_list,
                                                   publish=publish, accept=accept)

        def graphene_entity(cur_entity, cur_publishing):
            ent = Entity(id = (cur_entity.client_id, cur_entity.object_id))
            ent.dbObject = cur_entity
            ent.publishingentity = cur_publishing
            return ent

        def graphene_obj(dbobj, cur_cls):
            obj = cur_cls(id=(dbobj.client_id, dbobj.object_id))
            obj.dbObject = dbobj
            return obj

        entities = [graphene_entity(entity[0], entity[1]) for entity in entities]
        lexical_entries = [graphene_obj(lex, LexicalEntry) for lex in lexes]
        return LexicalEntriesAndEntities(entities=entities, lexical_entries=lexical_entries)



    def resolve_eaf_wordlist(self, info, id):
        # TODO: delete
        import tempfile
        import pympi
        import sys
        import os
        import random
        import string
        import requests
        from sqlalchemy.exc import IntegrityError
        from lingvodoc.exceptions import CommonException
        from lingvodoc.scripts.convert_rules import praat_to_elan

        # TODO: permission check
        """
        query myQuery {
            convert_markup(id: [742, 5494] )
        }
        """
        client_id = info.context.get('client_id')
        client = DBSession.query(Client).filter_by(id=client_id).first()
        user = DBSession.query(dbUser).filter_by(id=client.user_id).first()

        try:
            # out_type = req['out_type']
            client_id, object_id = id

            entity = DBSession.query(dbEntity).filter_by(client_id=client_id, object_id=object_id).first()
            if not entity:
                raise KeyError("No such file")
            resp = requests.get(entity.content)
            if not resp:
                raise ResponseError("Cannot access file")
            content = resp.content
            try:
                n = 10
                filename = (
                    time.asctime(time.gmtime()) + ''.join(
                        random.SystemRandom().choice(string.ascii_uppercase + string.digits)
                        for c in range(n)))
                # extension = os.path.splitext(blob.content)[1]
                f = open(filename, 'wb')
            except Exception as e:
                return ResponseError(message=str(e))
            try:
                f.write(content)
                f.close()
                if os.path.getsize(filename) / (10 * 1024 * 1024.0) < 1:
                    if 'data_type' in entity.additional_metadata :
                        if 'praat' in entity.additional_metadata['data_type']:
                            textgrid_obj = pympi.TextGrid(file_path=filename)
                            eaf_obj = textgrid_obj.to_eaf()
                            word_list = eaf_words(eaf_obj)
                            return word_list
                            #elan = to_eaf("-",eafobj)

                        elif 'elan' in entity.additional_metadata['data_type']:
                            #with open(filename, 'r') as f:
                            #    return f.read()
                            eaf_obj = pympi.Eaf(file_path=filename)
                            word_list = eaf_words(eaf_obj)
                            return word_list
                        else:
                            raise KeyError("Not allowed convert option")
                        raise KeyError('File too big')
                    raise KeyError("Not allowed convert option")
                raise KeyError('File too big')
            except Exception as e:
                raise ResponseError(message=e)
            finally:
                os.remove(filename)
                pass
        except KeyError as e:
            raise ResponseError(message=str(e))

        except IntegrityError as e:
            raise ResponseError(message=str(e))

        except CommonException as e:
            raise ResponseError(message=str(e))


    def resolve_convert_markup(self, info, id):

        # TODO: permission check
        """
        query myQuery {
            convert_markup(id: [742, 5494] )
        }
        """
        # client_id = info.context.get('client_id')
        # client = DBSession.query(Client).filter_by(id=client_id).first()
        # user = DBSession.query(dbUser).filter_by(id=client.user_id).first()
        client_id, object_id = id
        entity = DBSession.query(dbEntity).filter_by(client_id=client_id, object_id=object_id).first()

        if not entity:
            return ResponseError(f'No entity {client_id} / {object_id}.')

        try:

            storage = (
                info.context.request.registry.settings['storage'])

            with storage_file(
                storage, entity.content) as content_stream:

                content = content_stream.read()

        except:
            return ResponseError(f'Cannot access file \'{entity.content}\'.')

        return tgt_to_eaf(content, entity.additional_metadata)


    def resolve_parser_results(self, info, entity_id):
        entity_client_id, entity_object_id = entity_id
        results = DBSession.query(dbParserResult).filter_by(entity_client_id=entity_client_id,
                                                            entity_object_id=entity_object_id,
							    marked_for_deletion=False
                                                            ).all()
        return_list = list()
        for result in results:
            new_parser_result = ParserResult(id = [result.client_id, result.object_id])
            new_parser_result.dbObject = result
            return_list.append(new_parser_result)
        return return_list

    def resolve_parser_result(self, info, id):
        client_id, object_id = id
        result = DBSession.query(dbParserResult).filter_by(client_id=client_id,
                                                            object_id=object_id,
                                                            ).first()
        if not result or result.marked_for_deletion:
            return None
        parser_result = ParserResult(id=[result.client_id, result.object_id])
        parser_result.dbObject = result
        return parser_result

    def resolve_parsers(self, info):
        parsers = DBSession.query(dbParser).all()
        return_list = list()
        for parser in parsers:
            element = Parser(id=[parser.client_id, parser.object_id])
            element.dbObject = parser
            return_list.append(element)
        return return_list



class PerspectivesAndFields(graphene.InputObjectType):
    perspective_id = LingvodocID()
    field_id = LingvodocID()

class StarlingEtymologyObject(graphene.InputObjectType):
    starling_perspective_id = LingvodocID()
    perspectives_and_fields = graphene.List(PerspectivesAndFields)

class StarlingEtymology(graphene.Mutation):

    class Arguments:
        etymology_field_id = LingvodocID()
        complex_list = graphene.List(StarlingEtymologyObject)

    triumph = graphene.Boolean()

    @staticmethod
    @client_id_check()
    def mutate(root, info, **args):
        request = info.context.request

        client_id = info.context.get('client_id')

        client = DBSession.query(Client).filter_by(id=client_id).first()
        user = DBSession.query(dbUser).filter_by(id=client.user_id).first()
        if not user:
            raise ResponseError(message="This client id is orphaned. Try to logout and then login once more.")
        if user.id != 1:
            raise ResponseError(message="not admin")

        etymology_field_id = args['etymology_field_id']
        etymology_field = DBSession.query(dbField).filter_by(client_id=etymology_field_id[0],
                                                             object_id=etymology_field_id[1],
                                                             marked_for_deletion=False).first()
        if not etymology_field:
            raise ResponseError(message='no such field')
        timestamp = (
            time.asctime(time.gmtime()) + ''.join(
                random.SystemRandom().choice(string.ascii_uppercase + string.digits) for c in range(10)))
        for complex_element in args['complex_list']:
            starling_ids = complex_element['starling_perspective_id']
            starling_perspective = DBSession.query(dbPerspective).filter_by(client_id=starling_ids[0],
                                                                                      object_id=starling_ids[1],
                                                                                      marked_for_deletion=False).first()
            if not starling_perspective:
                raise ResponseError(message='no such starling perspective')
            for persp_and_field in complex_element['perspectives_and_fields']:
                persp_ids = persp_and_field['perspective_id']
                cur_persp = DBSession.query(dbPerspective).filter_by(client_id=persp_ids[0],
                                                                               object_id=persp_ids[1],
                                                                               marked_for_deletion=False).first()
                field_id = persp_and_field['field_id']
                field = DBSession.query(dbField).filter_by(client_id=field_id[0],
                                                           object_id=field_id[1],
                                                           marked_for_deletion=False).first()
                if not cur_persp:
                    raise ResponseError(message='no such perspective')

                StarlingEntity = aliased(dbEntity)
                StarlingLexicalEntry = aliased(dbLexicalEntry)
                result = DBSession.query(dbEntity, dbLexicalEntry, StarlingEntity, StarlingLexicalEntry).filter(
                    dbEntity.parent_client_id == dbLexicalEntry.client_id,
                    dbEntity.parent_object_id == dbLexicalEntry.object_id,
                    StarlingEntity.parent_client_id == StarlingLexicalEntry.client_id,
                    StarlingEntity.parent_object_id == StarlingLexicalEntry.object_id,
                    dbLexicalEntry.parent_client_id == cur_persp.client_id,
                    dbLexicalEntry.parent_object_id == cur_persp.object_id,
                    StarlingLexicalEntry.parent_client_id == starling_perspective.client_id,
                    StarlingLexicalEntry.parent_object_id == starling_perspective.object_id,
                    dbEntity.content == StarlingEntity.content,
                    dbEntity.field == field).all()
                for sub_res in result:
                    first_lex = sub_res[1]
                    second_lex = sub_res[3]
                    tag = sub_res[0].content + timestamp
                    first_tag = dbEntity(client_id = client.id, parent=first_lex, content=tag, field=etymology_field)
                    first_tag.publishingentity.accepted = True
                    first_tag.publishingentity.published = True
                    # DBSession.add(first_tag)
                    second_tag = dbEntity(client_id = client.id, parent=second_lex, content=tag, field=etymology_field)
                    second_tag.publishingentity.accepted = True
                    second_tag.publishingentity.published = True
                    # DBSession.add(second_tag)
                    CACHE.set(objects = [first_tag, second_tag], DBSession=DBSession)

        return StarlingEtymology(triumph=True)


class PhonemicAnalysis(graphene.Mutation):

    class Arguments:

        perspective_id = LingvodocID(required = True)
        transcription_field_id = LingvodocID(required = True)
        translation_field_id = LingvodocID(required = True)
        wrap_flag = graphene.Boolean()

        debug_flag = graphene.Boolean()
        intermediate_flag = graphene.Boolean()

    triumph = graphene.Boolean()
    entity_count = graphene.Int()
    result = graphene.String()

    intermediate_url_list = graphene.List(graphene.String)

    @staticmethod
    def mutate(self, info, **args):
        """
        mutation PhonemicAnalysis {
          phonemic_analysis(
            perspective_id: [70, 5],
            transcription_field_id: [66, 8],
            translation_field_id: [66, 10])
          {
            triumph
            entity_count
            result
          }
        }
        """

        perspective_cid, perspective_oid = args['perspective_id']

        transcription_field_cid, transcription_field_oid = args['transcription_field_id']
        translation_field_cid, translation_field_oid = args['translation_field_id']

        wrap_flag = args.get('wrap_flag', False)

        locale_id = info.context.get('locale_id') or 2

        __debug_flag__ = args.get('debug_flag', False)
        __intermediate_flag__ = args.get('intermediate_flag', False)

        try:

            perspective = DBSession.query(dbPerspective).filter_by(
                client_id = perspective_cid, object_id = perspective_oid).first()

            perspective_name = perspective.get_translation(locale_id)
            dictionary_name = perspective.parent.get_translation(locale_id)

            transcription_rules = (
                '' if not perspective.additional_metadata else
                    perspective.additional_metadata.get('transcription_rules', ''))

            # Showing phonemic analysis info, checking phonemic analysis library presence.

            log.debug(
                '\nphonemic_analysis {0}/{1}:'
                '\n  dictionary: {2}'
                '\n  perspective: {3}'
                '\n  transcription rules: {4}'
                '\n  transcription field: {5}/{6}'
                '\n  translation field: {7}/{8}'
                '\n  wrap_flag: {9}'
                '\n  __debug_flag__: {10}'
                '\n  __intermediate_flag__: {11}'
                '\n  locale_id: {12}'
                '\n  phonemic_analysis_f: {13}'.format(
                    perspective_cid, perspective_oid,
                    repr(dictionary_name.strip()),
                    repr(perspective_name.strip()),
                    repr(transcription_rules),
                    transcription_field_cid, transcription_field_oid,
                    translation_field_cid, translation_field_oid,
                    wrap_flag,
                    __debug_flag__,
                    __intermediate_flag__,
                    locale_id,
                    repr(phonemic_analysis_f)))

            if phonemic_analysis_f is None:

                return ResponseError(message =
                    'Analysis library is absent, please contact system administrator.')

            # Query for non-deleted, published and accepted entities of the specified perspective with the
            # specified field.

            dbTranslation = aliased(dbEntity, name = 'Translation')
            dbPublishingTranslation = aliased(dbPublishingEntity, name = 'PublishingTranslation')

            data_query = (

                DBSession.query(dbEntity).filter(
                    dbLexicalEntry.parent_client_id == perspective_cid,
                    dbLexicalEntry.parent_object_id == perspective_oid,
                    dbLexicalEntry.marked_for_deletion == False,
                    dbEntity.parent_client_id == dbLexicalEntry.client_id,
                    dbEntity.parent_object_id == dbLexicalEntry.object_id,
                    dbEntity.field_client_id == transcription_field_cid,
                    dbEntity.field_object_id == transcription_field_oid,
                    dbEntity.marked_for_deletion == False,
                    dbPublishingEntity.client_id == dbEntity.client_id,
                    dbPublishingEntity.object_id == dbEntity.object_id,
                    dbPublishingEntity.published == True,
                    dbPublishingEntity.accepted == True)

                .outerjoin(dbTranslation, and_(
                    dbTranslation.parent_client_id == dbEntity.parent_client_id,
                    dbTranslation.parent_object_id == dbEntity.parent_object_id,
                    dbTranslation.field_client_id == translation_field_cid,
                    dbTranslation.field_object_id == translation_field_oid,
                    dbTranslation.marked_for_deletion == False))

                .outerjoin(dbPublishingTranslation, and_(
                    dbPublishingTranslation.client_id == dbTranslation.client_id,
                    dbPublishingTranslation.object_id == dbTranslation.object_id,
                    dbPublishingTranslation.published == True,
                    dbPublishingTranslation.accepted == True))

                .add_columns(
                    func.array_agg(dbTranslation.content))

                .group_by(dbEntity))

            # Counting text entities we have. If we haven't got any, we return empty result.

            total_count = data_query.count()

            log.debug(
                'phonemic_analysis {0}/{1}: {2} transcription entities'.format(
                    perspective_cid, perspective_oid,
                    total_count,
                    transcription_field_cid, transcription_field_oid))

            if total_count <= 0:

                return PhonemicAnalysis(
                    triumph = True,
                    entity_count = total_count,
                    result = u'')

            # Otherwise we are going to perform phonemic analysis.

            data_list = [

                (entity.content,
                    translation_list[0] if translation_list else '')

                for entity, translation_list in data_query.all()
                if len(entity.content) > 0]

            if len(data_list) <= 0:

                return PhonemicAnalysis(
                    triumph = True,
                    entity_count = total_count,
                    result = u'No transcription entities with non-empty contents.')

            # Preparing analysis input.

            input = (

                '{0} - {1}\0{2}\0'.format(
                    dictionary_name,
                    perspective_name,
                    transcription_rules) +

                ''.join(

                    '{0}\0{1}\0'.format(
                        transcription, translation)

                    for transcription, translation in data_list))

            log.debug(
                'phonemic_analysis {0}/{1}: '
                'transcription field {2}/{3}, translation field {4}/{5}'
                '\ndata_list:\n{6}'
                '\ninput ({7} rows):\n{8}'.format(
                    perspective_cid, perspective_oid,
                    transcription_field_cid, transcription_field_oid,
                    translation_field_cid, translation_field_oid,
                    pprint.pformat(data_list, width = 108),
                    len(data_list) + 1,
                    pprint.pformat([input[i : i + 256]
                        for i in range(0, len(input), 256)], width = 144)))

            # Saving to a file, if required.

            intermediate_url_list = []

            if __debug_flag__ or __intermediate_flag__:

                perspective = DBSession.query(dbPerspective).filter_by(
                    client_id = perspective_cid, object_id = perspective_oid).first()

                perspective_name = (
                    perspective.get_translation(2).strip())

                if len(perspective_name) > 48:
                    perspective_name = perspective_name[:48] + '...'

                dictionary_name = (
                    perspective.parent.get_translation(2).strip())

                if len(dictionary_name) > 48:
                    dictionary_name = dictionary_name[:48] + '...'

                phonemic_name_str = (
                    'phonemic {0} {1} {2}'.format(
                    dictionary_name,
                    perspective_name,
                    len(data_list) + 1))

                # Initializing file storage directory, if required.

                if __intermediate_flag__:

                    storage = info.context.request.registry.settings['storage']
                    cur_time = time.time()

                    storage_dir = os.path.join(
                        storage['path'], 'phonemic', str(cur_time))

                for extension, encoding in (
                    ('utf8', 'utf-8'), ('utf16', 'utf-16')):

                    input_file_name = (

                        pathvalidate.sanitize_filename(
                            'input {0}.{1}'.format(
                                phonemic_name_str, extension)))

                    # Saving to the working directory...

                    if __debug_flag__:

                        with open(input_file_name, 'wb') as input_file:
                            input_file.write(input.encode(encoding))

                    # ...and / or to the file storage.

                    if __intermediate_flag__:

                        input_path = os.path.join(
                            storage_dir, input_file_name)

                        os.makedirs(
                            os.path.dirname(input_path),
                            exist_ok = True)

                        with open(input_path, 'wb') as input_file:
                            input_file.write(input.encode(encoding))

                        input_url = ''.join([
                            storage['prefix'],
                            storage['static_route'],
                            'phonemic', '/',
                            str(cur_time), '/',
                            input_file_name])

                        intermediate_url_list.append(input_url)

            # Calling analysis library, starting with getting required output buffer size and continuing
            # with analysis proper.

            output_buffer_size = phonemic_analysis_f(
                None, len(data_list) + 1, None, 0)

            if output_buffer_size <= 0:
                return ResponseError(message = 'Invalid output buffer size')

            log.debug(
                'phonemic_analysis {0}/{1}: output buffer size {2}'.format(
                perspective_cid, perspective_oid,
                output_buffer_size))

            input_buffer = ctypes.create_unicode_buffer(input)
            output_buffer = ctypes.create_unicode_buffer(output_buffer_size + 256)

            result = phonemic_analysis_f(
                input_buffer, len(data_list) + 1, output_buffer, 0)

            # If we don't have a good result, we return an error.

            log.debug(
                'phonemic_analysis {0}/{1}: '
                'transcription field {2}/{3}, translation field {4}/{5}: result {6}'.format(
                    perspective_cid, perspective_oid,
                    transcription_field_cid, transcription_field_oid,
                    translation_field_cid, translation_field_oid,
                    result))

            if result != 1:
                return ResponseError(message =
                    'Phonemic analysis library call error {0}'.format(result))

            output = output_buffer.value

            log.debug(
                'phonemic_analysis {0}/{1}: '
                'transcription field {2}/{3}, translation field {4}/{5}:'
                '\noutput:\n{6}'.format(
                    perspective_cid, perspective_oid,
                    transcription_field_cid, transcription_field_oid,
                    translation_field_cid, translation_field_oid,
                    repr(output)))

            # Reflowing output, if required.

            final_output = output

            if wrap_flag:

                line_list = output.split('\r\n')

                text_wrapper = textwrap.TextWrapper(
                    width = 108, tabsize = 4)

                reflow_list = []

                for line in line_list:
                    reflow_list.extend(text_wrapper.wrap(line))

                wrapped_output = '\n'.join(reflow_list)

                log.debug(
                    'phonemic_analysis {0}/{1}: '
                    'transcription field {2}/{3}, translation field {4}/{5}:'
                    '\nwrapped output:\n{6}'.format(
                        perspective_cid, perspective_oid,
                        transcription_field_cid, transcription_field_oid,
                        translation_field_cid, translation_field_oid,
                        wrapped_output))

                final_output = wrapped_output

            # Returning result.

            return PhonemicAnalysis(

                triumph = True,
                entity_count = total_count,
                result = final_output,

                intermediate_url_list =
                    intermediate_url_list if __intermediate_flag__ else None)

        except Exception as exception:

            traceback_string = ''.join(traceback.format_exception(
                exception, exception, exception.__traceback__))[:-1]

            log.warning('phonemic_analysis: exception')
            log.warning(traceback_string)

            return ResponseError(message =
                'Exception:\n' + traceback_string)


@celery.task
def async_cognate_analysis(
    language_str,
    source_perspective_id,
    base_language_id,
    base_language_name,
    group_field_id,
    perspective_info_list,
    multi_list,
    multi_name_list,
    mode,
    distance_flag,
    reference_perspective_id,
    figure_flag,
    distance_vowel_flag,
    distance_consonant_flag,
    match_translations_value,
    only_orphans_flag,
    locale_id,
    storage,
    task_key,
    cache_kwargs,
    sqlalchemy_url,
    __debug_flag__,
    __intermediate_flag__):
    """
    Sets up and launches cognate analysis in asynchronous mode.
    """

    # NOTE: copied from phonology.
    #
    # This is a no-op with current settings, we use it to enable logging inside celery tasks, because
    # somehow this does it, and otherwise we couldn't set it up.

    logging.debug('async_cognate_analysis')

    # Ok, and now we go on with task execution.

    engine = create_engine(sqlalchemy_url)
    DBSession.configure(bind = engine)
    initialize_cache(cache_kwargs)

    task_status = TaskStatus.get_from_cache(task_key)

    with transaction.manager:

        try:
            CognateAnalysis.perform_cognate_analysis(
                language_str,
                source_perspective_id,
                base_language_id,
                base_language_name,
                group_field_id,
                perspective_info_list,
                multi_list,
                multi_name_list,
                mode,
                None,
                None,
                None,
                None,
                None,
                match_translations_value,
                only_orphans_flag,
                locale_id,
                storage,
                task_status,
                __debug_flag__,
                __intermediate_flag__)

        # Some unknown external exception.

        except Exception as exception:

            traceback_string = ''.join(traceback.format_exception(
                exception, exception, exception.__traceback__))[:-1]

            log.warning(
                'cognate_analysis {0}: exception'.format(
                language_str))

            log.warning(traceback_string)

            if task_status is not None:

                task_status.set(5, 100,
                    'Finished (ERROR), exception:\n' + traceback_string)


class CognateAnalysis(graphene.Mutation):

    class Arguments:

        source_perspective_id = LingvodocID(required = True)
        base_language_id = LingvodocID(required = True)

        group_field_id = LingvodocID(required = True)
        perspective_info_list = graphene.List(graphene.List(LingvodocID), required = True)
        multi_list = graphene.List(ObjectVal)

        mode = graphene.String()

        distance_flag = graphene.Boolean()
        reference_perspective_id = LingvodocID()

        figure_flag = graphene.Boolean()
        distance_vowel_flag = graphene.Boolean()
        distance_consonant_flag = graphene.Boolean()

        match_translations_value = graphene.Int()
        only_orphans_flag = graphene.Boolean()

        debug_flag = graphene.Boolean()
        intermediate_flag = graphene.Boolean()

        synchronous = graphene.Boolean()

    triumph = graphene.Boolean()

    dictionary_count = graphene.Int()
    group_count = graphene.Int()
    not_enough_count = graphene.Int()
    transcription_count = graphene.Int()
    translation_count = graphene.Int()

    result = graphene.String()
    xlsx_url = graphene.String()
    distance_list = graphene.Field(ObjectVal)
    figure_url = graphene.String()

    minimum_spanning_tree = graphene.List(graphene.List(graphene.Int))
    embedding_2d = graphene.List(graphene.List(graphene.Float))
    embedding_3d = graphene.List(graphene.List(graphene.Float))
    perspective_name_list = graphene.List(graphene.String)

    suggestion_list = graphene.List(ObjectVal)
    suggestion_field_id = LingvodocID()

    intermediate_url_list = graphene.List(graphene.String)

    @staticmethod
    def tag_data_std(
        entry_already_set,
        group_list,
        perspective_id,
        field_client_id,
        field_object_id):
        """
        Gets lexical entry grouping data using current standard methods, computes elapsed time.
        """

        start_time = time.time()

        tag_data_list = (DBSession.query(
            dbLexicalEntry, func.count('*'))

            .filter(
                dbLexicalEntry.parent_client_id == perspective_id[0],
                dbLexicalEntry.parent_object_id == perspective_id[1],
                dbLexicalEntry.marked_for_deletion == False,
                dbEntity.parent_client_id == dbLexicalEntry.client_id,
                dbEntity.parent_object_id == dbLexicalEntry.object_id,
                dbEntity.field_client_id == field_client_id,
                dbEntity.field_object_id == field_object_id,
                dbEntity.marked_for_deletion == False,
                dbPublishingEntity.client_id == dbEntity.client_id,
                dbPublishingEntity.object_id == dbEntity.object_id,
                dbPublishingEntity.published == True,
                dbPublishingEntity.accepted == True)

            .group_by(
                dbLexicalEntry.client_id,
                dbLexicalEntry.object_id))

        # Processing each lexical entry with at least one tag.

        for entry, count in tag_data_list.all():

            if (entry.client_id, entry.object_id) in entry_already_set:
                continue

            tag_set = find_all_tags(
                entry, field_client_id, field_object_id, True, True)

            entry_list = find_lexical_entries_by_tags(
                tag_set, field_client_id, field_object_id, True, True)

            entry_id_set = set(
                (tag_entry.client_id, tag_entry.object_id)
                for tag_entry in entry_list)

            entry_already_set.update(entry_id_set)
            group_list.append(entry_id_set)

        return time.time() - start_time

    @staticmethod
    def find_group(
        entry_client_id,
        entry_object_id,
        field_client_id,
        field_object_id):
        """
        Retrieves all lexical entries grouped with a given id-specified entry.
        """

        entry_id_set = set((
            (entry_client_id, entry_object_id),))

        tag_query = (

            DBSession.query(
                dbEntity.content)

            .filter(
                dbEntity.parent_client_id == entry_client_id,
                dbEntity.parent_object_id == entry_object_id,
                dbEntity.field_client_id == field_client_id,
                dbEntity.field_object_id == field_object_id,
                dbEntity.marked_for_deletion == False,
                dbPublishingEntity.client_id == dbEntity.client_id,
                dbPublishingEntity.object_id == dbEntity.object_id,
                dbPublishingEntity.published == True,
                dbPublishingEntity.accepted == True))

        tag_list = list(tag_query.all())
        tag_set = set(tag_list)

        find_group_by_tags(
            DBSession,
            entry_id_set, tag_set, tag_list,
            field_client_id, field_object_id,
            True)

        return entry_id_set

    @staticmethod
    def tag_data_optimized(
        entry_already_set,
        group_list,
        perspective_id,
        field_client_id,
        field_object_id):
        """
        Gets lexical entry grouping data using (hopefully) optimized version of the current standard
        methods, computes elapsed time.
        """

        start_time = time.time()

        tag_data_list = (

            DBSession.query(
                dbLexicalEntry.client_id,
                dbLexicalEntry.object_id,
                func.count('*'))

            .filter(
                dbLexicalEntry.parent_client_id == perspective_id[0],
                dbLexicalEntry.parent_object_id == perspective_id[1],
                dbLexicalEntry.marked_for_deletion == False,
                dbEntity.parent_client_id == dbLexicalEntry.client_id,
                dbEntity.parent_object_id == dbLexicalEntry.object_id,
                dbEntity.field_client_id == field_client_id,
                dbEntity.field_object_id == field_object_id,
                dbEntity.marked_for_deletion == False,
                dbPublishingEntity.client_id == dbEntity.client_id,
                dbPublishingEntity.object_id == dbEntity.object_id,
                dbPublishingEntity.published == True,
                dbPublishingEntity.accepted == True)

            .group_by(
                dbLexicalEntry.client_id,
                dbLexicalEntry.object_id))

        # Processing each lexical entry with at least one tag.

        for entry_client_id, entry_object_id, count in tag_data_list.all():

            if (entry_client_id, entry_object_id) in entry_already_set:
                continue

            entry_id_set = CognateAnalysis.find_group(
                entry_client_id, entry_object_id,
                field_client_id, field_object_id)

            entry_already_set.update(entry_id_set)
            group_list.append(entry_id_set)

        return time.time() - start_time

    @staticmethod
    def tag_data_aggregated(
        perspective_info_list,
        tag_field_id,
        statistics_flag = False,
        optimize_flag = False):
        """
        Gets lexical entry grouping data using aggregated retrieval, computes elapsed time.
        """

        start_time = time.time()

        entry_id_dict = collections.defaultdict(set)
        tag_dict = collections.defaultdict(set)

        tag_set = set()

        # All tags for tagged lexical entries in specified perspectives.

        for perspective_id, transcription_field_id, translation_field_id in perspective_info_list:

            tag_query = (

                DBSession.query(
                    dbEntity.parent_client_id,
                    dbEntity.parent_object_id,
                    dbEntity.content)

                .filter(
                    dbLexicalEntry.parent_client_id == perspective_id[0],
                    dbLexicalEntry.parent_object_id == perspective_id[1],
                    dbLexicalEntry.marked_for_deletion == False,
                    dbEntity.parent_client_id == dbLexicalEntry.client_id,
                    dbEntity.parent_object_id == dbLexicalEntry.object_id,
                    dbEntity.field_client_id == tag_field_id[0],
                    dbEntity.field_object_id == tag_field_id[1],
                    dbEntity.marked_for_deletion == False,
                    dbPublishingEntity.client_id == dbEntity.client_id,
                    dbPublishingEntity.object_id == dbEntity.object_id,
                    dbPublishingEntity.published == True,
                    dbPublishingEntity.accepted == True))

            for entry_client_id, entry_object_id, tag in tag_query.all():

                entry_id_dict[(entry_client_id, entry_object_id)].add(tag)
                tag_set.add(tag)

        tag_list = tag_set

        # While we have tags we don't have all lexical entries for,
        # we get all entries of these tags...

        while tag_list:

            entry_id_query = (

                DBSession.query(
                    dbEntity.parent_client_id,
                    dbEntity.parent_object_id,
                    dbEntity.content)

                .filter(
                    dbLexicalEntry.marked_for_deletion == False,
                    dbEntity.parent_client_id == dbLexicalEntry.client_id,
                    dbEntity.parent_object_id == dbLexicalEntry.object_id,
                    dbEntity.field_client_id == tag_field_id[0],
                    dbEntity.field_object_id == tag_field_id[1],
                    dbEntity.marked_for_deletion == False,
                    dbEntity.content.in_(tag_list),
                    dbPublishingEntity.client_id == dbEntity.client_id,
                    dbPublishingEntity.object_id == dbEntity.object_id,
                    dbPublishingEntity.published == True,
                    dbPublishingEntity.accepted == True))

            entry_id_list = []

            for entry_client_id, entry_object_id, tag in entry_id_query.all():

                entry_id = entry_client_id, entry_object_id
                tag_dict[tag].add(entry_id)

                if entry_id not in entry_id_dict:
                    entry_id_list.append(entry_id)

            # And then get all tags for entries we haven't already done it for.

            tag_list = []

            entry_id_list.sort()
            log.debug('len(entry_id_list): {0}'.format(len(entry_id_list)))

            for i in range((len(entry_id_list) + 16383) // 16384):

                tag_query = (

                    DBSession.query(
                        dbEntity.parent_client_id,
                        dbEntity.parent_object_id,
                        dbEntity.content)

                    # We have to split entries into parts due to the danger of stack overflow in Postgres.

                    .filter(
                        tuple_(dbEntity.parent_client_id, dbEntity.parent_object_id)
                            .in_(entry_id_list[i * 16384 : (i + 1) * 16384]),
                        dbEntity.field_client_id == tag_field_id[0],
                        dbEntity.field_object_id == tag_field_id[1],
                        dbEntity.marked_for_deletion == False,
                        dbPublishingEntity.client_id == dbEntity.client_id,
                        dbPublishingEntity.object_id == dbEntity.object_id,
                        dbPublishingEntity.published == True,
                        dbPublishingEntity.accepted == True))

                for entry_client_id, entry_object_id, tag in tag_query.all():

                    entry_id = entry_client_id, entry_object_id
                    entry_id_dict[entry_id].add(tag)

                    if tag not in tag_dict:
                        tag_list.append(tag)

        # And now grouping lexical entries by tags.

        entry_already_set = set()
        group_list = []

        for entry_id, entry_tag_set in entry_id_dict.items():

            if entry_id in entry_already_set:
                continue

            entry_already_set.add(entry_id)

            group_entry_id_set = set((entry_id,))
            group_tag_set = set(entry_tag_set)

            current_tag_list = entry_tag_set

            # Recursively gathering grouped entries through their tags.

            while current_tag_list:

                tag_list = []

                for current_tag in current_tag_list:
                    for current_entry_id in tag_dict[current_tag]:

                        if current_entry_id in group_entry_id_set:
                            continue

                        group_entry_id_set.add(current_entry_id)

                        for tag in entry_id_dict[current_entry_id]:
                            if tag not in group_tag_set:

                                group_tag_set.add(tag)
                                tag_list.append(tag)

                current_tag_list = tag_list

            # Saving entry group data.

            entry_already_set.update(group_entry_id_set)
            group_list.append(group_entry_id_set)

        # Computing tag statistics for each entry group, if required.

        if statistics_flag:

            redundant_group_count = 0
            redundant_tag_count = 0
            redundant_entity_count = 0

            redundant_group_tag_count = 0
            redundant_group_entry_count = 0
            redundant_group_entity_count = 0

            delta_entity_count = 0

            # Processing each entry group.

            for entry_id_set in group_list:

                tag_set = set.union(
                    *(entry_id_dict[entry_id]
                        for entry_id in entry_id_set))

                tag_list = list(sorted(
                    (tag, len(tag_dict[tag]))
                        for tag in tag_set))

                partial_tag_set = set((tag_list[0][0],))
                partial_entry_id_set = set(tag_dict[tag_list[0][0]])

                index = 1

                # How many partial tag groups we need to cover the entry group completely?

                while (
                    len(partial_entry_id_set) < len(entry_id_set) and
                    index < len(tag_list)):

                    partial_tag_set.add(tag_list[index][0])
                    partial_entry_id_set.update(tag_dict[tag_list[index][0]])

                    index += 1

                group_entity_count = (

                    sum(len(tag_dict[tag])
                        for tag in tag_set))

                # Checking if we have some redundant partial tag groups.

                if index < len(tag_list):

                    redundant_group_count += 1
                    redundant_tag_count += len(tag_list) - index

                    for tag, count in tag_list[index:]:
                        redundant_entity_count += count

                    redundant_group_tag_count += len(tag_list)
                    redundant_group_entry_count += len(entry_id_set)

                    redundant_group_entity_count += group_entity_count

                delta_entity_count += group_entity_count - len(entry_id_set)

            # Showing gathered statistics.

            log.debug(
                '\ngroup_count: {0}/{1}, with {2} tags, {3} entries, {4} entities'
                '\ntag_count: {5}/{6}, {5}/{7} among redundant groups'
                '\nentity_count: {8}/{9}, {8}/{10} among redundant groups'
                '\ndelta(entity): {11}, minimum(entity): {12}'.format(
                    redundant_group_count, len(group_list),
                    redundant_group_tag_count,
                    redundant_group_entry_count,
                    redundant_group_entity_count,
                    redundant_tag_count, len(tag_dict),
                    redundant_group_tag_count,
                    redundant_entity_count,
                    sum(len(entry_id_set)
                        for entry_id_set in tag_dict.values()),
                    redundant_group_entity_count,
                    delta_entity_count,
                    len(entry_id_dict)))

        # If required, optimizing lexical entry groups by ensuring that each group has exactly one tag.

        if optimize_flag:

            redundant_tag_list = []

            for entry_id_set in group_list:

                tag_set = set.union(
                    *(entry_id_dict[entry_id]
                        for entry_id in entry_id_set))

                count, tag = max(
                    (len(tag_dict[tag]), tag)
                    for tag in tag_set)

                # Creating tag entities we need to link current group via selected tag.

                for entry_id in entry_id_set - tag_dict[tag]:

                    tag_entity = dbEntity(
                        client_id = entry_id[0],
                        parent_client_id = entry_id[0],
                        parent_object_id = entry_id[1],
                        field_client_id = tag_field_id[0],
                        field_object_id = tag_field_id[1],
                        content = tag)

                    tag_entity.publishingentity.published = True
                    tag_entity.publishingentity.accepted = True

                tag_set.remove(tag)
                redundant_tag_list.extend(tag_set)

            # Removing tag entities of the redundant tags.

            entity_id_query = (

                dbEntity.__table__
                    .update()
                    .where(and_(
                        dbEntity.marked_for_deletion == False,
                        dbEntity.content.in_(redundant_tag_list),
                        dbPublishingEntity.client_id == dbEntity.client_id,
                        dbPublishingEntity.object_id == dbEntity.object_id,
                        dbPublishingEntity.published == True,
                        dbPublishingEntity.accepted == True))
                    .values(marked_for_deletion = True)
                    .returning(dbEntity.client_id, dbEntity.object_id))

            entity_id_list = list(
                entity_id_query.execute())

            log.debug('entity_id_list: {0} entities\n{1}'.format(
                len(entity_id_list), pprint.pformat(entity_id_list)))

            # Optimization is not fully implemented at the moment.

            raise NotImplementedError

        return entry_already_set, group_list, time.time() - start_time

    @staticmethod
    def tag_data_plpgsql(
        perspective_info_list,
        tag_field_id,
        statistics_flag = False,
        optimize_flag = False):
        """
        Gets lexical entry grouping data using stored PL/pgSQL functions, computes elapsed time.
        """

        __debug_flag__ = False

        start_time = time.time()

        # Getting lexical entries with tag data of the specified tag field from all perspectives.

        perspective_id_list = [
            perspective_id
            for perspective_id, _, _ in perspective_info_list]

        entry_id_query = (

            DBSession.query(
                dbLexicalEntry.client_id,
                dbLexicalEntry.object_id)

            .filter(

                tuple_(
                    dbLexicalEntry.parent_client_id,
                    dbLexicalEntry.parent_object_id)

                    .in_(
                        ids_to_id_query(
                            perspective_id_list)),

                dbLexicalEntry.marked_for_deletion == False,
                dbEntity.parent_client_id == dbLexicalEntry.client_id,
                dbEntity.parent_object_id == dbLexicalEntry.object_id,
                dbEntity.field_client_id == tag_field_id[0],
                dbEntity.field_object_id == tag_field_id[1],
                dbEntity.marked_for_deletion == False,
                dbPublishingEntity.client_id == dbEntity.client_id,
                dbPublishingEntity.object_id == dbEntity.object_id,
                dbPublishingEntity.published == True,
                dbPublishingEntity.accepted == True)

            .group_by(
                dbLexicalEntry.client_id,
                dbLexicalEntry.object_id))

        entry_id_list = (
            entry_id_query.all())

        if __debug_flag__:

            log.debug(
                '\nentry_id_query:\n' +
                str(entry_id_query.statement.compile(compile_kwargs = {'literal_binds': True})))

            log.debug(
                f'len(entry_id_list): {len(entry_id_list)}')

        # Grouping lexical entries using stored PL/pgSQL function.

        entry_already_set = set()
        group_list = []

        sql_str = '''
            select * from linked_group(
                :field_client_id,
                :field_object_id,
                :client_id,
                :object_id)'''

        for entry_id in entry_id_list:

            if entry_id in entry_already_set:
                continue

            row_list = (

                DBSession.execute(sql_str, {
                    'field_client_id': tag_field_id[0],
                    'field_object_id': tag_field_id[1],
                    'client_id': entry_id[0],
                    'object_id': entry_id[1]})

                .fetchall())

            entry_id_set = set(
                map(tuple, row_list))

            entry_already_set.update(entry_id_set)
            group_list.append(entry_id_set)

        return entry_already_set, group_list, time.time() - start_time

    @staticmethod
    def export_xlsx(
        language_str,
        mode,
        output_str,
        d_output_str,
        perspective_count,
        __debug_flag__ = False,
        cognate_name_str = None):
        """
        Parses results of the cognate analysis and exports them as an XLSX file.
        """

        workbook_stream = io.BytesIO()

        workbook = xlsxwriter.Workbook(
            workbook_stream, {'in_memory': True})

        worksheet_results = (
            workbook.add_worksheet(
                utils.sanitize_worksheet_name('Results')))

        # 20% background yellow, 10% background gray.

        format_yellow = (
            workbook.add_format({'bg_color': '#ffffcc'}))

        format_gray = (
            workbook.add_format({'bg_color': '#e6e6e6'}))

        index = output_str.find('\0')
        size_list = list(map(int, output_str[:index].split(',')))

        log.debug(
            'cognate_analysis {0}: result table size {1}'.format(
            language_str,
            size_list))

        max_width = 0
        row_count = 0

        re_series = r'\s*(\[\S+\]|\?|0)(\s*—\s*(\[\S+\]|\?|0))+\s*'
        re_item_list = r'\s*(\[\S+\]|\?|0)\s*—(\s*—\s*(\[\S+\]|\?|0)\s*—)+\s*'

        def export_table(table_index, table_str, n_col, n_row, source_str):
            """
            Parses from the binary output and exports to the XLSX workbook a table with specified width and
            height.
            """

            nonlocal index

            nonlocal max_width
            nonlocal row_count

            if n_col > max_width:

                max_width = n_col
                worksheet_results.set_column(0, max_width - 1, 16)

            row_list = []

            for i in range(n_row):

                value_list = []

                # Another row of analysis result values.

                for j in range(n_col):

                    index_next = source_str.find('\0', index + 1)
                    value = source_str[index + 1 : index_next]

                    value_list.append(value)
                    index = index_next

                split_list_list = [
                    value.split('|') for value in value_list]

                item_list_count = max(map(len, split_list_list))

                # Checking if we need color formatting.

                cell_format = None

                if (re.match(re_series, value_list[0]) is not None or
                    re.match(re_item_list, '—'.join(value_list))):

                    cell_format = format_yellow

                emphasize_flag_list = [
                    value.startswith('(') and value.endswith(')')
                    for value in value_list[::2]]

                # Some values may actually be sequences, so we check and process them if they are.

                for i in range(item_list_count):

                    item_list = [
                        split_list[i] if i < len(split_list) else ''
                        for split_list in split_list_list]

                    row_list.append(item_list)

                    for j, (x_script, x_lat, emphasize_flag) in (
                        enumerate(zip(item_list[::2], item_list[1::2], emphasize_flag_list))):

                        worksheet_results.write_row(
                            row_count,
                            j * 2,
                            [x_script, x_lat],
                            format_gray
                                if emphasize_flag and (x_script + x_lat).strip() else
                                cell_format)

                    row_count += 1

                # Going on another row of analysis results.

                if source_str[index + 1] != '\0':
                    raise NotImplementedError

                index += 1

            log.debug(
                'cognate_analysis {0}: {1} table {2}:\n{3}'.format(
                language_str,
                table_str, table_index,
                pprint.pformat(row_list, width = 144)))

            # Returning table data.

            return row_list

        # Getting analysis result info, exporting it to the XLSX workbook.

        for table_index, howmany in enumerate(range(len(size_list) // 2)):

            n_col, n_row = size_list[howmany * 2 : howmany * 2 + 2]
            export_table(table_index, 'result', n_col, n_row, output_str)

        # And now we parse formant plot data, if we have any.

        if len(output_str) > index + 1:

            if output_str[index + 1] != '\0':
                raise NotImplementedError

            index += 1
            index_next = output_str.find('\0', index + 1)

            size_list = list(map(int,
                output_str[index + 1 : index_next].split(',')))

            index = index_next

            log.debug(
                'cognate_analysis {0}: plot table size {1}'.format(
                language_str,
                size_list))

            # Getting plot info, exporting it to the XLSX workbook, generating plots.

            worksheet_table_2d = (
                workbook.add_worksheet(
                    utils.sanitize_worksheet_name('F-table')))

            worksheet_chart = (
                workbook.add_worksheet(
                    utils.sanitize_worksheet_name('F-chart')))

            table_2d_row_index = 0
            chart_2d_count = 0

            for table_index, howmany in enumerate(range(len(size_list) // 2)):

                n_col, n_row = size_list[howmany * 2 : howmany * 2 + 2]

                row_list = export_table(
                    table_index, 'plot', n_col, n_row, output_str)

                plot_title = row_list[0][0]

                if not plot_title:
                    continue

                # Getting formant series info.

                series_title_list = []
                series_data_list = []

                for row in row_list[1:]:

                    if not row[0]:
                        continue

                    elif not row[1]:

                        series_title_list.append(row[0])
                        series_data_list.append([])

                    else:

                        series_data_list[-1].append(
                            tuple(map(float, row[1:])))

                log.debug(
                    'cognate_analysis {0}: plot data {1}:\n{2}\n{3}\n{4}'.format(
                    language_str,
                    table_index,
                    repr(plot_title),
                    pprint.pformat(series_title_list, width = 144),
                    pprint.pformat(series_data_list, width = 144)))

                # Proceeding with plot generation only if we have enough data.

                if sum(map(len, series_data_list)) <= 1:
                    continue

                chart_data_2d_list = []

                min_2d_f1, max_2d_f1 = None, None
                min_2d_f2, max_2d_f2 = None, None

                # Generating plot data.

                for series_index, (series_title, series_data) in enumerate(
                    zip(series_title_list, series_data_list)):

                    f_2d_list = list(map(
                        lambda f_tuple: numpy.array(f_tuple[:2]), series_data))

                    f_3d_list = list(map(
                        numpy.array, series_data))

                    if len(f_2d_list) <= 0:
                        continue

                    (filtered_2d_list, outlier_2d_list, mean_2d, ellipse_list,
                        filtered_3d_list, outlier_3d_list, mean_3d, sigma_2d, inverse_3d) = (

                        phonology.chart_data(
                            [(f_2d, ('', '')) for f_2d in f_2d_list],
                            [(f_3d, ('', '')) for f_3d in f_3d_list]))

                    chart_data_2d_list.append((
                        len(filtered_2d_list), len(f_2d_list), series_title,
                        filtered_2d_list, outlier_2d_list, mean_2d, ellipse_list))

                    # Updating F1/F2 maximum/minimum info.

                    f1_list, f2_list = zip(*(x[0] for x in filtered_2d_list))

                    min_f1_list, max_f1_list = min(f1_list), max(f1_list)
                    min_f2_list, max_f2_list = min(f2_list), max(f2_list)

                    if min_2d_f1 == None or min_f1_list < min_2d_f1:
                        min_2d_f1 = min_f1_list

                    if max_2d_f1 == None or max_f1_list > max_2d_f1:
                        max_2d_f1 = max_f1_list

                    if min_2d_f2 == None or min_f2_list < min_2d_f2:
                        min_2d_f2 = min_f2_list

                    if max_2d_f2 == None or max_f2_list > max_2d_f2:
                        max_2d_f2 = max_f2_list

                # Compiling info of the formant scatter chart data series.

                chart_dict_list, table_2d_row_index = (

                    phonology.chart_definition_list(
                        chart_data_2d_list, worksheet_table_2d,
                        min_2d_f1, max_2d_f1, min_2d_f2, max_2d_f2,
                        row_index = table_2d_row_index))

                if chart_dict_list:

                    # Generating the chart, if we have any data.

                    chart = workbook.add_chart({'type': 'scatter'})

                    chart.set_title({
                        'name': plot_title})

                    chart.set_x_axis({
                        'major_gridlines': {'visible': True},
                        'name': 'F2 (Hz)',
                        'reverse': True})

                    chart.set_y_axis({
                        'major_gridlines': {'visible': True},
                        'name': 'F1 (Hz)',
                        'reverse': True})

                    chart.set_legend({
                        'position': 'top'})

                    for chart_dict in chart_dict_list:
                        chart.add_series(chart_dict)

                    chart.set_style(11)
                    chart.set_size({'width': 1024, 'height': 768})

                    worksheet_chart.insert_chart(
                        'A{0}'.format(chart_2d_count * 40 + 1), chart)

                    chart_2d_count += 1

        # If we have distance matrix data, we also parse and export it.

        matrix_info_list = None

        if d_output_str != None:

            index = d_output_str.find('\0')
            size_list = list(map(int, d_output_str[:index].split(',')))

            log.debug(
                'cognate_analysis {0}: distance result table size {1}'.format(
                language_str,
                size_list))

            matrix_info_list = []

            # Parsing and exporting each distance matrix.

            for table_index, howmany in enumerate(range(len(size_list) // 2)):

                n_col, n_row = size_list[howmany * 2 : howmany * 2 + 2]

                row_list = export_table(
                    table_index, 'distance result', n_col, n_row, d_output_str)

                matrix_title = row_list[0][0]

                if not matrix_title:
                    continue

                # Getting distance matrix info.

                matrix_header_list = []
                matrix_data_list = []

                for row in row_list[1:]:

                    if row[0]:

                        matrix_header_list.append(row[0])
                        matrix_data_list.append(row[1 : 1 + perspective_count])

                # Getting distance matrix array, checking if we need to filter its parts.

                matrix_header_array = numpy.array(matrix_header_list)

                matrix_data_array = numpy.array([
                    tuple(map(float, value_list))
                    for value_list in matrix_data_list])

                where = matrix_data_array[:,0] >= 0

                if not all(where):

                    matrix_header_array = matrix_header_array[where]
                    matrix_data_array = matrix_data_array[where, :][:, where]

                    # If we do, we also export filtered version.

                    worksheet_results.write(
                        'A{0}'.format(row_count + 1), 'Filtered:')

                    row_count += 2

                    worksheet_results.write_row(
                        'A{0}'.format(row_count + 1),
                        [''] + list(matrix_header_array))

                    row_count += 1

                    for i, header in enumerate(matrix_header_array):

                        worksheet_results.write_row(
                            'A{0}'.format(row_count + 1),
                            [header] + [round(value) for value in matrix_data_array[i]])

                        row_count += 1

                # Showing info of the matrix we've got.

                matrix_info_list.append((
                    matrix_title,
                    matrix_header_list,
                    matrix_data_list,
                    matrix_header_array,
                    matrix_data_array))

                log.debug(
                    '\ncognate_analysis {0}:'
                    '\ndistance data {1}:'
                    '\n{2}\n{3}\n{4}'
                    '\nmatrix_header_array:\n{5}'
                    '\nmatrix_data_array:\n{6}'.format(
                    language_str,
                    table_index,
                    repr(matrix_title),
                    pprint.pformat(matrix_header_list, width = 144),
                    pprint.pformat(matrix_data_list, width = 144),
                    matrix_header_array,
                    matrix_data_array))

        workbook.close()

        # Saving resulting Excel workbook for debug purposes, if required.

        if __debug_flag__:

            xlsx_file_name = (
                cognate_name_str + '.xlsx')

            workbook_stream.seek(0)

            with open(xlsx_file_name, 'wb') as xlsx_file:
                shutil.copyfileobj(workbook_stream, xlsx_file)

        # Returning exported XLSX data as a binary stream and any parsed distance matrix info.

        return workbook_stream, matrix_info_list

    @staticmethod
    def parse_suggestions(
        language_str,
        output_str,
        perspective_count,
        perspective_source_index,
        entry_id_dict,
        __debug_flag__ = False,
        cognate_name_str = None,
        group_field_id = None):
        """
        Parses cognate suggestions.
        """

        index = -1
        row_list = []

        table_count = 0

        # Parsing result table rows.

        while index < len(output_str):

            value_list = []

            # Getting required number of values.

            for i in range(
                perspective_count * 2):

                index_next = (
                    output_str.find('\0', index + 1))

                # No more values, meaning we are at the end of the table.

                if index_next == -1:
                    break

                value = (
                    output_str[index + 1 : index_next])

                value_list.append(value)
                index = index_next

            if index_next == -1:
                break

            # Another row of analysis result values.

            log.debug(
                '\n{0} / {1}: {2}'.format(
                    len(row_list),
                    index,
                    repr(value_list)))

            row_list.append(
                value_list)

            if output_str[index + 1] != '\0':
                raise NotImplementedError

            index += 1

        # Parsing cognate suggestions table.

        row_index = 2
        suggestion_list = []

        header_str = row_list[0][0]

        begin_str = 'Предложения для '
        end_str = ': '

        while row_index < len(row_list):

            word_str = row_list[row_index][0]

            assert word_str.startswith(begin_str)

            if word_str.endswith('НЕТ'):

                row_index += 1
                continue

            assert word_str.endswith(end_str)

            word = (
                word_str[
                    len(begin_str) : -len(end_str)])

            word_entry_id = (

                entry_id_dict[(
                    perspective_source_index,
                    word)])

            # Parsing suggestions for the word.

            single_list = []
            group_list = []

            row_index += 1

            if row_list[row_index][perspective_source_index * 2] == header_str:
                row_index += 1

            value_list = row_list[row_index]

            while not (
                row_index >= len(row_list) or
                value_list[0].startswith(begin_str)):

                # Existing groups.

                if value_list[0] == 'Уже имеющиеся ряды: ':

                    row_index += 1
                    value_list = row_list[row_index]

                    while not (
                        value_list[0].startswith(begin_str) or
                        value_list[0] == 'Слова-сироты: '):

                        word_list = []

                        for i in range(perspective_count):

                            if value_list[i * 2] or value_list[i * 2 + 1]:

                                word_list.append(
                                    (i, value_list[i * 2 : i * 2 + 2]))

                        if word_list:
                            group_list.append(word_list)

                        row_index += 1

                        if row_index >= len(row_list):
                            break

                        value_list = row_list[row_index]

                # Single words.

                elif value_list[0] == 'Слова-сироты: ':

                    row_index += 1

                    if row_list[row_index][perspective_source_index * 2] == header_str:
                        row_index += 1

                    value_list = row_list[row_index]

                    while not (
                        value_list[0].startswith(begin_str) or
                        value_list[0] == 'Уже имеющиеся ряды: '):

                        for i in range(perspective_count):

                            if value_list[i * 2] or value_list[i * 2 + 1]:

                                single_list.append(
                                    (i, value_list[i * 2 : i * 2 + 2]))

                        row_index += 1

                        if row_index >= len(row_list):
                            break

                        value_list = row_list[row_index]

                # Something unexpected?

                else:

                    log.debug(value_list)
                    raise NotImplementedError

            # Getting lexical entry identifiers, if required.

            raw_list = single_list

            def f(index, tt_tuple):
                """
                Gets entry id by its perspective index, translations and transcriptions.
                """

                transcription_str, translation_str = tt_tuple

                return (

                    entry_id_dict[(
                        index,
                        transcription_str + (
                            ' ' + translation_str if translation_str else ''))])

            single_list = [
                (index, tt_tuple, f(index, tt_tuple))
                for index, tt_tuple in raw_list]

            # For lexical entry groups we get id of just the first entry.

            raw_list = group_list
            group_list = []

            word_group = None

            for word_list in raw_list:

                entry_id_list = list(f(*w) for w in word_list)

                if word_entry_id in entry_id_list:

                    if word_group is not None:
                        raise NotImplementedError

                    word_group = (word_list, entry_id_list[0])

                else:

                    group_list.append((
                        word_list, entry_id_list[0]))

            # Showing what we've got, saving suggestion if it is non-trivial.

            log.debug('\n' +
                pprint.pformat(
                    (word, word_entry_id, word_group, single_list, group_list),
                    width = 192))

            if single_list or group_list:

                suggestion_list.append((
                    perspective_source_index,
                    word,
                    word_entry_id,
                    word_group,
                    single_list,
                    group_list))

        # Maybe we need to gather suggestions info for debugging?

        if group_field_id is not None:

            data_list = []

            for index, word, word_entry_id, word_group, single_list, group_list in suggestion_list:

                entry_id_list = (
                    [word_entry_id] +
                    [single_info[-1] for single_info in single_list] +
                    [group_info[-1] for group_info in group_list])

                data_list.append(
                    (group_field_id, entry_id_list))

            log.debug(
                '\ndebug data list:\n{}'.format(
                    pprint.pformat(data_list, width = 192)))

        # Showing and returning what we've got.

        log.debug(
            '\nsuggestion_list (length {0}):\n{1}'.format(
                len(suggestion_list),
                pprint.pformat(
                    suggestion_list, width = 192)))

        return suggestion_list

    @staticmethod
    def acoustic_data(
        base_language_id,
        sound_entity_id,
        sound_url,
        markup_entity_id,
        markup_url,
        storage,
        __debug_flag__ = False):
        """
        Extracts acoustic data from a pair of sound recording and its markup, using cache in a manner
        compatible with phonological analysis.
        """

        log_str = (
            'cognate_analysis {0}/{1}: sound entity {2}/{3}, markup entity {4}/{5}'.format(
                base_language_id[0], base_language_id[1],
                sound_entity_id[0], sound_entity_id[1],
                markup_entity_id[0], markup_entity_id[1]))

        textgrid_result_list = (

            process_sound_markup(
                log_str,
                sound_entity_id,
                sound_url,
                markup_entity_id,
                markup_url,
                storage,
                __debug_flag__))

        if textgrid_result_list is None:
            return None

        # Extracting info of the first vowel from the first vowel-containing tier of the sound/markup
        # analysis results.

        def f():

            for tier_index, tier_name, tier_result_list in textgrid_result_list:

                if (tier_result_list == 'no_vowel' or
                    tier_result_list == 'no_vowel_selected'):
                    continue

                for tier_result in tier_result_list:

                    for (interval_str, interval_r_length, i_list, f_list,
                        sign_longest, sign_highest, source_index) in tier_result.interval_data_list:

                        interval_str_list = interval_str.split()

                        return [interval_str.split()[0], interval_str.split()[-3]] + f_list[:3]

        # Showing what we've finally got and returning it.

        result = f()

        log.debug(
            '{0}: {1}'.format(
            log_str, result))

        return result

    @staticmethod
    def graph_2d_embedding(d_ij, verbose = False):
        """
        Computes 2d embedding of a graph specified by non-negative simmetric distance matrix via stress
        minimization.

        Stress is based on relative strain for non-zero distances and absolute strain for zero distances.

        Let S_ij be source distances, D_ij be 2d distances, then stress is

          Sum[D_ij^2] for S_ij == 0 +
          Sum[D_ij^2 / S_ij^2 + S_ij^2 / D_ij^2] for S_ij > 0.

        Given D_ij^2 = (x_i - x_j)^2 + (y_i - y_j)^2, xy-gradient used for minimization can be computed
        using following:

          d[D_ij^2, x_i] = 2 (x_i - x_j)
          d[D_ij^2, x_j] = -2 (x_i - x_j)
          d[D_ij^2, y_i] = 2 (y_i - y_j)
          d[D_ij^2, y_j] = -2 (y_i - y_j)

        Obviously, d[D_ij^2 / S_ij^2, x_i] = 2 (x_i - x_j) / S_ij^2, and so on.

        And, with checking via WolframAlpha,

          d[S_ij^2 / D_ij^2, x_i] = -2 S_ij^2 (x_i - x_j) / D_ij^4, and so on.

        """

        N = numpy.size(d_ij, 0)

        def f(xy):
            """
            Computes stress given xy-coordinates.
            """

            x = xy[:N]
            y = xy[N:]

            result = 0.0

            for i in range(1, N):
                for j in range(i):

                    dr2 = (x[i] - x[j]) ** 2 + (y[i] - y[j]) ** 2

                    if d_ij[i,j] <= 0:
                        result += 4 * dr2

                    else:
                        d2_ij = d_ij[i,j] ** 2
                        result += dr2 / d2_ij + d2_ij / dr2

            return result

        def df(xy):
            """
            Computes gradient at the given xy-coordinates.
            """

            x = xy[:N]
            y = xy[N:]

            df_x = numpy.zeros(N)
            df_y = numpy.zeros(N)

            for i in range(1, N):
                for j in range(i):

                    dx = x[i] - x[j]
                    dy = y[i] - y[j]

                    dr2 = dx ** 2 + dy ** 2

                    if d_ij[i,j] <= 0:

                        df_x[i] += 4 * dx
                        df_x[j] -= 4 * dx

                        df_y[i] += 4 * dy
                        df_y[j] -= 4 * dy

                    else:

                        d2_ij = d_ij[i,j] ** 2
                        factor = (1 / d2_ij - d2_ij / dr2 ** 2)

                        df_x[i] += dx * factor
                        df_x[j] -= dx * factor

                        df_y[i] += dy * factor
                        df_y[j] -= dy * factor

            return numpy.concatenate((df_x, df_y))

        iter_count = 0

        def f_callback(xy):
            """
            Shows minimization progress, if enabled.
            """

            nonlocal iter_count

            log.debug(
                '\niteration {0}:\nxy:\n{1}\nf:\n{2}\ndf:\n{3}'.format(
                iter_count, xy, f(xy), df(xy)))

            iter_count += 1

        # Performing minization, returning minimization results.
        #
        # To get deterministic results we use the source distance matrix to seed the initial pseudo-random
        # coordinates we start optimization from.

        rng = (

            numpy.random.Generator(
                numpy.random.PCG64(

                    tuple(
                        hash(value)
                        for value in d_ij.flat))))

        result = (

            scipy.optimize.minimize(f,
                rng.random(N * 2),
                jac = df,
                callback = f_callback if verbose else None,
                options = {'disp': verbose}))

        result_x = numpy.stack((result.x[:N], result.x[N:])).T

        return result_x, f(result.x)

    @staticmethod
    def graph_3d_embedding(d_ij, verbose = False):
        """
        Computes 3d embedding of a graph specified by non-negative simmetric distance matrix via stress
        minimization.

        The same as with 2d embedding, see graph_2d_embedding.
        """

        N = numpy.size(d_ij, 0)
        N2 = N * 2

        def f(xyz):
            """
            Computes stress given xyz-coordinates.
            """

            x = xyz[:N]
            y = xyz[N:N2]
            z = xyz[N2:]

            result = 0.0

            for i in range(1, N):
                for j in range(i):

                    dr2 = (x[i] - x[j]) ** 2 + (y[i] - y[j]) ** 2 + (z[i] - z[j]) ** 2

                    if d_ij[i,j] <= 0:
                        result += 4 * dr2

                    else:
                        d2_ij = d_ij[i,j] ** 2
                        result += dr2 / d2_ij + d2_ij / dr2

            return result

        def df(xyz):
            """
            Computes gradient at a given xyz-coordinates.
            """

            x = xyz[:N]
            y = xyz[N:N2]
            z = xyz[N2:]

            df_x = numpy.zeros(N)
            df_y = numpy.zeros(N)
            df_z = numpy.zeros(N)

            for i in range(1, N):
                for j in range(i):

                    dx = x[i] - x[j]
                    dy = y[i] - y[j]
                    dz = z[i] - z[j]

                    dr2 = dx ** 2 + dy ** 2 + dz ** 2

                    if d_ij[i,j] <= 0:

                        df_x[i] += 4 * dx
                        df_x[j] -= 4 * dx

                        df_y[i] += 4 * dy
                        df_y[j] -= 4 * dy

                        df_z[i] += 4 * dz
                        df_z[j] -= 4 * dz

                    else:

                        d2_ij = d_ij[i,j] ** 2
                        factor = (1 / d2_ij - d2_ij / dr2 ** 2)

                        df_x[i] += dx * factor
                        df_x[j] -= dx * factor

                        df_y[i] += dy * factor
                        df_y[j] -= dy * factor

                        df_z[i] += dz * factor
                        df_z[j] -= dz * factor

            return numpy.concatenate((df_x, df_y, df_z))

        iter_count = 0

        def f_callback(xyz):
            """
            Shows minimization progress, if enabled.
            """

            nonlocal iter_count

            log.debug(
                '\niteration {0}:\nxyz:\n{1}\nf:\n{2}\ndf:\n{3}'.format(
                iter_count,
                numpy.stack((xyz[:N], xyz[N:N2], xyz[N2:])).T,
                f(xyz),
                df(xyz)))

            iter_count += 1

        # Performing minization, returning minimization results.
        #
        # To get deterministic results we use the source distance matrix to seed the initial pseudo-random
        # coordinates we start optimization from.

        rng = (

            numpy.random.Generator(
                numpy.random.PCG64(
                    
                    tuple(
                        hash(value)
                        for value in d_ij.flat))))

        result = (

            scipy.optimize.minimize(f,
                rng.random(N * 3),
                jac = df,
                callback = f_callback if verbose else None,
                options = {'disp': verbose}))

        result_x = numpy.stack((result.x[:N], result.x[N:N2], result.x[N2:])).T

        return result_x, f(result.x)

    @staticmethod
    def distance_graph(
            language_str,
            base_language_name,
            distance_data_array,
            distance_header_array,
            mode,
            storage,
            storage_dir,
            analysis_str = 'cognate_analysis',
            __debug_flag__ = False,
            __plot_flag__ = True):

        d_ij = (distance_data_array + distance_data_array.T) / 2

        log.debug(
            f'\n{analysis_str} {language_str}:'
            f'\ndistance_header_array:\n{distance_header_array}'
            f'\ndistance_data_array:\n{distance_data_array}'
            f'\nd_ij:\n{d_ij}')

        # Projecting the graph into a 2d plane via relative distance strain optimization, using PCA to
        # orient it left-right.

        if len(distance_data_array) > 1:

            embedding_2d, strain_2d = (
                CognateAnalysis.graph_2d_embedding(d_ij, verbose = __debug_flag__))

            embedding_2d_pca = (
                sklearn.decomposition.PCA(n_components = 2)
                    .fit_transform(embedding_2d))

            distance_2d = sklearn.metrics.euclidean_distances(embedding_2d)

        else:

            embedding_2d = numpy.zeros((1, 2))
            embedding_2d_pca = numpy.zeros((1, 2))

            strain_2d = 0.0

            distance_2d = numpy.zeros((1, 1))

        log.debug(
            f'\n{analysis_str} {language_str}:'
            f'\nembedding 2d:\n{embedding_2d}'
            f'\nembedding 2d (PCA-oriented):\n{embedding_2d_pca}'
            f'\nstrain 2d:\n{strain_2d}'
            f'\ndistances 2d:\n{distance_2d}')

        # And now the same with 3d embedding.

        if len(distance_data_array) > 1:

            embedding_3d, strain_3d = (
                CognateAnalysis.graph_3d_embedding(d_ij, verbose = __debug_flag__))

            # At least three points, standard PCA-based orientation.

            if len(distance_data_array) >= 3:

                embedding_3d_pca = (
                    sklearn.decomposition.PCA(n_components = 3)
                        .fit_transform(embedding_3d))

            # Only two points, so we take 2d embedding and extend it with zeros.

            else:

                embedding_3d_pca = (

                    numpy.hstack((
                        embedding_2d_pca,
                        numpy.zeros((embedding_2d_pca.shape[0], 1)))))

            # Making 3d embedding actually 3d, if required.

            if embedding_3d_pca.shape[1] <= 2:

                embedding_3d_pca = (

                    numpy.hstack((
                        embedding_3d_pca,
                        numpy.zeros((embedding_3d_pca.shape[0], 1)))))

            distance_3d = (
                sklearn.metrics.euclidean_distances(embedding_3d_pca))

        else:

            embedding_3d = numpy.zeros((1, 3))
            embedding_3d_pca = numpy.zeros((1, 3))

            strain_3d = 0.0

            distance_3d = numpy.zeros((1, 1))

        log.debug(
            f'\n{analysis_str} {language_str}:'
            f'\nembedding 3d:\n{embedding_3d}'
            f'\nembedding 3d (PCA-oriented):\n{embedding_3d_pca}'
            f'\nstrain 3d:\n{strain_3d}'
            f'\ndistances 3d:\n{distance_3d}')

        # Computing minimum spanning tree via standard Jarnik-Prim-Dijkstra algorithm using 2d and 3d
        # embedding distances to break ties.

        if len(distance_data_array) <= 1:
            mst_list = []

        else:

            d_min, d_extra_min, min_i, min_j = min(
                (d_ij[i,j], distance_2d[i,j] + distance_3d[i,j], i, j)
                for i in range(d_ij.shape[0] - 1)
                for j in range(i + 1, d_ij.shape[0]))

            mst_list = [(min_i, min_j)]
            mst_dict = {}

            # MST construction initialization.

            for i in range(d_ij.shape[0]):

                if i == min_i or i == min_j:
                    continue

                d_min_i = (d_ij[i, min_i], distance_2d[i, min_i] + distance_3d[i, min_i])
                d_min_j = (d_ij[i, min_j], distance_2d[i, min_j] + distance_3d[i, min_j])

                mst_dict[i] = (
                    (d_min_i, min_i) if d_min_i <= d_min_j else
                    (d_min_j, min_i))

            # Iterative MST construction.

            while len(mst_dict) > 0:

                (d_min, d_extra_min, i_min, i_from_min) = min(
                    (d, d_extra, i, i_from) for i, ((d, d_extra), i_from) in mst_dict.items())

                log.debug('\n' + pprint.pformat(mst_dict))
                log.debug('\n' + repr((i_from_min, i_min, d_min, d_extra_min)))

                mst_list.append((i_from_min, i_min))
                del mst_dict[i_min]

                # Updating shortest connection info.

                for i_to in mst_dict.keys():

                    d_to = (d_ij[i_min, i_to], distance_2d[i_min, i_to] + distance_3d[i_min, i_to])

                    if d_to < mst_dict[i_to][0]:
                        mst_dict[i_to] = (d_to, i_min)

        log.debug(
            f'\n{analysis_str} {language_str}:'
            f'\nminimum spanning tree:\n{pprint.pformat(mst_list)}')

        # Plotting with matplotlib.
        figure_url = None
        if __plot_flag__:

            figure = pyplot.figure(figsize = (10, 10))
            axes = figure.add_subplot(212)

            axes.set_title(
                'Etymological distance tree (relative distance embedding)',
                fontsize = 14, family = 'Gentium')

            axes.axis('equal')
            axes.axis('off')
            axes.autoscale()

            def f(axes, embedding_pca):
                """
                Plots specified graph embedding on a given axis.
                """

                flag_3d = numpy.size(embedding_pca, 1) > 2

                for index, (position, name) in enumerate(
                    zip(embedding_pca, distance_header_array)):

                    # Checking if any of the previous perspectives are already in this perspective's
                    # position.

                    same_position_index = None

                    for i, p in enumerate(embedding_pca[:index]):
                        if numpy.linalg.norm(position - p) <= 1e-3:

                            same_position_index = i
                            break

                    color = matplotlib.colors.hsv_to_rgb(
                        [(same_position_index or index) * 1.0 / len(distance_header_array), 0.5, 0.75])

                    label_same_str = (
                        '' if same_position_index is None else
                        ' (same as {0})'.format(same_position_index + 1))

                    kwargs = {
                        's': 35,
                        'color': color,
                        'label': '{0}) {1}{2}'.format(index + 1, name, label_same_str)}

                    axes.scatter(*position, **kwargs)

                    # Annotating position with its number, but only if we hadn't already annotated nearby.

                    if same_position_index is None:

                        if flag_3d:

                            axes.text(
                                position[0] + 0.01, position[1], position[2] + 0.01,
                                str(index + 1), None, fontsize = 14)

                        else:

                            axes.annotate(
                                str(index + 1),
                                (position[0] + 0.01, position[1] - 0.005),
                                fontsize = 14)

                # Plotting minimum spanning trees.

                line_list = [
                    (embedding_pca[i], embedding_pca[j])
                    for i, j in mst_list]

                line_collection = (
                    Line3DCollection if flag_3d else LineCollection)(
                        line_list, zorder = 0, color = 'gray')

                axes.add_collection(line_collection)

                pyplot.setp(axes.texts, family = 'Gentium')

            # Plotting our embedding, creating the legend.

            f(axes, embedding_2d_pca)

            pyplot.tight_layout()

            legend = axes.legend(
                scatterpoints = 1,
                loc = 'upper center',
                bbox_to_anchor = (0.5, -0.05),
                frameon = False,
                handlelength = 0.5,
                handletextpad = 0.75,
                fontsize = 14)

            pyplot.setp(legend.texts, family = 'Gentium')
            axes.autoscale_view()

            # Saving generated figure for debug purposes, if required.

            if __debug_flag__:

                figure_file_name = (
                    'figure cognate distance{0}.png'.format(
                    mode_name_str))

                with open(figure_file_name, 'wb') as figure_file:

                    pyplot.savefig(
                        figure_file,
                        bbox_extra_artists = (legend,),
                        bbox_inches = 'tight',
                        pad_inches = 0.25,
                        format = 'png')

                # Also generating 3d embedding figure.

                figure_3d = pyplot.figure()
                figure_3d.set_size_inches(16, 10)

                axes_3d = figure_3d.add_subplot(111, projection = '3d')

                axes_3d.axis('equal')
                axes_3d.view_init(elev = 30, azim = -75)

                f(axes_3d, embedding_3d_pca)

                # Setting up legend.

                axes_3d.set_xlabel('X')
                axes_3d.set_ylabel('Y')
                axes_3d.set_zlabel('Z')

                legend_3d = axes_3d.legend(
                    scatterpoints = 1,
                    loc = 'upper center',
                    bbox_to_anchor = (0.5, -0.05),
                    frameon = False,
                    handlelength = 0.5,
                    handletextpad = 0.75,
                    fontsize = 14)

                pyplot.setp(legend_3d.texts, family = 'Gentium')

                # Fake cubic bounding box to force axis aspect ratios, see
                # https://stackoverflow.com/a/13701747/2016856.

                X = embedding_3d_pca[:,0]
                Y = embedding_3d_pca[:,1]
                Z = embedding_3d_pca[:,2]

                max_range = numpy.array([
                    X.max() - X.min(), Y.max() - Y.min(), Z.max() - Z.min()]).max()

                Xb = (
                    0.5 * max_range * numpy.mgrid[-1:2:2,-1:2:2,-1:2:2][0].flatten() +
                    0.5 * (X.max() + X.min()))

                Yb = (
                    0.5 * max_range * numpy.mgrid[-1:2:2,-1:2:2,-1:2:2][1].flatten() +
                    0.5 * (Y.max() + Y.min()))

                Zb = (
                    0.5 * max_range * numpy.mgrid[-1:2:2,-1:2:2,-1:2:2][2].flatten() +
                    0.5 * (Z.max() + Z.min()))

                for xb, yb, zb in zip(Xb, Yb, Zb):
                   axes_3d.plot([xb], [yb], [zb], 'w')

                axes_3d.autoscale_view()

                # And saving it.

                figure_3d_file_name = (
                    'figure 3d cognate distance{0}.png'.format(
                    mode_name_str))

                with open(figure_3d_file_name, 'wb') as figure_3d_file:

                    figure_3d.savefig(
                        figure_3d_file,
                        bbox_extra_artists = (legend_3d,),
                        bbox_inches = 'tight',
                        pad_inches = 0.25,
                        format = 'png')

            # Storing generated figure as a PNG image.
            current_datetime = datetime.datetime.now(datetime.timezone.utc)
            figure_filename = pathvalidate.sanitize_filename(
                '{0} cognate{1} analysis {2:04d}.{3:02d}.{4:02d}.png'.format(
                    base_language_name[:64],
                    ' ' + mode if mode else '',
                    current_datetime.year,
                    current_datetime.month,
                    current_datetime.day))

            figure_path = os.path.join(storage_dir, figure_filename)
            os.makedirs(os.path.dirname(figure_path), exist_ok = True)

            with open(figure_path, 'wb') as figure_file:

                figure.savefig(
                    figure_file,
                    bbox_extra_artists = (legend,),
                    bbox_inches = 'tight',
                    pad_inches = 0.25,
                    format = 'png')

            cur_time = time.time()
            figure_url = ''.join([
                storage['prefix'], storage['static_route'],
                'cognate', '/', str(cur_time), '/', figure_filename])
        ### Plotting with matplotlib ends

        return (
            figure_url,
            mst_list,
            embedding_2d_pca,
            embedding_3d_pca
        )

    @staticmethod
    def perform_cognate_analysis(
        language_str,
        source_perspective_id,
        base_language_id,
        base_language_name,
        group_field_id,
        perspective_info_list,
        multi_list,
        multi_name_list,
        mode,
        distance_flag,
        reference_perspective_id,
        figure_flag,
        distance_vowel_flag,
        distance_consonant_flag,
        match_translations_value,
        only_orphans_flag,
        locale_id,
        storage,
        task_status = None,
        __debug_flag__ = False,
        __intermediate_flag__ = False):
        """
        Performs cognate analysis in either synchronous or asynchronous mode.
        """

        __result_flag__ = False

        if task_status is not None:
            task_status.set(1, 0, 'Gathering grouping data')

        # Sometimes in debugging mode we should return already computed results.

        if __debug_flag__:

            tag_data_digest = (

                hashlib.md5(

                    repr(list(group_field_id) +
                        [perspective_info[0] for perspective_info in perspective_info_list])

                    .encode('utf-8'))

                .hexdigest())

            result_file_name = (

                '__result_{0}_{1}__.gz'.format(

                    'multi{0}'.format(len(multi_list))
                        if mode == 'multi' else
                        '{0}_{1}'.format(*base_language_id),

                    tag_data_digest))

            if __result_flag__ and os.path.exists(result_file_name):

                with gzip.open(
                    result_file_name, 'rb') as result_file:

                    result_dict = pickle.load(result_file)

                return CognateAnalysis(**result_dict)

        # Gathering entry grouping data.

        perspective_dict = collections.defaultdict(dict)

        entry_already_set = set()
        group_list = []

        tag_dict = collections.defaultdict(set)

        text_dict = {}
        entry_id_dict = {}

        if not __debug_flag__:

            entry_already_set, group_list, group_time = (

                CognateAnalysis.tag_data_plpgsql(
                    perspective_info_list, group_field_id))

        else:

            # If we are in debug mode, we try to load existing tag data to reduce debugging time.

            tag_data_file_name = (

                '__tag_data_{0}_{1}__.gz'.format(

                    'multi{0}'.format(len(multi_list))
                        if mode == 'multi' else
                        '{0}_{1}'.format(*base_language_id),

                    tag_data_digest))

            # Checking if we have saved data.

            if os.path.exists(tag_data_file_name):

                with gzip.open(tag_data_file_name, 'rb') as tag_data_file:
                    entry_already_set, group_list, group_time = pickle.load(tag_data_file)

            else:

                # Don't have existing data, so we gather it and then save it for later use.

                entry_already_set, group_list, group_time = (

                    CognateAnalysis.tag_data_plpgsql(
                        perspective_info_list, group_field_id))

                with gzip.open(tag_data_file_name, 'wb') as tag_data_file:
                    pickle.dump((entry_already_set, group_list, group_time), tag_data_file)

        log.debug(
            '\ncognate_analysis {0}:'
            '\n{1} entries, {2} groups, {3:.2f}s elapsed time'.format(
            language_str,
            len(entry_already_set),
            len(group_list),
            group_time))

        if task_status is not None:
            task_status.set(2, 5, 'Gathering analysis source data')

        # Getting text data for each perspective.

        dbTranslation = aliased(dbEntity, name = 'Translation')
        dbSound = aliased(dbEntity, name = 'Sound')
        dbMarkup = aliased(dbEntity, name = 'Markup')

        dbPublishingTranslation = aliased(dbPublishingEntity, name = 'PublishingTranslation')
        dbPublishingSound = aliased(dbPublishingEntity, name = 'PublishingSound')
        dbPublishingMarkup = aliased(dbPublishingEntity, name = 'PublishingMarkup')

        phonemic_data_list = []
        suggestions_data_list = []

        sg_total_count = 0
        sg_xcript_count = 0
        sg_xlat_count = 0
        sg_both_count = 0

        source_perspective_index = None
        for index, (perspective_id, transcription_field_id, translation_field_id) in \
            enumerate(perspective_info_list):

            if perspective_id == source_perspective_id:
                source_perspective_index = index

            # Getting and saving perspective info.

            perspective = DBSession.query(dbPerspective).filter_by(
                client_id = perspective_id[0], object_id = perspective_id[1]).first()

            perspective_name = perspective.get_translation(locale_id)
            dictionary_name = perspective.parent.get_translation(locale_id)

            transcription_rules = (
                '' if not perspective.additional_metadata else
                    perspective.additional_metadata.get('transcription_rules', ''))

            perspective_data = perspective_dict[perspective_id]

            perspective_data['perspective_name'] = perspective_name
            perspective_data['dictionary_name'] = dictionary_name
            perspective_data['transcription_rules'] = transcription_rules

            # Preparing to save additional data, if required.

            if mode == 'phonemic':

                phonemic_data_list.append([
                    '{0} - {1}'.format(dictionary_name, perspective_name), ''])

            elif mode == 'suggestions':

                suggestions_data_list.append([])

            log.debug(
                '\ncognate_analysis {0}:'
                '\n  dictionary {1}/{2}: {3}'
                '\n  perspective {4}/{5}: {6}'
                '\n  transcription_rules: {7}'.format(
                language_str,
                perspective.parent_client_id, perspective.parent_object_id,
                repr(dictionary_name.strip()),
                perspective_id[0], perspective_id[1],
                repr(perspective_name.strip()),
                repr(transcription_rules)))

            # Getting text data.

            transcription_query = (

                DBSession.query(
                    dbLexicalEntry.client_id,
                    dbLexicalEntry.object_id).filter(
                        dbLexicalEntry.parent_client_id == perspective_id[0],
                        dbLexicalEntry.parent_object_id == perspective_id[1],
                        dbLexicalEntry.marked_for_deletion == False,
                        dbEntity.parent_client_id == dbLexicalEntry.client_id,
                        dbEntity.parent_object_id == dbLexicalEntry.object_id,
                        dbEntity.field_client_id == transcription_field_id[0],
                        dbEntity.field_object_id == transcription_field_id[1],
                        dbEntity.marked_for_deletion == False,
                        dbPublishingEntity.client_id == dbEntity.client_id,
                        dbPublishingEntity.object_id == dbEntity.object_id,
                        dbPublishingEntity.published == True,
                        dbPublishingEntity.accepted == True)

                .add_columns(
                    func.array_agg(dbEntity.content).label('transcription'))

                .group_by(dbLexicalEntry)).subquery()

            translation_query = (

                DBSession.query(
                    dbLexicalEntry.client_id,
                    dbLexicalEntry.object_id).filter(
                        dbLexicalEntry.parent_client_id == perspective_id[0],
                        dbLexicalEntry.parent_object_id == perspective_id[1],
                        dbLexicalEntry.marked_for_deletion == False,
                        dbEntity.parent_client_id == dbLexicalEntry.client_id,
                        dbEntity.parent_object_id == dbLexicalEntry.object_id,
                        dbEntity.field_client_id == translation_field_id[0],
                        dbEntity.field_object_id == translation_field_id[1],
                        dbEntity.marked_for_deletion == False,
                        dbPublishingEntity.client_id == dbEntity.client_id,
                        dbPublishingEntity.object_id == dbEntity.object_id,
                        dbPublishingEntity.published == True,
                        dbPublishingEntity.accepted == True)

                .add_columns(
                    func.array_agg(dbEntity.content).label('translation'))

                .group_by(dbLexicalEntry)).subquery()

            # Main query for transcription/translation data.

            data_query = (
                DBSession.query(transcription_query)

                .outerjoin(translation_query, and_(
                    transcription_query.c.client_id == translation_query.c.client_id,
                    transcription_query.c.object_id == translation_query.c.object_id))

                .add_columns(
                    translation_query.c.translation))

            # If we need to do an acoustic analysis, we also get sound/markup data.

            if mode == 'acoustic':

                sound_markup_query = (

                    DBSession.query(
                        dbLexicalEntry.client_id,
                        dbLexicalEntry.object_id).filter(
                            dbLexicalEntry.parent_client_id == perspective_id[0],
                            dbLexicalEntry.parent_object_id == perspective_id[1],
                            dbLexicalEntry.marked_for_deletion == False,
                            dbMarkup.parent_client_id == dbLexicalEntry.client_id,
                            dbMarkup.parent_object_id == dbLexicalEntry.object_id,
                            dbMarkup.marked_for_deletion == False,
                            dbMarkup.additional_metadata.contains({'data_type': 'praat markup'}),
                            dbPublishingMarkup.client_id == dbMarkup.client_id,
                            dbPublishingMarkup.object_id == dbMarkup.object_id,
                            dbPublishingMarkup.published == True,
                            dbPublishingMarkup.accepted == True,
                            dbSound.client_id == dbMarkup.self_client_id,
                            dbSound.object_id == dbMarkup.self_object_id,
                            dbSound.marked_for_deletion == False,
                            dbPublishingSound.client_id == dbSound.client_id,
                            dbPublishingSound.object_id == dbSound.object_id,
                            dbPublishingSound.published == True,
                            dbPublishingSound.accepted == True)

                    .add_columns(

                        func.jsonb_agg(func.jsonb_build_array(
                            dbSound.client_id, dbSound.object_id, dbSound.content,
                            dbMarkup.client_id, dbMarkup.object_id, dbMarkup.content))

                        .label('sound_markup'))

                    .group_by(dbLexicalEntry)).subquery()

                # Adding sound/markup retrieval to the main query.

                data_query = (
                    data_query

                    .outerjoin(sound_markup_query, and_(
                        transcription_query.c.client_id == sound_markup_query.c.client_id,
                        transcription_query.c.object_id == sound_markup_query.c.object_id))

                    .add_columns(
                        sound_markup_query.c.sound_markup))

            # If we are in asynchronous mode, we need to look up how many data rows we need
            # to process for this perspective.

            if task_status is not None:

                row_count = data_query.count()

                log.debug(
                    'cognate_analysis {0}: perspective {1}/{2}: {3} data rows'.format(
                    language_str,
                    perspective_id[0], perspective_id[1],
                    row_count))

            # Grouping transcriptions and translations by lexical entries.

            for row_index, row in enumerate(data_query.all()):

                entry_id = tuple(row[:2])
                transcription_list, translation_list = row[2:4]

                transcription_list = (
                    [] if not transcription_list else [
                        transcription.strip()
                        for transcription in transcription_list
                        if transcription.strip()])

                # If we have no trascriptions for this lexical entry, we skip it altogether.

                if not transcription_list:
                    continue

                translation_list = (
                    [] if not translation_list else [
                        translation.strip()
                        for translation in translation_list
                        if translation.strip()])

                # Saving transcription / translation data.

                translation_str = (
                    translation_list[0] if translation_list else '')

                if mode == 'phonemic':

                    for transcription in transcription_list:
                        phonemic_data_list[-1].extend([transcription, translation_str])

                elif mode == 'suggestions' and entry_id not in entry_already_set:

                    suggestions_data_list[-1].append([
                        '|'.join(transcription_list),
                        '|'.join(translation_list)])

                    sg_total_count += 1

                    # Counting how many instances of more than one transcription and / or translation
                    # we have.

                    if len(transcription_list) > 1:
                        sg_xcript_count += 1

                    if len(translation_list) > 1:
                        sg_xlat_count += 1

                    if len(transcription_list) > 1 and len(translation_list) > 1:
                        sg_both_count += 1

                # If we are fetching additional acoustic data, it's possible we have to process
                # sound recordings and markup this lexical entry has.

                if len(row) > 4 and row[4]:

                    row_list = row[4][0]

                    result = (
                        CognateAnalysis.acoustic_data(
                            base_language_id,
                            tuple(row_list[0:2]), row_list[2],
                            tuple(row_list[3:5]), row_list[5],
                            storage,
                            __debug_flag__))

                    # Updating task progress, if required.

                    if task_status is not None:

                        percent = int(math.floor(90.0 *
                            (index + float(row_index + 1) / row_count) /
                            len(perspective_info_list)))

                        task_status.set(2, 5 + percent, 'Gathering analysis source data')

                    entry_data_list = (index,
                        transcription_list,
                        translation_list,
                        result)

                # No additional acoustic data.

                else:
                    entry_data_list = (index, transcription_list, translation_list)

                text_dict[entry_id] = entry_data_list

                entry_id_key = (

                    index,
                    '|'.join(transcription_list) + (
                        ' ʽ' + '|'.join(translation_list) + 'ʼ' if translation_list else ''))

                entry_id_dict[entry_id_key] = entry_id

        # Showing some info on non-grouped entries, if required.

        if mode == 'suggestions':

            log.debug(
                '\ncognate_analysis {0}:'
                '\n{1} non-grouped entries'
                '\n{2} with multiple transcriptions'
                '\n{3} with multiple translations'
                '\n{4} with multiple transcriptions and translations'.format(
                language_str,
                sg_total_count,
                sg_xcript_count,
                sg_xlat_count,
                sg_both_count))

            # Also, if we are computing cognate suggestions, we should have a valid source perspective, it's
            # an error otherwise.

            if source_perspective_index is None:

                return ResponseError(message =
                    'Cognate suggestions require that the source perspective '
                    'is among the ones being analyzed.')

        if task_status is not None:
            task_status.set(3, 95, 'Performing analysis')

        # Ok, and now we form the source data for analysis.

        result_list = [[]]

        perspective_id_list = []
        perspective_name_list = []

        for perspective_id, transcription_field_id, translation_field_id in perspective_info_list:

            perspective_id_list.append(perspective_id)
            perspective_data = perspective_dict[perspective_id]

            perspective_str = '{0} - {1}'.format(
                perspective_data['dictionary_name'],
                perspective_data['perspective_name'])

            perspective_name_list.append(perspective_str)

            # Also going to use transcription transformation rules.

            result_list[0].extend([
                perspective_str,
                perspective_data['transcription_rules']])

        log.debug(
            '\ncognate_analysis {0}:'
            '\nsource_perspective_index: {1}'
            '\nperspective_list:\n{2}'
            '\nheader_list:\n{3}'.format(
            language_str,
            source_perspective_index,
            pprint.pformat(perspective_name_list, width = 108),
            pprint.pformat(result_list[0], width = 108)))

        # Each group of lexical entries.

        not_enough_count = 0

        total_transcription_count = 0
        total_translation_count = 0

        not_suggestions = mode != 'suggestions'

        for entry_id_set in group_list:

            group_entry_id_list = [[]
                for i in range(len(perspective_info_list))]

            group_transcription_list = [[]
                for i in range(len(perspective_info_list))]

            group_translation_list = [[]
                for i in range(len(perspective_info_list))]

            group_acoustic_list = [None
                for i in range(len(perspective_info_list))]

            transcription_count = 0
            translation_count = 0

            for entry_id in entry_id_set:

                if entry_id not in text_dict:
                    continue

                # Processing text data of each entry of the group.

                entry_data_list = text_dict[entry_id]

                (index,
                    transcription_list,
                    translation_list) = (

                    entry_data_list[:3])

                group_entry_id_list[index].append(entry_id)

                group_transcription_list[index].extend(transcription_list)
                group_translation_list[index].extend(translation_list)

                transcription_count += len(transcription_list)
                translation_count += len(translation_list)

                if (len(entry_data_list) > 3 and
                    entry_data_list[3] and
                    group_acoustic_list[index] is None):

                    group_acoustic_list[index] = entry_data_list[3]

            # Dropping groups with transcriptions from no more than a single dictionary, if required.

            if (not_suggestions and
                sum(min(1, len(transcription_list))
                    for transcription_list in group_transcription_list) <= 1):

                not_enough_count += 1
                continue

            total_transcription_count += transcription_count
            total_translation_count += translation_count

            result_list.append([])

            group_zipper = zip(
                group_entry_id_list,
                group_transcription_list,
                group_translation_list,
                group_acoustic_list)

            # Forming row of the source data table based on the entry group.

            for (
                index, (
                    entry_id_list,
                    transcription_list,
                    translation_list,
                    acoustic_list)) in (

                enumerate(group_zipper)):

                transcription_str = '|'.join(transcription_list)
                translation_str = '|'.join(translation_list)

                result_list[-1].append(transcription_str)
                result_list[-1].append(translation_str)

                if mode == 'acoustic':
                    result_list[-1].extend(acoustic_list or ['', '', '', '', ''])

                # Saving mapping from the translation / transcription info string to an id of one entry of
                # the group.

                if transcription_list or translation_list:

                    entry_id_key = (

                        index,
                        transcription_str + (
                            ' ʽ' + translation_str + 'ʼ' if translation_str else ''))

                    entry_id_dict[entry_id_key] = entry_id_list[0]

        # Showing what we've gathered.

        log.debug(
            '\ncognate_analysis {0}:'
            '\n  len(group_list): {1}'
            '\n  len(result_list): {2}'
            '\n  not_enough_count: {3}'
            '\n  transcription_count: {4}'
            '\n  translation_count: {5}'
            '\n  result_list:\n{6}'.format(
                language_str,
                len(group_list),
                len(result_list),
                not_enough_count,
                total_transcription_count,
                total_translation_count,
                pprint.pformat(result_list, width = 108)))

        # If we have no data at all, we return empty result.

        if len(result_list) <= 1 and not_suggestions:

            return CognateAnalysis(
                triumph = True,
                dictionary_count = len(perspective_info_list),
                group_count = len(group_list),
                not_enough_count = not_enough_count,
                transcription_count = total_transcription_count,
                translation_count = total_translation_count,
                result = '',
                xlsx_url = '',
                distance_list = [],
                figure_url = '',
                intermediate_url_list = None)

        analysis_f = (
            cognate_acoustic_analysis_f if mode == 'acoustic' else
            cognate_reconstruction_f if mode == 'reconstruction' else
            cognate_reconstruction_multi_f if mode == 'multi' else
            cognate_suggestions_f if mode == 'suggestions' else
            cognate_analysis_f)

        # Preparing analysis input.

        phonemic_input_list = [
            ''.join(text + '\0' for text in text_list)
            for text_list in phonemic_data_list]

        suggestions_result_list = []

        for tt_list in itertools.zip_longest(
            *suggestions_data_list, fillvalue = ['', '']):

            suggestions_result_list.append([])

            for tt in tt_list:
                suggestions_result_list[-1].extend(tt)

        if mode == 'suggestions':

            # Showing additional ungrouped input data, if required.

            log.debug(
                '\ncognate_analysis {0}:'
                '\nsuggestions_result_list:\n{1}'.format(
                    language_str,
                    pprint.pformat(suggestions_result_list, width = 144)))

        result_input = (

            ''.join(
                ''.join(text + '\0' for text in text_list)

                for text_list in (
                    result_list + suggestions_result_list)))

        input = '\0'.join(phonemic_input_list + [result_input])

        log.debug(
            '\ncognate_analysis {0}:'
            '\nanalysis_f: {1}'
            '\ninput ({2} columns, {3} rows{4}):\n{5}'.format(
                language_str,
                repr(analysis_f),
                len(perspective_info_list),
                len(result_list),
                '' if mode != 'suggestions' else
                    ', {0} ungrouped rows'.format(len(suggestions_result_list)),
                pprint.pformat([input[i : i + 256]
                    for i in range(0, len(input), 256)], width = 144)))

        # Saving input to a file, if required.

        storage_dir = None
        intermediate_url_list = []

        if __debug_flag__ or __intermediate_flag__:

            language_name_str = (
                ' '.join(multi_name_list) if mode == 'multi' else
                base_language_name.strip())

            if len(language_name_str) > 64:
                language_name_str = language_name_str[:64] + '...'

            mode_name_str = (

                '{0} {1} {2} {3}{4}'.format(

                    ' multi{0}'.format(len(multi_list))
                        if mode == 'multi' else
                        (' ' + mode if mode else ''),

                    language_name_str,

                    ' '.join(str(count) for id, count in multi_list)
                        if mode == 'multi' else
                        len(perspective_info_list),

                    len(result_list),

                    '' if not_suggestions else
                        ' {} {} {} {}'.format(
                            len(suggestions_result_list),
                            source_perspective_index,
                            match_translations_value,
                            int(only_orphans_flag))))

            cognate_name_str = (
                'cognate' + mode_name_str)

            # Initializing file storage directory, if required.

            if __intermediate_flag__ and storage_dir is None:

                cur_time = time.time()

                storage_dir = os.path.join(
                    storage['path'], 'cognate', str(cur_time))

            for extension, encoding in ('utf8', 'utf-8'), ('utf16', 'utf-16'):

                input_file_name = (

                    pathvalidate.sanitize_filename(
                        'input {0}.{1}'.format(
                            cognate_name_str, extension)))

                # Saving to the working directory...

                if __debug_flag__:

                    with open(input_file_name, 'wb') as input_file:
                        input_file.write(input.encode(encoding))

                # ...and / or to the file storage.

                if __intermediate_flag__:

                    input_path = os.path.join(
                        storage_dir, input_file_name)

                    os.makedirs(
                        os.path.dirname(input_path),
                        exist_ok = True)

                    with open(input_path, 'wb') as input_file:
                        input_file.write(input.encode(encoding))

                    input_url = ''.join([
                        storage['prefix'],
                        storage['static_route'],
                        'cognate', '/',
                        str(cur_time), '/',
                        input_file_name])

                    intermediate_url_list.append(input_url)

        # Calling analysis library, starting with getting required output buffer size and continuing
        # with analysis proper.

        if mode == 'multi':

            multi_count_list = [
                perspective_count
                for language_id, perspective_count in multi_list]

            perspective_count_array = (
                ctypes.c_int * len(multi_list))(*multi_count_list)

            # int CognateMultiReconstruct_GetAllOutput(
            #   LPTSTR bufIn, int* pnCols, int nGroups, int nRows, LPTSTR bufOut, int flags)

            output_buffer_size = analysis_f(
                None,
                perspective_count_array,
                len(multi_list),
                len(result_list),
                None,
                1)

        elif mode == 'suggestions':

            # int GuessCognates_GetAllOutput(
            #   LPTSTR bufIn, int nCols, int nRowsCorresp, int nRowsRest, int iDictThis, int lookMeaning,
            #   int onlyOrphans, LPTSTR bufOut, int flags)

            output_buffer_size = analysis_f(
                None,
                len(perspective_info_list),
                len(result_list),
                len(suggestions_result_list),
                source_perspective_index,
                match_translations_value,
                int(only_orphans_flag),
                None,
                1)

        else:

            # int CognateAnalysis_GetAllOutput(
            #   LPTSTR bufIn, int nCols, int nRows, LPTSTR bufOut, int flags)

            output_buffer_size = analysis_f(
                None,
                len(perspective_info_list),
                len(result_list),
                None,
                1)

        log.debug(
            '\ncognate_analysis {0}: output buffer size {1}'.format(
            language_str,
            output_buffer_size))

        input_buffer = ctypes.create_unicode_buffer(input)

        # Saving input buffer to a file, if required.

        if __debug_flag__:

            input_file_name = (
                'input {0}.buffer'.format(
                    cognate_name_str))

            with open(input_file_name, 'wb') as input_file:
                input_file.write(bytes(input_buffer))

        output_buffer = ctypes.create_unicode_buffer(output_buffer_size + 256)

        if mode == 'multi':

            result = analysis_f(
                input_buffer,
                perspective_count_array,
                len(multi_list),
                len(result_list),
                output_buffer,
                1)

        elif mode == 'suggestions':

            result = analysis_f(
                input_buffer,
                len(perspective_info_list),
                len(result_list),
                len(suggestions_result_list),
                source_perspective_index,
                match_translations_value,
                int(only_orphans_flag),
                output_buffer,
                1)

        else:

            result = analysis_f(
                input_buffer,
                len(perspective_info_list),
                len(result_list),
                output_buffer,
                1)

        log.debug(
            '\ncognate_analysis {0}: result {1}'.format(
            language_str,
            result))

        # If we don't have a good result, we return an error.

        if result <= 0:

            if task_status is not None:

                task_status.set(5, 100,
                    'Finished (ERROR): library call error {0}'.format(result))

            return ResponseError(message =
                'Cognate analysis library call error {0}'.format(result))

        output = output_buffer.value

        log.debug(
            '\ncognate_analysis {}:\noutput ({}):\n{}'.format(
            language_str,
            len(output),
            pprint.pformat([output[i : i + 256]
                for i in range(0, len(output), 256)], width = 144)))

        # Saving output buffer and output to files, if required.

        if __debug_flag__:

            output_file_name = (
                'output {0}.buffer'.format(
                    cognate_name_str))

            with open(output_file_name, 'wb') as output_file:
                output_file.write(bytes(output_buffer))

            for extension, encoding in ('utf8', 'utf-8'), ('utf16', 'utf-16'):

                output_file_name = (
                    'output {0}.{1}'.format(
                        cognate_name_str,
                        extension))

                with open(output_file_name, 'wb') as output_file:
                    output_file.write(output.encode(encoding))

        # Reflowing output.

        line_list = output.split('\r\n')

        text_wrapper = textwrap.TextWrapper(
            width = max(196, len(perspective_info_list) * 40), tabsize = 20)

        reflow_list = []

        for line in line_list:
            reflow_list.extend(text_wrapper.wrap(line))

        wrapped_output = '\n'.join(reflow_list)

        log.debug(
            'cognate_analysis {0}:\nwrapped output:\n{1}'.format(
            language_str,
            wrapped_output))

        # Getting binary output for parsing and exporting.

        if mode == 'multi':

            result_binary = analysis_f(
                input_buffer,
                perspective_count_array,
                len(multi_list),
                len(result_list),
                output_buffer,
                2)

        # If we are in the suggestions mode, we currently just return the output.

        elif mode == 'suggestions':

            result_binary = analysis_f(
                input_buffer,
                len(perspective_info_list),
                len(result_list),
                len(suggestions_result_list),
                source_perspective_index,
                match_translations_value,
                int(only_orphans_flag),
                output_buffer,
                2)

        else:

            result_binary = analysis_f(
                input_buffer,
                len(perspective_info_list),
                len(result_list),
                output_buffer,
                2)

        log.debug(
            'cognate_analysis {0}: result_binary {1}'.format(
            language_str,
            result_binary))

        if result_binary <= 0:

            if task_status is not None:

                task_status.set(5, 100,
                    'Finished (ERROR): library call (binary) error {0}'.format(result_binary))

            return ResponseError(message =
                'Cognate analysis library call (binary) error {0}'.format(result_binary))

        # Showing what we've got from the binary output call.

        output_binary = output_buffer[:result_binary]

        output_binary_list = [
            output_binary[i : i + 256]
            for i in range(0, len(output_binary), 256)]

        log.debug(
            '\ncognate_analysis {0}:'
            '\noutput_binary:\n{1}'.format(
            language_str,
            pprint.pformat(
                output_binary_list, width = 144)))

        # Saving binary output buffer and binary output to files, if required.

        if __debug_flag__:

            output_file_name = (
                'output binary {0}.buffer'.format(
                    cognate_name_str))

            with open(
                output_file_name, 'wb') as output_file:

                output_file.write(
                    bytes(output_buffer))

            for extension, encoding in ('utf8', 'utf-8'), ('utf16', 'utf-16'):

                output_file_name = (
                    'output binary {0}.{1}'.format(
                        cognate_name_str, extension))

                with open(
                    output_file_name, 'wb') as output_file:

                    output_file.write(
                        output_binary.encode(encoding))

        # For cognate suggestions we just parse and return suggestions.

        if mode == 'suggestions':

            suggestion_list = (

                CognateAnalysis.parse_suggestions(
                    language_str,
                    output_binary,
                    len(perspective_info_list),
                    source_perspective_index,
                    entry_id_dict,
                    __debug_flag__,
                    cognate_name_str if __debug_flag__ else None,
                    group_field_id if __debug_flag__ else None))

            result_dict = (

                dict(

                    triumph = True,

                    dictionary_count = len(perspective_info_list),
                    group_count = len(group_list),
                    not_enough_count = not_enough_count,
                    transcription_count = total_transcription_count,
                    translation_count = total_translation_count,

                    result = output,

                    perspective_name_list = perspective_name_list,

                    suggestion_list = suggestion_list,
                    suggestion_field_id = group_field_id,

                    intermediate_url_list =
                        intermediate_url_list if __intermediate_flag__ else None))

            if __debug_flag__ and __result_flag__:

                with gzip.open(
                    result_file_name, 'wb') as result_file:

                    pickle.dump(result_dict, result_file)

            return CognateAnalysis(**result_dict)

        # Performing etymological distance analysis, if required.

        d_output = None
        d_output_binary = None

        if distance_flag or figure_flag:

            d_output_buffer_size = cognate_distance_analysis_f(
                None, len(perspective_info_list), len(result_list), None, 1)

            log.debug(
                'cognate_analysis {0}: distance output buffer size {1}'.format(
                language_str,
                d_output_buffer_size))

            d_output_buffer = ctypes.create_unicode_buffer(d_output_buffer_size + 256)

            d_result = cognate_distance_analysis_f(
                input_buffer, len(perspective_info_list), len(result_list), d_output_buffer, 1)

            # If we don't have a good result, we return an error.

            log.debug(
                'cognate_analysis {0}: distance result {1}'.format(
                language_str,
                d_result))

            if d_result <= 0:

                if task_status is not None:

                    task_status.set(5, 100,
                        'Finished (ERROR): library call error {0}'.format(d_result))

                return ResponseError(message =
                    'Cognate analysis library call error {0}'.format(d_result))

            # Showing what we've got.

            d_output = d_output_buffer.value

            distance_output_list = [
                d_output[i : i + 256]
                for i in range(0, len(d_output), 256)]

            log.debug(
                'cognate_analysis {0}:\ndistance output:\n{1}'.format(
                language_str,
                pprint.pformat(
                    distance_output_list, width = 144)))

            # Saving distance output buffer and distance output to files, if required.

            if __debug_flag__:

                d_output_file_name = (
                    'output {0}.buffer'.format(
                        cognate_name_str))

                with open(
                    d_output_file_name, 'wb') as d_output_file:

                    d_output_file.write(
                        bytes(d_output_buffer))

                for extension, encoding in ('utf8', 'utf-8'), ('utf16', 'utf-16'):

                    d_output_file_name = (
                        'output {0}.{1}'.format(
                            cognate_name_str, extension))

                    with open(
                        d_output_file_name, 'wb') as d_output_file:

                        d_output_file.write(
                            d_output.encode(encoding))

            # Getting binary output for parsing and exporting.

            d_result_binary = cognate_distance_analysis_f(
                input_buffer, len(perspective_info_list), len(result_list), d_output_buffer, 2)

            log.debug(
                'cognate_analysis {0}: distance result_binary {1}'.format(
                language_str,
                d_result_binary))

            if d_result_binary <= 0:

                if task_status is not None:

                    task_status.set(5, 100,
                        'Finished (ERROR): library call (binary) error {0}'.format(d_result_binary))

                return ResponseError(message =
                    'Cognate analysis library call (binary) error {0}'.format(d_result_binary))

            # Showing what we've got from the binary output call.

            d_output_binary = d_output_buffer[:d_result_binary]

            d_output_binary_list = [
                d_output_binary[i : i + 256]
                for i in range(0, len(d_output_binary), 256)]

            log.debug(
                '\ncognate_analysis {0}:'
                '\ndistance output_binary:\n{1}'.format(
                language_str,
                pprint.pformat(
                    d_output_binary_list, width = 144)))

        # Indicating task's final stage, if required.

        if task_status is not None:
            task_status.set(4, 99, 'Exporting analysis results to XLSX')

        # Parsing analysis results and exporting them as an Excel file.

        workbook_stream, distance_matrix_list = (

            CognateAnalysis.export_xlsx(
                language_str,
                mode,
                output_binary,
                d_output_binary,
                len(perspective_info_list),
                __debug_flag__,
                cognate_name_str if __debug_flag__ else None))

        current_datetime = datetime.datetime.now(datetime.timezone.utc)

        xlsx_filename = pathvalidate.sanitize_filename(
            '{0} cognate{1} analysis {2:04d}.{3:02d}.{4:02d}.xlsx'.format(
                base_language_name[:64],
                ' ' + mode if mode else '',
                current_datetime.year,
                current_datetime.month,
                current_datetime.day))

        if storage_dir is None:

            cur_time = time.time()
            storage_dir = os.path.join(storage['path'], 'cognate', str(cur_time))

        # Storing Excel file with the results.

        xlsx_path = os.path.join(storage_dir, xlsx_filename)
        os.makedirs(os.path.dirname(xlsx_path), exist_ok = True)

        workbook_stream.seek(0)

        with open(xlsx_path, 'wb') as xlsx_file:
            shutil.copyfileobj(workbook_stream, xlsx_file)

        xlsx_url = ''.join([
            storage['prefix'], storage['static_route'],
            'cognate', '/', str(cur_time), '/', xlsx_filename])

        # Selecting one of the distance matrices, if we have any.

        distance_header_array = None

        if distance_matrix_list is not None:

            distance_matrix = distance_matrix_list[-1]

            if distance_vowel_flag and distance_consonant_flag:
                pass

            elif distance_vowel_flag:
                distance_matrix = distance_matrix_list[0]

            elif distance_consonant_flag:
                distance_matrix = distance_matrix_list[1]

            (distance_title,
                distance_header_list,
                distance_data_list,
                distance_header_array,
                distance_data_array) = distance_matrix

        # Generating list of etymological distances to the reference perspective, if required.

        distance_list = None

        if distance_flag and reference_perspective_id is not None:

            reference_index = None

            for index, perspective_id in enumerate(perspective_id_list):
                if perspective_id == reference_perspective_id:

                    reference_index = index
                    break

            if reference_index is not None:

                distance_value_list = list(map(
                    float, distance_data_list[reference_index]))

                max_distance = float(max(distance_value_list))

                # Compiling and showing relative distance list.

                if max_distance > 0:
                    distance_list = [
                        (perspective_id, distance / max_distance)

                        for perspective_id, distance in zip(
                            perspective_id_list, distance_value_list)]

                else:

                    distance_list = distance_value_list

                log.debug(
                    '\ncognate_analysis {0}:'
                    '\n  perspective_id_list: {1}'
                    '\n  perspective_name_list:\n{2}'
                    '\n  reference_perspective_id: {3}'
                    '\n  reference_index: {4}'
                    '\n  distance_value_list: {5}'
                    '\n  max_distance: {6}'
                    '\n  distance_list: {7}'.format(
                    language_str,
                    perspective_id_list,
                    pprint.pformat(perspective_name_list, width = 144),
                    reference_perspective_id,
                    reference_index,
                    distance_value_list,
                    max_distance,
                    distance_list))

        # Generating distance graph, if required.
        figure_url = None
        mst_list = None
        embedding_2d_pca = None
        embedding_3d_pca = None

        if figure_flag:
            figure_url, mst_list, embedding_2d_pca, embedding_3d_pca = \
                CognateAnalysis.distance_graph(
                    language_str,
                    base_language_name,
                    distance_data_array,
                    distance_header_array,
                    mode,
                    storage,
                    storage_dir,
                    __debug_flag__
                )

        # Finalizing task status, if required, returning result.
        if task_status is not None:

            result_link_list = (
                [xlsx_url] +
                ([] if figure_url is None else [figure_url]) +
                (intermediate_url_list if __intermediate_flag__ else []))

            task_status.set(5, 100, 'Finished',
                result_link_list = result_link_list)

        result_dict = (

            dict(

                triumph = True,

                dictionary_count = len(perspective_info_list),
                group_count = len(group_list),
                not_enough_count = not_enough_count,
                transcription_count = total_transcription_count,
                translation_count = total_translation_count,

                result = wrapped_output,
                xlsx_url = xlsx_url,
                distance_list = distance_list,
                figure_url = figure_url,

                minimum_spanning_tree = mst_list,
                embedding_2d = embedding_2d_pca,
                embedding_3d = embedding_3d_pca,
                perspective_name_list = distance_header_array,

                intermediate_url_list =
                    intermediate_url_list if __intermediate_flag__ else None))

        if __debug_flag__ and __result_flag__:

            with gzip.open(
                result_file_name, 'wb') as result_file:

                pickle.dump(result_dict, result_file)

        return CognateAnalysis(**result_dict)

    @staticmethod
    def mutate(
        self,
        info,

        source_perspective_id,
        base_language_id,

        group_field_id,
        perspective_info_list,
        multi_list = None,

        mode = None,

        distance_flag = None,
        reference_perspective_id = None,

        figure_flag = None,
        distance_vowel_flag = None,
        distance_consonant_flag = None,

        match_translations_value = 1,
        only_orphans_flag = True,

        debug_flag = False,
        intermediate_flag = False,
        synchronous = False):
        """
        mutation CognateAnalysis {
          cognate_analysis(
            base_language_id: [508, 41],
            group_field_id: [66, 25],
            perspective_info_list: [
              [[425, 4], [66, 8], [66, 10]],
              [[1552, 1759], [66, 8], [66, 10]],
              [[418, 4], [66, 8], [66, 10]]])
          {
            triumph
            entity_count
            dictionary_count
            group_count
            not_enough_count
            text_count
            result
          }
        }
        """

        # Administrator / perspective author / editing permission check.

        error_str = (
            'Only administrator, perspective author and users with perspective editing permissions '
            'can perform cognate analysis.')

        client_id = info.context.request.authenticated_userid

        if not client_id:
            return ResponseError(error_str)

        user = Client.get_user_by_client_id(client_id)

        author_client_id_set = (

            set(
                client_id
                for (client_id, _), _, _ in perspective_info_list))

        author_id_check = (

            DBSession

                .query(

                    DBSession
                        .query(literal(1))
                        .filter(
                            Client.id.in_(author_client_id_set),
                            Client.user_id == user.id)
                        .exists())

                .scalar())

        if (user.id != 1 and
            not author_id_check and
            not info.context.acl_check_if('edit', 'perspective', source_perspective_id)):

            return ResponseError(error_str)

        # Debug mode check.

        if debug_flag and user.id != 1:

            return (

                ResponseError(
                    message = 'Only administrator can use debug mode.'))

        language_str = (
            '{0}/{1}, language {2}/{3}'.format(
                source_perspective_id[0], source_perspective_id[1],
                base_language_id[0], base_language_id[1]))

        try:

            # Getting base language info.

            locale_id = info.context.get('locale_id') or 2

            base_language = DBSession.query(dbLanguage).filter_by(
                client_id = base_language_id[0], object_id = base_language_id[1]).first()

            base_language_name = base_language.get_translation(locale_id)

            request = info.context.request
            storage = request.registry.settings['storage']

            # Getting multi-language info, if required.

            if multi_list is None:
                multi_list = []

            multi_name_list = []

            for language_id, perspective_count in multi_list:

                language = DBSession.query(dbLanguage).filter_by(
                    client_id = language_id[0], object_id = language_id[1]).first()

                multi_name_list.append(
                    language.get_translation(locale_id))

            # Language tag.

            if mode == 'multi':

                multi_str = ', '.join(
                    '{0}/{1}'.format(*id)
                    for id, count in multi_list)

                language_str = (
                    '{0}/{1}, languages {2}'.format(
                        source_perspective_id[0], source_perspective_id[1],
                        multi_str))

            # Showing cognate analysis info, checking cognate analysis library presence.

            log.debug(
                 '\ncognate_analysis {}:'
                 '\n  base language: {}'
                 '\n  group field: {}/{}'
                 '\n  perspectives and transcription/translation fields: {}'
                 '\n  multi_list: {}'
                 '\n  multi_name_list: {}'
                 '\n  mode: {}'
                 '\n  distance_flag: {}'
                 '\n  reference_perspective_id: {}'
                 '\n  figure_flag: {}'
                 '\n  distance_vowel_flag: {}'
                 '\n  distance_consonant_flag: {}'
                 '\n  match_translations_value: {}'
                 '\n  only_orphans_flag: {} ({})'
                 '\n  debug_flag: {}'
                 '\n  intermediate_flag: {}'
                 '\n  cognate_analysis_f: {}'
                 '\n  cognate_acoustic_analysis_f: {}'
                 '\n  cognate_distance_analysis_f: {}'
                 '\n  cognate_reconstruction_f: {}'
                 '\n  cognate_reconstruction_multi_f: {}'
                 '\n  cognate_suggestions_f: {}'.format(
                    language_str,
                    repr(base_language_name.strip()),
                    group_field_id[0], group_field_id[1],
                    perspective_info_list,
                    multi_list,
                    multi_name_list,
                    repr(mode),
                    distance_flag,
                    reference_perspective_id,
                    figure_flag,
                    distance_vowel_flag,
                    distance_consonant_flag,
                    match_translations_value,
                    only_orphans_flag, int(only_orphans_flag),
                    debug_flag,
                    intermediate_flag,
                    repr(cognate_analysis_f),
                    repr(cognate_acoustic_analysis_f),
                    repr(cognate_distance_analysis_f),
                    repr(cognate_reconstruction_f),
                    repr(cognate_reconstruction_multi_f),
                    repr(cognate_suggestions_f)))

            # Checking if we have analysis function ready.

            analysis_f = (
                cognate_acoustic_analysis_f if mode == 'acoustic' else
                cognate_reconstruction_f if mode == 'reconstruction' else
                cognate_reconstruction_multi_f if mode == 'multi' else
                cognate_suggestions_f if mode == 'suggestions' else
                cognate_analysis_f)

            if analysis_f is None:

                return ResponseError(message =
                    'Analysis library fuction \'{0}()\' is absent, '
                    'please contact system administrator.'.format(
                        'CognateAcousticAnalysis_GetAllOutput' if mode == 'acoustic' else
                        'CognateReconstruct_GetAllOutput' if mode == 'reconstruction' else
                        'CognateMultiReconstruct_GetAllOutput' if mode == 'multi' else
                        'GuessCognates_GetAllOutput' if mode == 'suggestions' else
                        'CognateAnalysis_GetAllOutput'))

            # Transforming client/object pair ids from lists to 2-tuples.

            source_perspective_id = tuple(source_perspective_id)
            base_language_id = tuple(base_language_id)
            group_field_id = tuple(group_field_id)

            perspective_info_list = [

                (tuple(perspective_id),
                    tuple(transcription_field_id),
                    tuple(translation_field_id))

                for perspective_id,
                    transcription_field_id,
                    translation_field_id in perspective_info_list]

            multi_list = [
                [tuple(language_id), perspective_count]
                for language_id, perspective_count in multi_list]

            if reference_perspective_id is not None:
                reference_perspective_id = tuple(reference_perspective_id)

            # If we are to use acoustic data, we will launch cognate analysis in asynchronous mode.

            if mode == 'acoustic':

                client_id = request.authenticated_userid

                user_id = (
                    Client.get_user_by_client_id(client_id).id
                        if client_id else anonymous_userid(request))

                task_status = TaskStatus(
                    user_id, 'Cognate acoustic analysis', base_language_name, 5)

                # Launching cognate acoustic analysis asynchronously.

                request.response.status = HTTPOk.code

                if synchronous:
                    CognateAnalysis.perform_cognate_analysis(
                        language_str,
                        source_perspective_id,
                        base_language_id,
                        base_language_name,
                        group_field_id,
                        perspective_info_list,
                        multi_list,
                        multi_name_list,
                        mode,
                        None,
                        None,
                        None,
                        None,
                        None,
                        match_translations_value,
                        only_orphans_flag,
                        locale_id,
                        storage,
                        task_status,
                        debug_flag,
                        intermediate_flag)

                else:

                    async_cognate_analysis.delay(
                        language_str,
                        source_perspective_id,
                        base_language_id,
                        base_language_name,
                        group_field_id,
                        perspective_info_list,
                        multi_list,
                        multi_name_list,
                        mode,
                        distance_flag,
                        reference_perspective_id,
                        figure_flag,
                        distance_vowel_flag,
                        distance_consonant_flag,
                        match_translations_value,
                        only_orphans_flag,
                        locale_id,
                        storage,
                        task_status.key,
                        request.registry.settings['cache_kwargs'],
                        request.registry.settings['sqlalchemy.url'],
                        debug_flag,
                        intermediate_flag)

                # Signifying that we've successfully launched asynchronous cognate acoustic analysis.

                return CognateAnalysis(triumph = True)

            # We do not use acoustic data, so we perform cognate analysis synchronously.
            else:

                return CognateAnalysis.perform_cognate_analysis(
                    language_str,
                    source_perspective_id,
                    base_language_id,
                    base_language_name,
                    group_field_id,
                    perspective_info_list,
                    multi_list,
                    multi_name_list,
                    mode,
                    distance_flag,
                    reference_perspective_id,
                    figure_flag,
                    distance_vowel_flag,
                    distance_consonant_flag,
                    match_translations_value,
                    only_orphans_flag,
                    locale_id,
                    storage,
                    None,
                    debug_flag,
                    intermediate_flag)

        # Exception occured while we tried to perform cognate analysis.

        except Exception as exception:

            traceback_string = ''.join(traceback.format_exception(
                exception, exception, exception.__traceback__))[:-1]

            log.warning(
                'cognate_analysis {0}: exception'.format(
                language_str))

            log.warning(traceback_string)

            return ResponseError(message =
                'Exception:\n' + traceback_string)


class SwadeshAnalysis(graphene.Mutation):
    class Arguments:

        source_perspective_id = LingvodocID(required = True)
        base_language_id = LingvodocID(required = True)

        group_field_id = LingvodocID(required = True)
        perspective_info_list = graphene.List(graphene.List(LingvodocID), required = True)

        debug_flag = graphene.Boolean()

    triumph = graphene.Boolean()

    result = graphene.String()
    xlsx_url = graphene.String()
    minimum_spanning_tree = graphene.List(graphene.List(graphene.Int))
    embedding_2d = graphene.List(graphene.List(graphene.Float))
    embedding_3d = graphene.List(graphene.List(graphene.Float))
    perspective_name_list = graphene.List(graphene.String)

    @staticmethod
    def get_entry_text(entry):
        return f"{entry['swadesh']} [ {entry['transcription']} ] {entry['translation']}"

    @staticmethod
    def export_dataframe(result_pool, distance_data_array, bundles, get_entry_text):
        '''
        Keys:
        result_pool[perspective_id][entry_id]
        Fields:
        'group': group_index,
        'borrowed': bool,
        'swadesh': swadesh_lex,
        'transcription': transcription_list[0],
        'translation': translation_lex
        '''

        distances = pd.DataFrame(distance_data_array,
                                 columns=[perspective['name'] for perspective in result_pool.values()])
        # Start index for distances from 1 to match with dictionaries numbers
        distances.index += 1

        groups = pd.DataFrame()
        # Insert 'lines' column as the first one
        groups['lines'] = 0

        borrowed = pd.DataFrame()
        singles = pd.DataFrame()

        singles_index = 0
        borrowed_index = 0
        # re-group by group number and add joined values
        for perspective in result_pool.values():
            dict_name = perspective['name']
            for entry in perspective.values():
                # 'entry' iterator may present string value of 'name' field
                # but not a dictionary for one of entries. Continue in this case.
                if not isinstance(entry, dict):
                    continue
                group_num = entry['group']
                entry_text = get_entry_text(entry)
                if group_num is not None and group_num in bundles:
                    # Concatinate existing value if is and a new one, store the result to 'groups' dataframe
                    value = ""
                    if dict_name in groups:
                        cell = groups[dict_name].get(group_num)
                        value = cell if pd.notnull(cell) else value
                    value = f"{value}\n{entry_text}".strip()
                    groups.loc[group_num, dict_name] = value

                    # Count result lines to set rows height in xlsx after
                    lines = value.count('\n') + 1
                    cell = groups.loc[group_num].get('lines')
                    if pd.isnull(cell) or cell < lines:
                        groups.loc[group_num, 'lines'] = lines
                elif entry.get('borrowed'):
                    borrowed.loc[borrowed_index, dict_name] = entry_text
                    borrowed_index += 1
                else:
                    singles.loc[singles_index, dict_name] = entry_text
                    singles_index += 1

        return {
            'Cognates': groups if len(groups) < 2 else groups.sort_values(groups.columns[1]),
            'Singles': singles.sort_index(),
            'Borrowed': borrowed.sort_index(),
            'Distances': distances.sort_index()
        }

    @staticmethod
    def export_xlsx(
            result,
            base_language_name,
            storage
    ):
        # Exporting analysis results as an Excel file.

        current_datetime = datetime.datetime.now(datetime.timezone.utc)
        xlsx_filename = pathvalidate.sanitize_filename(
            '{0} {1} {2:04d}.{3:02d}.{4:02d}.xlsx'.format(
                base_language_name[:64],
                'glottochronology',
                current_datetime.year,
                current_datetime.month,
                current_datetime.day))

        cur_time = time.time()
        storage_dir = os.path.join(storage['path'], 'glottochronology', str(cur_time))

        # Storing Excel file with the results.

        xlsx_path = os.path.join(storage_dir, xlsx_filename)
        os.makedirs(os.path.dirname(xlsx_path), exist_ok=True)

        with pd.ExcelWriter(xlsx_path, engine='xlsxwriter') as writer:
            header_format = writer.book.add_format({'bold': True,
                                                    'text_wrap': True,
                                                    'valign': 'top',
                                                    'fg_color': '#D7E4BC',
                                                    'border': 1})
            for sheet_name, df in result.items():
                index = (sheet_name == 'Distances')
                startcol = int(index)
                # Exclude 'lines' column
                columns = df.columns[int(sheet_name == 'Cognates'):]
                # Check if the table is empty
                if columns.empty:
                    continue

                df.to_excel(writer,
                            sheet_name=sheet_name,
                            index=index,
                            startrow=1,
                            columns=columns,
                            header=False)

                worksheet = writer.sheets[sheet_name]
                worksheet.set_row(0, 70)
                worksheet.set_column(startcol, len(columns) - 1 + startcol, 30)
                # Write the column headers with the defined format.
                for col_num, value in enumerate(columns):
                    worksheet.write(0, col_num + startcol, value, header_format)
                # Set rows specific height
                if sheet_name == 'Cognates':
                    for row_num, coeff in enumerate(df['lines']):
                        if coeff > 1:
                            worksheet.set_row(row_num + 1, 14 * coeff)

        xlsx_url = ''.join([
            storage['prefix'], storage['static_route'],
            'glottochronology', '/', str(cur_time), '/', xlsx_filename])

        return xlsx_url

    @staticmethod
    def export_html(result, tiny_dicts=None, huge_size=1048576):
        result_tables = (
            build_table(result['Distances'], 'orange_light', width="300px", index=True),
            build_table(result['Cognates'], 'blue_light', width="300px").replace("\\n","<br>"),
            build_table(result['Singles'], 'green_light', width="300px"),
            build_table(result['Borrowed'], 'yellow_light', width="300px"))

        # Control output size
        spl = "<pre>\n\n</pre>"
        html_result = f"{result_tables[0]}" \
                      f"{spl}" \
                      f"{result_tables[1]}" \
                      f"{spl}" \
                      f"{result_tables[2]}" \
                      f"{spl}" \
                      f"{result_tables[3]}"

        if len(html_result) > huge_size:
            html_result = f"{result_tables[0]}" \
                          f"{spl}" \
                          f"{result_tables[1]}" \
                          f"<pre>\n\nNote: The table with single words is not shown due to huge summary size</pre>"

        if len(html_result) > huge_size:
            html_result = f"{result_tables[0]}" \
                          f"<pre>\n\nNote: The result tables with words are not shown due to huge summary size</pre>"

        html_result += ("<pre>Note: The following dictionaries contain too less words and were not processed: \n\n" +
                        '\n'.join(tiny_dicts) + "</pre>") if tiny_dicts else ""

        return html_result

    @staticmethod
    def swadesh_statistics(
            language_str,
            base_language_id,
            base_language_name,
            group_field_id,
            perspective_info_list,
            locale_id,
            storage,
            debug_flag = False):

        swadesh_list = ['я','ты','мы','этот, это','тот, то','кто','что','не','все','много','один','два','большой',
                        'долгий','маленький','женщина','мужчина','человек','рыба','птица','собака','вошь','дерево',
                        'семя','лист','корень','кора','кожа','мясо','кровь','кость','жир','яйцо','рог','хвост','перо',
                        'волосы','голова','ухо','глаз','нос','рот','зуб','язык (орган)','ноготь','нога (стопа)','колено',
                        'рука (кисть)','живот','горло','грудь','сердце','печень','пить','есть (кушать)','кусать','видеть',
                        'слышать','знать','спать','умирать','убивать','плавать','летать','гулять','приходить','лежать',
                        'сидеть','стоять','дать','сказать','солнце','луна','звезда','вода','дождь','камень','песок',
                        'земля','облако','дым','огонь','пепел','гореть','дорога,тропа','гора','красный','зелёный',
                        'жёлтый','белый','чёрный','ночь','тёплый','холодный','полный','новый','хороший','круглый',
                        'сухой','имя']

        def compare_translations(swadesh_lex, dictionary_lex):
            def split_lex(lex):
                # Split by commas and open brackets to separate
                # various forms of lexeme and extra note if is
                lex = ' '.join(lex.lower().split()) # reduce multi spaces
                if "убрать из стословника" in lex:
                    return set()

                return set(form.strip()
                           for form in lex.replace('(', ',').split(',')
                           if form.strip() and ')' not in form)  # exclude notes

            # return true if the intersection is not empty
            return bool(split_lex(swadesh_lex) & split_lex(dictionary_lex))


        # Gathering entry grouping data.

        if not debug_flag:

            _, group_list, _ = (
                CognateAnalysis.tag_data_plpgsql(
                    perspective_info_list, group_field_id))

        else:

            # If we are in debug mode, we try to load existing tag data to reduce debugging time.

            tag_data_digest = (

                hashlib.md5(

                    repr(list(group_field_id) +
                        [perspective_info[0] for perspective_info in perspective_info_list])

                    .encode('utf-8'))

                .hexdigest())

            tag_data_file_name = (
                f'__tag_data_{base_language_id[0]}_{base_language_id[1]}_{tag_data_digest}__.gz')

            # Checking if we have saved data.

            if os.path.exists(tag_data_file_name):

                with gzip.open(tag_data_file_name, 'rb') as tag_data_file:
                    _, group_list, _ = pickle.load(tag_data_file)

            else:

                # Don't have existing data, so we gather it and then save it for later use.

                r1, group_list, r3 = (

                    CognateAnalysis.tag_data_plpgsql(
                        perspective_info_list, group_field_id))

                with gzip.open(tag_data_file_name, 'wb') as tag_data_file:
                    pickle.dump((r1, group_list, r3), tag_data_file)


        # Getting text data for each perspective.
        # entries_set gathers entry_id(s) of words met in Swadesh' list
        # swadesh_total gathers numbers of words within Swadesh' list
        entries_set = {}
        swadesh_total = {}
        result_pool = {}
        tiny_dicts = set()
        for index, (perspective_id, transcription_field_id, translation_field_id) in \
                enumerate(perspective_info_list):

            # Getting and saving perspective info.
            perspective = (
                DBSession
                    .query(dbPerspective)
                    .filter_by(client_id=perspective_id[0], object_id=perspective_id[1])
                    .first()
            )
            dictionary_name = perspective.parent.get_translation(locale_id)

            # GC
            del perspective

            # Getting text data.
            transcription_query = (
                DBSession
                    .query(
                        dbLexicalEntry.client_id,
                        dbLexicalEntry.object_id)
                    .filter(
                        dbLexicalEntry.parent_client_id == perspective_id[0],
                        dbLexicalEntry.parent_object_id == perspective_id[1],
                        dbLexicalEntry.marked_for_deletion == False,
                        dbEntity.parent_client_id == dbLexicalEntry.client_id,
                        dbEntity.parent_object_id == dbLexicalEntry.object_id,
                        dbEntity.field_client_id == transcription_field_id[0],
                        dbEntity.field_object_id == transcription_field_id[1],
                        dbEntity.marked_for_deletion == False,
                        dbPublishingEntity.client_id == dbEntity.client_id,
                        dbPublishingEntity.object_id == dbEntity.object_id,
                        dbPublishingEntity.published == True,
                        dbPublishingEntity.accepted == True)
                    .add_columns(
                        func.array_agg(dbEntity.content).label('transcription'))
                    .group_by(dbLexicalEntry)
                    .subquery())

            translation_query = (
                DBSession
                    .query(
                        dbLexicalEntry.client_id,
                        dbLexicalEntry.object_id)
                    .filter(
                        dbLexicalEntry.parent_client_id == perspective_id[0],
                        dbLexicalEntry.parent_object_id == perspective_id[1],
                        dbLexicalEntry.marked_for_deletion == False,
                        dbEntity.parent_client_id == dbLexicalEntry.client_id,
                        dbEntity.parent_object_id == dbLexicalEntry.object_id,
                        dbEntity.field_client_id == translation_field_id[0],
                        dbEntity.field_object_id == translation_field_id[1],
                        dbEntity.marked_for_deletion == False,
                        dbPublishingEntity.client_id == dbEntity.client_id,
                        dbPublishingEntity.object_id == dbEntity.object_id,
                        dbPublishingEntity.published == True,
                        dbPublishingEntity.accepted == True)
                    .add_columns(
                        func.array_agg(dbEntity.content).label('translation'))
                    .group_by(dbLexicalEntry)
                    .subquery())

            # Main query for transcription/translation data.
            data_query = (
                DBSession
                    .query(transcription_query)
                    .outerjoin(translation_query, and_(
                        transcription_query.c.client_id == translation_query.c.client_id,
                        transcription_query.c.object_id == translation_query.c.object_id))
                    .add_columns(
                        translation_query.c.translation)
                    .all())

            # GC
            del transcription_query
            del translation_query

            # Grouping translations by lexical entries.
            entries_set[perspective_id] = set()
            swadesh_total[perspective_id] = set()
            result_pool[perspective_id] = {'name': dictionary_name}
            for row_index, row in enumerate(data_query):
                entry_id = tuple(row[:2])
                transcription_list, translation_list = row[2:4]

                # If we have no transcriptions for this lexical entry, we skip it altogether.
                if not transcription_list:
                    continue

                translation_list = (
                    [] if not translation_list else [
                        translation.strip()
                        for translation in translation_list
                        if translation.strip()])

                # Parsing translations and matching with Swadesh's words
                transcription_lex = ', '.join(transcription_list)
                for swadesh_num, swadesh_lex in enumerate(swadesh_list):
                    for translation_lex in translation_list:
                        if compare_translations(swadesh_lex, translation_lex):
                            # Store the entry's content in human-readable format
                            result_pool[perspective_id][entry_id] = {
                                'group': None,
                                'borrowed': (" заим." in f" {transcription_lex} {translation_lex}"),
                                'swadesh': swadesh_lex,
                                'transcription': transcription_lex,
                                'translation': translation_lex
                            }
                            # Store entry_id and number of the lex within Swadesh's list
                            entries_set[perspective_id].add(entry_id)
                            if not result_pool[perspective_id][entry_id]['borrowed']:
                                # Total list of Swadesh's words in the perspective,
                                # they can have not any etymological links
                                swadesh_total[perspective_id].add(swadesh_num)

            # Forget the dictionary if it contains less than 50 Swadesh words
            if len(swadesh_total[perspective_id]) < 50:
                del entries_set[perspective_id]
                del swadesh_total[perspective_id]
                del result_pool[perspective_id]
                tiny_dicts.add(dictionary_name)

            # GC
            del data_query

        # Checking if found entries have links
        means = collections.OrderedDict()
        for perspective_id, entries in entries_set.items():
            means[perspective_id] = collections.defaultdict(set)
            for group_index, group in enumerate(group_list):
                # Select etymologically linked entries
                linked = entries & group
                for entry_id in linked:
                    result_pool[perspective_id][entry_id]['group'] = group_index
                    swadesh = result_pool[perspective_id][entry_id]['swadesh']
                    # Store the correspondence: perspective { meanings(1/2/3) { etymological_groups(1.1/1.2/2.1/3.1)
                    if not result_pool[perspective_id][entry_id]['borrowed']:
                        means[perspective_id][swadesh].add(group_index)

        dictionary_count = len(means)
        distance_data_array = numpy.full((dictionary_count, dictionary_count), 50, dtype='float')
        complex_data_array = numpy.full((dictionary_count, dictionary_count), "n/a", dtype='object')
        distance_header_array = numpy.full(dictionary_count, "<noname>", dtype='object')

        # Calculate intersection between lists of linked meanings (Swadesh matching)
        # So length of this intersection is the similarity of corresponding perspectives
        # means_total is amount of Swadesh's lexemes met in the both perspectives
        bundles = set()
        # Calculate each-to-each distances, exclude self-to-self
        for n1, (perspective1, means1) in enumerate(means.items()):
            # Numerate dictionaries
            result_pool[perspective1]['name'] = f"{n1 + 1}. {result_pool[perspective1]['name']}"
            distance_header_array[n1] = result_pool[perspective1]['name']
            for n2, (perspective2, means2) in enumerate(means.items()):
                if n1 == n2:
                    distance_data_array[n1][n2] = 0
                    complex_data_array[n1][n2] = "n/a"
                else:
                    # Common meanings of entries which have etymological links
                    # but this links may be not mutual
                    means_common = means1.keys() & means2.keys()
                    means_linked = 0
                    # Checking if the found meanings have common links
                    for swadesh in means_common:
                        links_common = means1[swadesh] & means2[swadesh]
                        if links_common:
                            # Bundles are links with two or more entries in the result table
                            bundles.update(links_common)
                            means_linked += 1

                    means_total = len(swadesh_total[perspective1] & swadesh_total[perspective2])

                    if n2 > n1 and means_linked >= means_total:
                        log.debug(f"{n1+1},{n2+1} : "
                                  f"{len(means_common)} but {means_linked} of {means_total} : "
                                  f"{', '.join(sorted(means_common))}")

                    c = means_linked / means_total if means_total > 0 else 0
                    distance = math.sqrt( math.log(c) / -0.1 / math.sqrt(c) ) if c > 0 else 25
                    percent = means_linked * 100 // means_total if means_total > 0 else 0
                    distance_data_array[n1][n2] = round(distance, 2)
                    complex_data_array[n1][n2] = f"{distance_data_array[n1][n2]:.2f} ({percent}%)"

        result = SwadeshAnalysis.export_dataframe(result_pool, complex_data_array, bundles, SwadeshAnalysis.get_entry_text)

        # GC
        del result_pool

        xlsx_url = SwadeshAnalysis.export_xlsx(result, base_language_name, storage)

        # 'lines' field is not needed any more
        del result['Cognates']['lines']

        html_result = SwadeshAnalysis.export_html(result, tiny_dicts)

        _, mst_list, embedding_2d_pca, embedding_3d_pca = \
            CognateAnalysis.distance_graph(
                language_str,
                base_language_name,
                distance_data_array,
                distance_header_array,
                None,
                None,
                None,
                analysis_str = 'swadesh_analysis',
                __debug_flag__ = debug_flag,
                __plot_flag__ = False
            )

        result_dict = (
            dict(
                triumph = True,

                result = html_result,
                xlsx_url = xlsx_url,
                minimum_spanning_tree = mst_list,
                embedding_2d = embedding_2d_pca,
                embedding_3d = embedding_3d_pca,
                perspective_name_list = distance_header_array))

        return SwadeshAnalysis(**result_dict)

    @staticmethod
    def mutate(
        self,
        info,
        source_perspective_id,
        base_language_id,
        group_field_id,
        perspective_info_list,
        debug_flag = False):
        """
        mutation SwadeshAnalysis {
          swadesh_analysis(
            base_language_id: [508, 41],
            group_field_id: [66, 25],
            perspective_info_list: [
              [[425, 4], [66, 8], [66, 10]],
              [[1552, 1759], [66, 8], [66, 10]],
              [[418, 4], [66, 8], [66, 10]]])
          {
            triumph          }
        }
        """

        # Administrator / perspective author / editing permission check.
        error_str = (
            'Only administrator, perspective author and users with perspective editing permissions '
            'can perform Swadesh analysis.')

        client_id = info.context.request.authenticated_userid

        if not client_id:
            return ResponseError(error_str)

        user = Client.get_user_by_client_id(client_id)

        author_client_id_set = (

            set(
                client_id
                for (client_id, _), _, _ in perspective_info_list))

        author_id_check = (

            DBSession

                .query(

                    DBSession
                        .query(literal(1))
                        .filter(
                            Client.id.in_(author_client_id_set),
                            Client.user_id == user.id)
                        .exists())

                .scalar())

        if (user.id != 1 and
            not author_id_check and
            not info.context.acl_check_if('edit', 'perspective', source_perspective_id)):

            return ResponseError(error_str)

        # Debug mode check.

        if debug_flag and user.id != 1:

            return (

                ResponseError(
                    message = 'Only administrator can use debug mode.'))

        language_str = (
            '{0}/{1}, language {2}/{3}'.format(
                source_perspective_id[0], source_perspective_id[1],
                base_language_id[0], base_language_id[1]))

        try:

            # Getting base language info.

            locale_id = info.context.get('locale_id') or 2

            base_language = DBSession.query(dbLanguage).filter_by(
                client_id = base_language_id[0], object_id = base_language_id[1]).first()

            base_language_name = base_language.get_translation(locale_id)

            request = info.context.request
            storage = request.registry.settings['storage']

            # Transforming client/object pair ids from lists to 2-tuples.

            source_perspective_id = tuple(source_perspective_id)
            base_language_id = tuple(base_language_id)
            group_field_id = tuple(group_field_id)

            perspective_info_list = [

                (tuple(perspective_id),
                    tuple(transcription_field_id),
                    tuple(translation_field_id))

                for perspective_id,
                    transcription_field_id,
                    translation_field_id in perspective_info_list]

            return SwadeshAnalysis.swadesh_statistics(
                language_str,
                base_language_id,
                base_language_name,
                group_field_id,
                perspective_info_list,
                locale_id,
                storage,
                debug_flag)

        # Exception occured while we tried to perform swadesh analysis.
        except Exception as exception:

            traceback_string = ''.join(traceback.format_exception(
                exception, exception, exception.__traceback__))[:-1]

            log.warning(
                'swadesh_analysis {0}: exception'.format(
                language_str))

            log.warning(traceback_string)

            return ResponseError(message =
                'Exception:\n' + traceback_string)


class MorphCognateAnalysis(graphene.Mutation):
    class Arguments:

        source_perspective_id = LingvodocID(required = True)
        base_language_id = LingvodocID(required = True)

        group_field_id = LingvodocID(required = True)
        perspective_info_list = graphene.List(graphene.List(LingvodocID), required = True)

        debug_flag = graphene.Boolean()

    triumph = graphene.Boolean()

    result = graphene.String()
    xlsx_url = graphene.String()
    minimum_spanning_tree = graphene.List(graphene.List(graphene.Int))
    embedding_2d = graphene.List(graphene.List(graphene.Float))
    embedding_3d = graphene.List(graphene.List(graphene.Float))
    perspective_name_list = graphene.List(graphene.String)

    @staticmethod
    def get_entry_text(entry):
        return f"{'; '.join(entry['affix'])} ( {'; '.join(entry['meaning'])} )"

    @staticmethod
    def morph_cognate_statistics(
            language_str,
            base_language_id,
            base_language_name,
            group_field_id,
            perspective_info_list,
            locale_id,
            storage,
            debug_flag = False):

        # Gathering entry grouping data.

        if not debug_flag:

            _, group_list, _ = (
                CognateAnalysis.tag_data_plpgsql(
                    perspective_info_list, group_field_id))

        else:

            # If we are in debug mode, we try to load existing tag data to reduce debugging time.

            tag_data_digest = (
                hashlib.md5(
                    repr(list(group_field_id) +
                        [perspective_info[0] for perspective_info in perspective_info_list])
                    .encode('utf-8'))
                .hexdigest())

            tag_data_file_name = (
                f'__tag_data_{base_language_id[0]}_{base_language_id[1]}_{tag_data_digest}__.gz')

            # Checking if we have saved data.

            if os.path.exists(tag_data_file_name):

                with gzip.open(tag_data_file_name, 'rb') as tag_data_file:
                    _, group_list, _ = pickle.load(tag_data_file)

            else:

                # Don't have existing data, so we gather it and then save it for later use.

                r1, group_list, r3 = (

                    CognateAnalysis.tag_data_plpgsql(
                        perspective_info_list, group_field_id))

                with gzip.open(tag_data_file_name, 'wb') as tag_data_file:
                    pickle.dump((r1, group_list, r3), tag_data_file)

        # Getting text data for each perspective.
        to_canon_meaning = collections.defaultdict(dict)
        meaning_to_links = {}
        result_pool = {}
        tiny_dicts = set()
        clean_meaning_re = re.compile('[.\dA-Z]+')
        for index, (perspective_id, affix_field_id, meaning_field_id) in \
                enumerate(perspective_info_list):

            # Getting and saving perspective info.
            perspective = (
                DBSession
                    .query(dbPerspective)
                    .filter_by(client_id=perspective_id[0], object_id=perspective_id[1])
                    .first()
            )
            dictionary_name = perspective.parent.get_translation(locale_id)

            # GC
            del perspective

            # Getting text data.
            affix_query = (
                DBSession
                    .query(
                        dbLexicalEntry.client_id,
                        dbLexicalEntry.object_id)
                    .filter(
                        dbLexicalEntry.parent_client_id == perspective_id[0],
                        dbLexicalEntry.parent_object_id == perspective_id[1],
                        dbLexicalEntry.marked_for_deletion == False,
                        dbEntity.parent_client_id == dbLexicalEntry.client_id,
                        dbEntity.parent_object_id == dbLexicalEntry.object_id,
                        dbEntity.field_client_id == affix_field_id[0],
                        dbEntity.field_object_id == affix_field_id[1],
                        dbEntity.marked_for_deletion == False,
                        dbPublishingEntity.client_id == dbEntity.client_id,
                        dbPublishingEntity.object_id == dbEntity.object_id,
                        dbPublishingEntity.published == True,
                        dbPublishingEntity.accepted == True)
                    .add_columns(
                        func.array_agg(dbEntity.content).label('affix'))
                    .group_by(dbLexicalEntry)
                    .subquery())

            meaning_query = (
                DBSession
                    .query(
                        dbLexicalEntry.client_id,
                        dbLexicalEntry.object_id)
                    .filter(
                        dbLexicalEntry.parent_client_id == perspective_id[0],
                        dbLexicalEntry.parent_object_id == perspective_id[1],
                        dbLexicalEntry.marked_for_deletion == False,
                        dbEntity.parent_client_id == dbLexicalEntry.client_id,
                        dbEntity.parent_object_id == dbLexicalEntry.object_id,
                        dbEntity.field_client_id == meaning_field_id[0],
                        dbEntity.field_object_id == meaning_field_id[1],
                        dbEntity.marked_for_deletion == False,
                        dbPublishingEntity.client_id == dbEntity.client_id,
                        dbPublishingEntity.object_id == dbEntity.object_id,
                        dbPublishingEntity.published == True,
                        dbPublishingEntity.accepted == True)
                    .add_columns(
                        func.array_agg(dbEntity.content).label('meaning'))
                    .group_by(dbLexicalEntry)
                    .subquery())

            # Main query for transcription/translation data.
            data_query = (
                DBSession
                    .query(affix_query)
                    .outerjoin(meaning_query, and_(
                        affix_query.c.client_id == meaning_query.c.client_id,
                        affix_query.c.object_id == meaning_query.c.object_id))
                    .add_columns(
                        meaning_query.c.meaning)
                    .all())

            # GC
            del affix_query
            del meaning_query

            meaning_to_links[perspective_id] = {}
            result_pool[perspective_id] = {'name': dictionary_name}

            for row in data_query:
                entry_id = tuple(row[:2])

                if not (affix_list := row[2]) or not (meaning_list := row[3]):
                    continue

                affix = list(map(lambda a: a.strip(), affix_list))
                meaning = []
                for m in meaning_list:
                    if clean_m := re.search(clean_meaning_re, m):
                        meaning.append(clean_m.group(0))

                if not meaning:
                    continue

                # Compounding a dictionary to convert every meaning to the first one within each row
                # Initialize group of links for every sub_meaning to have a full set of sub_meanings
                for sub_meaning in meaning:
                    to_canon_meaning[perspective_id][sub_meaning] = meaning[0]
                    meaning_to_links[perspective_id][sub_meaning] = set()

                # Grouping affixes and meanings by lexical entries.
                result_pool[perspective_id][entry_id] = {
                    'group': None,
                    'affix': affix,
                    'meaning': meaning
                }

            # Forget the dictionary if it contains less than 30 sub-meanings
            if len(to_canon_meaning[perspective_id]) < 30:
                del meaning_to_links[perspective_id]
                del result_pool[perspective_id]
                tiny_dicts.add(dictionary_name)

            # GC
            del data_query

        # Checking if found entries have links
        for perspective_id, entries in result_pool.items():
            for group_index, group in enumerate(group_list):
                # Select etymologically linked entries
                linked = entries.keys() & group
                for entry_id in linked:
                    result_pool[perspective_id][entry_id]['group'] = group_index
                    meaning = result_pool[perspective_id][entry_id]['meaning']
                    for sub_meaning in meaning:
                        meaning_to_links[perspective_id][sub_meaning].add(group_index)

        dictionary_count = len(result_pool)
        distance_data_array = numpy.full((dictionary_count, dictionary_count), 50, dtype='float')
        complex_data_array = numpy.full((dictionary_count, dictionary_count), "n/a", dtype='object')
        distance_header_array = numpy.full(dictionary_count, "<noname>", dtype='object')

        bundles = set()
        # Calculate each-to-each distances, exclude self-to-self
        for n1, (perspective1, meaning_to_links1) in enumerate(meaning_to_links.items()):
            # Numerate dictionaries
            result_pool[perspective1]['name'] = f"{n1 + 1}. {result_pool[perspective1]['name']}"
            distance_header_array[n1] = result_pool[perspective1]['name']

            to_canon_meaning1 = to_canon_meaning[perspective1]
            canon_meanings1_set = set(to_canon_meaning1.values())

            for n2, (perspective2, meaning_to_links2) in enumerate(meaning_to_links.items()):
                if n1 == n2:
                    distance_data_array[n1][n2] = 0
                    complex_data_array[n1][n2] = "n/a"
                else:
                    # Compile new meaning_to_links2 using canon_meanings instead of sub_meanings
                    canon_meaning_to_links2 = collections.defaultdict(set)
                    for sub_meaning, links in meaning_to_links2.items():
                        if canon_meaning := to_canon_meaning1.get(sub_meaning):
                            canon_meaning_to_links2[canon_meaning].update(links)

                    # Common canonical meanings of perspective1 and perspective2
                    meanings_common = canon_meanings1_set & canon_meaning_to_links2.keys()
                    meanings_total = len(meanings_common)

                    meanings_linked = 0
                    # Checking if the found meanings have common links
                    for meaning in meanings_common:
                        links_common = meaning_to_links1[meaning] & canon_meaning_to_links2[meaning]
                        if links_common:
                            # Bundles are links with two or more entries in the result table
                            bundles.update(links_common)
                            meanings_linked += 1

                    '''
                    if debug_flag and n2 > n1 and meanings_linked >= meanings_total:
                        log.debug(f"{n1+1},{n2+1} : "
                                  f"{len(meanings_common)} but {meanings_linked} of {meanings_total} : "
                                  f"{', '.join(sorted(meanings_common))}")
                    '''

                    # meanings_linked > 0 meanings that meanings_total > 0 even more so
                    distance = math.log(meanings_linked / meanings_total) / -0.14 if meanings_linked > 0 else 50
                    singles = meanings_total - meanings_linked
                    percent = meanings_linked * 100 // meanings_total if meanings_total > 0 else 0
                    distance_data_array[n1][n2] = round(distance, 2)
                    complex_data_array[n1][n2] = f"{distance_data_array[n1][n2]:.2f} ({percent}%)"

        result = SwadeshAnalysis.export_dataframe(result_pool, complex_data_array, bundles, MorphCognateAnalysis.get_entry_text)

        # GC
        del result_pool

        xlsx_url = SwadeshAnalysis.export_xlsx(result, base_language_name, storage)

        # 'lines' field is not needed any more
        del result['Cognates']['lines']

        html_result = SwadeshAnalysis.export_html(result, tiny_dicts)

        _, mst_list, embedding_2d_pca, embedding_3d_pca = \
            CognateAnalysis.distance_graph(
                language_str,
                base_language_name,
                distance_data_array,
                distance_header_array,
                None,
                None,
                None,
                analysis_str = 'morph_cognate_analysis',
                __debug_flag__ = debug_flag,
                __plot_flag__ = False
            )

        result_dict = (
            dict(
                triumph = True,

                result = html_result,
                xlsx_url = xlsx_url,
                minimum_spanning_tree = mst_list,
                embedding_2d = embedding_2d_pca,
                embedding_3d = embedding_3d_pca,
                perspective_name_list = distance_header_array))

        return MorphCognateAnalysis(**result_dict)

    @staticmethod
    def mutate(
        self,
        info,
        source_perspective_id,
        base_language_id,
        group_field_id,
        perspective_info_list,
        debug_flag = False):
        """
        mutation MorphCognateAnalysis {
          morph_cognate_analysis(
            base_language_id: [508, 41],
            group_field_id: [66, 25],
            perspective_info_list: [
              [[425, 4], [66, 8], [66, 10]],
              [[1552, 1759], [66, 8], [66, 10]],
              [[418, 4], [66, 8], [66, 10]]])
          {
            triumph          }
        }
        """

        # Administrator / perspective author / editing permission check.
        error_str = (
            'Only administrator, perspective author and users with perspective editing permissions '
            'can perform Swadesh analysis.')

        client_id = info.context.request.authenticated_userid

        if not client_id:
            return ResponseError(error_str)

        user = Client.get_user_by_client_id(client_id)

        author_client_id_set = (

            set(
                client_id
                for (client_id, _), _, _ in perspective_info_list))

        author_id_check = (

            DBSession

                .query(

                    DBSession
                        .query(literal(1))
                        .filter(
                            Client.id.in_(author_client_id_set),
                            Client.user_id == user.id)
                        .exists())

                .scalar())

        if (user.id != 1 and
            not author_id_check and
            not info.context.acl_check_if('edit', 'perspective', source_perspective_id)):

            return ResponseError(error_str)

        # Debug mode check.

        if debug_flag and user.id != 1:

            return (

                ResponseError(
                    message = 'Only administrator can use debug mode.'))

        language_str = (
            '{0}/{1}, language {2}/{3}'.format(
                source_perspective_id[0], source_perspective_id[1],
                base_language_id[0], base_language_id[1]))

        try:

            # Getting base language info.

            locale_id = info.context.get('locale_id') or 2

            base_language = DBSession.query(dbLanguage).filter_by(
                client_id = base_language_id[0], object_id = base_language_id[1]).first()

            base_language_name = base_language.get_translation(locale_id)

            request = info.context.request
            storage = request.registry.settings['storage']

            # Transforming client/object pair ids from lists to 2-tuples.

            source_perspective_id = tuple(source_perspective_id)
            base_language_id = tuple(base_language_id)
            group_field_id = tuple(group_field_id)

            perspective_info_list = [

                (tuple(perspective_id),
                    tuple(affix_field_id),
                    tuple(meaning_field_id))

                for perspective_id,
                    affix_field_id,
                    meaning_field_id in perspective_info_list]

            return MorphCognateAnalysis.morph_cognate_statistics(
                language_str,
                base_language_id,
                base_language_name,
                group_field_id,
                perspective_info_list,
                locale_id,
                storage,
                debug_flag)

        # Exception occured while we tried to perform swadesh analysis.
        except Exception as exception:

            traceback_string = ''.join(traceback.format_exception(
                exception, exception, exception.__traceback__))[:-1]

            log.warning(
                'morph_cognate_analysis {0}: exception'.format(
                language_str))

            log.warning(traceback_string)

            return ResponseError(message =
                'Exception:\n' + traceback_string)


class Phonology(graphene.Mutation):

    class Arguments:
        perspective_id=LingvodocID(required=True)
        limit=graphene.Int()
        limit_exception=graphene.Int()
        limit_no_vowel=graphene.Int()
        limit_result=graphene.Int()
        no_cache=graphene.Boolean()
        interval_only=graphene.Boolean()
        group_by_description=graphene.Boolean(required=True)
        maybe_translation_field=LingvodocID()
        only_first_translation=graphene.Boolean(required=True)
        vowel_selection=graphene.Boolean(required=True)
        maybe_tier_list=graphene.List(graphene.String)
        keep_list=graphene.List(graphene.Int)
        join_list=graphene.List(graphene.Int)
        chart_threshold=graphene.Int()
        generate_csv=graphene.Boolean()
        link_field_list=graphene.List(LingvodocID)
        link_perspective_list=graphene.List(graphene.List(LingvodocID))
        use_fast_track=graphene.Boolean()
        synchronous=graphene.Boolean()
        debug_flag=graphene.Boolean()

    triumph = graphene.Boolean()

    @staticmethod
    # @client_id_check()

    def mutate(self, info, **args):
        """
        query MyQuery {
          phonology(
            perspective_id: [126, 5],
            group_by_description: false,
            only_first_translation: false,
            vowel_selection: false,
            maybe_tier_list: [],
            maybe_translation_field: [66, 19])
        }
        """

        parameters = (
            Phonology_Parameters.from_graphql(args))

        parameters.__debug_flag__ = (
            args.get('debug_flag', False))

        locale_id = info.context.get('locale_id')
        request = info.context.get('request')

        utils_phonology(request, locale_id, parameters)

        return Phonology(triumph=True)


@celery.task
def async_phonological_statistical_distance(
    id_list,
    vowel_selection,
    chart_threshold,
    locale_id,
    storage,
    task_key,
    cache_kwargs,
    sqlalchemy_url,
    __debug_flag__):
    """
    Sets up and launches cognate analysis in asynchronous mode.
    """

    # NOTE: copied from phonology.
    #
    # As far as we know, this is a no-op with current settings, we use it to enable logging inside celery
    # tasks, because somehow this does it, and otherwise we couldn't set it up.

    logging.debug('async_phonological_statistical_distance')

    # Ok, and now we go on with task execution.

    engine = create_engine(sqlalchemy_url)
    DBSession.configure(bind = engine)
    initialize_cache(cache_kwargs)
    global CACHE
    from lingvodoc.cache.caching import CACHE
    task_status = TaskStatus.get_from_cache(task_key)

    with transaction.manager:

        try:
            PhonologicalStatisticalDistance.perform_phonological_statistical_distance(
                id_list,
                vowel_selection,
                chart_threshold,
                locale_id,
                storage,
                task_status,
                __debug_flag__)

        # Some unknown external exception.

        except Exception as exception:

            traceback_string = ''.join(traceback.format_exception(
                exception, exception, exception.__traceback__))[:-1]

            log.warning('phonological_statistical_distance: exception')
            log.warning(traceback_string)

            if task_status is not None:

                task_status.set(1, 100,
                    'Finished (ERROR), exception:\n' + traceback_string)


class PhonologicalStatisticalDistance(graphene.Mutation):

    class Arguments:

        id_list = graphene.List(LingvodocID, required = True)
        vowel_selection = graphene.Boolean(required = True)
        chart_threshold = graphene.Int()

        synchronous_flag = graphene.Boolean()
        debug_flag = graphene.Boolean()

    triumph = graphene.Boolean()

    def perform_phonological_statistical_distance(
        id_list,
        vowel_selection,
        chart_threshold,
        locale_id,
        storage,
        task_status,
        __debug_flag__):
        """
        Performs phonological statistical distance computation in either synchronous or asynchronous mode.
        """

        Markup = aliased(dbEntity, name = 'Markup')
        Sound = aliased(dbEntity, name = 'Sound')

        PublishingMarkup = aliased(dbPublishingEntity, name = 'PublishingMarkup')
        PublishingSound = aliased(dbPublishingEntity, name = 'PublishingSound')

        info_list = []
        total_count = 0

        # Getting preliminary perspective info.

        for perspective_id in id_list:

            perspective = DBSession.query(dbPerspective).filter_by(
                client_id = perspective_id[0], object_id = perspective_id[1]).first()

            dictionary = perspective.parent

            name = '{0} - {1}'.format(
                dictionary.get_translation(locale_id),
                perspective.get_translation(locale_id))

            query = DBSession.query(
                Markup, Sound).filter(
                    dbLexicalEntry.parent_client_id == perspective_id[0],
                    dbLexicalEntry.parent_object_id == perspective_id[1],
                    dbLexicalEntry.marked_for_deletion == False,
                    Markup.parent_client_id == dbLexicalEntry.client_id,
                    Markup.parent_object_id == dbLexicalEntry.object_id,
                    Markup.marked_for_deletion == False,
                    Markup.additional_metadata.contains({'data_type': 'praat markup'}),
                    PublishingMarkup.client_id == Markup.client_id,
                    PublishingMarkup.object_id == Markup.object_id,
                    PublishingMarkup.published == True,
                    PublishingMarkup.accepted == True,
                    Sound.client_id == Markup.self_client_id,
                    Sound.object_id == Markup.self_object_id,
                    Sound.marked_for_deletion == False,
                    PublishingSound.client_id == Sound.client_id,
                    PublishingSound.object_id == Sound.object_id,
                    PublishingSound.published == True,
                    PublishingSound.accepted == True)

            count = query.count()

            info_list.append((perspective, dictionary, name, query, count))
            total_count += count

        # Showing what we've got.

        log.debug(
            'phonological_statistical_distance ({0} perspectives): {1} sound/markup pairs\n{2}'.format(
            len(id_list),
            total_count,
            '\n'.join('{0}: {1} sound/markup pairs'.format(name, count)
                for p, d, name, q, count in info_list)))

        # And now we are going to gather formant distribution data.

        formant_data_list = []
        row_count = 0

        x_min, x_max = None, None
        y_min, y_max = None, None

        no_index_list = []
        yes_index_list = []

        for perspective_index, (perspective, dictionary, name, query, count) in enumerate(info_list):

            formant_list = []
            vowel_counter = collections.Counter()

            for row_index, row in enumerate(query.yield_per(100)):

                # Showing what sound/markup pair we are to process.

                sound_id = (row.Sound.client_id, row.Sound.object_id)
                sound_url = row.Sound.content

                markup_id = (row.Markup.client_id, row.Markup.object_id)
                markup_url = row.Markup.content

                log_str = (
                    '\nphonological_statistical_distance ({0}, {1}): '
                    'sound entity {2}/{3}, markup entity {4}/{5}'.format(
                        perspective_index, row_index,
                        sound_id[0], sound_id[1],
                        markup_id[0], markup_id[1]))

                # Getting phonology data of another sound/markup pair.

                textgrid_result_list = (

                    process_sound_markup(
                        log_str,
                        sound_id,
                        sound_url,
                        markup_id,
                        markup_url,
                        storage,
                        __debug_flag__))

                row_count += 1

                if task_status is not None:

                    task_status.set(1,
                        int(math.floor(95.0 * row_count / total_count)),
                        'Gathering formant distribution data')

                # Extracting F1/F2 formant data.

                if textgrid_result_list is None:
                    continue

                for tier_number, tier_name, tier_result_list in textgrid_result_list:

                    if tier_result_list == 'no_vowel' or tier_result_list == 'no_vowel_selected':
                        continue

                    for tier_result in tier_result_list:

                        if vowel_selection:

                            # Either only for longest interval and interval with highest intensity, or...

                            f_list_a = tuple(map(float, tier_result.max_length_f_list[:2]))
                            f_list_b = tuple(map(float, tier_result.max_intensity_f_list[:2]))

                            text_a_list = tier_result.max_length_str.split()
                            text_b_list = tier_result.max_intensity_str.split()

                            vowel_a = get_vowel_class(
                                tier_result.max_length_source_index, tier_result.source_interval_list)

                            vowel_b = get_vowel_class(
                                tier_result.max_intensity_source_index, tier_result.source_interval_list)

                            formant_list.append(f_list_a)
                            vowel_counter[vowel_a] += 1

                            if text_b_list[2] != text_a_list[2]:

                                formant_list.append(f_list_b)
                                vowel_counter[vowel_b] += 1

                        else:

                            # ...for all intervals.

                            for (interval_str, interval_r_length, i_list, f_list,
                                sign_longest, sign_highest, source_index) in tier_result.interval_data_list:

                                formant_list.append(tuple(map(float, f_list[:2])))

                                vowel_counter[get_vowel_class(
                                    source_index, tier_result.source_interval_list)] += 1

            # If we do not have enough formant data, we skip modelling of this perspective.

            if len(formant_list) < 2:

                formant_data_list.append((vowel_counter, formant_list, None))
                no_index_list.append(perspective_index)

                log.debug(
                    '\nphonological_statistical_distance:'
                    '\nperspective {0} ({1}, {2}/{3}):'
                    '\n{4} vowels, {5} formant vectors'
                    '\nno model, not enough data'.format(
                    perspective_index,
                    name,
                    perspective.client_id,
                    perspective.object_id,
                    len(vowel_counter),
                    len(formant_list)))

                continue

            yes_index_list.append(perspective_index)

            # And now we are going to model the distribution as a gaussian mixture.

            formant_array = numpy.array(formant_list)
            n_components = min(len(vowel_counter) * 2, len(formant_list))

            mixture_kwargs = {
                'n_components': n_components,
                'weight_concentration_prior_type': 'dirichlet_distribution',
                'weight_concentration_prior': 1.0,
                'reg_covar': 1e-9,
                'max_iter': 1024}

            cache_key = 'psd.' + (
                base64.b85encode(hashlib.md5(pickle.dumps(
                    (formant_array, mixture_kwargs)))
                        .digest()).decode('ascii'))

            cache_result = caching.CACHE.get(cache_key)

            # If we have cached model, we use it, otherwise we build it anew.

            if cache_result:

                mixture, component_count = cache_result

            else:

                mixture = sklearn.mixture.BayesianGaussianMixture(
                    verbose = 2 if __debug_flag__ else 0,
                    **mixture_kwargs)

                mixture.fit(formant_array)

                min_weight = min(mixture.weights_)

                component_where = mixture.weights_ > min_weight
                component_count = component_where.sum()

                # Removing any non-significant components, if required.

                if component_count < n_components - 2:

                    for name in [
                        'weights_',
                        'weight_concentration_',
                        'means_',
                        'covariances_',
                        'precisions_cholesky_',
                        'degrees_of_freedom_',
                        'mean_precision_']:

                        setattr(mixture, name,
                            getattr(mixture, name)[component_where])

                    mixture.weights_ /= sum(mixture.weights_)

                # Caching newly built model.

                caching.CACHE.set(cache_key,
                    (mixture, component_count))

            # Updating distribution data.

            if perspective_index <= 0:

                x_min, x_max = formant_array[:,0].min(), formant_array[:,0].max()
                y_min, y_max = formant_array[:,1].min(), formant_array[:,1].max()

            else:

                x_min = min(x_min, formant_array[:,0].min())
                x_max = max(x_max, formant_array[:,0].max())

                y_min = min(y_min, formant_array[:,1].min())
                y_max = max(y_max, formant_array[:,1].max())

            formant_data_list.append((vowel_counter, formant_array, mixture))

            # Another distribution of F1/F2 formant vectors we've got.

            log.debug(
                '\nphonological_statistical_distance:'
                '\nperspective {0} ({1}, {2}/{3}):'
                '\n{4} vowels, {5} formant vectors'
                '\n{6}'
                '\n{7}model with {8} / {9} significant components'.format(
                perspective_index,
                name,
                perspective.client_id,
                perspective.object_id,
                len(vowel_counter),
                len(formant_list),
                pprint.pformat(vowel_counter, width = 144),
                'CACHEd ' if cache_result else '',
                component_count if component_count <= n_components - 2 else n_components,
                n_components))

            if task_status is not None:

                task_status.set(1,
                    int(math.floor(95.0 + 2.0 * (perspective_index + 1) / len(info_list))),
                    'Creating formant distribution models')

        # Preparing for compilation of modelling results.

        workbook_stream = io.BytesIO()

        workbook = xlsxwriter.Workbook(
            workbook_stream, {'in_memory': True})

        worksheet_distance = (
            workbook.add_worksheet(
                utils.sanitize_worksheet_name('Distance')))

        worksheet_list = []

        # Getting integrable density grids for each distribution model.

        x_min = x_min or 500
        x_max = x_max or 1000

        y_min = y_min or 500
        y_max = y_max or 1000

        x_low = max(0, x_min - (x_max - x_min) * 0.25)
        x_high = x_max + (x_max - x_min) * 0.25

        y_low = max(0, y_min - (y_max - y_min) * 0.25)
        y_high = y_max + (y_max - y_min) * 0.25

        n_step = 1024

        x, y = pylab.meshgrid(
            numpy.linspace(x_low, x_high, n_step),
            numpy.linspace(y_low, y_high, n_step))

        # Showing grid parameters.

        log.debug(
            '\nphonological_statistical_distance: '
            '\nmin/max: {0:.3f}, {1:.3f}, {2:.3f}, {3:.3f}'
            '\nlow/high: {4:.3f}, {5:.3f}, {6:.3f}, {7:.3f}'
            '\n{8} steps'.format(
            x_min, x_max, y_min, y_max,
            x_low, x_high, y_low, y_high,
            n_step))

        grid_data_list = []

        for perspective_index, (
            (perspective, dictionary, name, query, count),
            (vowel_counter, formant_array, mixture)) in (

            enumerate(zip(info_list, formant_data_list))):

            # Checking if we have no data for this perspective.

            if mixture is None:

                grid_data_list.append(None)
                continue

            z = None

            worksheet_list.append(
                workbook.add_worksheet(
                    utils.sanitize_worksheet_name(
                        'Figure {0}'.format(len(worksheet_list) + 1))))

            # If we are in debug mode, we try to load saved grid data we might have.

            if __debug_flag__:

                z_digest = (
                    hashlib.md5(pickle.dumps(
                        (formant_array, mixture_kwargs, x_low, x_high, y_low, y_high, n_step)))
                            .hexdigest())

                z_file_name = (
                    '__grid_data_{0}_{1}_{2}__.gz'.format(
                        perspective.client_id, perspective.object_id, z_digest))

                if os.path.exists(z_file_name):

                    with gzip.open(z_file_name, 'rb') as z_file:
                        z, elapsed_time, integral, check = pickle.load(z_file)

            # Computing density grid data.

            if z is None:

                x.shape = (n_step ** 2,)
                y.shape = (n_step ** 2,)

                start_time = time.time()

                z = numpy.exp(mixture.score_samples(numpy.stack((x, y), 1)))
                z.shape = (n_step, n_step)

                elapsed_time = time.time() - start_time

                x.shape = (n_step, n_step)
                y.shape = (n_step, n_step)

                # Integrating its interpolation via Simpson's rule, normalizing if required.

                integral = scipy.integrate.simps(scipy.integrate.simps(z, x), y[:,0])
                z /= integral

                check = scipy.integrate.simps(scipy.integrate.simps(z, x), y[:,0])

                if __debug_flag__:

                    with gzip.open(z_file_name, 'wb') as z_file:
                        pickle.dump((z, elapsed_time, integral, check), z_file)

            # And we have another density grid data.

            log.debug(
                '\nphonological_statistical_distance:'
                '\nperspective {0} ({1}, {2}/{3}):'
                '\n{4} components, {5} ** 2 ({6}) samples in {7:.3f}s'
                '\nintegral of interpolation {8:.6f}, {8:.6f} after normalization'.format(
                perspective_index, name, perspective.client_id, perspective.object_id,
                len(mixture.weights_), n_step, n_step ** 2, elapsed_time,
                integral, check))

            grid_data_list.append(z)

            # Preparing figures with modelling results.

            figure = pyplot.figure(figsize = (16, 10))

            axes = figure.add_subplot(111)
            axes.set_title(name, fontsize = 14, family = 'Gentium')

            axes.invert_xaxis()
            axes.invert_yaxis()

            # F1/F2 formant data points.

            axes.scatter(
                formant_array[:,1],
                formant_array[:,0],
                marker = 'o', color = 'black', s = 4)

            x_lim = axes.get_xlim()
            y_lim = axes.get_ylim()

            x_lim = (min(x_lim[0], y_high), max(x_lim[1], y_low))
            y_lim = (min(y_lim[0], x_high), max(y_lim[1], x_low))

            # Probability density heat map.

            axes.imshow(z.T[::-1,:],
                aspect = 'auto',
                extent = (y_low, y_high, x_low, x_high),
                cmap = 'YlOrRd',
                alpha = 0.5)

            axes.set_xlim(x_lim)
            axes.set_ylim(y_lim)

            # Plotting standard deviation ellipses, see
            # https://scikit-learn.org/stable/auto_examples/mixture/plot_concentration_prior.html.

            for n in range(mixture.means_.shape[0]):

                eig_vals, eig_vecs = numpy.linalg.eigh(mixture.covariances_[n])
                unit_eig_vec = eig_vecs[0] / numpy.linalg.norm(eig_vecs[0])

                angle = numpy.arctan2(unit_eig_vec[1], unit_eig_vec[0])
                angle = 180 * angle / numpy.pi

                eig_vals = 2 * numpy.sqrt(2) * numpy.sqrt(eig_vals)

                ellipse = (
                    matplotlib.patches.Ellipse(
                        mixture.means_[n][::-1],
                        eig_vals[1],
                        eig_vals[0],
                        180 - angle,
                        edgecolor = 'black'))

                ellipse.set_alpha(max(mixture.weights_[n], 0.0625))
                ellipse.set_facecolor('#E0E0E0')

                axes.add_artist(ellipse)

            # Saving figure as a PNG stream.

            pyplot.setp(axes.texts, family = 'Gentium')

            pyplot.xticks(fontsize = 14, family = 'Gentium')
            pyplot.yticks(fontsize = 14, family = 'Gentium')

            figure_stream = io.BytesIO()

            figure.savefig(
                figure_stream,
                bbox_inches = 'tight',
                pad_inches = 0.25,
                format = 'png')

            # Adding figure to the workbook.

            figure_file_name = (
                'figure phonology distribution 2d {0} {1}.png'.format(
                perspective.client_id,
                perspective.object_id))

            worksheet_list[-1].insert_image(
                'A1', figure_file_name, {'image_data': figure_stream})

            if __debug_flag__:

                figure_stream.seek(0)

                with open(figure_file_name, 'wb') as figure_file:
                    shutil.copyfileobj(figure_stream, figure_file)

            # And now 3d with the density plot.

            figure_3d = pyplot.figure()
            figure_3d.set_size_inches(16, 10)

            axes_3d = figure_3d.add_subplot(111, projection = '3d')
            axes_3d.set_title(name, fontsize = 14, family = 'Gentium')

            axes_3d.autoscale(tight = True)
            axes_3d.autoscale_view(tight = True)

            axes_3d.invert_xaxis()
            axes_3d.view_init(elev = 40, azim = -165)

            axes_3d.scatter(
                formant_array[:,0],
                formant_array[:,1],
                numpy.zeros(formant_array.shape[0]),
                color = 'black', s = 4, depthshade = False)

            # Plotting density.

            x_where = numpy.logical_and(x[0,:] >= y_lim[1], x[0,:] <= y_lim[0])
            y_where = numpy.logical_and(y[:,0] >= x_lim[1], y[:,0] <= x_lim[0])

            axes_3d.plot_surface(
                x[:,x_where][y_where,:],
                y[:,x_where][y_where,:],
                z[:,x_where][y_where,:],
                rstride = 8, cstride = 8,
                color = '#56B4E9', alpha = 0.5)

            axes_3d.set_xlim(y_lim)
            axes_3d.set_ylim((x_lim[1], x_lim[0]))

            grid_sum = z.sum() * (x_high - x_low) * (y_high - y_low) / n_step ** 2

            log.debug('\ndensity grid sum: {0:.6f}'.format(grid_sum))

            # Saving 3d modelling figure as a PNG stream.

            pyplot.setp(axes_3d.texts, family = 'Gentium')

            pyplot.xticks(fontsize = 14, family = 'Gentium')
            pyplot.yticks(fontsize = 14, family = 'Gentium')

            for label in axes_3d.get_zticklabels():

                label.set_fontsize(14)
                label.set_family('Gentium')

            figure_3d_stream = io.BytesIO()

            figure_3d.savefig(
                figure_3d_stream,
                bbox_inches = 'tight',
                pad_inches = 0.25,
                format = 'png')

            # Adding figure to the workbook.

            figure_3d_file_name = (
                'figure phonology distribution 3d {0} {1}.png'.format(
                perspective.client_id,
                perspective.object_id))

            worksheet_list[-1].insert_image(
                'A44', figure_3d_file_name, {'image_data': figure_3d_stream})

            if __debug_flag__:

                figure_3d_stream.seek(0)

                with open(figure_3d_file_name, 'wb') as figure_3d_file:
                    shutil.copyfileobj(figure_3d_stream, figure_3d_file)

            # Updating task status, if required.

            if task_status is not None:

                task_status.set(1,
                    int(math.floor(97.0 + 2.0 * (perspective_index + 1) / len(info_list))),
                    'Computing formant distribution density data')

        # And now we should compute pairwise total variation statistical distances between computed
        # approximations of formant distributions.

        d_ij = numpy.zeros((len(yes_index_list), len(yes_index_list)))

        for i in range(len(yes_index_list) - 1):
            for j in range(i + 1, len(yes_index_list)):

                index_i = yes_index_list[i]
                index_j = yes_index_list[j]

                perspective_i, name_i = info_list[index_i][0], info_list[index_i][2]
                perspective_j, name_j = info_list[index_j][0], info_list[index_j][2]

                z_i = grid_data_list[index_i]
                z_j = grid_data_list[index_j]

                delta_abs = numpy.abs(z_i - z_j)

                d_value = (0.5 *
                    scipy.integrate.simps(
                        scipy.integrate.simps(delta_abs, x), y[:,0]))

                # Saving and showing another statistical distance.

                d_ij[i,j] = d_value
                d_ij[j,i] = d_value

                log.debug(
                    '\nphonological_statistical_distance:'
                    '\nperspective {0} ({1}, {2}/{3}),'
                    '\nperspective {4} ({5}, {6}/{7}):'
                    '\ntotal variation distance {8:.6f}'.format(
                    index_i, name_i, perspective_i.client_id, perspective_i.object_id,
                    index_j, name_j, perspective_j.client_id, perspective_j.object_id,
                    d_value))

        # Adding distance data to the workbook.

        name_list = [info_list[i][2] for i in yes_index_list]

        worksheet_distance.set_column(0, 0, 64)

        worksheet_distance.write_row('B1', name_list)
        worksheet_distance.write_column('A2', name_list)

        for i in range(len(yes_index_list)):

            worksheet_distance.write_row(
                'B{0}'.format(i + 2),
                [round(value, 4) for value in d_ij[i,:]])

        if no_index_list:

            worksheet_distance.write_column(
                'A{0}'.format(len(yes_index_list) + 3),
                ['Insufficient formant data:'] +
                    [info_list[i][2] for i in no_index_list])

        workbook.close()

        xlsx_file_name = pathvalidate.sanitize_filename(
            'Phonological statistical distance ({0} perspectives).xlsx'.format(len(info_list)))

        if __debug_flag__:

            workbook_stream.seek(0)

            with open(xlsx_file_name, 'wb') as xlsx_file:
                shutil.copyfileobj(workbook_stream, xlsx_file)

        # Storing XLSX file with distribution comparison results.

        cur_time = time.time()

        storage_dir = os.path.join(storage['path'], 'psd', str(cur_time))

        xlsx_path = os.path.join(storage_dir, xlsx_file_name)
        os.makedirs(os.path.dirname(xlsx_path), exist_ok = True)

        workbook_stream.seek(0)

        with open(xlsx_path, 'wb') as xlsx_file:
            shutil.copyfileobj(workbook_stream, xlsx_file)

        xlsx_url = ''.join([
            storage['prefix'], storage['static_route'],
            'psd', '/', str(cur_time), '/', xlsx_file_name])

        # Finalizing task status, if required.

        if task_status is not None:
            task_status.set(1, 100, 'Finished', result_link = xlsx_url)

        return PhonologicalStatisticalDistance(triumph = True)

    @staticmethod
    def mutate(self, info, **args):
        """
        query MyQuery {
          phonological_statistical_distance(
            id_list: [[656, 3]],
            vowel_selection: false,
            chart_threshold: 1)
        }
        """

        id_list = args['id_list']
        vowel_selection = args['vowel_selection']
        chart_threshold = args['chart_threshold']

        synchronous_flag = args.get('synchronous_flag', False)

        __debug_flag__ = args.get('debug_flag', False)

        try:

            # Showing our arguments.

            log.debug(
                'phonological_statistical_distance ({0} perspectives):\n'
                'id_list: {1}\n'
                'vowel_selection: {2}\n'
                'chart_threshold: {3}\n'
                'synchronous_flag: {4}\n'
                'debug_flag: {5}\n'.format(
                len(id_list),
                id_list,
                vowel_selection,
                chart_threshold,
                synchronous_flag,
                __debug_flag__))

            locale_id = info.context.get('locale_id')

            request = info.context.get('request')
            storage = request.registry.settings['storage']

            # Simple synchronous phonological statistical distance computation.

            if synchronous_flag:

                return PhonologicalStatisticalDistance.perform_phonological_statistical_distance(
                    id_list,
                    vowel_selection,
                    chart_threshold,
                    locale_id,
                    storage,
                    None,
                    __debug_flag__)

            # Asynchronous phonological statistical distance computation with task status setup.

            client_id = request.authenticated_userid

            user_id = (
                Client.get_user_by_client_id(client_id).id
                    if client_id else anonymous_userid(request))

            if __debug_flag__ and user_id != 1:

                return (

                    ResponseError(
                        message = 'Only administrator can use debug mode.'))

            task_status = TaskStatus(
                user_id,
                'Phonological statistical distance computation',
                '{0} perspectives'.format(len(id_list)),
                1)

            # Launching asynchronous phonological statistical distance computation.

            request.response.status = HTTPOk.code

            async_phonological_statistical_distance.delay(
                id_list,
                vowel_selection,
                chart_threshold,
                locale_id,
                storage,
                task_status.key,
                request.registry.settings['cache_kwargs'],
                request.registry.settings['sqlalchemy.url'],
                __debug_flag__)

            return PhonologicalStatisticalDistance(triumph = True)

        # Some unknown external exception.

        except Exception as exception:

            traceback_string = ''.join(traceback.format_exception(
                exception, exception, exception.__traceback__))[:-1]

            log.warning('phonological_statistical_distance: exception')
            log.warning(traceback_string)

            return ResponseError(message =
                'Exception:\n' + traceback_string)


class SoundAndMarkup(graphene.Mutation):

    class Arguments:
        perspective_id = LingvodocID(required = True)
        published_mode = graphene.String(required = True)

    triumph = graphene.Boolean()

    def mutate(self, info, **args):
        """
        query MyQuery {
          sound_and_markup(
            perspective_id: [657, 4])
        }
        """

        locale_id = info.context.get('locale_id')
        request = info.context.get('request')

        gql_sound_and_markup(request, locale_id, args['perspective_id'], args['published_mode'])

        return SoundAndMarkup(triumph = True)


def save_dictionary(
    dict_id,
    dictionary_obj,
    request,
    user_id,
    locale_id,
    publish,
    sound_flag,
    markup_flag,
    synchronous = False,
    debug_flag = False):

    my_args = dict()
    my_args["client_id"] = dict_id[0]
    my_args["object_id"] = dict_id[1]
    my_args["locale_id"] = locale_id
    my_args["storage"] = request.registry.settings["storage"]
    my_args['sqlalchemy_url'] = request.registry.settings["sqlalchemy.url"]
    try:
        gist = DBSession.query(dbTranslationGist). \
            filter_by(client_id=dictionary_obj.translation_gist_client_id,
                      object_id=dictionary_obj.translation_gist_object_id).first()
        dict_name = gist.get_translation(locale_id)

        if not synchronous:

            task_name_str = 'Saving dictionary'

            modifier_list = [
                'published only' if publish else 'with unpublished']

            if sound_flag:
                modifier_list.append('with sound')

            task_name_str += (
                ' (' + ', '.join(modifier_list) + ')')

            task = TaskStatus(user_id, task_name_str, dict_name, 4)

    except:
        raise ResponseError('bad request')
    my_args['dict_name'] = dict_name
    my_args["task_key"] = task.key if not synchronous else None
    my_args["cache_kwargs"] = request.registry.settings["cache_kwargs"]
    my_args["published"] = publish
    my_args['sound_flag'] = sound_flag
    my_args['markup_flag'] = markup_flag
    my_args['__debug_flag__'] = debug_flag

    res = (sync_save_dictionary if synchronous else async_save_dictionary.delay)(**my_args)


class SaveDictionary(graphene.Mutation):

    class Arguments:
        id = LingvodocID(required=True)
        mode = graphene.String(required=True)
        sound_flag = graphene.Boolean()
        markup_flag = graphene.Boolean()
        synchronous = graphene.Boolean()
        debug_flag = graphene.Boolean()

    triumph = graphene.Boolean()

    @staticmethod
    def mutate(
        root,
        info,
        id,
        mode,
        sound_flag = False,
        markup_flag = False,
        synchronous = False,
        debug_flag = False):

        request = info.context.request
        locale_id = int(request.cookies.get('locale_id') or 2)

        client_id = authenticated_userid(request)

        user_id = (
            Client.get_user_by_client_id(client_id).id
                if client_id else anonymous_userid(request))

        if debug_flag and user_id != 1:

            return (

                ResponseError(
                    message = 'Only administrator can use debug mode.'))

        dictionary = (

            DBSession
                .query(dbDictionary)
                .filter_by(client_id = id[0], object_id = id[1])
                .first())

        if mode == 'published':
            publish = True
        elif mode == 'all':
            publish = None
        else:
            return ResponseError(message="mode: <all|published>")

        if mode == 'all':

            for perspective in dictionary.dictionaryperspective:

                info.context.acl_check(
                    'view',
                    'lexical_entries_and_entities',
                    perspective.id)

        save_dictionary(
            id,
            dictionary,
            request,
            user_id,
            locale_id,
            publish,
            sound_flag,
            markup_flag,
            synchronous,
            debug_flag)

        return DownloadDictionary(triumph=True)

class SaveAllDictionaries(graphene.Mutation):

    class Arguments:
        mode = graphene.String(required=True)

    triumph = graphene.Boolean()

    @staticmethod
    # @client_id_check()
    def mutate(root, info, **args):
        request = info.context.request
        locale_id = int(request.cookies.get('locale_id') or 2)
        mode = args['mode']
        variables = {'auth': authenticated_userid(request)}
        client = DBSession.query(Client).filter_by(id=variables['auth']).first()
        user = DBSession.query(dbUser).filter_by(id=client.user_id).first()
        user_id = user.id
        if user_id != 1:
            raise ResponseError(message="not admin")
        # counter = 0
        dictionaries = DBSession.query(dbDictionary).filter_by(marked_for_deletion=False).all()
        if mode == 'published':
            publish = True
        elif mode == 'all':
            publish = None

        else:
            raise ResponseError(message="mode: <all|published>")

        for dictionary in dictionaries:

            save_dictionary(
                [dictionary.client_id, dictionary.object_id],
                dictionary_obj,
                request,
                user_id,
                locale_id,
                publish)

            # if not counter % 5:
            #     time.sleep(5)
            # counter += 1

        return DownloadDictionary(triumph=True)


class MoveColumn(graphene.Mutation):
    class Arguments:
        perspective_id = LingvodocID(required=True)
        from_id = LingvodocID(required=True)
        to_id = LingvodocID(required=True)

    triumph = graphene.Boolean()

    @staticmethod
    # @client_id_check()
    def mutate(root, info, **args):
        request = info.context.request
        locale_id = int(request.cookies.get('locale_id') or 2)
        perspective_id = args['perspective_id']
        from_id = args['from_id']
        to_id = args['to_id']
        variables = {'auth': authenticated_userid(request)}
        client = DBSession.query(Client).filter_by(id=variables['auth']).first()
        user = DBSession.query(dbUser).filter_by(id=client.user_id).first()
        user_id = user.id
        client_ids = DBSession.query(Client.id).filter(Client.user_id==user_id).all()
        # if user_id != 1:
        #     raise ResponseError(message="not admin")
        # counter = 0
        perspective = DBSession.query(dbPerspective).filter_by(client_id=perspective_id[0],
                                                                         object_id=perspective_id[1],
                                                                         marked_for_deletion=False).first()
        if not perspective:
            raise ResponseError('No such perspective')
        info.context.acl_check('edit', 'perspective',
                           (perspective.client_id, perspective.object_id))

        lexes = DBSession.query(dbLexicalEntry).join(dbEntity).join(dbPublishingEntity).filter(
            dbLexicalEntry.parent_client_id == perspective_id[0],
            dbLexicalEntry.parent_object_id == perspective_id[1],
            dbLexicalEntry.marked_for_deletion == False,
            dbEntity.client_id.in_(client_ids),
            dbEntity.field_client_id == from_id[0],
            dbEntity.field_object_id == from_id[1],
            dbEntity.marked_for_deletion == False,
            dbPublishingEntity.accepted == True).all()

        for lex in lexes:
            entities = DBSession.query(dbEntity).join(dbPublishingEntity).filter(
                dbEntity.client_id.in_(client_ids),
                dbEntity.parent_client_id == lex.client_id,
                dbEntity.parent_object_id == lex.object_id,
                dbEntity.field_client_id == from_id[0],
                dbEntity.field_object_id == from_id[1],
                dbEntity.marked_for_deletion == False,
                dbPublishingEntity.accepted == True).all()
            for entity in entities:
                existing = DBSession.query(dbEntity).join(dbPublishingEntity).filter(
                    dbEntity.parent_client_id == lex.client_id,
                    dbEntity.parent_object_id == lex.object_id,
                    dbEntity.field_client_id == to_id[0],
                    dbEntity.field_object_id == to_id[1],
                    dbEntity.marked_for_deletion == False,
                    dbPublishingEntity.accepted == True,
                    dbEntity.content == entity.content).first()
                if not existing:
                    self_id = None
                    if entity.self_client_id and entity.self_object_id:
                        self_id = [entity.self_client_id, entity.self_object_id]
                    link_id = None
                    if entity.link_client_id and entity.link_object_id:
                        link_id = [entity.link_client_id, entity.link_object_id]
                    create_entity(id=[client.id, None],
                                  parent_id=[lex.client_id, lex.object_id],
                                  field_id=to_id,
                                  self_id=self_id,
                                  additional_metadata=entity.additional_metadata,
                                  link_id=link_id,
                                  locale_id=entity.locale_id,
                                  filename=None,
                                  content=entity.content,
                                  registry=None,
                                  request=request,
                                  )
                del_object(entity, "move_column", info.context.get('client_id'))

        return MoveColumn(triumph=True)

class AddRolesBulk(graphene.Mutation):
    """
    Adds specified roles for a specified user for all dictionaries and perspectives of a specified language,
    maybe with all its descendants.

    mutation addRolesBulk {
      add_roles_bulk(user_id: 5, language_id: [508, 36]) {
        triumph } }
    """

    class Arguments:
        user_id = graphene.Int(required=True)
        language_id = LingvodocID(required=True)

    language_count = graphene.Int()
    dictionary_count = graphene.Int()
    perspective_count = graphene.Int()

    @staticmethod
    def mutate(root, info, **args):

        user_id = args.get('user_id')
        language_id = args.get('language_id')

        request_client_id = info.context.get('client_id')
        request_user = Client.get_user_by_client_id(request_client_id)

        if request_user.id != 1:
            raise ResponseError('Only administrator can perform bulk roles additions.')

        user = DBSession.query(dbUser).filter_by(id = user_id).first()

        if not user:
            raise ResponseError('No user with id {0}'.format(user_id))

        # Getting permission groups info.

        dictionary_group_list = [row[0] for row in DBSession.query(
            dbBaseGroup.id).filter_by(dictionary_default = True).all()]

        perspective_group_list = [row[0] for row in DBSession.query(
            dbBaseGroup.id).filter_by(perspective_default = True).all()]

        log.debug(
            'add_roles_bulk (user_id: {0}, language_id: {1}/{2}):'
            '\ndictionary_group_list: {3}\nperspective_group_list: {4}'.format(
                user_id,
                language_id[0], language_id[1],
                dictionary_group_list,
                perspective_group_list))

        def process_language(language):
            """
            Recursively adds required permissions for a specified user for all dictionaries and perspectives
            of a given language and its descendants.
            """

            dictionary_list = DBSession.query(dbDictionary).filter_by(
                parent_client_id = language.client_id,
                parent_object_id = language.object_id,
                marked_for_deletion = False).all()

            language_count = 1
            dictionary_count = len(dictionary_list)
            perspective_count = 0

            # All dictionaries of the language.

            for dictionary in dictionary_list:

                for basegroup_id in dictionary_group_list:

                    edit_role(dictionary, user_id, basegroup_id, request_client_id,
                        dictionary_default = True, action = 'add')

                perspective_list = DBSession.query(dbPerspective).filter_by(
                    parent_client_id = dictionary.client_id,
                    parent_object_id = dictionary.object_id,
                    marked_for_deletion = False).all()

                perspective_count += len(perspective_list)

                # All perspectives of each dictionary.

                for perspective in perspective_list:

                    for basegroup_id in perspective_group_list:

                        edit_role(perspective, user_id, basegroup_id, request_client_id,
                            perspective_default = True, action = 'add')

            language_list = DBSession.query(dbLanguage).filter_by(
                parent_client_id = language.client_id,
                parent_object_id = language.object_id,
                marked_for_deletion = False).all()

            # All child languages.

            for child_language in language_list:
                l_count, d_count, p_count = process_language(child_language)

                language_count += l_count
                dictionary_count += d_count
                perspective_count += p_count

            return language_count, dictionary_count, perspective_count

        # Adding required permissions.

        language = DBSession.query(dbLanguage).filter_by(
            marked_for_deletion = False,
            client_id = language_id[0],
            object_id = language_id[1]).first()

        if not language:
            raise ResponseError('Language {0}/{1} not found'.format(*language_id))

        language_count, dictionary_count, perspective_count = process_language(language)

        return AddRolesBulk(
            language_count = language_count,
            dictionary_count = dictionary_count,
            perspective_count = perspective_count)


class XlsxBulkDisconnect(graphene.Mutation):
    """
    Parses uploaded XLSX file, disconnects highlighted cognates.
    """

    class Arguments:
        xlsx_file = Upload()
        debug_flag = graphene.Boolean()

    entry_info_count = graphene.Int()
    skip_count = graphene.Int()

    group_count = graphene.Int()
    disconnect_count = graphene.Int()

    triumph = graphene.Boolean()

    sql_search_str = ('''

        select
        L.client_id,
        L.object_id

        from
        dictionary D,
        dictionaryperspective P,
        dictionaryperspectivetofield Fc,
        dictionaryperspectivetofield Fl,
        lexicalentry L

        where

        (D.translation_gist_client_id, D.translation_gist_object_id) in (

          select distinct
          parent_client_id,
          parent_object_id

          from
          translationatom

          where
          content = :d_name and
          marked_for_deletion = false) and

        D.marked_for_deletion = false and

        P.parent_client_id = D.client_id and
        P.parent_object_id = D.object_id and
        P.marked_for_deletion = false and

        exists (

          select 1

          from
          translationatom A

          where
          A.parent_client_id = P.translation_gist_client_id and
          A.parent_object_id = P.translation_gist_object_id and
          A.marked_for_deletion = false and
          A.content = :p_name) and

        Fc.parent_client_id = P.client_id and
        Fc.parent_object_id = P.object_id and
        Fc.marked_for_deletion = false and

        Fl.parent_client_id = P.client_id and
        Fl.parent_object_id = P.object_id and
        Fl.marked_for_deletion = false and

        (
          Fc.field_client_id = 66 and Fc.field_object_id = 8 and
          Fl.field_client_id = 66 and Fl.field_object_id = 10 or

          exists (

            select 1

            from
            field F,
            translationatom A

            where
            F.client_id = Fc.field_client_id and
            F.object_id = Fc.field_object_id and
            A.parent_client_id = F.translation_gist_client_id and
            A.parent_object_id = F.translation_gist_object_id and
            A.marked_for_deletion = false and
            A.locale_id = 2 and
            lower(regexp_replace(A.content, '\W+', '', 'g')) =
              'phonemictranscription') and

          exists (

            select 1

            from
            field F,
            translationatom A

            where
            F.client_id = Fl.field_client_id and
            F.object_id = Fl.field_object_id and
            A.parent_client_id = F.translation_gist_client_id and
            A.parent_object_id = F.translation_gist_object_id and
            A.marked_for_deletion = false and
            A.locale_id = 2 and
            lower(regexp_replace(A.content, '\W+', '', 'g')) =
              'meaning') or

          exists (

            select 1

            from
            field F,
            translationatom A

            where
            F.client_id = Fc.field_client_id and
            F.object_id = Fc.field_object_id and
            A.parent_client_id = F.translation_gist_client_id and
            A.parent_object_id = F.translation_gist_object_id and
            A.marked_for_deletion = false and
            A.locale_id = 1 and
            lower(regexp_replace(A.content, '\W+', '', 'g')) =
              'фонологическаятранскрипция') and

          exists (

            select 1

            from
            field F,
            translationatom A

            where
            F.client_id = Fl.field_client_id and
            F.object_id = Fl.field_object_id and
            A.parent_client_id = F.translation_gist_client_id and
            A.parent_object_id = F.translation_gist_object_id and
            A.marked_for_deletion = false and
            A.locale_id = 1 and
            lower(regexp_replace(A.content, '\W+', '', 'g')) =
              'значение') or

          exists (

            select 1

            from
            field F,
            translationatom A

            where
            F.client_id = Fc.field_client_id and
            F.object_id = Fc.field_object_id and
            A.parent_client_id = F.translation_gist_client_id and
            A.parent_object_id = F.translation_gist_object_id and
            A.marked_for_deletion = false and
            A.locale_id = 2 and
            lower(regexp_replace(A.content, '\W+', '', 'g')) =
              'transcriptionofparadigmaticforms') and

          exists (

            select 1

            from
            field F,
            translationatom A

            where
            F.client_id = Fl.field_client_id and
            F.object_id = Fl.field_object_id and
            A.parent_client_id = F.translation_gist_client_id and
            A.parent_object_id = F.translation_gist_object_id and
            A.marked_for_deletion = false and
            A.locale_id = 2 and
            lower(regexp_replace(A.content, '\W+', '', 'g')) =
              'translationofparadigmaticforms') or

          exists (

            select 1

            from
            field F,
            translationatom A

            where
            F.client_id = Fc.field_client_id and
            F.object_id = Fc.field_object_id and
            A.parent_client_id = F.translation_gist_client_id and
            A.parent_object_id = F.translation_gist_object_id and
            A.marked_for_deletion = false and
            A.locale_id = 1 and
            lower(regexp_replace(A.content, '\W+', '', 'g')) =
              'транскрипцияпарадигматическихформ') and

          exists (

            select 1

            from
            field F,
            translationatom A

            where
            F.client_id = Fl.field_client_id and
            F.object_id = Fl.field_object_id and
            A.parent_client_id = F.translation_gist_client_id and
            A.parent_object_id = F.translation_gist_object_id and
            A.marked_for_deletion = false and
            A.locale_id = 1 and
            lower(regexp_replace(A.content, '\W+', '', 'g')) =
              'переводпарадигматическихформ') or

          :p_name ~* '.*starling.*' and

          (
            exists (

              select 1

              from
              field F,
              translationatom A

              where
              F.client_id = Fc.field_client_id and
              F.object_id = Fc.field_object_id and
              A.parent_client_id = F.translation_gist_client_id and
              A.parent_object_id = F.translation_gist_object_id and
              A.marked_for_deletion = false and
              A.locale_id = 2 and
              lower(regexp_replace(A.content, '\W+', '', 'g')) =
                'protoform') and

            exists (

              select 1

              from
              field F,
              translationatom A

              where
              F.client_id = Fl.field_client_id and
              F.object_id = Fl.field_object_id and
              A.parent_client_id = F.translation_gist_client_id and
              A.parent_object_id = F.translation_gist_object_id and
              A.marked_for_deletion = false and
              A.locale_id = 2 and
              lower(regexp_replace(A.content, '\W+', '', 'g')) =
                'protoformmeaning') or

            exists (

              select 1

              from
              field F,
              translationatom A

              where
              F.client_id = Fc.field_client_id and
              F.object_id = Fc.field_object_id and
              A.parent_client_id = F.translation_gist_client_id and
              A.parent_object_id = F.translation_gist_object_id and
              A.marked_for_deletion = false and
              A.locale_id = 1 and
              lower(regexp_replace(A.content, '\W+', '', 'g')) =
                'праформа') and

            exists (

              select 1

              from
              field F,
              translationatom A

              where
              F.client_id = Fl.field_client_id and
              F.object_id = Fl.field_object_id and
              A.parent_client_id = F.translation_gist_client_id and
              A.parent_object_id = F.translation_gist_object_id and
              A.marked_for_deletion = false and
              A.locale_id = 1 and
              lower(regexp_replace(A.content, '\W+', '', 'g')) =
                'значениепраформы'))) and

        L.parent_client_id = P.client_id and
        L.parent_object_id = P.object_id and
        L.marked_for_deletion = false and

        exists (

          select 1

          from
          public.entity E

          where
          E.parent_client_id = L.client_id and
          E.parent_object_id = L.object_id and
          E.field_client_id = 66 and
          E.field_object_id = 25 and
          E.marked_for_deletion = false and
          E.content in {}){}{};

        ''')

    sql_tag_str = ('''

        (

          select distinct
          E.content

          from
          lexicalentry L,
          public.entity E

          where

          L.parent_client_id = {} and
          L.parent_object_id = {} and
          L.marked_for_deletion = false{}

          and

          E.parent_client_id = L.client_id and
          E.parent_object_id = L.object_id and
          E.field_client_id = 66 and
          E.field_object_id = 25 and
          E.marked_for_deletion = false)

        ''')

    sql_xcript_str = ('''

        and

        (
          select count(*)

          from
          public.entity E

          where
          E.parent_client_id = L.client_id and
          E.parent_object_id = L.object_id and
          E.field_client_id = Fc.field_client_id and
          E.field_object_id = Fc.field_object_id and
          E.marked_for_deletion = false and
          E.content ~ :xc_regexp
        )
        = {}

        ''')

    sql_xlat_str = ('''

        and

        (
          select count(*)

          from
          public.entity E

          where
          E.parent_client_id = L.client_id and
          E.parent_object_id = L.object_id and
          E.field_client_id = Fl.field_client_id and
          E.field_object_id = Fl.field_object_id and
          E.marked_for_deletion = false and
          E.content ~ :xl_regexp
        )
        = {}

        ''')

    def escape(string):
        """
        Escapes special regexp characters in literal strings for PostgreSQL regexps, see
        https://stackoverflow.com/questions/4202538/escape-regex-special-characters-in-a-python-string/12012114.
        """

        return re.sub(r'([!$()*+.:<=>?[\\\]^{|}-])', r'\\\1', string)

    @staticmethod
    def get_entry_id(
        perspective_id,
        entry_info,
        row_index_set):
        """
        Tries to get id of a cognate entry.
        """

        (content_info,
            dp_name,
            xcript_tuple,
            xlat_tuple) = (

            entry_info)

        # We have to have at least some entry info.

        if not xcript_tuple and not xlat_tuple and not content_info:
            return None

        d_name, p_name = (
            dp_name.split(' › '))

        param_dict = {
            'd_name': d_name,
            'p_name': p_name}

        if xcript_tuple:

            param_dict.update({

                'xc_regexp':

                    r'^\s*(' +

                    '|'.join(
                        XlsxBulkDisconnect.escape(xcript)
                        for xcript in xcript_tuple) +

                    r')\s*$'})

        if xlat_tuple:

            param_dict.update({

                'xl_regexp':

                    r'^\s*(' +

                    '|'.join(
                        XlsxBulkDisconnect.escape(xlat)
                        for xlat in xlat_tuple) +

                    r')\s*$'})

        # Trying to find the entry.

        if content_info:

            sql_str_a_list = []

            for i, (field_id_set, content_str) in enumerate(content_info):

                sql_str_a_list.append(

                    '''
                    and

                    exists (

                      select 1

                      from
                      public.entity E

                      where
                      E.parent_client_id = L.client_id and
                      E.parent_object_id = L.object_id and
                      E.marked_for_deletion = false and
                      (E.field_client_id, E.field_object_id) in ({}) and
                      E.content = :content_{})
                    '''

                    .format(
                        ', '.join(map(str, field_id_set)),
                        i))

                param_dict[
                    'content_{}'.format(i)] = content_str

            sql_str_a = (

                XlsxBulkDisconnect.sql_tag_str.format(
                    perspective_id[0],
                    perspective_id[1],
                    ''.join(sql_str_a_list)))

        else:

            sql_str_a = (

                '(select * from {})'.format(
                    tag_table_name))

        sql_str_b = (

            XlsxBulkDisconnect.sql_xcript_str.format(len(xcript_tuple))
                if xcript_tuple else '')

        sql_str_c = (

            XlsxBulkDisconnect.sql_xlat_str.format(len(xlat_tuple))
                if xlat_tuple else '')

        sql_str = (

            XlsxBulkDisconnect.sql_search_str.format(
                sql_str_a,
                sql_str_b,
                sql_str_c))

        result_list = (

            DBSession

                .execute(
                    sql_str,
                    param_dict)

                .fetchall())

        result_list = [

            (entry_cid, entry_oid)
            for entry_cid, entry_oid in result_list]

        log.debug(
            '\nresult_list: {}'.format(
                result_list))

        # If we haven't found anything, no problem, just going on ahead.

        if not result_list:
            return None

        # We shouldn't have any duplicate results.

        result_set = set(result_list)

        if len(result_set) < len(result_list):

            log.warning(

                '\n' +

                str(
                    sqlalchemy
                        .text(sql_str)
                        .bindparams(**param_dict)
                        .compile(compile_kwargs = {'literal_binds': True})) +

                '\nresult_list: {}'
                '\nresult_set: {}'.format(
                    result_list,
                    result_set))

            result_list = list(result_set)

        # If we've got the unambiguous entry info, ok, cool, otherwise no problem, skipping this and going
        # ahead.

        if len(result_list) <= len(row_index_set):
            return result_list

        return None

    @staticmethod
    def mutate(root, info, **args):

        __debug_flag__ = args.get('debug_flag', False)

        try:

            client_id = info.context.get('client_id')
            client = DBSession.query(Client).filter_by(id = client_id).first()

            if not client or client.user_id != 1:
                return ResponseError('Only administrator can bulk disconnect.')

            request = info.context.request

            if '1' not in request.POST:
                return ResponseError('XLSX file is required.')

            multipart = request.POST.pop('1')

            xlsx_file_name = multipart.filename
            xlsx_file = multipart.file

            log.debug(
                '\n{}\n{}'.format(
                    xlsx_file_name,
                    type(xlsx_file)))

            settings = (
                request.registry.settings)

            # Processing XLSX workbook, assuming each worksheet has data of a single perspective.

            workbook = (
                openpyxl.load_workbook(xlsx_file))

            entry_info_count = 0
            skip_count = 0

            group_count = 0
            disconnect_count = 0

            tag_table_name = None

            for sheet_name in workbook.sheetnames:

                worksheet = workbook[sheet_name]

                # Assuming the first row has field names.

                field_name_list = []
                cognates_index = None

                for i in itertools.count(1):

                    cell = worksheet.cell(1, i)

                    if cell.value:

                        field_name_list.append(cell.value)
                        cognates_index = i

                    else:
                        break

                # Trying to parse perspective's fields.

                (perspective_name,
                    perspective_cid,
                    perspective_oid) = (

                    re.match(
                        r'^(.*)_(\d+)_(\d+)$',
                        sheet_name)

                        .groups())

                perspective_id = (
                    perspective_cid, perspective_oid)

                log.debug(
                    '\nperspective: \'{}\' {}/{}'
                    '\nfield_name_list:\n{}'.format(
                        perspective_name,
                        perspective_cid,
                        perspective_oid,
                        field_name_list))

                field_id_set_list = []

                for field_name in field_name_list[:-1]:

                    result_list = (

                        DBSession

                            .query(
                                dbField.client_id,
                                dbField.object_id)

                            .filter(
                                dbColumn.parent_client_id == perspective_cid,
                                dbColumn.parent_object_id == perspective_oid,
                                dbColumn.marked_for_deletion == False,

                                tuple_(
                                    dbColumn.field_client_id,
                                    dbColumn.field_object_id)

                                    .in_(
                                        sqlalchemy.text('select * from text_field_id_view')),

                                dbField.client_id == dbColumn.field_client_id,
                                dbField.object_id == dbColumn.field_object_id,
                                dbField.marked_for_deletion == False,
                                dbTranslationAtom.parent_client_id == dbField.translation_gist_client_id,
                                dbTranslationAtom.parent_object_id == dbField.translation_gist_object_id,
                                dbTranslationAtom.marked_for_deletion == False,
                                dbTranslationAtom.content == field_name)

                            .distinct()

                            .all())

                    field_id_set_list.append(

                        tuple(set(
                            tuple(field_id) for field_id in result_list)

                            if result_list else None))

                # Getting all possible cognate tags of the entries of the perspective.

                if tag_table_name is None:

                    tag_table_name = (

                        'tag_table_' +
                        str(uuid.uuid4()).replace('-', '_'))

                    DBSession.execute('''

                        create temporary table

                        {} (
                          tag TEXT,
                          primary key (tag))

                        on commit drop;

                        '''.format(
                            tag_table_name))

                else:

                    DBSession.execute(
                        'truncate table {};'.format(
                            tag_table_name))

                DBSession.execute('''

                    insert into {}

                    select
                    E.content

                    from
                    lexicalentry L,
                    public.entity E

                    where
                    L.parent_client_id = {} and
                    L.parent_object_id = {} and
                    L.marked_for_deletion = false and
                    E.parent_client_id = L.client_id and
                    E.parent_object_id = L.object_id and
                    E.field_client_id = 66 and
                    E.field_object_id = 25 and
                    E.marked_for_deletion = false

                    on conflict do nothing;

                    '''.format(
                        tag_table_name,
                        perspective_cid,
                        perspective_oid))

                # Processing cognate groups.

                color_set = set()

                content_info = None

                entry_info_list = []
                entry_info_dict = collections.defaultdict(set)

                entry_content_info = None
                entry_dp_name = None

                for i in range(2, worksheet.max_row):

                    # Getting text field info if we have any.

                    previous_content_info = content_info

                    content_info = []
                    content_flag = False

                    for j, field_id_set in enumerate(field_id_set_list):

                        field_cell = worksheet.cell(i, j + 1)

                        if field_cell.value:

                            content_flag = True

                            if field_id_set is not None:

                                content_info.append(
                                    (field_id_set, field_cell.value))

                    if content_flag:

                        content_info = (
                            tuple(content_info) if content_info else None)

                    else:

                        content_info = (
                            previous_content_info)

                    # Do we have a beginning of another cognate entry info?

                    dp_cell = worksheet.cell(i, cognates_index)

                    if dp_cell.value:

                        if (entry_dp_name and
                            entry_highlight_list and
                            len(entry_highlight_list) >= len(entry_xcript_list)):

                            entry_info = (
                                entry_content_info,
                                entry_dp_name,
                                tuple(xcript for xcript in entry_xcript_list if xcript),
                                tuple(xlat[1:-1] for xlat in entry_xlat_list))

                            if entry_info not in entry_info_dict:
                                entry_info_list.append(entry_info)

                            entry_info_dict[entry_info].add(entry_row_index)

                        entry_dp_name = dp_cell.value
                        entry_content_info = content_info

                        entry_row_index = i

                        entry_xcript_list = []
                        entry_xlat_list = []

                        entry_highlight_list = []

                    # Do we have transcription and / or translation, is transcription highlighted?

                    xc_cell = worksheet.cell(i, cognates_index + 1)

                    if xc_cell.value or dp_cell.value:

                        entry_xcript_list.append(xc_cell.value)

                        color = xc_cell.fill.fgColor.rgb

                        if color != '00000000':
                            color_set.add(color)

                        if color == 'FFFFFF00':
                            entry_highlight_list.append(xc_cell.value)

                    xl_cell = worksheet.cell(i, cognates_index + 2)

                    if xl_cell.value:
                        entry_xlat_list.append(xl_cell.value)

                log.debug(
                    '\n' +
                    pprint.pformat(
                        entry_info_list, width = 192))

                # Processing highlighted entries.

                entry_id_list = []
                skip_list = []

                for entry_info in entry_info_list:

                    id_list = (

                        XlsxBulkDisconnect.get_entry_id(
                            perspective_id,
                            entry_info,
                            entry_info_dict[entry_info]))

                    if id_list:

                        entry_id_list.extend(
                            id_list)

                    else:

                        skip_list.append((
                            entry_info,
                            entry_info_dict[entry_info]))

                log.debug(
                    '\nentry_id_list:\n{}'
                    '\nskip_list:\n{}'
                    '\nlen(entry_info_list): {}'
                    '\nlen(entry_id_list): {}'
                    '\nlen(skip_list): {}'.format(

                        pprint.pformat(
                            entry_id_list, width = 192),

                        pprint.pformat(
                            skip_list, width = 192),

                        len(entry_info_list),
                        len(entry_id_list),
                        len(skip_list)))

                # Performing disconnects.

                entry_id_set = set(entry_id_list)
                already_set = set()

                perspective_group_count = 0
                perspective_disconnect_count = 0

                for entry_id in entry_id_list:

                    if entry_id in already_set:
                        continue

                    result_list = (

                        DBSession

                            .execute(
                                'select * from linked_group(66, 25, {}, {})'.format(
                                    *entry_id))

                            .fetchall())

                    cognate_id_set = (

                        set(
                            (entry_cid, entry_oid)
                            for entry_cid, entry_oid in result_list))

                    disconnect_set = (
                        cognate_id_set & entry_id_set)

                    leave_set = (
                        cognate_id_set - disconnect_set)

                    log.debug(
                        '\ncognate_id_set ({}):\n{}'
                        '\ndisconnect_set ({}):\n{}'
                        '\nleave_set ({}):\n{}'.format(
                        len(cognate_id_set),
                        cognate_id_set,
                        len(disconnect_set),
                        disconnect_set,
                        len(leave_set),
                        leave_set))

                    # Disconnecting highlighted entries, see `class DeleteGroupingTags()`.

                    entity_list = (

                        DBSession

                            .query(dbEntity)

                            .filter(

                                tuple_(
                                    dbEntity.parent_client_id,
                                    dbEntity.parent_object_id)
                                    .in_(disconnect_set),

                                dbEntity.field_client_id == 66,
                                dbEntity.field_object_id == 25,
                                dbEntity.marked_for_deletion == False)

                            .all())

                    for entity in entity_list:

                        if 'desktop' in settings:

                            real_delete_entity(
                                entity,
                                settings)

                        else:

                            del_object(
                                entity,
                                'xlsx_bulk_disconnect',
                                client_id)

                    # Connecting disconnected entries together, if there is more than one, see
                    # `class ConnectLexicalEntries()`.

                    n = 10

                    rnd = (
                        random.SystemRandom())

                    choice_str = (
                        string.digits + string.ascii_letters)

                    tag_str = (

                        time.asctime(time.gmtime()) +

                        ''.join(
                            rnd.choice(choice_str)
                            for c in range(n)))

                    for entry_id in disconnect_set:

                        dbEntity(
                            client_id = client_id,
                            parent_client_id = entry_id[0],
                            parent_object_id = entry_id[1],
                            field_client_id = 66,
                            field_object_id = 25,
                            content = tag_str,
                            published = True,
                            accepted = True)

                    already_set.update(disconnect_set)

                    perspective_group_count += 1
                    perspective_disconnect_count += len(disconnect_set)

                # Finished this perspective.

                log.debug(
                    '\n\'{}\' {}/{}:'
                    '\nperspective_group_count: {}'
                    '\nperspective_disconnect_count: {}'.format(
                        perspective_name,
                        perspective_cid,
                        perspective_oid,
                        perspective_group_count,
                        perspective_disconnect_count))

                entry_info_count += len(entry_info_list)
                skip_count += len(skip_list)

                group_count += perspective_group_count
                disconnect_count += perspective_disconnect_count

            # Finished bulk disconnects.

            log.debug(
                '\n{} perspectives'
                '\nentry_info_count: {}'
                '\nskip_count: {}'
                '\ngroup_count: {}'
                '\ndisconnect_count: {}'.format(
                    len(workbook.sheetnames),
                    entry_info_count,
                    skip_count,
                    group_count,
                    disconnect_count))

            return (

                XlsxBulkDisconnect(
                    entry_info_count = entry_info_count,
                    skip_count = skip_count,
                    group_count = group_count,
                    disconnect_count = disconnect_count,
                    triumph = True))

        except Exception as exception:

            traceback_string = (

                ''.join(
                    traceback.format_exception(
                        exception, exception, exception.__traceback__))[:-1])

            log.warning('xlsx_bulk_disconnect: exception')
            log.warning(traceback_string)

            return (

                ResponseError(
                    'Exception:\n' + traceback_string))


class NewUnstructuredData(graphene.Mutation):
    """
    Creates new unstructured data entry, returns its id.
    """

    class Arguments:

        data = ObjectVal(required = True)
        metadata = ObjectVal()

    triumph = graphene.Boolean()
    id = graphene.String()

    @staticmethod
    def get_random_unstructured_data_id():
        """
        Returns reasonably short random unused base59 string unstructured data id.
        """

        # Not using 'l', 'I' and 'O' which in some cases can cause confusion.

        base59_str = '0123456789abcdefghijkmnopqrstuvwxyzABCDEFGHJKLMNPQRSTUVWXYZ'

        rng = (
            random.Random(
                int(time.time() * 1000000000)))

        id_count = DBSession.query(dbUnstructuredData).count()

        id_length = 5

        id_length_limit = (
            5 if id_count <= 0 else
            min(5, int(math.ceil(math.log(2 * id_count, 59)))))

        def random_id_generator(how_many):

            nonlocal id_length

            for i in range(how_many):

                yield (

                    ''.join(
                        rng.choice(base59_str)
                        for i in range(id_length)))

                if id_length < id_length_limit:
                    id_length += 1

        sql_str = ('''

            select id
            from (values {}) T(id)

            where id not in (
              select id from unstructured_data)

            order by length(id)
            limit 1;

            ''')

        def get_another_query_str():

            return (

                sql_str.format(

                    ', '.join(
                        '(\'{}\')'.format(random_id)
                        for random_id in random_id_generator(8))))

        result = (

            DBSession
                .execute(get_another_query_str())
                .first())

        while not result:

            result = (

                DBSession
                    .execute(get_another_query_str())
                    .first())

        return result[0]

    @staticmethod
    def mutate(root, info, **args):

        try:

            client_id = info.context.get('client_id')
            client = DBSession.query(Client).filter_by(id = client_id).first()

            if not client:
                return ResponseError('Only registered users can create unstructured data entries.')

            data = args.get('data')
            metadata = args.get('metadata')

            log.debug(
                '\nnew_unstructured_data'
                '\nclient_id: {}'
                '\ndata:\n{}'
                '\nmetadata:\n{}'.format(
                    client_id,
                    pprint.pformat(data, width = 144),
                    pprint.pformat(metadata, width = 144)))

            id_str = NewUnstructuredData.get_random_unstructured_data_id()

            unstructured_data = (

                dbUnstructuredData(
                    id = id_str,
                    client_id = client_id,
                    data = data,
                    additional_metadata = metadata))

            DBSession.add(unstructured_data)

            return (

                NewUnstructuredData(
                    id = id_str,
                    triumph = True))

        except Exception as exception:

            traceback_string = (

                ''.join(
                    traceback.format_exception(
                        exception, exception, exception.__traceback__))[:-1])

            log.warning('new_unstructured_data: exception')
            log.warning(traceback_string)

            return (

                ResponseError(
                    'Exception:\n' + traceback_string))


class Docx2Eaf(graphene.Mutation):
    """
    Tries to convert a table-containing .docx to .eaf.

    curl 'http://localhost:6543/graphql'
      -H 'Cookie: locale_id=2; auth_tkt=_!userid_type:int; client_id=4211'
      -H 'Content-Type: multipart/form-data'
      -F operations='{
         "query": "mutation docx2eaf($docxFile: Upload, $separateFlag: Boolean) {
           docx2eaf(docx_file: $docxFile, separate_flag: $separateFlag, debug_flag: true) {
             triumph eaf_url alignment_url check_txt_url check_docx_url message } }",
         "variables": { "docxFile": null, "separateFlag": false } }'
      -F map='{ "1": ["variables.docx_file"] }'
      -F 1=@"/root/lingvodoc-extra/Чертыкова_Беседы_14.09.2019.docx"
    """

    class Arguments:

        docx_file = Upload()
        separate_flag = graphene.Boolean()
        all_tables_flag = graphene.Boolean()
        no_header_flag = graphene.Boolean()
        no_parsing_flag = graphene.Boolean()
        debug_flag = graphene.Boolean()

    triumph = graphene.Boolean()

    eaf_url = graphene.String()
    alignment_url = graphene.String()
    check_txt_url = graphene.String()
    check_docx_url = graphene.String()

    message = graphene.String()

    @staticmethod
    def mutate(root, info, **args):

        try:

            client_id = info.context.get('client_id')
            client = DBSession.query(Client).filter_by(id = client_id).first()

            if not client:

                return (

                    Docx2Eaf(
                        triumph = False,
                        message = 'Only registered users can convert .docx to .eaf.'))

            request = info.context.request

            if '1' not in request.POST:
                return ResponseError('.docx file is required.')

            multipart = request.POST.pop('1')

            docx_file_name = multipart.filename
            docx_file = multipart.file

            separate_flag = args.get('separate_flag', False)
            all_tables_flag = args.get('all_tables_flag', False)
            no_header_flag = args.get('no_header_flag', False)
            no_parsing_flag = args.get('no_parsing_flag', False)

            __debug_flag__ = args.get('debug_flag', False)

            if __debug_flag__ and client.user_id != 1:

                return (

                    ResponseError(
                        message = 'Only administrator can use debug mode.'))

            log.debug(
                '\n{}\n{}'.format(
                    docx_file_name,
                    type(docx_file)))

            if __debug_flag__:

                with open('docx2eaf_input.docx', 'wb') as input_file:

                    shutil.copyfileobj(docx_file, input_file)
                    docx_file.seek(0)

            url_list = []

            with tempfile.TemporaryDirectory() as tmp_dir_path:

                tmp_docx_file_path = (
                    os.path.join(tmp_dir_path, 'docx2eaf_input.docx'))

                tmp_eaf_file_path = (
                    os.path.join(tmp_dir_path, 'docx2eaf_output.eaf'))

                tmp_check_txt_file_path = (
                    os.path.join(tmp_dir_path, 'docx2eaf_check.txt'))

                tmp_check_docx_file_path = (
                    os.path.join(tmp_dir_path, 'docx2eaf_check.docx'))

                with open(tmp_docx_file_path, 'wb') as tmp_docx_file:
                    shutil.copyfileobj(docx_file, tmp_docx_file)

                result = (

                    docx_import.docx2eaf(
                        tmp_docx_file_path,
                        tmp_eaf_file_path,
                        separate_by_paragraphs_flag = separate_flag,
                        modify_docx_flag = True,
                        all_tables_flag = all_tables_flag,
                        no_header_flag = no_header_flag,
                        no_parsing_flag = no_parsing_flag,
                        check_file_path = tmp_check_txt_file_path,
                        check_docx_file_path = tmp_check_docx_file_path,
                        __debug_flag__ = __debug_flag__))

                # Saving local copies, if required.

                if __debug_flag__:

                    shutil.copyfile(tmp_eaf_file_path, 'docx2eaf_output.eaf')
                    shutil.copyfile(tmp_check_txt_file_path, 'docx2eaf_check.txt')

                    if not separate_flag and not all_tables_flag:
                        shutil.copyfile(tmp_check_docx_file_path, 'docx2eaf_check.docx')

                # Saving processed files.

                storage = (
                    request.registry.settings['storage'])

                storage_temporary = storage['temporary']

                host = storage_temporary['host']
                bucket = storage_temporary['bucket']

                minio_client = (

                    minio.Minio(
                        host,
                        access_key = storage_temporary['access_key'],
                        secret_key = storage_temporary['secret_key'],
                        secure = True))

                current_time = time.time()

                input_file_name = (

                    pathvalidate.sanitize_filename(
                        os.path.splitext(os.path.basename(docx_file_name))[0]))

                for file_path, suffix in (

                    (tmp_eaf_file_path, '.eaf'),
                    (tmp_docx_file_path, ' alignment.docx'),
                    (tmp_check_txt_file_path, ' check.txt'),
                    (tmp_check_docx_file_path, ' check.docx')):

                    if ((separate_flag or all_tables_flag) and
                        (suffix == ' check.docx' or suffix == ' alignment.docx')):

                        url_list.append(None)
                        continue

                    object_name = (

                        storage_temporary['prefix'] +

                        '/'.join((
                            'docx2eaf',
                            '{:.6f}'.format(current_time),
                            input_file_name + suffix)))

                    (etag, version_id) = (

                        minio_client.fput_object(
                            bucket,
                            object_name,
                            file_path))

                    url = (

                        '/'.join((
                            'https:/',
                            host,
                            bucket,
                            object_name)))

                    log.debug(
                        '\nobject_name:\n{}'
                        '\netag:\n{}'
                        '\nversion_id:\n{}'
                        '\nurl:\n{}'.format(
                            object_name,
                            etag,
                            version_id,
                            url))

                    url_list.append(url)

                log.debug(
                    '\nurl_list:\n' +
                    pprint.pformat(url_list, width = 192))

            return (

                Docx2Eaf(
                    triumph = True,
                    eaf_url = url_list[0],
                    alignment_url = url_list[1],
                    check_txt_url = url_list[2],
                    check_docx_url = url_list[3]))

        except docx_import.Docx2EafError as exception:

            return (

                Docx2Eaf(
                    triumph = False,
                    message = exception.args[0]))

        except Exception as exception:

            traceback_string = (

                ''.join(
                    traceback.format_exception(
                        exception, exception, exception.__traceback__))[:-1])

            log.warning('docx2eaf: exception')
            log.warning(traceback_string)

            return (

                ResponseError(
                    'Exception:\n' + traceback_string))


@celery.task
def async_valency_compute(
    perspective_id,
    debug_flag,
    full_name,
    task_key,
    storage,
    cache_kwargs,
    sqlalchemy_url):

    # NOTE: copied from phonology.
    #
    # This is a no-op with current settings, we use it to enable logging inside celery tasks, because
    # somehow this does it, and otherwise we couldn't set it up.

    logging.debug('async_valency')

    engine = create_engine(sqlalchemy_url)
    DBSession.configure(bind = engine)
    initialize_cache(cache_kwargs)

    task_status = TaskStatus.get_from_cache(task_key)

    with transaction.manager:

        try:

            Valency.compute(
                perspective_id,
                debug_flag,
                full_name,
                task_key,
                storage,
                cache_kwargs,
                sqlalchemy_url)

        # Some unknown external exception.

        except Exception as exception:

            traceback_string = ''.join(traceback.format_exception(
                exception, exception, exception.__traceback__))[:-1]

            log.warning(
                'valency \'{}\' {}/{}: exception'.format(
                    full_name,
                    perspective_id[0],
                    perspective_id[1]))

            log.warning(traceback_string)

            if task_status is not None:

                task_status.set(1, 100,
                    'ERROR, exception:\n' + traceback_string)


class Valency(graphene.Mutation):
    """
    Extracts valency info.

    curl 'http://localhost:6543/graphql' \
      -H 'Content-Type: application/json' \
      -H 'Cookie: auth_tkt=_!userid_type:int; client_id=4217; locale_id=2' \
      --data-raw '{ \
        "operationName": "valency", \
        "variables":{"perspectiveId":[3648,8]}, \
        "query": \
          "mutation valency($perspectiveId: LingvodocID!) { \
            valency(perspective_id: $perspectiveId, synchronous: true, debug_flag: true) { \
              triumph }}"}'
    """

    class Arguments:

        perspective_id = LingvodocID(required = True)

        debug_flag = graphene.Boolean()
        synchronous = graphene.Boolean()

    triumph = graphene.Boolean()

    @staticmethod
    def get_parser_result_data(
        perspective_id,
        debug_flag):

        entry_dict = collections.defaultdict(dict)
        entry_list = []

        entity_list = (

            DBSession

                .query(
                    dbEntity)

                .filter(
                    dbLexicalEntry.parent_client_id == perspective_id[0],
                    dbLexicalEntry.parent_object_id == perspective_id[1],
                    dbLexicalEntry.marked_for_deletion == False,
                    dbEntity.parent_client_id == dbLexicalEntry.client_id,
                    dbEntity.parent_object_id == dbLexicalEntry.object_id,
                    dbEntity.marked_for_deletion == False,
                    dbPublishingEntity.client_id == dbEntity.client_id,
                    dbPublishingEntity.object_id == dbEntity.object_id,
                    dbPublishingEntity.published == True,
                    dbPublishingEntity.accepted == True)

                .order_by(
                    dbLexicalEntry.created_at,
                    dbLexicalEntry.client_id,
                    dbLexicalEntry.object_id,
                    dbEntity.created_at,
                    dbEntity.client_id,
                    dbEntity.object_id)

                .all())

        # Processing entities.

        for entity in entity_list:

            entry_id = (
                (entity.parent_client_id, entity.parent_object_id))

            new_entry_flag = (
                entry_id not in entry_dict)

            entry_info_dict = entry_dict[entry_id]

            if new_entry_flag:

                entry_info_dict['parser_result_list'] = []
                entry_list.append(entry_info_dict)

            if entity.field_id == (674, 5):

                entry_info_dict['comment'] = entity.content
                continue

            elif not is_subject_for_parsing(entity.content):
                continue

            parser_result_list = (

                DBSession

                    .query(dbParserResult)

                    .filter_by(
                        entity_client_id = entity.client_id,
                        entity_object_id = entity.object_id,
                        marked_for_deletion = False)

                    .order_by(
                        dbParserResult.created_at,
                        dbParserResult.client_id,
                        dbParserResult.object_id)

                    .all())

            # Processing all parser results of this entity.

            for parser_result_index, parser_result in enumerate(parser_result_list):

                paragraph_list, token_count, _ = (

                    export_parser_result.process_parser_result(
                        parser_result.content,
                        debug_flag = debug_flag,
                        format_flag = True))

                entry_info_dict['parser_result_list'].append({
                    'index': parser_result_index,
                    'id': parser_result.id,
                    'hash': hashlib.sha256(parser_result.content.encode('utf-8')).hexdigest(),
                    'paragraphs': paragraph_list})

        # Adding titles to parser results, if we have them.

        parser_result_list = []

        for entry_info_dict in entry_list:

            title_str = (
                entry_info_dict.get('comment'))

            for parser_result in entry_info_dict['parser_result_list']:

                parser_result['title'] = title_str
                parser_result_list.append(parser_result)

        return parser_result_list

    @staticmethod
    def compute(
        perspective_id,
        debug_flag,
        full_name,
        task_key,
        storage,
        cache_kwargs,
        sqlalchemy_url):

        log.debug(
            '\nvalency \'{}\' {}/{}:'
            '\n  debug_flag: {}'.format(
                full_name,
                perspective_id[0],
                perspective_id[1],
                debug_flag))

        task_status = (
            None if task_key is None else
            TaskStatus.get_from_cache(task_key))

        if task_status:
            task_status.set(1, 0, 'Compiling corpus')

        parser_result_list = (

            Valency.get_parser_result_data(
                perspective_id, debug_flag))

        # If we have no parser results, we won't do anything.

        if not parser_result_list:

            if task_status:
                task_status.set(1, 100, 'Finished, no parser result data')

            return

        # Processing parser results.

        if task_status:
            task_status.set(1, 50, 'Processing data')

        sentence_data = (
            valency.corpus_to_sentences(parser_result_list))

        arx_data = (
            valency.corpus_to_arx(parser_result_list))

        valence_data = (
            valency.sentences_arx_to_valencies(sentence_data, arx_data))

        result_data = (
            valency.sentences_valencies_to_result(sentence_data, valence_data))

        # Saving processed data as zipped JSON.

        current_time = time.time()

        date = datetime.datetime.utcfromtimestamp(current_time)

        zip_date = (
            date.year,
            date.month,
            date.day,
            date.hour,
            date.minute,
            date.second)

        storage_temporary = storage['temporary']

        host = storage_temporary['host']
        bucket = storage_temporary['bucket']

        minio_client = (

            minio.Minio(
                host,
                access_key = storage_temporary['access_key'],
                secret_key = storage_temporary['secret_key'],
                secure = True))

        url_list = []

        for data_value, data_name in [
            (parser_result_list, 'corpus'),
            (sentence_data, 'sentences'),
            (arx_data, 'arx'),
            (valence_data, 'valencies'),
            (result_data, 'result')]:

            data_json_str = (

                json.dumps(
                    data_value,
                    ensure_ascii = False,
                    indent = 2))

            temporary_file = (

                tempfile.NamedTemporaryFile(
                    delete = False))

            zip_file = (

                zipfile.ZipFile(
                    temporary_file, 'w',
                    compression = zipfile.ZIP_DEFLATED,
                    compresslevel = 9))

            zip_info = (
                zipfile.ZipInfo(data_name + '.json', zip_date))

            zip_info.compress_type = zipfile.ZIP_DEFLATED

            zip_file.writestr(
                zip_info, data_json_str)

            zip_file.close()
            temporary_file.close()

            if debug_flag:

                shutil.copy(
                    temporary_file.name,
                    '__valency__' + data_name + '.json.zip')

            object_name = (

                storage_temporary['prefix'] +

                '/'.join((
                    'valency',
                    '{:.6f}'.format(current_time),
                    data_name + '.json.zip')))

            (etag, version_id) = (

                minio_client.fput_object(
                    bucket,
                    object_name,
                    temporary_file.name))

            os.remove(
                temporary_file.name)

            url = (

                '/'.join((
                    'https:/',
                    host,
                    bucket,
                    object_name)))

            url_list.append(url)

            log.debug(
                '\nobject_name: {}'
                '\netag: {}'
                '\nversion_id: {}'
                '\nurl: {}'.format(
                    object_name,
                    etag,
                    version_id,
                    url))

        if task_status:
            task_status.set(1, 100, 'Finished', result_link_list = url_list)

    @staticmethod
    def mutate(
        root,
        info,
        perspective_id,
        debug_flag = False,
        synchronous = False):

        try:

            client_id = info.context.get('client_id')
            client = DBSession.query(Client).filter_by(id = client_id).first()

            if not client:

                return (

                    ResponseError(
                        message = 'Only registered users can compute valency information.'))

            if debug_flag and client.user_id != 1:

                return (

                    ResponseError(
                        message = 'Only administrator can use debug mode.'))

            perspective = (
                DBSession.query(dbPerspective).filter_by(
                    client_id = perspective_id[0], object_id = perspective_id[1]).first())

            if not perspective:

                return (

                    ResponseError(
                        message = 'No perspective {}/{} in the system.'.format(*perspective_id)))

            dictionary = perspective.parent

            locale_id = info.context.get('locale_id') or 2

            dictionary_name = dictionary.get_translation(locale_id)
            perspective_name = perspective.get_translation(locale_id)

            full_name = dictionary_name + ' \u203a ' + perspective_name

            if dictionary.marked_for_deletion:

                return (

                    ResponseError(message =
                        'Dictionary \'{}\' {}/{} is deleted.'.format(
                            dictionary_name,
                            perspective.parent_client_id,
                            perspective.parent_object_id)))

            if perspective.marked_for_deletion:

                return (

                    ResponseError(message =
                        'Perspective \'{}\' {}/{} is deleted.'.format(
                            full_name,
                            perspective_id[0],
                            perspective_id[1])))

            if not synchronous:

                task = TaskStatus(client.user_id, 'Valency', 'Valency: ' + full_name, 1)

            settings = info.context.request.registry.settings

            (Valency.compute if synchronous else async_valency_compute.delay)(
                perspective_id,
                debug_flag,
                full_name,
                task.key if not synchronous else None,
                settings['storage'],
                settings['cache_kwargs'],
                settings['sqlalchemy.url'])

            return (

                Valency(
                    triumph = True))

        except Exception as exception:

            traceback_string = (

                ''.join(
                    traceback.format_exception(
                        exception, exception, exception.__traceback__))[:-1])

            log.warning('valency: exception')
            log.warning(traceback_string)

            return (

                ResponseError(
                    'Exception:\n' + traceback_string))


diacritic_re = (

    re.compile(
        ''.join([
            '[', '\u0300', '\u0301', '\u0302', '\u0303', '\u0304', '\u0305', '\u0306', '\u0307',
            '\u0308', '\u0309', '\u030a', '\u030b', '\u030c', '\u030d', '\u030e', '\u030f', '\u0310',
            '\u0311', '\u0312', '\u0313', '\u0314', '\u0315', '\u0316', '\u0317', '\u0318', '\u0319',
            '\u031a', '\u031b', '\u031c', '\u031d', '\u031e', '\u031f', '\u0320', '\u0321', '\u0322',
            '\u0323', '\u0324', '\u0325', '\u0326', '\u0327', '\u0328', '\u0329', '\u032a', '\u032b',
            '\u032c', '\u032d', '\u032e', '\u032f', '\u0330', '\u0331', '\u0332', '\u0333', '\u0334',
            '\u0335', '\u0336', '\u0337', '\u0338', '\u0339', '\u033a', '\u033b', '\u033c', '\u033d',
            '\u033e', '\u033f', '\u0340', '\u0341', '\u0342', '\u0343', '\u0344', '\u0345', '\u0346',
            '\u0347', '\u0348', '\u0349', '\u034a', '\u034b', '\u034c', '\u034d', '\u034e', '\u034f',
            '\u0350', '\u0351', '\u0352', '\u0353', '\u0354', '\u0355', '\u0356', '\u0357', '\u0358',
            '\u0359', '\u035a', '\u035b', '\u035c', '\u035d', '\u035e', '\u035f', '\u0360', '\u0361',
            '\u0362', '\u0363', '\u0364', '\u0365', '\u0366', '\u0367', '\u0368', '\u0369', '\u036a',
            '\u036b', '\u036c', '\u036d', '\u036e', '\u036f', '\u0483', '\u0484', '\u0485', '\u0486',
            '\u0487', '\u0488', '\u0489', '\u0591', '\u0592', '\u0593', '\u0594', '\u0595', '\u0596',
            '\u0597', '\u0598', '\u0599', '\u059a', '\u059b', '\u059c', '\u059d', '\u059e', '\u059f',
            '\u05a0', '\u05a1', '\u05a2', '\u05a3', '\u05a4', '\u05a5', '\u05a6', '\u05a7', '\u05a8',
            '\u05a9', '\u05aa', '\u05ab', '\u05ac', '\u05ad', '\u05ae', '\u05af', '\u05b0', '\u05b1',
            '\u05b2', '\u05b3', '\u05b4', '\u05b5', '\u05b6', '\u05b7', '\u05b8', '\u05b9', '\u05ba',
            '\u05bb', '\u05bc', '\u05bd', '\u05bf', '\u05c1', '\u05c2', '\u05c4', '\u05c5', '\u05c7',
            '\u0610', '\u0611', '\u0612', '\u0613', '\u0614', '\u0615', '\u0616', '\u0617', '\u0618',
            '\u0619', '\u061a', '\u064b', '\u064c', '\u064d', '\u064e', '\u064f', '\u0650', '\u0651',
            '\u0652', '\u0653', '\u0654', '\u0655', '\u0656', '\u0657', '\u0658', '\u0659', '\u065a',
            '\u065b', '\u065c', '\u065d', '\u065e', '\u065f', '\u0670', '\u06d6', '\u06d7', '\u06d8',
            '\u06d9', '\u06da', '\u06db', '\u06dc', '\u06df', '\u06e0', '\u06e1', '\u06e2', '\u06e3',
            '\u06e4', '\u06e7', '\u06e8', '\u06ea', '\u06eb', '\u06ec', '\u06ed', '\u0711', '\u0730',
            '\u0731', '\u0732', '\u0733', '\u0734', '\u0735', '\u0736', '\u0737', '\u0738', '\u0739',
            '\u073a', '\u073b', '\u073c', '\u073d', '\u073e', '\u073f', '\u0740', '\u0741', '\u0742',
            '\u0743', '\u0744', '\u0745', '\u0746', '\u0747', '\u0748', '\u0749', '\u074a', '\u07a6',
            '\u07a7', '\u07a8', '\u07a9', '\u07aa', '\u07ab', '\u07ac', '\u07ad', '\u07ae', '\u07af',
            '\u07b0', '\u07eb', '\u07ec', '\u07ed', '\u07ee', '\u07ef', '\u07f0', '\u07f1', '\u07f2',
            '\u07f3', '\u0816', '\u0817', '\u0818', '\u0819', '\u081b', '\u081c', '\u081d', '\u081e',
            '\u081f', '\u0820', '\u0821', '\u0822', '\u0823', '\u0825', '\u0826', '\u0827', '\u0829',
            '\u082a', '\u082b', '\u082c', '\u082d', '\u0859', '\u085a', '\u085b', '\u08d4', '\u08d5',
            '\u08d6', '\u08d7', '\u08d8', '\u08d9', '\u08da', '\u08db', '\u08dc', '\u08dd', '\u08de',
            '\u08df', '\u08e0', '\u08e1', '\u08e3', '\u08e4', '\u08e5', '\u08e6', '\u08e7', '\u08e8',
            '\u08e9', '\u08ea', '\u08eb', '\u08ec', '\u08ed', '\u08ee', '\u08ef', '\u08f0', '\u08f1',
            '\u08f2', '\u08f3', '\u08f4', '\u08f5', '\u08f6', '\u08f7', '\u08f8', '\u08f9', '\u08fa',
            '\u08fb', '\u08fc', '\u08fd', '\u08fe', '\u08ff', '\u0900', '\u0901', '\u0902', '\u0903',
            '\u093a', '\u093b', '\u093c', '\u093e', '\u093f', '\u0940', '\u0941', '\u0942', '\u0943',
            '\u0944', '\u0945', '\u0946', '\u0947', '\u0948', '\u0949', '\u094a', '\u094b', '\u094c',
            '\u094d', '\u094e', '\u094f', '\u0951', '\u0952', '\u0953', '\u0954', '\u0955', '\u0956',
            '\u0957', '\u0962', '\u0963', '\u0981', '\u0982', '\u0983', '\u09bc', '\u09be', '\u09bf',
            '\u09c0', '\u09c1', '\u09c2', '\u09c3', '\u09c4', '\u09c7', '\u09c8', '\u09cb', '\u09cc',
            '\u09cd', '\u09d7', '\u09e2', '\u09e3', '\u0a01', '\u0a02', '\u0a03', '\u0a3c', '\u0a3e',
            '\u0a3f', '\u0a40', '\u0a41', '\u0a42', '\u0a47', '\u0a48', '\u0a4b', '\u0a4c', '\u0a4d',
            '\u0a51', '\u0a70', '\u0a71', '\u0a75', '\u0a81', '\u0a82', '\u0a83', '\u0abc', '\u0abe',
            '\u0abf', '\u0ac0', '\u0ac1', '\u0ac2', '\u0ac3', '\u0ac4', '\u0ac5', '\u0ac7', '\u0ac8',
            '\u0ac9', '\u0acb', '\u0acc', '\u0acd', '\u0ae2', '\u0ae3', '\u0b01', '\u0b02', '\u0b03',
            '\u0b3c', '\u0b3e', '\u0b3f', '\u0b40', '\u0b41', '\u0b42', '\u0b43', '\u0b44', '\u0b47',
            '\u0b48', '\u0b4b', '\u0b4c', '\u0b4d', '\u0b56', '\u0b57', '\u0b62', '\u0b63', '\u0b82',
            '\u0bbe', '\u0bbf', '\u0bc0', '\u0bc1', '\u0bc2', '\u0bc6', '\u0bc7', '\u0bc8', '\u0bca',
            '\u0bcb', '\u0bcc', '\u0bcd', '\u0bd7', '\u0c00', '\u0c01', '\u0c02', '\u0c03', '\u0c3e',
            '\u0c3f', '\u0c40', '\u0c41', '\u0c42', '\u0c43', '\u0c44', '\u0c46', '\u0c47', '\u0c48',
            '\u0c4a', '\u0c4b', '\u0c4c', '\u0c4d', '\u0c55', '\u0c56', '\u0c62', '\u0c63', '\u0c81',
            '\u0c82', '\u0c83', '\u0cbc', '\u0cbe', '\u0cbf', '\u0cc0', '\u0cc1', '\u0cc2', '\u0cc3',
            '\u0cc4', '\u0cc6', '\u0cc7', '\u0cc8', '\u0cca', '\u0ccb', '\u0ccc', '\u0ccd', '\u0cd5',
            '\u0cd6', '\u0ce2', '\u0ce3', '\u0d01', '\u0d02', '\u0d03', '\u0d3e', '\u0d3f', '\u0d40',
            '\u0d41', '\u0d42', '\u0d43', '\u0d44', '\u0d46', '\u0d47', '\u0d48', '\u0d4a', '\u0d4b',
            '\u0d4c', '\u0d4d', '\u0d57', '\u0d62', '\u0d63', '\u0d82', '\u0d83', '\u0dca', '\u0dcf',
            '\u0dd0', '\u0dd1', '\u0dd2', '\u0dd3', '\u0dd4', '\u0dd6', '\u0dd8', '\u0dd9', '\u0dda',
            '\u0ddb', '\u0ddc', '\u0ddd', '\u0dde', '\u0ddf', '\u0df2', '\u0df3', '\u0e31', '\u0e34',
            '\u0e35', '\u0e36', '\u0e37', '\u0e38', '\u0e39', '\u0e3a', '\u0e47', '\u0e48', '\u0e49',
            '\u0e4a', '\u0e4b', '\u0e4c', '\u0e4d', '\u0e4e', '\u0eb1', '\u0eb4', '\u0eb5', '\u0eb6',
            '\u0eb7', '\u0eb8', '\u0eb9', '\u0ebb', '\u0ebc', '\u0ec8', '\u0ec9', '\u0eca', '\u0ecb',
            '\u0ecc', '\u0ecd', '\u0f18', '\u0f19', '\u0f35', '\u0f37', '\u0f39', '\u0f3e', '\u0f3f',
            '\u0f71', '\u0f72', '\u0f73', '\u0f74', '\u0f75', '\u0f76', '\u0f77', '\u0f78', '\u0f79',
            '\u0f7a', '\u0f7b', '\u0f7c', '\u0f7d', '\u0f7e', '\u0f7f', '\u0f80', '\u0f81', '\u0f82',
            '\u0f83', '\u0f84', '\u0f86', '\u0f87', '\u0f8d', '\u0f8e', '\u0f8f', '\u0f90', '\u0f91',
            '\u0f92', '\u0f93', '\u0f94', '\u0f95', '\u0f96', '\u0f97', '\u0f99', '\u0f9a', '\u0f9b',
            '\u0f9c', '\u0f9d', '\u0f9e', '\u0f9f', '\u0fa0', '\u0fa1', '\u0fa2', '\u0fa3', '\u0fa4',
            '\u0fa5', '\u0fa6', '\u0fa7', '\u0fa8', '\u0fa9', '\u0faa', '\u0fab', '\u0fac', '\u0fad',
            '\u0fae', '\u0faf', '\u0fb0', '\u0fb1', '\u0fb2', '\u0fb3', '\u0fb4', '\u0fb5', '\u0fb6',
            '\u0fb7', '\u0fb8', '\u0fb9', '\u0fba', '\u0fbb', '\u0fbc', '\u0fc6', '\u102b', '\u102c',
            '\u102d', '\u102e', '\u102f', '\u1030', '\u1031', '\u1032', '\u1033', '\u1034', '\u1035',
            '\u1036', '\u1037', '\u1038', '\u1039', '\u103a', '\u103b', '\u103c', '\u103d', '\u103e',
            '\u1056', '\u1057', '\u1058', '\u1059', '\u105e', '\u105f', '\u1060', '\u1062', '\u1063',
            '\u1064', '\u1067', '\u1068', '\u1069', '\u106a', '\u106b', '\u106c', '\u106d', '\u1071',
            '\u1072', '\u1073', '\u1074', '\u1082', '\u1083', '\u1084', '\u1085', '\u1086', '\u1087',
            '\u1088', '\u1089', '\u108a', '\u108b', '\u108c', '\u108d', '\u108f', '\u109a', '\u109b',
            '\u109c', '\u109d', '\u135d', '\u135e', '\u135f', '\u1712', '\u1713', '\u1714', '\u1732',
            '\u1733', '\u1734', '\u1752', '\u1753', '\u1772', '\u1773', '\u17b4', '\u17b5', '\u17b6',
            '\u17b7', '\u17b8', '\u17b9', '\u17ba', '\u17bb', '\u17bc', '\u17bd', '\u17be', '\u17bf',
            '\u17c0', '\u17c1', '\u17c2', '\u17c3', '\u17c4', '\u17c5', '\u17c6', '\u17c7', '\u17c8',
            '\u17c9', '\u17ca', '\u17cb', '\u17cc', '\u17cd', '\u17ce', '\u17cf', '\u17d0', '\u17d1',
            '\u17d2', '\u17d3', '\u17dd', '\u180b', '\u180c', '\u180d', '\u1885', '\u1886', '\u18a9',
            '\u1920', '\u1921', '\u1922', '\u1923', '\u1924', '\u1925', '\u1926', '\u1927', '\u1928',
            '\u1929', '\u192a', '\u192b', '\u1930', '\u1931', '\u1932', '\u1933', '\u1934', '\u1935',
            '\u1936', '\u1937', '\u1938', '\u1939', '\u193a', '\u193b', '\u1a17', '\u1a18', '\u1a19',
            '\u1a1a', '\u1a1b', '\u1a55', '\u1a56', '\u1a57', '\u1a58', '\u1a59', '\u1a5a', '\u1a5b',
            '\u1a5c', '\u1a5d', '\u1a5e', '\u1a60', '\u1a61', '\u1a62', '\u1a63', '\u1a64', '\u1a65',
            '\u1a66', '\u1a67', '\u1a68', '\u1a69', '\u1a6a', '\u1a6b', '\u1a6c', '\u1a6d', '\u1a6e',
            '\u1a6f', '\u1a70', '\u1a71', '\u1a72', '\u1a73', '\u1a74', '\u1a75', '\u1a76', '\u1a77',
            '\u1a78', '\u1a79', '\u1a7a', '\u1a7b', '\u1a7c', '\u1a7f', '\u1ab0', '\u1ab1', '\u1ab2',
            '\u1ab3', '\u1ab4', '\u1ab5', '\u1ab6', '\u1ab7', '\u1ab8', '\u1ab9', '\u1aba', '\u1abb',
            '\u1abc', '\u1abd', '\u1abe', '\u1b00', '\u1b01', '\u1b02', '\u1b03', '\u1b04', '\u1b34',
            '\u1b35', '\u1b36', '\u1b37', '\u1b38', '\u1b39', '\u1b3a', '\u1b3b', '\u1b3c', '\u1b3d',
            '\u1b3e', '\u1b3f', '\u1b40', '\u1b41', '\u1b42', '\u1b43', '\u1b44', '\u1b6b', '\u1b6c',
            '\u1b6d', '\u1b6e', '\u1b6f', '\u1b70', '\u1b71', '\u1b72', '\u1b73', '\u1b80', '\u1b81',
            '\u1b82', '\u1ba1', '\u1ba2', '\u1ba3', '\u1ba4', '\u1ba5', '\u1ba6', '\u1ba7', '\u1ba8',
            '\u1ba9', '\u1baa', '\u1bab', '\u1bac', '\u1bad', '\u1be6', '\u1be7', '\u1be8', '\u1be9',
            '\u1bea', '\u1beb', '\u1bec', '\u1bed', '\u1bee', '\u1bef', '\u1bf0', '\u1bf1', '\u1bf2',
            '\u1bf3', '\u1c24', '\u1c25', '\u1c26', '\u1c27', '\u1c28', '\u1c29', '\u1c2a', '\u1c2b',
            '\u1c2c', '\u1c2d', '\u1c2e', '\u1c2f', '\u1c30', '\u1c31', '\u1c32', '\u1c33', '\u1c34',
            '\u1c35', '\u1c36', '\u1c37', '\u1cd0', '\u1cd1', '\u1cd2', '\u1cd4', '\u1cd5', '\u1cd6',
            '\u1cd7', '\u1cd8', '\u1cd9', '\u1cda', '\u1cdb', '\u1cdc', '\u1cdd', '\u1cde', '\u1cdf',
            '\u1ce0', '\u1ce1', '\u1ce2', '\u1ce3', '\u1ce4', '\u1ce5', '\u1ce6', '\u1ce7', '\u1ce8',
            '\u1ced', '\u1cf2', '\u1cf3', '\u1cf4', '\u1cf8', '\u1cf9', '\u1dc0', '\u1dc1', '\u1dc2',
            '\u1dc3', '\u1dc4', '\u1dc5', '\u1dc6', '\u1dc7', '\u1dc8', '\u1dc9', '\u1dca', '\u1dcb',
            '\u1dcc', '\u1dcd', '\u1dce', '\u1dcf', '\u1dd0', '\u1dd1', '\u1dd2', '\u1dd3', '\u1dd4',
            '\u1dd5', '\u1dd6', '\u1dd7', '\u1dd8', '\u1dd9', '\u1dda', '\u1ddb', '\u1ddc', '\u1ddd',
            '\u1dde', '\u1ddf', '\u1de0', '\u1de1', '\u1de2', '\u1de3', '\u1de4', '\u1de5', '\u1de6',
            '\u1de7', '\u1de8', '\u1de9', '\u1dea', '\u1deb', '\u1dec', '\u1ded', '\u1dee', '\u1def',
            '\u1df0', '\u1df1', '\u1df2', '\u1df3', '\u1df4', '\u1df5', '\u1dfb', '\u1dfc', '\u1dfd',
            '\u1dfe', '\u1dff', '\u20d0', '\u20d1', '\u20d2', '\u20d3', '\u20d4', '\u20d5', '\u20d6',
            '\u20d7', '\u20d8', '\u20d9', '\u20da', '\u20db', '\u20dc', '\u20dd', '\u20de', '\u20df',
            '\u20e0', '\u20e1', '\u20e2', '\u20e3', '\u20e4', '\u20e5', '\u20e6', '\u20e7', '\u20e8',
            '\u20e9', '\u20ea', '\u20eb', '\u20ec', '\u20ed', '\u20ee', '\u20ef', '\u20f0', '\u2cef',
            '\u2cf0', '\u2cf1', '\u2d7f', '\u2de0', '\u2de1', '\u2de2', '\u2de3', '\u2de4', '\u2de5',
            '\u2de6', '\u2de7', '\u2de8', '\u2de9', '\u2dea', '\u2deb', '\u2dec', '\u2ded', '\u2dee',
            '\u2def', '\u2df0', '\u2df1', '\u2df2', '\u2df3', '\u2df4', '\u2df5', '\u2df6', '\u2df7',
            '\u2df8', '\u2df9', '\u2dfa', '\u2dfb', '\u2dfc', '\u2dfd', '\u2dfe', '\u2dff', '\u302a',
            '\u302b', '\u302c', '\u302d', '\u302e', '\u302f', '\u3099', '\u309a', '\ua66f', '\ua670',
            '\ua671', '\ua672', '\ua674', '\ua675', '\ua676', '\ua677', '\ua678', '\ua679', '\ua67a',
            '\ua67b', '\ua67c', '\ua67d', '\ua69e', '\ua69f', '\ua6f0', '\ua6f1', '\ua802', '\ua806',
            '\ua80b', '\ua823', '\ua824', '\ua825', '\ua826', '\ua827', '\ua880', '\ua881', '\ua8b4',
            '\ua8b5', '\ua8b6', '\ua8b7', '\ua8b8', '\ua8b9', '\ua8ba', '\ua8bb', '\ua8bc', '\ua8bd',
            '\ua8be', '\ua8bf', '\ua8c0', '\ua8c1', '\ua8c2', '\ua8c3', '\ua8c4', '\ua8c5', '\ua8e0',
            '\ua8e1', '\ua8e2', '\ua8e3', '\ua8e4', '\ua8e5', '\ua8e6', '\ua8e7', '\ua8e8', '\ua8e9',
            '\ua8ea', '\ua8eb', '\ua8ec', '\ua8ed', '\ua8ee', '\ua8ef', '\ua8f0', '\ua8f1', '\ua926',
            '\ua927', '\ua928', '\ua929', '\ua92a', '\ua92b', '\ua92c', '\ua92d', '\ua947', '\ua948',
            '\ua949', '\ua94a', '\ua94b', '\ua94c', '\ua94d', '\ua94e', '\ua94f', '\ua950', '\ua951',
            '\ua952', '\ua953', '\ua980', '\ua981', '\ua982', '\ua983', '\ua9b3', '\ua9b4', '\ua9b5',
            '\ua9b6', '\ua9b7', '\ua9b8', '\ua9b9', '\ua9ba', '\ua9bb', '\ua9bc', '\ua9bd', '\ua9be',
            '\ua9bf', '\ua9c0', '\ua9e5', '\uaa29', '\uaa2a', '\uaa2b', '\uaa2c', '\uaa2d', '\uaa2e',
            '\uaa2f', '\uaa30', '\uaa31', '\uaa32', '\uaa33', '\uaa34', '\uaa35', '\uaa36', '\uaa43',
            '\uaa4c', '\uaa4d', '\uaa7b', '\uaa7c', '\uaa7d', '\uaab0', '\uaab2', '\uaab3', '\uaab4',
            '\uaab7', '\uaab8', '\uaabe', '\uaabf', '\uaac1', '\uaaeb', '\uaaec', '\uaaed', '\uaaee',
            '\uaaef', '\uaaf5', '\uaaf6', '\uabe3', '\uabe4', '\uabe5', '\uabe6', '\uabe7', '\uabe8',
            '\uabe9', '\uabea', '\uabec', '\uabed', '\ufb1e', '\ufe00', '\ufe01', '\ufe02', '\ufe03',
            '\ufe04', '\ufe05', '\ufe06', '\ufe07', '\ufe08', '\ufe09', '\ufe0a', '\ufe0b', '\ufe0c',
            '\ufe0d', '\ufe0e', '\ufe0f', '\ufe20', '\ufe21', '\ufe22', '\ufe23', '\ufe24', '\ufe25',
            '\ufe26', '\ufe27', '\ufe28', '\ufe29', '\ufe2a', '\ufe2b', '\ufe2c', '\ufe2d', '\ufe2e',
            '\ufe2f', '\U000101fd', '\U000102e0', '\U00010376', '\U00010377', '\U00010378',
            '\U00010379', '\U0001037a', '\U00010a01', '\U00010a02', '\U00010a03', '\U00010a05',
            '\U00010a06', '\U00010a0c', '\U00010a0d', '\U00010a0e', '\U00010a0f', '\U00010a38',
            '\U00010a39', '\U00010a3a', '\U00010a3f', '\U00010ae5', '\U00010ae6', '\U00011000',
            '\U00011001', '\U00011002', '\U00011038', '\U00011039', '\U0001103a', '\U0001103b',
            '\U0001103c', '\U0001103d', '\U0001103e', '\U0001103f', '\U00011040', '\U00011041',
            '\U00011042', '\U00011043', '\U00011044', '\U00011045', '\U00011046', '\U0001107f',
            '\U00011080', '\U00011081', '\U00011082', '\U000110b0', '\U000110b1', '\U000110b2',
            '\U000110b3', '\U000110b4', '\U000110b5', '\U000110b6', '\U000110b7', '\U000110b8',
            '\U000110b9', '\U000110ba', '\U00011100', '\U00011101', '\U00011102', '\U00011127',
            '\U00011128', '\U00011129', '\U0001112a', '\U0001112b', '\U0001112c', '\U0001112d',
            '\U0001112e', '\U0001112f', '\U00011130', '\U00011131', '\U00011132', '\U00011133',
            '\U00011134', '\U00011173', '\U00011180', '\U00011181', '\U00011182', '\U000111b3',
            '\U000111b4', '\U000111b5', '\U000111b6', '\U000111b7', '\U000111b8', '\U000111b9',
            '\U000111ba', '\U000111bb', '\U000111bc', '\U000111bd', '\U000111be', '\U000111bf',
            '\U000111c0', '\U000111ca', '\U000111cb', '\U000111cc', '\U0001122c', '\U0001122d',
            '\U0001122e', '\U0001122f', '\U00011230', '\U00011231', '\U00011232', '\U00011233',
            '\U00011234', '\U00011235', '\U00011236', '\U00011237', '\U0001123e', '\U000112df',
            '\U000112e0', '\U000112e1', '\U000112e2', '\U000112e3', '\U000112e4', '\U000112e5',
            '\U000112e6', '\U000112e7', '\U000112e8', '\U000112e9', '\U000112ea', '\U00011300',
            '\U00011301', '\U00011302', '\U00011303', '\U0001133c', '\U0001133e', '\U0001133f',
            '\U00011340', '\U00011341', '\U00011342', '\U00011343', '\U00011344', '\U00011347',
            '\U00011348', '\U0001134b', '\U0001134c', '\U0001134d', '\U00011357', '\U00011362',
            '\U00011363', '\U00011366', '\U00011367', '\U00011368', '\U00011369', '\U0001136a',
            '\U0001136b', '\U0001136c', '\U00011370', '\U00011371', '\U00011372', '\U00011373',
            '\U00011374', '\U00011435', '\U00011436', '\U00011437', '\U00011438', '\U00011439',
            '\U0001143a', '\U0001143b', '\U0001143c', '\U0001143d', '\U0001143e', '\U0001143f',
            '\U00011440', '\U00011441', '\U00011442', '\U00011443', '\U00011444', '\U00011445',
            '\U00011446', '\U000114b0', '\U000114b1', '\U000114b2', '\U000114b3', '\U000114b4',
            '\U000114b5', '\U000114b6', '\U000114b7', '\U000114b8', '\U000114b9', '\U000114ba',
            '\U000114bb', '\U000114bc', '\U000114bd', '\U000114be', '\U000114bf', '\U000114c0',
            '\U000114c1', '\U000114c2', '\U000114c3', '\U000115af', '\U000115b0', '\U000115b1',
            '\U000115b2', '\U000115b3', '\U000115b4', '\U000115b5', '\U000115b8', '\U000115b9',
            '\U000115ba', '\U000115bb', '\U000115bc', '\U000115bd', '\U000115be', '\U000115bf',
            '\U000115c0', '\U000115dc', '\U000115dd', '\U00011630', '\U00011631', '\U00011632',
            '\U00011633', '\U00011634', '\U00011635', '\U00011636', '\U00011637', '\U00011638',
            '\U00011639', '\U0001163a', '\U0001163b', '\U0001163c', '\U0001163d', '\U0001163e',
            '\U0001163f', '\U00011640', '\U000116ab', '\U000116ac', '\U000116ad', '\U000116ae',
            '\U000116af', '\U000116b0', '\U000116b1', '\U000116b2', '\U000116b3', '\U000116b4',
            '\U000116b5', '\U000116b6', '\U000116b7', '\U0001171d', '\U0001171e', '\U0001171f',
            '\U00011720', '\U00011721', '\U00011722', '\U00011723', '\U00011724', '\U00011725',
            '\U00011726', '\U00011727', '\U00011728', '\U00011729', '\U0001172a', '\U0001172b',
            '\U00011c2f', '\U00011c30', '\U00011c31', '\U00011c32', '\U00011c33', '\U00011c34',
            '\U00011c35', '\U00011c36', '\U00011c38', '\U00011c39', '\U00011c3a', '\U00011c3b',
            '\U00011c3c', '\U00011c3d', '\U00011c3e', '\U00011c3f', '\U00011c92', '\U00011c93',
            '\U00011c94', '\U00011c95', '\U00011c96', '\U00011c97', '\U00011c98', '\U00011c99',
            '\U00011c9a', '\U00011c9b', '\U00011c9c', '\U00011c9d', '\U00011c9e', '\U00011c9f',
            '\U00011ca0', '\U00011ca1', '\U00011ca2', '\U00011ca3', '\U00011ca4', '\U00011ca5',
            '\U00011ca6', '\U00011ca7', '\U00011ca9', '\U00011caa', '\U00011cab', '\U00011cac',
            '\U00011cad', '\U00011cae', '\U00011caf', '\U00011cb0', '\U00011cb1', '\U00011cb2',
            '\U00011cb3', '\U00011cb4', '\U00011cb5', '\U00011cb6', '\U00016af0', '\U00016af1',
            '\U00016af2', '\U00016af3', '\U00016af4', '\U00016b30', '\U00016b31', '\U00016b32',
            '\U00016b33', '\U00016b34', '\U00016b35', '\U00016b36', '\U00016f51', '\U00016f52',
            '\U00016f53', '\U00016f54', '\U00016f55', '\U00016f56', '\U00016f57', '\U00016f58',
            '\U00016f59', '\U00016f5a', '\U00016f5b', '\U00016f5c', '\U00016f5d', '\U00016f5e',
            '\U00016f5f', '\U00016f60', '\U00016f61', '\U00016f62', '\U00016f63', '\U00016f64',
            '\U00016f65', '\U00016f66', '\U00016f67', '\U00016f68', '\U00016f69', '\U00016f6a',
            '\U00016f6b', '\U00016f6c', '\U00016f6d', '\U00016f6e', '\U00016f6f', '\U00016f70',
            '\U00016f71', '\U00016f72', '\U00016f73', '\U00016f74', '\U00016f75', '\U00016f76',
            '\U00016f77', '\U00016f78', '\U00016f79', '\U00016f7a', '\U00016f7b', '\U00016f7c',
            '\U00016f7d', '\U00016f7e', '\U00016f8f', '\U00016f90', '\U00016f91', '\U00016f92',
            '\U0001bc9d', '\U0001bc9e', '\U0001d165', '\U0001d166', '\U0001d167', '\U0001d168',
            '\U0001d169', '\U0001d16d', '\U0001d16e', '\U0001d16f', '\U0001d170', '\U0001d171',
            '\U0001d172', '\U0001d17b', '\U0001d17c', '\U0001d17d', '\U0001d17e', '\U0001d17f',
            '\U0001d180', '\U0001d181', '\U0001d182', '\U0001d185', '\U0001d186', '\U0001d187',
            '\U0001d188', '\U0001d189', '\U0001d18a', '\U0001d18b', '\U0001d1aa', '\U0001d1ab',
            '\U0001d1ac', '\U0001d1ad', '\U0001d242', '\U0001d243', '\U0001d244', '\U0001da00',
            '\U0001da01', '\U0001da02', '\U0001da03', '\U0001da04', '\U0001da05', '\U0001da06',
            '\U0001da07', '\U0001da08', '\U0001da09', '\U0001da0a', '\U0001da0b', '\U0001da0c',
            '\U0001da0d', '\U0001da0e', '\U0001da0f', '\U0001da10', '\U0001da11', '\U0001da12',
            '\U0001da13', '\U0001da14', '\U0001da15', '\U0001da16', '\U0001da17', '\U0001da18',
            '\U0001da19', '\U0001da1a', '\U0001da1b', '\U0001da1c', '\U0001da1d', '\U0001da1e',
            '\U0001da1f', '\U0001da20', '\U0001da21', '\U0001da22', '\U0001da23', '\U0001da24',
            '\U0001da25', '\U0001da26', '\U0001da27', '\U0001da28', '\U0001da29', '\U0001da2a',
            '\U0001da2b', '\U0001da2c', '\U0001da2d', '\U0001da2e', '\U0001da2f', '\U0001da30',
            '\U0001da31', '\U0001da32', '\U0001da33', '\U0001da34', '\U0001da35', '\U0001da36',
            '\U0001da3b', '\U0001da3c', '\U0001da3d', '\U0001da3e', '\U0001da3f', '\U0001da40',
            '\U0001da41', '\U0001da42', '\U0001da43', '\U0001da44', '\U0001da45', '\U0001da46',
            '\U0001da47', '\U0001da48', '\U0001da49', '\U0001da4a', '\U0001da4b', '\U0001da4c',
            '\U0001da4d', '\U0001da4e', '\U0001da4f', '\U0001da50', '\U0001da51', '\U0001da52',
            '\U0001da53', '\U0001da54', '\U0001da55', '\U0001da56', '\U0001da57', '\U0001da58',
            '\U0001da59', '\U0001da5a', '\U0001da5b', '\U0001da5c', '\U0001da5d', '\U0001da5e',
            '\U0001da5f', '\U0001da60', '\U0001da61', '\U0001da62', '\U0001da63', '\U0001da64',
            '\U0001da65', '\U0001da66', '\U0001da67', '\U0001da68', '\U0001da69', '\U0001da6a',
            '\U0001da6b', '\U0001da6c', '\U0001da75', '\U0001da84', '\U0001da9b', '\U0001da9c',
            '\U0001da9d', '\U0001da9e', '\U0001da9f', '\U0001daa1', '\U0001daa2', '\U0001daa3',
            '\U0001daa4', '\U0001daa5', '\U0001daa6', '\U0001daa7', '\U0001daa8', '\U0001daa9',
            '\U0001daaa', '\U0001daab', '\U0001daac', '\U0001daad', '\U0001daae', '\U0001daaf',
            '\U0001e000', '\U0001e001', '\U0001e002', '\U0001e003', '\U0001e004', '\U0001e005',
            '\U0001e006', '\U0001e008', '\U0001e009', '\U0001e00a', '\U0001e00b', '\U0001e00c',
            '\U0001e00d', '\U0001e00e', '\U0001e00f', '\U0001e010', '\U0001e011', '\U0001e012',
            '\U0001e013', '\U0001e014', '\U0001e015', '\U0001e016', '\U0001e017', '\U0001e018',
            '\U0001e01b', '\U0001e01c', '\U0001e01d', '\U0001e01e', '\U0001e01f', '\U0001e020',
            '\U0001e021', '\U0001e023', '\U0001e024', '\U0001e026', '\U0001e027', '\U0001e028',
            '\U0001e029', '\U0001e02a', '\U0001e8d0', '\U0001e8d1', '\U0001e8d2', '\U0001e8d3',
            '\U0001e8d4', '\U0001e8d5', '\U0001e8d6', '\U0001e944', '\U0001e945', '\U0001e946',
            '\U0001e947', '\U0001e948', '\U0001e949', '\U0001e94a', '\U000e0100', '\U000e0101',
            '\U000e0102', '\U000e0103', '\U000e0104', '\U000e0105', '\U000e0106', '\U000e0107',
            '\U000e0108', '\U000e0109', '\U000e010a', '\U000e010b', '\U000e010c', '\U000e010d',
            '\U000e010e', '\U000e010f', '\U000e0110', '\U000e0111', '\U000e0112', '\U000e0113',
            '\U000e0114', '\U000e0115', '\U000e0116', '\U000e0117', '\U000e0118', '\U000e0119',
            '\U000e011a', '\U000e011b', '\U000e011c', '\U000e011d', '\U000e011e', '\U000e011f',
            '\U000e0120', '\U000e0121', '\U000e0122', '\U000e0123', '\U000e0124', '\U000e0125',
            '\U000e0126', '\U000e0127', '\U000e0128', '\U000e0129', '\U000e012a', '\U000e012b',
            '\U000e012c', '\U000e012d', '\U000e012e', '\U000e012f', '\U000e0130', '\U000e0131',
            '\U000e0132', '\U000e0133', '\U000e0134', '\U000e0135', '\U000e0136', '\U000e0137',
            '\U000e0138', '\U000e0139', '\U000e013a', '\U000e013b', '\U000e013c', '\U000e013d',
            '\U000e013e', '\U000e013f', '\U000e0140', '\U000e0141', '\U000e0142', '\U000e0143',
            '\U000e0144', '\U000e0145', '\U000e0146', '\U000e0147', '\U000e0148', '\U000e0149',
            '\U000e014a', '\U000e014b', '\U000e014c', '\U000e014d', '\U000e014e', '\U000e014f',
            '\U000e0150', '\U000e0151', '\U000e0152', '\U000e0153', '\U000e0154', '\U000e0155',
            '\U000e0156', '\U000e0157', '\U000e0158', '\U000e0159', '\U000e015a', '\U000e015b',
            '\U000e015c', '\U000e015d', '\U000e015e', '\U000e015f', '\U000e0160', '\U000e0161',
            '\U000e0162', '\U000e0163', '\U000e0164', '\U000e0165', '\U000e0166', '\U000e0167',
            '\U000e0168', '\U000e0169', '\U000e016a', '\U000e016b', '\U000e016c', '\U000e016d',
            '\U000e016e', '\U000e016f', '\U000e0170', '\U000e0171', '\U000e0172', '\U000e0173',
            '\U000e0174', '\U000e0175', '\U000e0176', '\U000e0177', '\U000e0178', '\U000e0179',
            '\U000e017a', '\U000e017b', '\U000e017c', '\U000e017d', '\U000e017e', '\U000e017f',
            '\U000e0180', '\U000e0181', '\U000e0182', '\U000e0183', '\U000e0184', '\U000e0185',
            '\U000e0186', '\U000e0187', '\U000e0188', '\U000e0189', '\U000e018a', '\U000e018b',
            '\U000e018c', '\U000e018d', '\U000e018e', '\U000e018f', '\U000e0190', '\U000e0191',
            '\U000e0192', '\U000e0193', '\U000e0194', '\U000e0195', '\U000e0196', '\U000e0197',
            '\U000e0198', '\U000e0199', '\U000e019a', '\U000e019b', '\U000e019c', '\U000e019d',
            '\U000e019e', '\U000e019f', '\U000e01a0', '\U000e01a1', '\U000e01a2', '\U000e01a3',
            '\U000e01a4', '\U000e01a5', '\U000e01a6', '\U000e01a7', '\U000e01a8', '\U000e01a9',
            '\U000e01aa', '\U000e01ab', '\U000e01ac', '\U000e01ad', '\U000e01ae', '\U000e01af',
            '\U000e01b0', '\U000e01b1', '\U000e01b2', '\U000e01b3', '\U000e01b4', '\U000e01b5',
            '\U000e01b6', '\U000e01b7', '\U000e01b8', '\U000e01b9', '\U000e01ba', '\U000e01bb',
            '\U000e01bc', '\U000e01bd', '\U000e01be', '\U000e01bf', '\U000e01c0', '\U000e01c1',
            '\U000e01c2', '\U000e01c3', '\U000e01c4', '\U000e01c5', '\U000e01c6', '\U000e01c7',
            '\U000e01c8', '\U000e01c9', '\U000e01ca', '\U000e01cb', '\U000e01cc', '\U000e01cd',
            '\U000e01ce', '\U000e01cf', '\U000e01d0', '\U000e01d1', '\U000e01d2', '\U000e01d3',
            '\U000e01d4', '\U000e01d5', '\U000e01d6', '\U000e01d7', '\U000e01d8', '\U000e01d9',
            '\U000e01da', '\U000e01db', '\U000e01dc', '\U000e01dd', '\U000e01de', '\U000e01df',
            '\U000e01e0', '\U000e01e1', '\U000e01e2', '\U000e01e3', '\U000e01e4', '\U000e01e5',
            '\U000e01e6', '\U000e01e7', '\U000e01e8', '\U000e01e9', '\U000e01ea', '\U000e01eb',
            '\U000e01ec', '\U000e01ed', '\U000e01ee', '\U000e01ef', ']'])))

def diacritic_xform(value_str):

    return (

        re.sub(
            diacritic_re,
            '',
            unicodedata.normalize('NFKD', value_str)))


class CreateValencyData(graphene.Mutation):

    case_list = [
        'nom', 'acc', 'gen', 'ad', 'abl', 'dat', 'ab', 'ins', 'car', 'term', 'cns', 'com', 'comp',
        'trans', 'sim', 'par', 'loc', 'prol', 'in', 'ill', 'el', 'egr',  'lat', 'allat']

    case_index_dict = {
        case: index
        for index, case in enumerate(case_list)}

    class Arguments:

        perspective_id = LingvodocID(required = True)
        debug_flag = graphene.Boolean()

    triumph = graphene.Boolean()

    @staticmethod
    def align(
        token_list,
        word_list,
        debug_flag):
        """
        Aligns words to text tokens via Levenshtein matching.
        """

        token_chr_list = []

        for i, token in enumerate(token_list):
            for j, chr in enumerate(token):

                token_chr_list.append((i, j, chr.lower()))

        word_chr_list = []

        for i, word in enumerate(word_list):
            for j, chr in enumerate(word or ''):

                word_chr_list.append((i, j, chr.lower()))

        if debug_flag:

            log.debug(
                f'\ntoken_chr_list: {token_chr_list}'
                f'\nword_chr_list: {word_chr_list}'
                f'\n{"".join(token_list)}'
                f'\n{"".join(word or "" for word in word_list)}')

        # Standard 2-row Wagner-Fischer with addition of substitution tracking.

        value_list = [
            (i, None)
            for i in range(len(token_chr_list) + 1)]

        for i in range(len(word_chr_list)):

            before_list = value_list
            value_list = [(i + 1, None)]

            for j in range(len(token_chr_list)):

                delete_source = before_list[j + 1]
                delete_value = (delete_source[0] + 1, delete_source[1])

                insert_source = value_list[j]
                insert_value = (insert_source[0] + 1, insert_source[1])

                substitute_source = before_list[j]
                substitute_cost = substitute_source[0]

                if token_chr_list[j][2] != word_chr_list[i][2]:
                    substitute_cost += 1

                substitute_value = (

                    substitute_cost,
                    (token_chr_list[j][:2], word_chr_list[i][:2], substitute_source[1]))

                value_list.append(
                    min(delete_value, insert_value, substitute_value))

        result = (
            value_list[len(token_chr_list)])

        log.debug(f'\n{result[0]}')

        # Matching words to tokens by number of substitution and token ordering.

        map_tuple = result[1]
        map_list = []

        while map_tuple:

            map_list.append(map_tuple[:2])
            map_tuple = map_tuple[2]

        log_list = []

        map_counter = (
            collections.defaultdict(collections.Counter))

        for index_from, index_to in reversed(map_list):

            if debug_flag:

                token = token_list[index_from[0]]
                word = word_list[index_to[0]]

                log_list.append(
                    f'\n{index_from}, {index_to}: '
                    f'{token[: index_from[1]]}[{token[index_from[1]]}]{token[index_from[1] + 1 :]} / '
                    f'{word[: index_to[1]]}[{word[index_to[1]]}]{word[index_to[1] + 1 :]}')

            map_counter[index_to[0]][index_from[0]] += 1

        if debug_flag:

            log.debug(
                ''.join(log_list))

        word_token_dict = {}
        token_already_set = set()

        for word_index in range(len(word_list)):

            token_result = (

                max(
                    ((count, -token_index)
                        for token_index, count in map_counter[word_index].items()
                        if token_index not in token_already_set),
                    default = None))

            if token_result is not None:

                token_count, token_index_value = token_result

                token_index = -token_index_value

                word_token_dict[word_index] = token_index
                token_already_set.add(token_index)

        if debug_flag:

            log.debug(
                ''.join(
                    f'\n{repr(word_list[word_index])} ({word_index}) -> '
                    f'{repr(token_list[token_index])} ({token_index})'
                    for word_index, token_index in word_token_dict.items()))

        return (
            result[0], word_token_dict)

    @staticmethod
    def process_parser(
        perspective_id,
        data_case_set,
        instance_insert_list,
        debug_flag):

        # Getting parser result data.

        parser_result_list = (

            Valency.get_parser_result_data(
                perspective_id, debug_flag))

        sentence_data_list = (
            valency.corpus_to_sentences(parser_result_list))

        if debug_flag:

            parser_result_file_name = (
                f'create valency {perspective_id[0]} {perspective_id[1]} parser result.json')

            with open(
                parser_result_file_name, 'w') as parser_result_file:

                json.dump(
                    parser_result_list,
                    parser_result_file,
                    ensure_ascii = False,
                    indent = 2)

            sentence_data_file_name = (
                f'create valency {perspective_id[0]} {perspective_id[1]} sentence data.json')

            with open(
                sentence_data_file_name, 'w') as sentence_data_file:

                json.dump(
                    sentence_data_list,
                    sentence_data_file,
                    ensure_ascii = False,
                    indent = 2)

        # Initializing annotation data from parser results.

        for i in sentence_data_list:

            parser_result_id = i['id']

            # Checking if we already have such parser result valency data.

            valency_parser_data = (

                DBSession

                    .query(
                        dbValencyParserData)

                    .filter(
                        dbValencySourceData.perspective_client_id == perspective_id[0],
                        dbValencySourceData.perspective_object_id == perspective_id[1],
                        dbValencySourceData.id == dbValencyParserData.id,
                        dbValencyParserData.parser_result_client_id == parser_result_id[0],
                        dbValencyParserData.parser_result_object_id == parser_result_id[1])

                    .first())

            if valency_parser_data:

                # The same hash, we just skip it.

                if valency_parser_data.hash == i['hash']:
                    continue

                # Not the same hash, we actually should update it, but for now we leave it for later.

                continue

            valency_source_data = (

                dbValencySourceData(
                    perspective_client_id = perspective_id[0],
                    perspective_object_id = perspective_id[1]))

            DBSession.add(valency_source_data)
            DBSession.flush()

            valency_parser_data = (

                dbValencyParserData(
                    id = valency_source_data.id,
                    parser_result_client_id = parser_result_id[0],
                    parser_result_object_id = parser_result_id[1],
                    hash = i['hash']))

            DBSession.add(valency_parser_data)
            DBSession.flush()

            for p in i['paragraphs']:

                for s in p['sentences']:

                    instance_list = []

                    for index, (lex, cs, indent, ind, r, animacy) in (
                        enumerate(valency.sentence_instance_gen(s))):

                        instance_list.append({
                            'index': index,
                            'location': (ind, r),
                            'case': cs})

                        data_case_set.add(cs)

                    sentence_data = {
                        'tokens': s,
                        'instances': instance_list}

                    valency_sentence_data = (

                        dbValencySentenceData(
                            source_id = valency_source_data.id,
                            data = sentence_data,
                            instance_count = len(instance_list)))

                    DBSession.add(valency_sentence_data)
                    DBSession.flush()

                    for instance in instance_list:

                        instance_insert_list.append({
                            'sentence_id': valency_sentence_data.id,
                            'index': instance['index'],
                            'verb_lex': s[instance['location'][0]]['lex'].lower(),
                            'case_str': instance['case'].lower()})

                    log.debug(
                        '\n' +
                        pprint.pformat(
                            (valency_source_data.id, len(instance_list), sentence_data),
                            width = 192))

    @staticmethod
    def process_eaf(
        info,
        perspective_id,
        data_case_set,
        instance_insert_list,
        debug_flag):

        # Getting ELAN corpus data, processing each ELAN file.

        entity_list = (

            DBSession

                .query(
                    dbEntity)

                .filter(
                    dbLexicalEntry.parent_client_id == perspective_id[0],
                    dbLexicalEntry.parent_object_id == perspective_id[1],
                    dbLexicalEntry.marked_for_deletion == False,
                    dbEntity.parent_client_id == dbLexicalEntry.client_id,
                    dbEntity.parent_object_id == dbLexicalEntry.object_id,
                    dbEntity.marked_for_deletion == False,
                    dbEntity.content.ilike('%.eaf'),
                    dbEntity.additional_metadata.contains({'data_type': 'elan markup'}),
                    dbPublishingEntity.client_id == dbEntity.client_id,
                    dbPublishingEntity.object_id == dbEntity.object_id,
                    dbPublishingEntity.published == True,
                    dbPublishingEntity.accepted == True)

                .order_by(
                    dbLexicalEntry.created_at,
                    dbLexicalEntry.client_id,
                    dbLexicalEntry.object_id,
                    dbEntity.created_at,
                    dbEntity.client_id,
                    dbEntity.object_id)

                .all())

        storage = (
            info.context.request.registry.settings['storage'])

        storage_f = (
            as_storage_file if debug_flag else storage_file)

        delimiter_re = '[-\xad\xaf\x96\u2013\x97\u2014.]'

        verb_gloss_str_list = [

            'PRS',
            'FUT',
            f'1{delimiter_re}?PST',
            f'2{delimiter_re}?PST',

            'IMP',
            'COND',
            'SBJV',

            'CAUS',
            'REFL',
            'IMPERS',
            'ITER',

            'OPT',
            'INF',

            'NPST',
            'NST',
            'PST1',
            'PST2',
            'PST3']

        verb_re = (

            re.compile(
                f'{delimiter_re}\\b({"|".join(verb_gloss_str_list)})\\b',
                re.IGNORECASE))

        case_re = (

            re.compile(
                f'{delimiter_re}\\b({"|".join(CreateValencyData.case_list)})\\b',
                re.IGNORECASE))

        lex_xlat_dict = collections.defaultdict(set)
        xlat_lex_dict = collections.defaultdict(set)

        for entity in entity_list:

            # Checking if we already have such EAF corpus valency data.

            valency_eaf_data = (

                DBSession

                    .query(dbValencyEafData)

                    .filter(
                        dbValencySourceData.perspective_client_id == perspective_id[0],
                        dbValencySourceData.perspective_object_id == perspective_id[1],
                        dbValencySourceData.id == dbValencyEafData.id,
                        dbValencyEafData.entity_client_id == entity.client_id,
                        dbValencyEafData.entity_object_id == entity.object_id)

                    .first())

            if valency_eaf_data:

                # The same hash, we just skip it.

                if valency_eaf_data.hash == entity.additional_metadata['hash']:
                    continue

                # Not the same hash, we actually should update it, but for now we leave it for later.

                continue

            valency_source_data = (

                dbValencySourceData(
                    perspective_client_id = perspective_id[0],
                    perspective_object_id = perspective_id[1]))

            DBSession.add(valency_source_data)
            DBSession.flush()

            valency_eaf_data = (

                dbValencyEafData(
                    id = valency_source_data.id,
                    entity_client_id = entity.client_id,
                    entity_object_id = entity.object_id,
                    hash = entity.additional_metadata['hash']))

            DBSession.add(valency_eaf_data)
            DBSession.flush()

            create_valency_str = (
                f'create valency '
                f'{perspective_id[0]} {perspective_id[1]} '
                f'{entity.client_id} {entity.object_id}')

            # Getting and parsing corpus file.

            log.debug(
                f'\nentity.content:\n{entity.content}')

            try:
                with storage_f(storage, entity.content) as eaf_stream:
                    content = eaf_stream.read()

            except:
                raise ResponseError(f'Cannot access {entity.content}')

            with tempfile.NamedTemporaryFile() as temporary_file:

                temporary_file.write(
                    tgt_to_eaf(content, entity.additional_metadata).encode('utf-8'))

                temporary_file.flush()

                elan_check = elan_parser.ElanCheck(temporary_file.name)
                elan_check.parse()

                if not elan_check.check():
                    continue

                elan_reader = elan_parser.Elan(temporary_file.name)
                elan_reader.parse()

                eaf_data = elan_reader.proc()

                # Saving corpus file if required.

                if debug_flag:

                    shutil.copyfile(
                        temporary_file.name,
                        f'{create_valency_str}.eaf')

            if debug_flag:

                log.debug(
                    '\n' +
                    pprint.pformat(eaf_data, width = 192))

                with open(
                    f'{create_valency_str}.pprint', 'w') as pprint_file:

                    pprint_file.write(
                        pprint.pformat(eaf_data, width = 192))

            # Processing extracted data.

            for eaf_item in eaf_data:

                if len(eaf_item) < 3:
                    continue

                if debug_flag:

                    log.debug(
                        '\neaf_item:\n{}\n{}\n{}'.format(
                            pprint.pformat(eaf_item[0], width = 192),
                            pprint.pformat(eaf_item[1], width = 192),
                            pprint.pformat(eaf_item[2], width = 192)))

                if not isinstance(eaf_item[2], collections.OrderedDict):
                    continue

                # Text tokenization.

                if (len(eaf_item[0]) < 1 or
                    eaf_item[0][0].tier != 'text'):

                    raise NotImplementedError

                if eaf_item[0][0].text is None:
                    continue

                text_source_str = (

                    ' '.join(
                        item.text
                        for item in eaf_item[0]
                        if item.text))

                token_list = []

                for text_str in re.split(r'\s+', text_source_str):

                    last_index = 0

                    for match in re.finditer(r'\b\S+\b', text_str):

                        if match.start() > last_index:
                            token_list.append(text_str[last_index : match.start()])

                        if match.end() > match.start():
                            token_list.append(match.group())

                        last_index = match.end()

                    if last_index < len(text_str):
                        token_list.append(text_str[last_index :])

                xlat_list = []
                xcript_list = []
                word_list = []

                for key, value in eaf_item[2].items():

                    if (key.text is None and
                        all(item.text is None for item in value)):

                        continue

                    if key.tier != 'translation':

                        log.warning(f'\nkey: {key}')
                        raise NotImplementedError

                    xlat = key.text

                    xcript = None
                    word = None

                    for item in value:

                        if item.tier == 'transcription':
                            xcript = item.text

                        elif item.tier == 'word':
                            word = item.text

                    xlat_list.append(xlat)
                    xcript_list.append(xcript)
                    word_list.append(word)

                log.debug(
                    f'\ntext: {repr(eaf_item[0][0].text)}'
                    f'\ntoken_list: {token_list}'
                    f'\nxlat_list: {xlat_list}'
                    f'\nxcript_list: {xcript_list}'
                    f'\nword_list: {word_list}')

                # Aligning words to text tokens.

                xcript_alignment = (

                    CreateValencyData.align(
                        token_list, xcript_list, debug_flag))

                word_alignment = (

                    CreateValencyData.align(
                        token_list, word_list, debug_flag))

                (distance, word_token_dict) = (

                    min(
                        word_alignment,
                        xcript_alignment,
                        key = lambda x: (x[0], -len(x[1]))))

                log.debug(
                    f'\ndistance: {distance}'
                    f'\nword_token_dict: {word_token_dict}' +
                    ''.join(
                        f'\n{repr(word_list[word_index])} ({word_index}) -> '
                        f'{repr(token_list[token_index])} ({token_index})'
                        for word_index, token_index in word_token_dict.items()))

                token_word_dict = {
                    token_index: word_index
                    for word_index, token_index in word_token_dict.items()}

                # Constructing phrase's data.

                token_data_list = []

                for i, token in enumerate(token_list):

                    token_dict = {'token': token}
                    token_data_list.append(token_dict)

                    word_index = token_word_dict.get(i)

                    if word_index is None:
                        continue

                    xlat = xlat_list[word_index]

                    token_dict.update({
                        'translation': xlat,
                        'transcription': xcript_list[word_index],
                        'word': word_list[word_index]})

                    if xlat is None:
                        continue

                    if verb_re.search(xlat) is not None:
                        token_dict['gr'] = 'V'

                    case_match = case_re.search(xlat)

                    if case_match is not None:

                        case_str = (
                            case_match.group(1).lower())

                        if 'gr' in token_dict:
                            token_dict['gr'] += ',' + case_str

                        else:
                            token_dict['gr'] = case_str

                instance_list = []

                for index, (lex, cs, indent, ind, r, animacy) in (

                    enumerate(
                        valency.sentence_instance_gen(
                            token_data_list, False))):

                    instance_list.append({
                        'index': index,
                        'location': (ind, r),
                        'case': cs})

                    data_case_set.add(cs)

                sentence_data = {
                    'tokens': token_data_list,
                    'instances': instance_list}

                valency_sentence_data = (

                    dbValencySentenceData(
                        source_id = valency_source_data.id,
                        data = sentence_data,
                        instance_count = len(instance_list)))

                DBSession.add(valency_sentence_data)
                DBSession.flush()

                # Generating instance info.

                for instance in instance_list:

                    token = (
                        token_data_list[instance['location'][0]])

                    xcript = token['transcription']

                    xcript_split_list = (
                        re.split('(-|\u2013|\u2014)', xcript, maxsplit = 1))

                    xcript_str = (
                        xcript_split_list[0] if len(xcript_split_list) <= 1 else
                        xcript_split_list[0] + xcript_split_list[1])

                    match = (
                        re.search('([-.][\dA-Z]+)+$', xcript_str))

                    if match:
                        xcript_str = xcript_str[:match.start() + 1]

                    verb_lex = (

                        (xcript_str or
                            token['word'] or
                            token['token'] or
                            xcript or
                            '')

                            .strip()
                            .lower())

                    if not verb_lex:
                        continue

                    instance_insert_list.append({
                        'sentence_id': valency_sentence_data.id,
                        'index': instance['index'],
                        'verb_lex': verb_lex,
                        'case_str': instance['case'].lower()})

                    # Verb grouping by translations.

                    xlat = token['translation']
                    match = re.search('[-.][\dA-Z]+', xlat)

                    if match:
                        xlat = xlat[:match.start()]

                    lex_xlat_dict[verb_lex].add(xlat)
                    xlat_lex_dict[xlat].add(verb_lex)

                log.debug(
                    '\n' +
                    pprint.pformat(
                        (valency_source_data.id, len(instance_list), sentence_data),
                        width = 192))

        # Computing and saving translation-based verb mergings.
        #
        # Bipartite graph connected components depth-first search.

        if debug_flag:

            log.debug(

                f'\nlex_xlat_dict ({len(lex_xlat_dict)}):\n' +

                pprint.pformat(
                    lex_xlat_dict, width = 144) + 

                f'\nxlat_lex_dict ({len(xlat_lex_dict)}):\n' + 

                pprint.pformat(
                    xlat_lex_dict, width = 144))

        lex_set_list = []

        lex_index_set = set()
        xlat_index_set = set()

        def f(lex_from, lex_prefix, lex_set):

            lex_index_set.add(lex_from)
            lex_set.add(lex_from)

            for xlat in lex_xlat_dict[lex_from]:

                if xlat in xlat_index_set:
                    continue

                xlat_index_set.add(xlat)

                for lex_to in xlat_lex_dict[xlat]:

                    if lex_to in lex_index_set:
                        continue

                    if diacritic_xform(lex_to)[:2] != lex_prefix:
                        continue

                    f(lex_to, lex_prefix, lex_set)

        for lex in lex_xlat_dict.keys():

            if lex in lex_index_set:
                continue

            lex_prefix = (
                diacritic_xform(lex)[:2])

            lex_set = set()

            xlat_index_set.clear()

            f(lex, lex_prefix, lex_set)

            lex_set_list.append(lex_set)

        if debug_flag:

            lex_set_list = (

                sorted(
                    tuple(lex_set)
                    for lex_set in lex_set_list))

            log.debug(

                '\nlex_set_list:\n' +

                pprint.pformat(
                    lex_set_list, width = 144))

        merge_insert_list = []

        for lex_set in lex_set_list:

            if len(lex_set) <= 1:
                continue

            merge_id = (
                DBSession.execute(dbValencyMergeIdSequence))

            for verb_lex in lex_set:

                merge_insert_list.append({
                    'perspective_client_id': perspective_id[0],
                    'perspective_object_id': perspective_id[1],
                    'verb_lex': verb_lex,
                    'merge_id': merge_id})

        if merge_insert_list:

            DBSession.execute(

                dbValencyMergeData.__table__
                    .insert()
                    .values(merge_insert_list))

    @staticmethod
    def process(
        info,
        perspective_id,
        debug_flag):

        order_case_set = (

            set([
                'nom', 'acc', 'gen', 'ad', 'abl', 'dat', 'ab', 'ins', 'car', 'term', 'cns', 'com',
                'comp', 'trans', 'sim', 'par', 'loc', 'prol', 'in', 'ill', 'el', 'egr', 'lat',
                'allat']))

        data_case_set = set()
        instance_insert_list = []

        CreateValencyData.process_parser(
            perspective_id,
            data_case_set,
            instance_insert_list,
            debug_flag)

        CreateValencyData.process_eaf(
            info,
            perspective_id,
            data_case_set,
            instance_insert_list,
            debug_flag)

        if instance_insert_list:

            DBSession.execute(

                dbValencyInstanceData.__table__
                    .insert()
                    .values(instance_insert_list))

        log.debug(
            f'\ndata_case_set:\n{data_case_set}'
            f'\ndata_case_set - order_case_set:\n{data_case_set - order_case_set}'
            f'\norder_case_set - data_case_set:\n{order_case_set - data_case_set}')

        return len(instance_insert_list)

    @staticmethod
    def test(
        info, debug_flag):

        parser_result_query = (

            DBSession

                .query(
                    dbLexicalEntry.parent_client_id,
                    dbLexicalEntry.parent_object_id)

                .filter(
                    dbLexicalEntry.marked_for_deletion == False,
                    dbEntity.parent_client_id == dbLexicalEntry.client_id,
                    dbEntity.parent_object_id == dbLexicalEntry.object_id,
                    dbEntity.marked_for_deletion == False,
                    dbEntity.content.op('~*')('.*\.(doc|docx|odt)'),
                    dbPublishingEntity.client_id == dbEntity.client_id,
                    dbPublishingEntity.object_id == dbEntity.object_id,
                    dbPublishingEntity.published == True,
                    dbPublishingEntity.accepted == True,
                    dbParserResult.entity_client_id == dbEntity.client_id,
                    dbParserResult.entity_object_id == dbEntity.object_id,
                    dbParserResult.marked_for_deletion == False)

                .group_by(
                    dbLexicalEntry.parent_client_id,
                    dbLexicalEntry.parent_object_id))

        eaf_corpus_query = (

            DBSession

                .query(
                    dbLexicalEntry.parent_client_id,
                    dbLexicalEntry.parent_object_id)

                .filter(
                    dbLexicalEntry.marked_for_deletion == False,
                    dbEntity.parent_client_id == dbLexicalEntry.client_id,
                    dbEntity.parent_object_id == dbLexicalEntry.object_id,
                    dbEntity.marked_for_deletion == False,
                    dbEntity.content.ilike('%.eaf'),
                    dbEntity.additional_metadata.contains({'data_type': 'elan markup'}),
                    dbPublishingEntity.client_id == dbEntity.client_id,
                    dbPublishingEntity.object_id == dbEntity.object_id,
                    dbPublishingEntity.published == True,
                    dbPublishingEntity.accepted == True)

                .group_by(
                    dbLexicalEntry.parent_client_id,
                    dbLexicalEntry.parent_object_id))

        valency_data_query = (

            DBSession

                .query(
                    dbValencySourceData.perspective_client_id,
                    dbValencySourceData.perspective_object_id)

                .distinct())

        perspective_list = (

            DBSession

                .query(
                    dbPerspective)

                .filter(
                    dbPerspective.marked_for_deletion == False,

                    tuple_(
                        dbPerspective.client_id,
                        dbPerspective.object_id)

                        .notin_(
                            DBSession.query(valency_data_query.cte())),

                    tuple_(
                        dbPerspective.client_id,
                        dbPerspective.object_id)

                        .in_(
                            union(
                                DBSession.query(parser_result_query.cte()),
                                DBSession.query(eaf_corpus_query.cte()))))

                .order_by(
                    dbPerspective.client_id,
                    dbPerspective.object_id)

                .all())

        import random
        random.shuffle(perspective_list)

        for perspective in perspective_list:

            log.debug(
                f'\nperspective_id: {perspective.id}')

            CreateValencyData.process(
                info, perspective.id, debug_flag)

            if utils.get_resident_memory() > 2 * 2**30:
                break

    @staticmethod
    def mutate(
        root,
        info,
        perspective_id,
        debug_flag = False):

        try:

            client_id = info.context.get('client_id')
            client = DBSession.query(Client).filter_by(id = client_id).first()

            if not client:

                return (

                    ResponseError(
                        message = 'Only registered users can create valency data.'))

            if debug_flag and client.user_id != 1:

                return (

                    ResponseError(
                        message = 'Only administrator can use debug mode.'))

            perspective = (
                DBSession.query(dbPerspective).filter_by(
                    client_id = perspective_id[0], object_id = perspective_id[1]).first())

            if not perspective:

                return (

                    ResponseError(
                        message = 'No perspective {}/{} in the system.'.format(*perspective_id)))

            dictionary = perspective.parent

            locale_id = info.context.get('locale_id') or 2

            dictionary_name = dictionary.get_translation(locale_id)
            perspective_name = perspective.get_translation(locale_id)

            full_name = dictionary_name + ' \u203a ' + perspective_name

            if dictionary.marked_for_deletion:

                return (

                    ResponseError(message =
                        'Dictionary \'{}\' {}/{} of perspective \'{}\' {}/{} is deleted.'.format(
                            dictionary_name,
                            dictionary.client_id,
                            dictionary.object_id,
                            perspective_name,
                            perspective.client_id,
                            perspective.object_id)))

            if perspective.marked_for_deletion:

                return (

                    ResponseError(message =
                        'Perspective \'{}\' {}/{} is deleted.'.format(
                            full_name,
                            perspective.client_id,
                            perspective.object_id)))

            CreateValencyData.process(
                info,
                perspective_id,
                debug_flag)

            if False:

                CreateValencyData.test(
                    info,
                    debug_flag)

            return (

                CreateValencyData(
                    triumph = True))

        except Exception as exception:

            traceback_string = (

                ''.join(
                    traceback.format_exception(
                        exception, exception, exception.__traceback__))[:-1])

            log.warning('create_valency_data: exception')
            log.warning(traceback_string)

            transaction.abort()

            return (

                ResponseError(
                    'Exception:\n' + traceback_string))


class SaveValencyData(graphene.Mutation):

    class Arguments:

        perspective_id = LingvodocID(required = True)
        debug_flag = graphene.Boolean()

    triumph = graphene.Boolean()
    data_url = graphene.String()

    @staticmethod
    def get_annotation_data(perspective_id):

        # Getting valency annotation data.

        annotation_list = (

            DBSession

                .query(
                    dbValencyAnnotationData)

                .filter(
                    dbValencyAnnotationData.accepted != None,
                    dbValencyAnnotationData.instance_id == dbValencyInstanceData.id,
                    dbValencyInstanceData.sentence_id == dbValencySentenceData.id,
                    dbValencySentenceData.source_id == dbValencySourceData.id,
                    dbValencySourceData.perspective_client_id == perspective_id[0],
                    dbValencySourceData.perspective_object_id == perspective_id[1])

                .all())

        instance_id_set = set()
        user_id_set = set()

        for annotation in annotation_list:

            instance_id_set.add(annotation.instance_id)
            user_id_set.add(annotation.user_id)

        instance_list = []

        if instance_id_set:

            instance_list = (

                DBSession

                    .query(
                        dbValencyInstanceData)

                    .filter(
                        dbValencyInstanceData.id.in_(

                            utils.values_query(
                                instance_id_set, models.SLBigInteger)))

                    .all())

        user_list = []

        if user_id_set:

            user_list = (

                DBSession

                    .query(
                        dbUser.id, dbUser.name)

                    .filter(
                        dbUser.id.in_(

                            utils.values_query(
                                user_id_set, models.SLBigInteger)))

                    .all())

        sentence_id_set = (
            set(instance.sentence_id for instance in instance_list))

        sentence_list = []

        if sentence_id_set:

            sentence_list = (

                DBSession

                    .query(
                        dbValencySentenceData)

                    .filter(
                        dbValencySentenceData.id.in_(

                            utils.values_query(
                                sentence_id_set, models.SLBigInteger)))

                    .all())

        # Preparing valency annotation data.

        sentence_data_list = []

        for sentence in sentence_list:

            sentence_data = sentence.data
            sentence_data['id'] = sentence.id

            sentence_data_list.append(sentence_data)

        instance_data_list = [

            {'id': instance.id,
                'sentence_id': instance.sentence_id,
                'index': instance.index,
                'verb_lex': instance.verb_lex,
                'case_str': instance.case_str}

                for instance in instance_list]

        annotation_data_list = [

            {'instance_id': annotation.instance_id,
                'user_id': annotation.user_id,
                'accepted': annotation.accepted}

                for annotation in annotation_list]

        user_data_list = [

            {'id': user.id,
                'name': user.name}

                for user in user_list]

        return {
            'sentence_list': sentence_data_list,
            'instance_list': instance_data_list,
            'annotation_list': annotation_data_list,
            'user_list': user_data_list}

    @staticmethod
    def mutate(
        root,
        info,
        perspective_id,
        debug_flag = False):

        try:

            client_id = info.context.get('client_id')
            client = DBSession.query(Client).filter_by(id = client_id).first()

            if not client:

                return (

                    ResponseError(
                        message = 'Only registered users can create valency data.'))

            if debug_flag and client.user_id != 1:

                return (

                    ResponseError(
                        message = 'Only administrator can use debug mode.'))

            perspective = (
                DBSession.query(dbPerspective).filter_by(
                    client_id = perspective_id[0], object_id = perspective_id[1]).first())

            if not perspective:

                return (

                    ResponseError(
                        message = 'No perspective {}/{} in the system.'.format(*perspective_id)))

            dictionary = perspective.parent

            locale_id = info.context.get('locale_id') or 2

            dictionary_name = dictionary.get_translation(locale_id)
            perspective_name = perspective.get_translation(locale_id)

            full_name = dictionary_name + ' \u203a ' + perspective_name

            if dictionary.marked_for_deletion:

                return (

                    ResponseError(message =
                        'Dictionary \'{}\' {}/{} of perspective \'{}\' {}/{} is deleted.'.format(
                            dictionary_name,
                            dictionary.client_id,
                            dictionary.object_id,
                            perspective_name,
                            perspective.client_id,
                            perspective.object_id)))

            if perspective.marked_for_deletion:

                return (

                    ResponseError(message =
                        'Perspective \'{}\' {}/{} is deleted.'.format(
                            full_name,
                            perspective.client_id,
                            perspective.object_id)))

            # Getting valency annotation data.

            data_dict = (
                SaveValencyData.get_annotation_data(perspective_id))

            # Saving valency annotation data as zipped JSON.

            current_time = (
                time.time())

            current_date = (
                datetime.datetime.utcfromtimestamp(current_time))

            zip_date = (
                current_date.year,
                current_date.month,
                current_date.day,
                current_date.hour,
                current_date.minute,
                current_date.second)

            storage_temporary = (
                info.context.request.registry.settings['storage']['temporary'])

            host = storage_temporary['host']
            bucket = storage_temporary['bucket']

            minio_client = (
                    
                minio.Minio(
                    host,
                    access_key = storage_temporary['access_key'],
                    secret_key = storage_temporary['secret_key'],
                    secure = True))

            temporary_file = (
                    
                tempfile.NamedTemporaryFile(
                    delete = False))

            zip_file = (

                zipfile.ZipFile(
                    temporary_file,
                    'w',
                    compression = zipfile.ZIP_DEFLATED,
                    compresslevel = 9))

            zip_info = (

                zipfile.ZipInfo(
                    'data.json', zip_date))

            zip_info.compress_type = zipfile.ZIP_DEFLATED

            with zip_file.open(
                zip_info, 'w') as binary_data_file:

                with io.TextIOWrapper(
                    binary_data_file, 'utf-8') as text_data_file:

                    json.dump(
                        data_dict,
                        text_data_file,
                        ensure_ascii = False,
                        sort_keys = True,
                        indent = 2)

            zip_file.close()
            temporary_file.close()

            if debug_flag:

                shutil.copy(
                    temporary_file.name,
                    '__data__.json.zip')

            object_name = (

                storage_temporary['prefix'] +
            
                '/'.join((
                    'valency_data',
                    '{:.6f}'.format(current_time),
                    'data.json.zip')))

            (etag, version_id) = (

                minio_client.fput_object(
                    bucket,
                    object_name,
                    temporary_file.name))

            os.remove(
                temporary_file.name)

            url = (

                '/'.join((
                    'https:/',
                    host,
                    bucket,
                    object_name)))

            return (

                SaveValencyData(
                    triumph = True,
                    data_url = url))

        except Exception as exception:

            traceback_string = (

                ''.join(
                    traceback.format_exception(
                        exception, exception, exception.__traceback__))[:-1])

            log.warning('save_valency_data: exception')
            log.warning(traceback_string)

            transaction.abort()

            return (

                ResponseError(
                    'Exception:\n' + traceback_string))


class SetValencyAnnotation(graphene.Mutation):

    class ValencyInstanceAnnotation(graphene.types.Scalar):

        @staticmethod
        def identity(value):
            return value

        serialize = identity
        parse_value = identity

        @staticmethod
        def parse_literal(ast):

            if not isinstance(ast, ListValue) or len(ast.values) != 2:
                return None

            a_value, b_value = ast.values

            if (not isinstance(a_value, IntValue) or
                not isinstance(b_value, BooleanValue)):
                return None

            return [int(a_value.value), bool(b_value.value)]

    class Arguments:
        pass

    Arguments.annotation_list = (
        graphene.List(ValencyInstanceAnnotation, required = True))

    triumph = graphene.Boolean()

    @staticmethod
    def mutate(root, info, **args):

        try:

            client_id = info.context.get('client_id')
            client = DBSession.query(Client).filter_by(id = client_id).first()

            if not client:

                return (

                    ResponseError(
                        message = 'Only registered users can set valency annotations.'))

            annotation_list = args['annotation_list']

            log.debug(
                f'\nuser_id: {client.user_id}'
                f'\nannotation_list: {annotation_list}')

            # NOTE:
            #
            # Directly formatting arguments in in general can be unsafe, but here it's ok because we are
            # relying on GraphQL's argument validation.

            value_list_str = (

                ', '.join(
                    '({}, {}, {})'.format(
                        instance_id, client.user_id, 'true' if accepted else 'false')
                    for instance_id, accepted in annotation_list))

            sql_str = (

                f'''
                insert into
                valency_annotation_data
                values {value_list_str}
                on conflict on constraint valency_annotation_data_pkey
                do update set accepted = excluded.accepted;
                ''')

            DBSession.execute(sql_str)

            mark_changed(DBSession())

            return (

                SetValencyAnnotation(
                    triumph = True))

        except Exception as exception:

            traceback_string = (

                ''.join(
                    traceback.format_exception(
                        exception, exception, exception.__traceback__))[:-1])

            log.warning('set_valency_annotation: exception')
            log.warning(traceback_string)

            return (

                ResponseError(
                    'Exception:\n' + traceback_string))


class ValencyVerbCases(graphene.Mutation):
    """
    Compiles valency verb and cases info.
    """

    class Arguments:

        perspective_id = LingvodocID()
        language_arg_list = ObjectVal()

        debug_flag = graphene.Boolean()

    triumph = graphene.Boolean()
    verb_data_list = graphene.List(ObjectVal)
    xlsx_url = graphene.String()

    @staticmethod
    def process_perspective(
        perspective_id,
        locale_id,
        verb_data_dict,
        language_str,
        by_perspective_flag,
        debug_flag):
        """
        Gathers and accumulates valency verb case data of a perspective.
        """

        perspective = (

            DBSession

                .query(dbPerspective)

                .filter_by(
                    client_id = perspective_id[0],
                    object_id = perspective_id[1])

                .first())

        if not perspective:

            raise (

                ResponseError(
                    message = 'No perspective {}/{} in the system.'.format(*perspective_id)))

        dictionary = perspective.parent

        dictionary_name = dictionary.get_translation(locale_id)
        perspective_name = perspective.get_translation(locale_id)

        full_name = (
            dictionary_name + ' \u203a ' + perspective_name)

        if dictionary.marked_for_deletion:

            raise (

                ResponseError(message =
                    'Dictionary \'{}\' {}/{} of perspective \'{}\' is deleted.'.format(
                        dictionary_name,
                        dictionary.client_id,
                        dictionary.object_id,
                        perspective_name,
                        perspective.client_id,
                        perspective.object_id)))

        if perspective.marked_for_deletion:

            raise (

                ResponseError(message =
                    'Perspective \'{}\' {}/{} is deleted.'.format(
                        full_name,
                        perspective_id[0],
                        perspective_id[1])))

        if debug_flag:

            log.debug(
                '\nperspective '
                f'{dictionary.id} {repr(dictionary_name)} '
                f'{perspective.id} {repr(perspective_name)}')

        # We might need data separate by perspective.

        if by_perspective_flag:

            language_str = (
                language_str + ' \u203a\u203a ' + full_name)

        # Getting data of accepted instances and their sentences.
        #
        # Instance is accepted if it has at least one acceptance annotation.

        instance_list = (

            DBSession

                .query(
                    dbValencyInstanceData)

                .filter(
                    dbValencyAnnotationData.accepted == True,
                    dbValencyAnnotationData.instance_id == dbValencyInstanceData.id,
                    dbValencyInstanceData.sentence_id == dbValencySentenceData.id,
                    dbValencySentenceData.source_id == dbValencySourceData.id,
                    dbValencySourceData.perspective_client_id == perspective_id[0],
                    dbValencySourceData.perspective_object_id == perspective_id[1])

                .group_by(
                    dbValencyInstanceData.id)

                .all())

        sentence_id_set = (

            set(
                instance.sentence_id
                for instance in instance_list))

        sentence_list = []

        if sentence_id_set:

            sentence_list = (

                DBSession

                    .query(
                        dbValencySentenceData)

                    .filter(
                        dbValencySentenceData.id.in_(

                            utils.values_query(
                                sentence_id_set, models.SLBigInteger)))

                    .all())

        sentence_dict = {
            sentence.id: sentence
            for sentence in sentence_list}

        sentence_str_dict = {}

        # Processing each instance.

        for instance in instance_list:

            sentence_id = (
                instance.sentence_id)

            sentence = (
                sentence_dict[sentence_id])

            token_list = (
                sentence.data['tokens'])

            token_info_dict = (

                token_list[
                    sentence.data['instances'][instance.index]['location'][0]])

            verb_xlat = (

                token_info_dict.get('trans_ru') or
                token_info_dict.get('translation') or
                instance.verb_lex)

            sentence_str = (
                sentence_str_dict.get(sentence_id))

            if sentence_str is None:

                sentence_str = (

                    valency_verb_cases.get_sentence_str(
                        token_list))

                sentence_str_dict[sentence_id] = (
                    sentence_str)

            # Relying on Python dictionary insertion order preservation (assumption valid starting with
            # Python 3.7).

            verb_data_dict[
                verb_xlat][
                language_str][
                instance.case_str][
                instance.verb_lex][
                sentence_str] = None

    @staticmethod
    def get_case_verb_sentence_list(
        case_verb_dict):
        """
        Sorts case verb sentence data by cases and verbs.
        """

        case_verb_sentence_list = []

        for case, verb_sentence_dict in (

            sorted(
                case_verb_dict.items(),
                key = lambda item: CreateValencyData.case_index_dict[item[0]])):

            verb_sentence_list = (

                sorted(
                    (verb, list(sentence_dict.keys()))
                    for verb, sentence_dict in verb_sentence_dict.items()))

            verb_list = [
                verb
                for verb, sentence_list in verb_sentence_list]

            case_verb_sentence_list.append(
                (case, verb_list, verb_sentence_list))

        return case_verb_sentence_list

    @staticmethod
    def save_xlsx_file(
        info,
        workbook_stream,
        debug_flag):
        """
        Saves XLSX file.
        """

        if debug_flag:

            workbook_stream.seek(0)

            with open('valency_verb_cases.xlsx', 'wb') as xlsx_file:
                shutil.copyfileobj(workbook_stream, xlsx_file)

        storage = (
            info.context.request.registry.settings['storage'])

        storage_temporary = storage['temporary']

        host = storage_temporary['host']
        bucket = storage_temporary['bucket']

        minio_client = (

            minio.Minio(
                host,
                access_key = storage_temporary['access_key'],
                secret_key = storage_temporary['secret_key'],
                secure = True))

        current_time = time.time()

        object_name = (

            storage_temporary['prefix'] +

            '/'.join((
                'valency_verb_cases',
                '{:.6f}'.format(current_time),
                'valency_verb_cases.xlsx')))

        object_length = (
            workbook_stream.tell())

        workbook_stream.seek(0)

        (etag, version_id) = (

            minio_client.put_object(
                bucket,
                object_name,
                workbook_stream,
                object_length))

        url = (

            '/'.join((
                'https:/',
                host,
                bucket,
                object_name)))

        log.debug(
            '\nobject_name:\n{}'
            '\netag:\n{}'
            '\nversion_id:\n{}'
            '\nurl:\n{}'.format(
                object_name,
                etag,
                version_id,
                url))

        return url

    @staticmethod
    def mutate(
        root,
        info,
        perspective_id = None,
        language_arg_list = None,
        debug_flag = False):

        try:

            # Checking user and arguments.

            client_id = info.context.get('client_id')
            client = DBSession.query(Client).filter_by(id = client_id).first()

            if not client:

                return (

                    ResponseError(
                        message = 'Only registered users can get valency verb cases info.'))

            if debug_flag and client.user_id != 1:

                return (

                    ResponseError(
                        message = 'Only administrator can use debug mode.'))

            if perspective_id is None and language_arg_list is None:

                return (

                    ResponseError(
                        message = 'Please specify either perspective or a set of languages.'))

            locale_id = info.context.get('locale_id') or 2

            # For a single corpus.

            if perspective_id is not None:

                verb_data_dict = (

                    collections.defaultdict(
                        lambda: collections.defaultdict(
                            lambda: collections.defaultdict(
                                lambda: collections.defaultdict(dict)))))

                try:

                    ValencyVerbCases.process_perspective(
                        perspective_id,
                        locale_id,
                        verb_data_dict,
                        None,
                        False,
                        debug_flag)

                except ResponseError as error:

                    return error

                if debug_flag:

                    log.debug(
                        '\nverb_data_dict:\n' +
                        pprint.pformat(
                            verb_data_dict, width = 192))

                verb_data_list = []

                for verb_xlat, language_dict in sorted(verb_data_dict.items()):

                    case_verb_sentence_dict = language_dict[None]

                    case_verb_sentence_list = (
                        ValencyVerbCases.get_case_verb_sentence_list(case_verb_sentence_dict))

                    verb_list = (

                        sorted(
                            set.union(
                                *(set(verb_list)
                                    for _, verb_list, _ in case_verb_sentence_list))))

                    verb_data_list.append(
                        (verb_xlat, case_verb_sentence_list, verb_list))

                if debug_flag:

                    log.debug(
                        '\nverb_data_list:\n' + 
                        pprint.pformat(
                            verb_data_list, width = 192))

                # Saving as XLSX file.

                workbook_stream = (
                    io.BytesIO())

                workbook = (
                    xlsxwriter.Workbook(
                        workbook_stream, {'in_memory': True}))

                worksheet = (
                    workbook.add_worksheet(
                        utils.sanitize_worksheet_name('Verb cases')))

                worksheet.set_column(
                    0, 0, 32)

                worksheet.set_column(
                    2, 2, 16)

                row_count = 0

                for verb_xlat, case_verb_sentence_list, _ in verb_data_list:

                    case, verb_list, _ = case_verb_sentence_list[0]

                    worksheet.write_row(
                        row_count, 0, [verb_xlat, case, ', '.join(verb_list)])

                    row_count += 1

                    for case, verb_list, _ in case_verb_sentence_list[1:]:

                        worksheet.write_row(
                            row_count, 1, [case, ', '.join(verb_list)])

                        row_count += 1

                # With sentences.

                worksheet = (

                    workbook.add_worksheet(
                        utils.sanitize_worksheet_name('Verb cases and sentences')))

                worksheet.set_column(
                    0, 0, 32)

                worksheet.set_column(
                    2, 2, 16)

                worksheet.set_column(
                    3, 3, 64)

                row_count = 0

                for verb_xlat, case_verb_sentence_list, _ in verb_data_list:

                    for case, _, verb_sentence_list in case_verb_sentence_list:

                        for verb, sentence_list in verb_sentence_list:

                            for sentence in sentence_list:

                                worksheet.write_row(
                                    row_count, 0, [verb_xlat, case, verb, sentence])

                                if debug_flag:

                                    log.debug([verb_xlat, case, verb, sentence])

                                row_count += 1

                                verb_xlat = ''
                                case = ''
                                verb = ''

                workbook.close()

                # Saving XLSX file, returning data.

                xlsx_url = (

                    ValencyVerbCases.save_xlsx_file(
                        info,
                        workbook_stream,
                        debug_flag))

                return (

                    ValencyVerbCases(
                        triumph = True,
                        verb_data_list = verb_data_list,
                        xlsx_url = xlsx_url))

            # For corpora in a set of languages.

            if language_arg_list is not None:

                log.debug(
                    '\nlanguage_arg_list:\n' +
                    pprint.pformat(
                        language_arg_list, width = 192))

                # Processing all languages with their perspective sets.

                verb_data_dict = (

                    collections.defaultdict(
                        lambda: collections.defaultdict(
                            lambda: collections.defaultdict(
                                lambda: collections.defaultdict(dict)))))

                for language_id, perspective_id_list, by_perspective_flag in language_arg_list:

                    language = (

                        DBSession

                            .query(dbLanguage)

                            .filter_by(
                                client_id = language_id[0],
                                object_id = language_id[1])

                            .first())

                    if not language:

                        return (

                            ResponseError(
                                message = 'No language {}/{} in the system.'.format(*language_id)))

                    language_name = (

                        language.get_translation(
                            locale_id))

                    if debug_flag:

                        log.debug(
                            f'\nlanguage {language.id} {repr(language_name)}')
                    
                    for perspective_id in perspective_id_list:

                        try:

                            ValencyVerbCases.process_perspective(
                                perspective_id,
                                locale_id,
                                verb_data_dict,
                                language_name,
                                by_perspective_flag,
                                debug_flag)

                        except ResponseError as error:

                            return error

                # Processing gathered verb case data, sorting by verb translations, languaged, cases and
                # verb lexemes.

                verb_data_list = []

                for verb_xlat, language_dict in sorted(verb_data_dict.items()):

                    language_list = []

                    for language_str, case_verb_sentence_dict in language_dict.items():

                        case_verb_sentence_list = (
                            ValencyVerbCases.get_case_verb_sentence_list(case_verb_sentence_dict))

                        language_list.append(
                            (language_str, case_verb_sentence_list))

                    verb_list = (

                        sorted(
                            set.union(
                                *(set(verb_list)
                                    for _, case_verb_sentence_list in language_list
                                    for _, verb_list, _ in case_verb_sentence_list))))

                    verb_data_list.append(
                        (verb_xlat, language_list, verb_list))

                if debug_flag:

                    log.debug(
                        '\nverb_data_list:\n' + 
                        pprint.pformat(
                            verb_data_list, width = 192))

                # Saving as XLSX file.

                workbook_stream = (
                    io.BytesIO())

                workbook = (
                    xlsxwriter.Workbook(
                        workbook_stream, {'in_memory': True}))

                worksheet = (
                    workbook.add_worksheet(
                        utils.sanitize_worksheet_name('Verb cases')))

                worksheet.set_column(
                    0, 0, 32)

                worksheet.set_column(
                    1, 1, 12)

                worksheet.set_column(
                    3, 3, 16)

                row_count = 0

                for verb_xlat, language_list, _ in verb_data_list:

                    for language_str, case_verb_sentence_list in language_list:

                        for case, verb_list, _ in case_verb_sentence_list:

                            worksheet.write_row(
                                row_count, 0, [verb_xlat, language_str, case, ', '.join(verb_list)])

                            row_count += 1

                            verb_xlat = ''
                            language_str = ''
                            case = ''

                # With sentences.

                worksheet = (

                    workbook.add_worksheet(
                        utils.sanitize_worksheet_name('Verb cases and sentences')))

                worksheet.set_column(
                    0, 0, 32)

                worksheet.set_column(
                    1, 1, 12)

                worksheet.set_column(
                    3, 3, 16)

                worksheet.set_column(
                    4, 4, 64)

                row_count = 0

                for verb_xlat, language_list, _ in verb_data_list:

                    for language_str, case_verb_sentence_list in language_list:

                        for case, _, verb_sentence_list in case_verb_sentence_list:

                            for verb, sentence_list in verb_sentence_list:

                                for sentence in sentence_list:

                                    worksheet.write_row(
                                        row_count, 0, [verb_xlat, language_str, case, verb, sentence])

                                    if debug_flag:

                                        log.debug([verb_xlat, language_str, case, verb, sentence])

                                    row_count += 1

                                    verb_xlat = ''
                                    language_str = ''
                                    case = ''
                                    verb = ''

                workbook.close()

                # Saving XLSX file, returning data.

                xlsx_url = (

                    ValencyVerbCases.save_xlsx_file(
                        info,
                        workbook_stream,
                        debug_flag))

                return (

                    ValencyVerbCases(
                        triumph = True,
                        verb_data_list = verb_data_list,
                        xlsx_url = xlsx_url))

        except Exception as exception:

            traceback_string = (

                ''.join(
                    traceback.format_exception(
                        exception, exception, exception.__traceback__))[:-1])

            log.warning('valency_verb_cases: exception')
            log.warning(traceback_string)

            return (

                ResponseError(
                    'Exception:\n' + traceback_string))


class BidirectionalLinks(graphene.Mutation):
    """
    Ensures that links between Lexical Entries and Paradigms perspectives of each specified dictionary are
    bidirectional.

    curl 'http://localhost:6543/graphql' \
      -H 'Content-Type: application/json' \
      -H 'Cookie: auth_tkt=_tkn_!userid_type:int' \
      --data-raw '{ \
        "operationName": "bidirectionalLinks", \
        "variables":{"dictionaryIdList":[[4755,333300]]}, \
        "query": \
          "mutation bidirectionalLinks($dictionaryIdList: [LingvodocID]!) { \
            bidirectional_links(dictionary_id_list: $dictionaryIdList, debug_flag: true) { \
              triumph }}"}'

    #! Remove any newline after --data-raw to call this from Ubuntu.
    #! Change _tkn_ to actual token for administrator.
    #! Be careful, used token may be not actual after database restart.
    """

    class Arguments:

        dictionary_id_list = graphene.List(LingvodocID, required = True)
        debug_flag = graphene.Boolean()

    triumph = graphene.Boolean()

    @staticmethod
    def mutate(
        root,
        info,
        dictionary_id_list,
        debug_flag = False):

        try:
            client_id = info.context.get('client_id')
            log.debug(f"client_id: {client_id}")

            client = DBSession.query(Client).filter_by(id = client_id).first()

            if not client or client.user_id != 1:
                return ResponseError('Only administrator can fix links using request.')

            backref_fid = (
                get_id_to_field_dict()['Backref'])

            # Processing each dictionary.

            for dictionary_id in dictionary_id_list:

                log.debug(
                    f'\ndictionary_id: {dictionary_id}')

                dictionary = (

                    DBSession
                        .query(dbDictionary)
                        .get((dictionary_id[1], dictionary_id[0])))

                if dictionary is None:
                    return (
                        ResponseError(
                            f'Dictionary {dictionary_id} does not exist.'))

                if dictionary.marked_for_deletion:
                    return (
                        ResponseError(
                            f'Dictionary {dictionary_id} is deleted.'))

                if debug_flag:

                    log.debug(
                        f'\ndictionary.get_translation():\n{repr(dictionary.get_translation())}')

                # Processing all linked lexical entries of all perspectives of the dictionary, intentionally
                # counting even deleted links.

                link_list = (

                    DBSession

                        .query(
                            dbPerspective.client_id,
                            dbPerspective.object_id,
                            dbLexicalEntry.client_id,
                            dbLexicalEntry.object_id,
                            dbEntity.client_id,
                            dbEntity.object_id,
                            dbEntity.link_client_id,
                            dbEntity.link_object_id,
                            dbEntity.marked_for_deletion,
                            dbPublishingEntity.published,
                            dbPublishingEntity.accepted)

                        .filter(
                            dbPerspective.parent_client_id == dictionary_id[0],
                            dbPerspective.parent_object_id == dictionary_id[1],
                            dbPerspective.marked_for_deletion == False,
                            dbLexicalEntry.parent_client_id == dbPerspective.client_id,
                            dbLexicalEntry.parent_object_id == dbPerspective.object_id,
                            dbLexicalEntry.marked_for_deletion == False,
                            dbEntity.parent_client_id == dbLexicalEntry.client_id,
                            dbEntity.parent_object_id == dbLexicalEntry.object_id,
                            dbEntity.field_client_id == backref_fid[0],
                            dbEntity.field_object_id == backref_fid[1],
                            dbPublishingEntity.client_id == dbEntity.client_id,
                            dbPublishingEntity.object_id == dbEntity.object_id)

                        .all())

                log.debug(
                    f'\nlen(link_list): {len(link_list)}')

                link_list = [

                    ((p_cid, p_oid),
                        (l_cid, l_oid),
                        (e_cid, e_oid),
                        (k_cid, k_oid),
                        (d, p, a))

                    for p_cid, p_oid, l_cid, l_oid, e_cid, e_oid, k_cid, k_oid, d, p, a in link_list]

                # Compiling info of existing links.

                link_dict = {
                    (entry_id, link_id): perspective_id
                    for perspective_id, entry_id, _, link_id, _ in link_list}

                entry_list = (

                    DBSession

                        .query(
                            dbPerspective.client_id,
                            dbPerspective.object_id,
                            dbLexicalEntry.client_id,
                            dbLexicalEntry.object_id)

                        .filter(
                            dbPerspective.parent_client_id == dictionary_id[0],
                            dbPerspective.parent_object_id == dictionary_id[1],
                            dbPerspective.marked_for_deletion == False,
                            dbLexicalEntry.parent_client_id == dbPerspective.client_id,
                            dbLexicalEntry.parent_object_id == dbPerspective.object_id,
                            dbLexicalEntry.marked_for_deletion == False)

                        .all())

                entry_dict = {
                    (l_cid, l_oid): (p_cid, p_oid)
                    for p_cid, p_oid, l_cid, l_oid in entry_list}

                log.debug(
                    f'\nlen(link_dict): {len(link_dict)}'
                    f'\nlen(entry_dict): {len(entry_dict)}')

                if debug_flag:

                    perspective_dict = {}

                    perspective_list = (

                        DBSession

                            .query(
                                dbPerspective)

                            .filter(
                                dbPerspective.parent_client_id == dictionary_id[0],
                                dbPerspective.parent_object_id == dictionary_id[1],
                                dbPerspective.marked_for_deletion == False)

                            .all())

                    perspective_dict = {
                        perspective.id: perspective
                        for perspective in perspective_list}

                    translation_dict = {
                        perspective.id: perspective.get_translation()
                        for perspective in perspective_list}

                fixed = 0
                for (from_id, to_id) in link_dict:

                    if (to_id, from_id) in link_dict:
                        continue

                    # if the linked lexical entry is deleted
                    if  to_id not in entry_dict:
                        continue

                    if debug_flag:

                        log.debug(f'\nfrom perspective: {translation_dict[entry_dict[from_id]]}'
                                  f'\nto perspective: {translation_dict[entry_dict[to_id]]}')

                    create_entity(id=[entry_dict[to_id][0], None],
                                  parent_id=to_id,
                                  field_id=backref_fid,
                                  link_id=from_id)
                    fixed += 1

                print(f'\nTotal fixed links: {fixed}')

        except Exception as exception:

            traceback_string = (

                ''.join(
                    traceback.format_exception(
                        exception, exception, exception.__traceback__))[:-1])

            log.warning('bidirectional_links: exception')
            log.warning(traceback_string)

            return (

                ResponseError(
                    'Exception:\n' + traceback_string))


class MyMutations(graphene.ObjectType):
    """
    Mutation classes.
    It may be replaced by
    create_field = gql_field.CreateField.Field()
    for more beautiful imports
    """
    convert_starling = starling_converter.GqlStarling.Field()
    convert_dialeqt = ConvertDictionary.Field()
    convert_corpus = ConvertFiveTiers.Field()
    create_field = CreateField.Field()
    # update_field = UpdateField.Field()
    # delete_field = DeleteField.Field()
    copy_sound_markup_fields = CopySoundMarkupFields.Field()
    copy_single_field = CopySingleField.Field()
    create_entity = CreateEntity.Field()
    update_entity = UpdateEntity.Field()
    delete_entity = DeleteEntity.Field()
    update_entity_content = UpdateEntityContent.Field()
    bulk_update_entity_content = BulkUpdateEntityContent.Field()
    approve_all_for_user = ApproveAllForUser.Field()
    bulk_create_entity = BulkCreateEntity.Field()
    create_user = CreateUser.Field()
    update_user = UpdateUser.Field()
    activate_deactivate_user = ActivateDeactivateUser.Field()
    create_language = CreateLanguage.Field()
    update_language = UpdateLanguage.Field()
    update_language_atom = UpdateLanguageAtom.Field()
    move_language = MoveLanguage.Field()
    delete_language = DeleteLanguage.Field()
    create_dictionary = CreateDictionary.Field()
    update_dictionary = UpdateDictionary.Field()
    update_dictionary_status = UpdateDictionaryStatus.Field()
    update_dictionary_atom = UpdateDictionaryAtom.Field()
    add_dictionary_roles = AddDictionaryRoles.Field()
    delete_dictionary_roles = DeleteDictionaryRoles.Field()
    delete_dictionary = DeleteDictionary.Field()
    undelete_dictionary = UndeleteDictionary.Field()
    create_organization = CreateOrganization.Field()
    update_organization = UpdateOrganization.Field()
    delete_organization = DeleteOrganization.Field()
    create_translationatom = CreateTranslationAtom.Field()
    update_translationatom = UpdateTranslationAtom.Field()
    delete_translationatom = DeleteTranslationAtom.Field()
    create_translationgist = CreateTranslationGist.Field()
    delete_translationgist = DeleteTranslationGist.Field()
    create_lexicalentry = CreateLexicalEntry.Field()
    delete_lexicalentry = DeleteLexicalEntry.Field()
    bulk_delete_lexicalentry = BulkDeleteLexicalEntry.Field()
    bulk_undelete_lexicalentry = BulkUndeleteLexicalEntry.Field()
    bulk_create_lexicalentry = BulkCreateLexicalEntry.Field()
    join_lexical_entry_group = ConnectLexicalEntries.Field()
    leave_lexical_entry_group = DeleteGroupingTags.Field()
    create_perspective = CreateDictionaryPerspective.Field()
    update_perspective = UpdateDictionaryPerspective.Field()
    update_perspective_status = UpdatePerspectiveStatus.Field()
    update_perspective_atom = UpdatePerspectiveAtom.Field()
    add_perspective_roles = AddPerspectiveRoles.Field()
    delete_perspective_roles = DeletePerspectiveRoles.Field()
    delete_perspective = DeleteDictionaryPerspective.Field()
    undelete_perspective = UndeleteDictionaryPerspective.Field()
    create_column = CreateColumn.Field()
    update_column = UpdateColumn.Field()
    delete_column = DeleteColumn.Field()
    create_grant = CreateGrant.Field()
    update_grant = UpdateGrant.Field()
    # delete_grant = DeleteGrant.Field()
    create_userblob = CreateUserBlob.Field()
    delete_userblob = DeleteUserBlob.Field()
    create_grant_permission = CreateGrantPermission.Field()
    add_dictionary_to_grant = AddDictionaryToGrant.Field()
    administrate_org = AdministrateOrg.Field()
    participate_org = ParticipateOrg.Field()
    add_dictionary_to_organization = AddDictionaryToOrganization.Field()
    accept_userrequest = AcceptUserRequest.Field()
    #delete_userrequest = DeleteUserRequest.Field()
    download_dictionary = DownloadDictionary.Field()
    save_dictionary = SaveDictionary.Field()
    save_all_dictionaries = SaveAllDictionaries.Field()
    download_dictionaries = DownloadDictionaries.Field()
    synchronize = Synchronize.Field()
    delete_task = DeleteTask.Field()
    starling_etymology = StarlingEtymology.Field()
    phonemic_analysis = PhonemicAnalysis.Field()
    cognate_analysis = CognateAnalysis.Field()
    swadesh_analysis = SwadeshAnalysis.Field()
    morph_cognate_analysis = MorphCognateAnalysis.Field()
    phonology = Phonology.Field()
    phonological_statistical_distance = PhonologicalStatisticalDistance.Field()
    sound_and_markup = SoundAndMarkup.Field()
    merge_bulk = MergeBulk.Field()
    move_column = MoveColumn.Field()
    add_roles_bulk = AddRolesBulk.Field()
    create_basegroup = CreateBasegroup.Field()
    add_user_to_basegroup = AddUserToBasegroup.Field()
    execute_parser = ExecuteParser.Field()
    delete_parser_result = DeleteParserResult.Field()
    update_parser_result = UpdateParserResult.Field()
    xlsx_bulk_disconnect = XlsxBulkDisconnect.Field()
    new_unstructured_data = NewUnstructuredData.Field()
    docx2eaf = Docx2Eaf.Field()
    valency = Valency.Field()
    create_valency_data = CreateValencyData.Field()
    save_valency_data = SaveValencyData.Field()
    set_valency_annotation = SetValencyAnnotation.Field()
    valency_verb_cases = ValencyVerbCases.Field()
    bidirectional_links = BidirectionalLinks.Field()

schema = graphene.Schema(query=Query, auto_camelcase=False, mutation=MyMutations)

class Context(dict):
    """
    Context for graphene query execution, works as dict for arbitrary key-value associations and supports
    ACL-based permission checking.
    """

    def __init__(self, context_dict):
        """
        Initializes query execution context by initializing context key-value dictionary and, hopefully,
        getting client id and request data.
        """

        dict.__init__(self, context_dict)

        self.client_id = context_dict.get('client_id')
        self.request = context_dict.get('request')
        self.locale_id = context_dict.get('locale_id')
        self.headers = context_dict.get('headers')
        self.cookies = context_dict.get('cookies')

        self.cache = {}

    def acl_check_if(self, action, subject, subject_id):
        """
        Checks if the client has permission to perform given action on a specified subject via ACL.
        """
        if type(subject_id) is list:
            subject_id = tuple(subject_id)

        if (action, subject, subject_id) in self.cache:
            return self.cache[(action, subject, subject_id)]

        result = acl.check_direct(self.client_id, self.request, action, subject, subject_id)
        self.cache[(action, subject, subject_id)] = result

        return result

    def acl_check(self, action, subject, subject_id):
        """
        Checks if the client has permission to perform given action on a specified subject via ACL, raises
        permission exception otherwise.
        """

        if not self.acl_check_if(action, subject, subject_id):
            raise PermissionException(self.client_id, action, subject, subject_id)

    def acl_check_with_id_if(self, action, subject, args):
        """
        Checks via ACL if the client has permission to perform given action on a specified subject, with
        subject identifier extracted from query execution arguments.
        """

        return self.acl_check_if(action, subject, args.get('id'))


if __name__ == '__main__':

    # Cognate analysis debugging script.

    import pdb
    import re
    import sys

    dictionary_count, line_count = map(
        int, re.findall(r'\d+', sys.argv[1])[:2])

    print(dictionary_count, line_count)

    with open(sys.argv[1], 'rb') as input_file:
        input = input_file.read().decode('utf-16')

    input_buffer = ctypes.create_unicode_buffer(input)

    output_buffer_size = cognate_analysis_f(
        None, dictionary_count, line_count, None, 1)

    print(output_buffer_size)

    output_buffer = ctypes.create_unicode_buffer(output_buffer_size + 256)

    result = cognate_analysis_f(
        input_buffer, dictionary_count, line_count, output_buffer, 1)

    print(result)
